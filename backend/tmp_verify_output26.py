"""Verify the output of sequence 26 is correct - check for data integrity."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl

output_path = r"g:\FCCS\backend\uploads\reports\26\outputs\ICM_Output_26.xlsx"
if not os.path.exists(output_path):
    print(f"Output not found at {output_path}")
    sys.exit(1)

wb = openpyxl.load_workbook(output_path, data_only=True)
ws = wb.active

print(f"Sheet: '{ws.title}'")
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols")

# Check headers at row 32
headers_r32 = []
for c in range(1, ws.max_column + 1):
    v = str(ws.cell(32, c).value or '').strip()
    headers_r32.append(v)

print(f"\nTotal header columns: {len(headers_r32)}")

# Check section labels at row 29
headers_r29 = []
for c in range(1, ws.max_column + 1):
    v = str(ws.cell(29, c).value or '').strip()
    if v:
        headers_r29.append((c, v))

print(f"\nSection labels (row 29):")
for c, v in headers_r29:
    print(f"  Col {c}: {v}")

# Count data rows
data_count = 0
for r in range(33, ws.max_row + 1):
    e = ws.cell(r, 1).value
    p = ws.cell(r, 2).value
    if e or p:
        data_count += 1

print(f"\nData rows (from row 33): {data_count}")

# Check entity 013011 rows in Parent block
print(f"\n=== Entity 013011 in Parent block ===")
# Find Parent block columns
parent_start = None
contrib_start = None
for c, v in headers_r29:
    if v == "Parent Input":
        parent_start = c
    elif v == "Contribution Input":
        contrib_start = c

if parent_start:
    # Find account 430015 in Parent block
    for c in range(parent_start, (contrib_start or ws.max_column) + 1):
        h = str(ws.cell(32, c).value or '')
        if '430015' in h:
            print(f"\n  430015 column found at col {c}: '{h[:60]}'")
            for r in range(33, min(ws.max_row + 1, 33 + 20)):
                v = ws.cell(r, c).value
                if v is not None:
                    e = ws.cell(r, 1).value
                    p = ws.cell(r, 2).value
                    print(f"    Row {r}: Entity='{e}', Partner='{p}', Value={v}")

# Check first 5 rows of data
print(f"\n=== First 5 data rows ===")
for r in range(33, min(38, ws.max_row + 1)):
    e = ws.cell(r, 1).value
    p = ws.cell(r, 2).value
    print(f"  Row {r}: Entity='{e}', Partner='{p}'")

print("\nDone.")
