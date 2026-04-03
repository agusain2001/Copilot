"""Verify strict 3-way matching: Entity + ICP + Account must all match."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from app.ic_processor import (
    read_journal_report, read_icm_data, detect_icm_header_row,
    read_icm_account_columns, read_icm_headers, parse_report_inputs,
    match_journal_to_icm
)

icm_path = r"g:\FCCS\New folder (3)\Intercompany Report_1150_Intercompany Report.xlsx"
# Journal Report (1) = Parent (15135 bytes = same as old Journal Report (5))
# Journal Report (2) = Contribution (21168 bytes = same as old Journal Report (6))
parent_journal = r"g:\FCCS\New folder (3)\Journal Report (1).xlsx"
contrib_journal = r"g:\FCCS\New folder (3)\Journal Report (2).xlsx"
plug_journal = r"g:\FCCS\Update files\Journal Report (4).xlsx"
report_inputs = r"g:\FCCS\AI\report_inputs.xlsx"

plug_mapping = parse_report_inputs(report_inputs)
print(f"Plug mapping: {plug_mapping}")

# Read ICM source
wb_icm = openpyxl.load_workbook(icm_path, data_only=True)
candidates = []
for sheet in wb_icm.worksheets:
    hdr_row, data_start = detect_icm_header_row(sheet)
    c1 = str(sheet.cell(hdr_row, 1).value or "").strip().lower()
    c2 = str(sheet.cell(hdr_row, 2).value or "").strip().lower()
    if c1 == "entity" and c2 == "partner":
        candidates.append((sheet, hdr_row, data_start))

if candidates:
    ws_icm, icm_hdr_row, icm_data_start = min(candidates, key=lambda t: t[0].max_column)
else:
    ws_icm = wb_icm.active
    icm_hdr_row, icm_data_start = detect_icm_header_row(ws_icm)

ent_cols, par_cols = read_icm_account_columns(ws_icm, header_row=icm_hdr_row)
icm_header_map = read_icm_headers(ws_icm, header_row=icm_hdr_row)
data_rows = read_icm_data(ws_icm, data_start=icm_data_start)

print(f"\n=== ICM SOURCE ===")
print(f"Entity-side accounts: {[c[0] for c in ent_cols]}")
print(f"Partner-side accounts: {[c[0] for c in par_cols]}")
print(f"430015 in entity cols: {'430015' in [c[0] for c in ent_cols]}")
print(f"430015 in partner cols: {'430015' in [c[0] for c in par_cols]}")
print(f"Total ICM data rows: {len(data_rows)}")

# Check ICM rows with entity 013011
print(f"\n=== ICM ROWS: entity 013011 ===")
for r in data_rows:
    if r["entity_code"] == "013011":
        print(f"  Entity={r['entity']}, Partner={r['partner']}, eCode={r['entity_code']}, pCode={r['partner_code']}")

# Read journals and show entries with account 430015
for jname, jpath, is_plug in [
    ("PARENT (1)", parent_journal, False),
    ("CONTRIBUTION (2)", contrib_journal, False),
    ("PLUG (4)", plug_journal, True),
]:
    print(f"\n=== {jname} Journal - Entries with account 430015 ===")
    jmap = plug_mapping if is_plug else None
    p_pri, p_fb = read_journal_report(jpath, jmap)
    found = False
    for (e, icp, acct), lines in p_pri.items():
        if acct == "430015":
            found = True
            for l in lines:
                print(f"  PRIMARY: Entity={e}, ICP={icp}, Acct={acct}, Debit={l['debit']}, Credit={l['credit']}")
    for (e, icp, acct), lines in p_fb.items():
        if acct == "430015":
            found = True
            for l in lines:
                print(f"  FALLBACK: Entity={e}, ICP={icp}, Acct={acct}, Debit={l['debit']}, Credit={l['credit']}")
    if not found:
        print("  (no entries with account 430015)")

# Now check the generated output
print(f"\n=== OUTPUT FILE ===")
output_path = r"g:\FCCS\New folder (3)\Fixed_ICM_Output.xlsx"
if os.path.exists(output_path):
    out_wb = openpyxl.load_workbook(output_path, data_only=True)
    out_ws = out_wb.active
    
    # Find 430015 columns
    for c in range(1, out_ws.max_column + 1):
        hdr = str(out_ws.cell(32, c).value or "")
        if "430015" in hdr:
            print(f"\n  Column {c}: '{hdr}'")
            for r in range(33, min(out_ws.max_row + 1, 33 + len(data_rows) + 50)):
                v = out_ws.cell(r, c).value
                if v is not None:
                    ent = out_ws.cell(r, 1).value
                    prt = out_ws.cell(r, 2).value
                    print(f"    Row {r}: Entity='{ent}', Partner='{prt}', Value={v}")
else:
    print(f"  Output file not found at {output_path}")
