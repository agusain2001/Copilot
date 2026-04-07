import openpyxl
wb = openpyxl.load_workbook(r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_FINAL.xlsx", data_only=True)
ws = wb.active
out = open("tmp_label_check.txt", "w", encoding="utf-8")
out.write("=== ALL E117100 rows ===\n")
for r in range(33, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or "").strip()
    prt = str(ws.cell(r, 2).value or "").strip()
    if "117100" in ent or "E117100" in ent:
        vals = []
        for c in range(3, ws.max_column + 1):
            v = ws.cell(r, c).value
            hdr = str(ws.cell(32, c).value or f"Col{c}")[:40]
            if v is not None and v != 0:
                vals.append(f"{hdr}={v}")
        out.write(f"Row {r}: Entity={ent[:55]} | Partner={prt[:55]}\n")
        if vals:
            out.write(f"  Values: {vals}\n")
        else:
            out.write(f"  [NO VALUES]\n")

out.write("\n\n=== ALL E101000 rows ===\n")
for r in range(33, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or "").strip()
    prt = str(ws.cell(r, 2).value or "").strip()
    if "101000" in ent or "E101000" in ent:
        out.write(f"Row {r}: Entity={ent[:55]} | Partner={prt[:55]}\n")

out.close()
print("Done")
