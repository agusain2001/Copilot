"""Quick diagnostic: compare old vs new output to show what was added."""
import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl

# Read headers and data from the new output
out_path = r'G:\FCCS\backend\uploads\reports\22\outputs\ICM_Output_22_v2.xlsx'
wb = openpyxl.load_workbook(out_path, data_only=True)
ws = wb.active

# Header row 32
print("=" * 80)
print("OUTPUT FILE ANALYSIS")
print("=" * 80)

# Count data rows
data_start = 33
row_count = 0
for r in range(data_start, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or "").strip()
    par = str(ws.cell(r, 2).value or "").strip()
    if ent or par:
        row_count += 1

print(f"\nTotal data rows: {row_count}")
print(f"Max column used: {ws.max_column}")

# Show header row
print(f"\nHeader row (row 32) columns:")
for c in range(1, ws.max_column + 1):
    v = ws.cell(32, c).value
    if v:
        print(f"  Col {c}: {str(v)[:60]}")

# Show all entity/partner pairs
print(f"\nAll Entity/Partner pairs in output:")
for r in range(data_start, data_start + row_count):
    ent = str(ws.cell(r, 1).value or "").strip()
    par = str(ws.cell(r, 2).value or "").strip()
    print(f"  Row {r}: {ent[:30]:30s}  |  {par[:30]}")

# Now check journals for coverage
print("\n" + "=" * 80)
print("JOURNAL ANALYSIS - finding all valid (entity, icp, account) combos")
print("=" * 80)

from app.ic_processor import (
    read_journal_report, extract_entity_code_journal,
    extract_icp_code, extract_account_code, is_detail_row,
    J_ENTITY, J_ACCT, J_ICP, JOURNAL_DATA_START
)

journals = {
    "Parent":       r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (1).xlsx',
    "Contribution": r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (2).xlsx',
    "Plug":         r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (4).xlsx',
}

for jname, jpath in journals.items():
    wb_j = openpyxl.load_workbook(jpath, data_only=True)
    ws_j = wb_j.active
    
    pairs = set()
    accounts = set()
    combos = set()
    skipped = 0
    
    for row in ws_j.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws_j.max_row):
        vals = [cell.value for cell in row]
        if str(vals[0] or "").strip() == "Grand Total":
            break
        if not is_detail_row(vals):
            continue
        
        entity_code = extract_entity_code_journal(str(vals[J_ENTITY - 1] or "").strip())
        icp_code = extract_icp_code(str(vals[J_ICP - 1] or "").strip())
        acct_code = extract_account_code(str(vals[J_ACCT - 1] or "").strip())
        
        if not entity_code or not icp_code or not acct_code:
            skipped += 1
            continue
        
        pairs.add((entity_code, icp_code))
        accounts.add(acct_code)
        combos.add((entity_code, icp_code, acct_code))
    
    print(f"\n{jname} Journal:")
    print(f"  Valid entity/ICP pairs: {len(pairs)}")
    print(f"  Unique account codes: {sorted(accounts)}")
    print(f"  Unique (ent, icp, acct) combos: {len(combos)}")
    print(f"  Skipped (missing field): {skipped}")
