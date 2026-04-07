"""
Trace exactly how journal keys map to ICM rows for entity-side and partner-side.
Compare what the code expects vs what the journal files actually have.
"""
import openpyxl
import re
from collections import defaultdict

def extract_entity_code_journal(raw):
    raw = str(raw or "").strip()
    m = re.match(r"(E?\d+\w*)", raw)
    return m.group(1) if m else ""

def normalize_entity_code(code):
    code = str(code or "").strip()
    if code.startswith("E") and len(code) > 1 and code[1:].replace("_", "").isdigit():
        return code[1:]
    return code

def extract_icp_code(raw):
    m = re.match(r"(ICP_\w+)", str(raw or "").strip())
    return m.group(1) if m else ""

def extract_6digit_from_icp(icp_code):
    s = str(icp_code or "").strip()
    return s[4:] if s.startswith("ICP_") else s

# Read Parent journal
wb = openpyxl.load_workbook(r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx", data_only=True)
ws = wb.active
headers = [str(ws.cell(30, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
col_map = {h: i for i, h in enumerate(headers) if h}
eidx = col_map.get('entity', 2)
aidx = col_map.get('account', 3)
icpidx = col_map.get('intercompany', 4)

print("=== Parent Journal ALL unique (entity, icp) keys ===")
journal_keys = set()
for r in range(31, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    label = str(vals[0] or "").strip()
    if label == "Grand Total": break
    try:
        ent_raw = str(vals[eidx] or "").strip()
        icp_raw = str(vals[icpidx] or "").strip()
    except IndexError:
        continue
    ent_code_raw = extract_entity_code_journal(ent_raw)
    ent_code = normalize_entity_code(ent_code_raw)
    icp_code = extract_icp_code(icp_raw)
    if ent_code and icp_code:
        journal_keys.add((ent_code, icp_code))

for k in sorted(journal_keys):
    print(f"  {k}")

print(f"\nTotal unique journal keys: {len(journal_keys)}")

# Read ICM source to see what entity/partner pairs exist
wb_icm = openpyxl.load_workbook(r"g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx", data_only=True)

print("\n=== ICM Source Entity/Partner pairs ===")
icm_pairs = []
for sheet in wb_icm.worksheets:
    for r in range(1, min(50, sheet.max_row + 1)):
        v1 = str(sheet.cell(r, 1).value or "").strip().lower()
        v2 = str(sheet.cell(r, 2).value or "").strip().lower()
        if v1 == "entity" and v2 == "partner":
            hdr_row = r
            for dr in range(r+1, sheet.max_row + 1):
                ent_raw = str(sheet.cell(dr, 1).value or "").strip()
                prt_raw = str(sheet.cell(dr, 2).value or "").strip()
                if not ent_raw and not prt_raw:
                    continue
                # Extract standard entity code
                m = re.match(r"(\d{6})", ent_raw)
                ent_code = m.group(1) if m else ""
                prt_code = extract_icp_code(prt_raw)
                if ent_code and prt_code:
                    icm_pairs.append((ent_code, prt_code))
            break

print(f"Total ICM pairs: {len(icm_pairs)}")
for p in sorted(set(icm_pairs)):
    print(f"  {p}")

# Now trace the matching logic
print("\n=== MATCHING TRACE ===")
print("For each ICM row (ent, prt), show:")
print("  [Entity-side] direct key: (ent, prt, acct) → found in journal?")
print("  [Partner-side] reverse key: (prt_digit, ICP_{ent}, acct) → found in journal?")
print()

for ent, prt in sorted(set(icm_pairs)):
    # What the code does for entity-side: uses (ent, prt, acct_code)
    entity_key_base = (ent, prt)
    
    # What the code does for partner-side:
    prt_entity = normalize_entity_code(extract_6digit_from_icp(prt))
    reverse_icp_plain = f"ICP_{ent}"
    reverse_icp_e = f"ICP_E{ent}"
    partner_key_base = (prt_entity, reverse_icp_plain)
    partner_key_e_base = (prt_entity, reverse_icp_e)
    
    # Check matches
    entity_found = entity_key_base in journal_keys
    partner_found = partner_key_base in journal_keys or partner_key_e_base in journal_keys
    
    status_e = "YES" if entity_found else "NO"
    status_p = "YES" if partner_found else "NO"
    
    which_p = ""
    if partner_key_base in journal_keys:
        which_p = f" via {partner_key_base}"
    elif partner_key_e_base in journal_keys:
        which_p = f" via {partner_key_e_base}"
    
    if not entity_found or not partner_found:
        print(f"ICM row ({ent}, {prt}):")
        print(f"  Entity-side  direct:  ({ent}, {prt}) → {status_e}")
        print(f"  Partner-side reverse: ({prt_entity}, {reverse_icp_plain}) or ({prt_entity}, {reverse_icp_e}) → {status_p}{which_p}")
        print()
