import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import extract_entity_code_icm, extract_icp_code, extract_account_code

# Compare old output vs new enhanced output
old_path = r'G:\FCCS\Update files\ICM_Output_NEW_RUN.xlsx'
new_path = r'G:\FCCS\Update files\ICM_Output_ENHANCED.xlsx'

for label, path in [("OLD (ICM_Output_NEW_RUN)", old_path), ("NEW (ICM_Output_ENHANCED)", new_path)]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    
    # Count data rows with valid entity codes
    total_rows = 0
    valid_rows = 0
    for r in range(33, ws.max_row + 1):
        ent = str(ws.cell(r, 1).value or '').strip()
        if ent:
            total_rows += 1
            if extract_entity_code_icm(ent):
                valid_rows += 1
    
    # Count headers with account codes
    accts = set()
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(32, c).value or '').strip()
        code = extract_account_code(v)
        if code:
            accts.add(code)
    
    print(f"\n{label}:")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
    print(f"  Total text rows (from 33): {total_rows}")
    print(f"  Valid numeric entity rows: {valid_rows}")
    print(f"  Account codes in headers: {sorted(accts)}")

# Show the NEW output's extra accounts details
print("\n\n--- NEW output: All header columns ---")
wb_new = openpyxl.load_workbook(new_path, data_only=True)
ws_new = wb_new.active
for c in range(1, ws_new.max_column + 1):
    v = ws_new.cell(32, c).value
    if v:
        print(f"  Col {c}: {str(v)[:70]}")

# Check that 155020, 310001, 315005 are in the headers
print("\n--- Verifying extra accounts in headers ---")
target_codes = {'155020', '310001', '315005', '189501'}
for c in range(1, ws_new.max_column + 1):
    v = str(ws_new.cell(32, c).value or '').strip()
    code = extract_account_code(v)
    if code in target_codes:
        print(f"  FOUND {code} at Col {c}: {v[:70]}")
