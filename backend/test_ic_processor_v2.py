import os
import sys
import unittest
import uuid
import csv

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "venv", "Lib", "site-packages"))
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl

from app.ic_processor import (
    compare_v1_v2_report31,
    process_icm_report,
)
from app.ic_refactor.facts import normalize_account_code, normalize_journal_account_code, normalize_party_code
from app.ic_refactor.ledger import build_cell_ledger
from app.ic_refactor.models import LayoutColumn, LayoutSpec, NormalizedFact, PairOwner
from app.ic_refactor.pairs import build_pair_registry
from app.ic_refactor.pipeline import run_pipeline_v2
from app.ic_refactor.plug import build_plug_section_facts, derive_plug_facts


class IcProcessorV2Tests(unittest.TestCase):
    def test_normalize_party_code_rejects_suffix_code(self):
        diagnostics = {"bad_codes": []}
        self.assertEqual(normalize_party_code("E117100"), "117100")
        self.assertEqual(normalize_party_code("ICP_007009"), "007009")
        self.assertEqual(normalize_party_code("-E117100"), "117100")
        self.assertEqual(normalize_party_code("27009"), "027009")
        self.assertIsNone(normalize_party_code("117000A", diagnostics=diagnostics, field_name="entity"))
        self.assertEqual(len(diagnostics["bad_codes"]), 1)

    def test_normalize_account_code_extracts_numeric(self):
        self.assertEqual(normalize_account_code("534018 - 534018:Interest expense - related parties"), "534018")
        self.assertEqual(normalize_account_code("[111000].[111001]:111001:Land for undertermined use"), "111001")
        self.assertIsNone(normalize_account_code("[FCCS_Long Term Assets].[Plug_InvSh]:Investment-share capital plug account"))

    def test_normalize_journal_account_code_requires_direct_leading_code(self):
        self.assertEqual(
            normalize_journal_account_code("189001:189001:Inter company account - receivables"),
            "189001",
        )
        self.assertEqual(
            normalize_journal_account_code("534018 - 534018:Interest expense - related parties"),
            "534018",
        )
        self.assertIsNone(
            normalize_journal_account_code("[165000].[189501]:189501:Due from Related Party - Non-current")
        )
        self.assertIsNone(
            normalize_journal_account_code("[155000].[155020]:155020:Diar Club Al Houra")
        )

    def test_pair_registry_prefers_trusted_icm_direction(self):
        base_rows = [
            type("Row", (), {
                "pair_key": ("007009", "117100"),
                "pair_group": ("007009", "117100"),
                "row_num": 10,
                "row_direction": "partner_to_entity",
                "entity_label": "007009 - Example",
                "partner_label": "ICP_E117100 - Example ICP",
            })(),
            type("Row", (), {
                "pair_key": ("117100", "007009"),
                "pair_group": ("007009", "117100"),
                "row_num": 11,
                "row_direction": "entity_to_partner",
                "entity_label": "E117100 - Example",
                "partner_label": "ICP_007009 - Example ICP",
            })(),
        ]
        pair_registry, row_registry = build_pair_registry(
            base_rows,
            [],
            {},
            {"entity_labels": {}, "partner_labels": {}, "pair_labels": {}, "journal_pair_votes": {}},
            {"pairs": []},
        )
        owner = pair_registry[("007009", "117100")]
        self.assertEqual((owner.canonical_entity, owner.canonical_partner), ("117100", "007009"))
        self.assertEqual(owner.owner_reason, "icm_trusted_direction")
        self.assertEqual(row_registry[0].row_key, ("117100", "007009"))

    def test_derive_plug_facts_and_plug_section(self):
        fact_registry = {
            "fact_000001": NormalizedFact(
                family="parent",
                source_file="parent.xlsx",
                source_row=31,
                entity_num="117100",
                partner_num="007009",
                account_code="188600",
                amount=100.0,
                direction="entity_to_partner",
                raw_entity="E117100:Entity",
                raw_partner="ICP_007009:Partner",
                raw_account="188600",
                meta={},
            ),
            "fact_000002": NormalizedFact(
                family="ic_elim",
                source_file="elim.xlsx",
                source_row=33,
                entity_num="117100",
                partner_num="007009",
                account_code="188600",
                amount=75.0,
                direction="entity_to_partner",
                raw_entity="E117100:Entity",
                raw_partner="ICP_007009:Partner",
                raw_account="Total",
                meta={"column_role": "total"},
            ),
        }
        parent_plug = derive_plug_facts(["fact_000001"], fact_registry, {"534018"}, "188600", "parent")
        standalone = build_plug_section_facts(["fact_000002"], fact_registry, "188600")
        self.assertEqual(len(parent_plug), 1)
        self.assertEqual(parent_plug[0].account_code, "188600")
        self.assertEqual(parent_plug[0].meta["derived_from"], "parent")
        self.assertEqual(len(standalone), 1)
        self.assertEqual(standalone[0].meta["derived_from"], "ic_elim_total")

    def test_cell_ledger_assigns_exactly_once(self):
        fact_registry = {
            "fact_000001": NormalizedFact(
                family="parent",
                source_file="parent.xlsx",
                source_row=31,
                entity_num="117100",
                partner_num="007009",
                account_code="534018",
                amount=25.0,
                direction="entity_to_partner",
                raw_entity="E117100:Entity",
                raw_partner="ICP_007009:Partner",
                raw_account="534018",
                meta={},
            ),
            "fact_000002": NormalizedFact(
                family="parent",
                source_file="parent.xlsx",
                source_row=32,
                entity_num="007009",
                partner_num="117100",
                account_code="534018",
                amount=40.0,
                direction="partner_to_entity",
                raw_entity="007009:Partner",
                raw_partner="ICP_E117100:Entity ICP",
                raw_account="534018",
                meta={},
            ),
        }
        pair_registry = {
            ("007009", "117100"): PairOwner(
                pair_group=("007009", "117100"),
                canonical_entity="117100",
                canonical_partner="007009",
                owner_type="icm",
                owner_reason="icm_trusted_direction",
                icm_row_num=5,
                display_entity="E117100 - Entity",
                display_partner="ICP_007009 - Partner",
            )
        }
        layout = LayoutSpec(
            ent_cols=[LayoutColumn(code="534018", description="534018", series="S1", tag="Entity")],
            par_cols=[LayoutColumn(code="534018", description="534018", series="S1", tag="Partner")],
            plug_code="188600",
        )
        ledger, assignment = build_cell_ledger(
            ["fact_000001", "fact_000002"],
            fact_registry,
            pair_registry,
            layout,
            {"unmatched_facts": []},
        )
        self.assertEqual(ledger[(("117100", "007009"), "parent", "entity_side", "534018")].amount, 25.0)
        self.assertEqual(ledger[(("117100", "007009"), "parent", "partner_side", "534018")].amount, 40.0)
        self.assertEqual(len(assignment), 2)
        self.assertTrue(all(payload["reason"] == "routed" for payload in assignment.values()))

    def test_process_icm_report_v2_runs_small_fixture(self):
        from test_ic_processor_bidirectional import _ic_elim_workbook, _journal_workbook, _report_inputs_workbook

        base_dir = os.path.dirname(__file__)
        suffix = uuid.uuid4().hex
        paths = {
            "icm": os.path.join(base_dir, f"tmp_v2_ic_elim_{suffix}.xlsx"),
            "parent": os.path.join(base_dir, f"tmp_v2_parent_{suffix}.xlsx"),
            "contrib": os.path.join(base_dir, f"tmp_v2_contrib_{suffix}.xlsx"),
            "plug": os.path.join(base_dir, f"tmp_v2_plug_{suffix}.xlsx"),
            "inputs": os.path.join(base_dir, f"tmp_v2_inputs_{suffix}.xlsx"),
            "output": os.path.join(base_dir, f"tmp_v2_output_{suffix}.xlsx"),
        }
        for path in paths.values():
            self.addCleanup(lambda p=path: os.path.exists(p) and os.remove(p))

        _ic_elim_workbook(
            paths["icm"],
            [
                {
                    "entity": "E117100 - Entity",
                    "partner": "ICP_007009 - Partner ICP",
                    "entity_value": 100,
                    "partner_value": -25,
                    "total": 75,
                }
            ],
            plug_code="188600",
        )
        _journal_workbook(
            paths["parent"],
            [
                {
                    "entity": "E117100:Entity",
                    "account": "534018 - 534018:Interest expense - related parties",
                    "icp": "ICP_007009:Partner ICP",
                    "debit": 100,
                    "credit": 0,
                }
            ],
        )
        _journal_workbook(paths["contrib"], [])
        _journal_workbook(paths["plug"], [])
        _report_inputs_workbook(paths["inputs"], plug_code="188600", elim_code="534018")

        process_icm_report(
            paths["icm"],
            {
                "parent_journal": paths["parent"],
                "contribution_journal": paths["contrib"],
                "plugaccount_journal": paths["plug"],
            },
            paths["output"],
            report_inputs_path=paths["inputs"],
        )

        wb = openpyxl.load_workbook(paths["output"], data_only=True)
        self.assertIn("ICM Matched", wb.sheetnames)
        self.assertEqual(wb.sheetnames, ["ICM Matched"])

    def test_parent_ownership_then_qar_fx_roundup(self):
        from test_ic_processor_bidirectional import _journal_workbook

        base_dir = os.path.dirname(__file__)
        suffix = uuid.uuid4().hex
        paths = {
            "icm": os.path.join(base_dir, f"tmp_v2_icm_{suffix}.xlsx"),
            "parent": os.path.join(base_dir, f"tmp_v2_parent_fx_{suffix}.xlsx"),
            "plug": os.path.join(base_dir, f"tmp_v2_plug_fx_{suffix}.xlsx"),
            "entity_currency": os.path.join(base_dir, f"tmp_v2_entity_currency_{suffix}.csv"),
            "exchange_rates": os.path.join(base_dir, f"tmp_v2_exchange_{suffix}.xlsx"),
            "ownership": os.path.join(base_dir, f"tmp_v2_ownership_{suffix}.xlsx"),
            "output": os.path.join(base_dir, f"tmp_v2_output_fx_{suffix}.xlsx"),
        }
        for path in paths.values():
            self.addCleanup(lambda p=path: os.path.exists(p) and os.remove(p))

        wb_icm = openpyxl.Workbook()
        ws_icm = wb_icm.active
        ws_icm.cell(4, 1).value = "Entity"
        ws_icm.cell(4, 2).value = "Partner"
        ws_icm.cell(4, 3).value = "534018 - 534018:Interest expense - related parties Entity"
        ws_icm.cell(4, 4).value = "534018 - 534018:Interest expense - related parties Partner"
        ws_icm.cell(5, 1).value = "E117100 - Entity"
        ws_icm.cell(5, 2).value = "ICP_007009 - Partner ICP"
        ws_icm.cell(5, 3).value = 0
        ws_icm.cell(5, 4).value = 0
        wb_icm.save(paths["icm"])

        _journal_workbook(
            paths["parent"],
            [
                {
                    "entity": "E117100:Entity",
                    "account": "534018 - 534018:Interest expense - related parties",
                    "icp": "ICP_007009:Partner ICP",
                    "debit": 100,
                    "credit": 0,
                }
            ],
        )
        _journal_workbook(paths["plug"], [])

        with open(paths["entity_currency"], "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Entity", "Alias", "Base Currency"])
            writer.writerow(["E117100", "Entity 117100", "EUR"])

        wb_rate = openpyxl.Workbook()
        ws_rate = wb_rate.active
        ws_rate.cell(3, 3).value = "Dec"
        ws_rate.cell(3, 4).value = "Currency"
        ws_rate.cell(3, 5).value = "Type"
        ws_rate.cell(4, 3).value = 3.92245
        ws_rate.cell(4, 4).value = "EUR"
        ws_rate.cell(4, 5).value = "Average"
        ws_rate.cell(5, 3).value = 3.8087
        ws_rate.cell(5, 4).value = "EUR"
        ws_rate.cell(5, 5).value = "Ending"
        wb_rate.save(paths["exchange_rates"])

        wb_own = openpyxl.Workbook()
        ws_own = wb_own.active
        ws_own.cell(1, 1).value = "Entity"
        ws_own.cell(1, 2).value = "FCCS_Percent Ownership"
        ws_own.cell(2, 1).value = "-E117100"
        ws_own.cell(2, 2).value = 0.5
        wb_own.save(paths["ownership"])

        process_icm_report(
            paths["icm"],
            {
                "parent_journal": paths["parent"],
                "plugaccount_journal": paths["plug"],
            },
            paths["output"],
            lookup_paths={
                "entity_with_currency": paths["entity_currency"],
                "exchange_rates": paths["exchange_rates"],
                "ownership_structure": paths["ownership"],
            },
        )

        wb_out = openpyxl.load_workbook(paths["output"], data_only=True)
        ws_out = wb_out["ICM Matched"]

        parent_start = None
        qar_start = None
        for col in range(1, ws_out.max_column + 1):
            value = ws_out.cell(29, col).value
            if value == "Parent Input":
                parent_start = col
            elif value == "QAR Currency":
                qar_start = col
        self.assertIsNotNone(parent_start)
        self.assertIsNotNone(qar_start)

        target_row = None
        for row_num in range(33, ws_out.max_row + 1):
            entity = str(ws_out.cell(row_num, 1).value or "")
            partner = str(ws_out.cell(row_num, 2).value or "")
            if "117100" in entity and "007009" in partner:
                target_row = row_num
                break
        self.assertIsNotNone(target_row)

        self.assertEqual(ws_out.cell(target_row, parent_start).value, 50)
        self.assertEqual(ws_out.cell(target_row, qar_start).value, 197)

    def test_parent_pipeline_skips_bracketed_account_formats(self):
        from test_ic_processor_bidirectional import _journal_workbook

        base_dir = os.path.dirname(__file__)
        suffix = uuid.uuid4().hex
        paths = {
            "icm": os.path.join(base_dir, f"tmp_v2_icm_bracket_{suffix}.xlsx"),
            "parent": os.path.join(base_dir, f"tmp_v2_parent_bracket_{suffix}.xlsx"),
        }
        for path in paths.values():
            self.addCleanup(lambda p=path: os.path.exists(p) and os.remove(p))

        wb_icm = openpyxl.Workbook()
        ws_icm = wb_icm.active
        ws_icm.cell(32, 1).value = "Entity"
        ws_icm.cell(32, 2).value = "Partner"
        ws_icm.cell(32, 3).value = "189501 - 189501:Due from Related Party - Non-current Entity"
        ws_icm.cell(32, 4).value = "189501 - 189501:Due from Related Party - Non-current Partner"
        ws_icm.cell(32, 5).value = "Total"
        ws_icm.cell(33, 1).value = "E101000 - Example Entity"
        ws_icm.cell(33, 2).value = "ICP_001001 - Example Partner ICP"
        ws_icm.cell(33, 3).value = 0
        ws_icm.cell(33, 4).value = 0
        wb_icm.save(paths["icm"])

        _journal_workbook(
            paths["parent"],
            [
                {
                    "entity": "E101000:Example Entity",
                    "account": "[165000].[189501]:189501:Due from Related Party - Non-current",
                    "icp": "ICP_001001:Example Partner ICP",
                    "debit": 0,
                    "credit": 1555338.34,
                }
            ],
        )

        result = run_pipeline_v2(paths["icm"], {"parent_journal": paths["parent"]})

        self.assertEqual(result.fact_build.facts_parent, [])
        self.assertEqual(
            [item["reason"] for item in result.diagnostics["unmatched_facts"]],
            ["unsupported_journal_account_format"],
        )

    def test_lookup_fallback_defaults_keep_pipeline_running(self):
        from test_ic_processor_bidirectional import _journal_workbook

        base_dir = os.path.dirname(__file__)
        suffix = uuid.uuid4().hex
        paths = {
            "icm": os.path.join(base_dir, f"tmp_v2_icm_fallback_{suffix}.xlsx"),
            "parent": os.path.join(base_dir, f"tmp_v2_parent_fallback_{suffix}.xlsx"),
            "plug": os.path.join(base_dir, f"tmp_v2_plug_fallback_{suffix}.xlsx"),
            "output": os.path.join(base_dir, f"tmp_v2_output_fallback_{suffix}.xlsx"),
        }
        for path in paths.values():
            self.addCleanup(lambda p=path: os.path.exists(p) and os.remove(p))

        wb_icm = openpyxl.Workbook()
        ws_icm = wb_icm.active
        ws_icm.cell(4, 1).value = "Entity"
        ws_icm.cell(4, 2).value = "Partner"
        ws_icm.cell(4, 3).value = "534018 - 534018:Interest expense - related parties Entity"
        ws_icm.cell(4, 4).value = "534018 - 534018:Interest expense - related parties Partner"
        ws_icm.cell(5, 1).value = "E117100 - Entity"
        ws_icm.cell(5, 2).value = "ICP_007009 - Partner ICP"
        wb_icm.save(paths["icm"])

        _journal_workbook(
            paths["parent"],
            [
                {
                    "entity": "E117100:Entity",
                    "account": "534018 - 534018:Interest expense - related parties",
                    "icp": "ICP_007009:Partner ICP",
                    "debit": 100,
                    "credit": 0,
                }
            ],
        )
        _journal_workbook(paths["plug"], [])

        process_icm_report(
            paths["icm"],
            {
                "parent_journal": paths["parent"],
                "plugaccount_journal": paths["plug"],
            },
            paths["output"],
        )

        wb_out = openpyxl.load_workbook(paths["output"], data_only=True)
        self.assertEqual(wb_out.sheetnames, ["ICM Matched"])

    def test_no_plug_journal_keeps_standalone_plug_section_blank(self):
        from test_ic_processor_bidirectional import _ic_elim_workbook, _journal_workbook, _report_inputs_workbook

        base_dir = os.path.dirname(__file__)
        suffix = uuid.uuid4().hex
        paths = {
            "icm": os.path.join(base_dir, f"tmp_v2_icm_noplug_{suffix}.xlsx"),
            "parent": os.path.join(base_dir, f"tmp_v2_parent_noplug_{suffix}.xlsx"),
            "inputs": os.path.join(base_dir, f"tmp_v2_inputs_noplug_{suffix}.xlsx"),
            "output": os.path.join(base_dir, f"tmp_v2_output_noplug_{suffix}.xlsx"),
        }
        for path in paths.values():
            self.addCleanup(lambda p=path: os.path.exists(p) and os.remove(p))

        _ic_elim_workbook(
            paths["icm"],
            [
                {
                    "entity": "E117100 - Entity",
                    "partner": "ICP_007009 - Partner ICP",
                    "entity_value": 0,
                    "partner_value": 0,
                    "total": 75,
                }
            ],
            plug_code="188600",
        )
        _journal_workbook(
            paths["parent"],
            [
                {
                    "entity": "E117100:Entity",
                    "account": "534018 - 534018:Interest expense - related parties",
                    "icp": "ICP_007009:Partner ICP",
                    "debit": 100,
                    "credit": 0,
                }
            ],
        )
        _report_inputs_workbook(paths["inputs"], plug_code="188600", elim_code="534018")

        process_icm_report(
            paths["icm"],
            {
                "parent_journal": paths["parent"],
            },
            paths["output"],
            report_inputs_path=paths["inputs"],
        )

        wb_out = openpyxl.load_workbook(paths["output"], data_only=True)
        ws_out = wb_out["ICM Matched"]

        plug_section_col = None
        for col in range(1, ws_out.max_column + 1):
            if ws_out.cell(29, col).value == "Plug Account":
                plug_section_col = col
                break
        self.assertIsNotNone(plug_section_col)

        target_row = None
        for row_num in range(33, ws_out.max_row + 1):
            entity = str(ws_out.cell(row_num, 1).value or "")
            partner = str(ws_out.cell(row_num, 2).value or "")
            if "117100" in entity and "007009" in partner:
                target_row = row_num
                break
        self.assertIsNotNone(target_row)

        self.assertIsNone(ws_out.cell(target_row, plug_section_col).value)
        self.assertIsNone(ws_out.cell(target_row, plug_section_col + 1).value)

    def test_report31_comparison_harness(self):
        outputs = compare_v1_v2_report31()
        self.assertTrue(os.path.exists(outputs["current_output"]))
        self.assertTrue(os.path.exists(outputs["refactor_output"]))
        self.assertTrue(os.path.exists(outputs["comparison_output"]))

    def test_run_pipeline_v2_assigns_or_diagnoses_every_routed_fact(self):
        base_dir = r"g:\FCCS\backend\uploads\reports\31\inputs"
        result = run_pipeline_v2(
            os.path.join(base_dir, "IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx"),
            {
                "parent_journal": os.path.join(base_dir, "Parent report.xlsx"),
                "contribution_journal": os.path.join(base_dir, "Contribution report.xlsx"),
                "plugaccount_journal": os.path.join(base_dir, "Journal Report (4).xlsx"),
            },
            report_inputs_path=os.path.join(base_dir, "report Inputs.xlsx"),
        )
        routed_or_diagnosed = len(result.fact_assignment_log)
        self.assertGreater(routed_or_diagnosed, 0)
        self.assertGreaterEqual(
            routed_or_diagnosed,
            len(result.fact_build.facts_parent) + len(result.fact_build.facts_contrib) + len(result.fact_build.facts_plug_source),
        )
        self.assertTrue(
            all(
                payload["reason"] in {"routed", "diagnosed_unmatched", "plug_source_reconciliation_only"}
                for payload in result.fact_assignment_log.values()
            )
        )


if __name__ == "__main__":
    unittest.main()
