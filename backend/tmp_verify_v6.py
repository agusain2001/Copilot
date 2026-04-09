import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import extract_account_code

path = r'G:\FCCS\Update files\ICM_Output_ENHANCED_v6.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Max col: {ws.max_column}")

# Header row 32
all_codes = set()
for c in range(1, ws.max_column + 1):
    v = str(ws.cell(32, c).value or '').strip()
    if v:
        code = extract_account_code(v)
        if code: all_codes.add(code)

expected_codes = {'165001', '165002', '165003', '165004', '165005', '187052', '188800', '189001', '189014', '189015', '189501', '224000', '224001', '224003', '224009', '224024'}

print("\nAccounts present in output columns:")
print(sorted(all_codes))

matched = expected_codes.intersection(all_codes)
missing = expected_codes - all_codes

print(f"\nExpected match count: {len(matched)}")
if missing:
    print(f"MISSING ACCOUNTS!: {missing}")
else:
    print("ALL ACCOUNTS CONFIRMED PRESENT IN OUTPUT COLUMNS ✅")
