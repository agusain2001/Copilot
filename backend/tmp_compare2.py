import openpyxl, re

f = open('tmp_compare_out.txt', 'w')

# ICM source entities
wb = openpyxl.load_workbook(r'g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx', data_only=True)
ws = wb['Sheet1']

hdr_row = None
for r in range(1, 50):
    v1 = str(ws.cell(r, 1).value or '').strip().lower()
    v2 = str(ws.cell(r, 2).value or '').strip().lower()
    if v1 == 'entity' and v2 == 'partner':
        hdr_row = r
        break

icm_entities = set()
icm_pairs = []
for dr in range(hdr_row+1, ws.max_row + 1):
    e = str(ws.cell(dr, 1).value or '').strip()
    p = str(ws.cell(dr, 2).value or '').strip()
    if not e and not p:
        continue
    ecode = ''
    m = re.match(r'(\d{6})', e)
    if m:
        ecode = m.group(1)
    elif e.startswith('E'):
        m2 = re.match(r'(E\d+)', e)
        if m2:
            ecode = m2.group(1)
    if ecode:
        icm_entities.add(ecode)
    
    pcode = ''
    pm = re.match(r'(ICP_\w+)', p)
    if pm:
        pcode = pm.group(1)
    if ecode and pcode:
        icm_pairs.append((ecode, pcode))

f.write(f"ICM source unique entities:\n")
for e in sorted(icm_entities):
    f.write(f"  {e}\n")

# Journal entities
wb2 = openpyxl.load_workbook(r'g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx', data_only=True)
ws2 = wb2.active
headers = [str(ws2.cell(30, c).value or '').strip().lower() for c in range(1, ws2.max_column + 1)]
col_map = {h: i for i, h in enumerate(headers) if h}
eidx = col_map.get('entity', 2)

journal_entities_norm = set()
for r2 in range(31, ws2.max_row + 1):
    vals = [ws2.cell(r2, c).value for c in range(1, ws2.max_column + 1)]
    label = str(vals[0] or '').strip()
    if label == 'Grand Total':
        break
    try:
        e = str(vals[eidx] or '').strip()
    except IndexError:
        continue
    m = re.match(r'(E?\d+\w*)', e)
    if m:
        code = m.group(1)
        if code.startswith('E') and len(code) > 1 and code[1:].replace('_', '').isdigit():
            journal_entities_norm.add(code[1:])
        else:
            journal_entities_norm.add(code)

f.write(f"\nJournal entity codes (normalized):\n")
for e in sorted(journal_entities_norm):
    f.write(f"  {e}\n")

f.write(f"\nICM entities NOT in journal:\n")
for e in sorted(icm_entities - journal_entities_norm):
    f.write(f"  {e}\n")

f.write(f"\nJournal entities NOT in ICM:\n")
for e in sorted(journal_entities_norm - icm_entities):
    f.write(f"  {e}\n")

f.write(f"\nIntersection:\n")
for e in sorted(icm_entities & journal_entities_norm):
    f.write(f"  {e}\n")

# Now the KEY analysis - for each ICM pair, can we find matching journal data?
f.write(f"\n{'='*70}\n")
f.write(f"KEY ANALYSIS: Journal key matching for ICM pairs\n")
f.write(f"{'='*70}\n\n")

# Read journal keys
icpidx = col_map.get('intercompany', 4)
journal_keys = set()
for r2 in range(31, ws2.max_row + 1):
    vals = [ws2.cell(r2, c).value for c in range(1, ws2.max_column + 1)]
    label = str(vals[0] or '').strip()
    if label == 'Grand Total':
        break
    try:
        ent_raw = str(vals[eidx] or '').strip()
        icp_raw = str(vals[icpidx] or '').strip()
    except IndexError:
        continue
    e_code_raw = re.match(r'(E?\d+\w*)', ent_raw)
    e_code = e_code_raw.group(1) if e_code_raw else ''
    if e_code.startswith('E') and len(e_code) > 1 and e_code[1:].replace('_', '').isdigit():
        e_code = e_code[1:]
    
    i_code = ''
    im = re.match(r'(ICP_\w+)', icp_raw)
    if im:
        i_code = im.group(1)
    if e_code and i_code:
        journal_keys.add((e_code, i_code))

f.write(f"Total journal (entity, icp) keys: {len(journal_keys)}\n")
for k in sorted(journal_keys):
    f.write(f"  {k}\n")

f.write(f"\nTotal ICM (entity, partner) pairs: {len(icm_pairs)}\n\n")

# The code currently matches:
# Entity-side: key = (icm_ent, icm_prt, acct) => looks for (ent, prt) in journal
# Partner-side: key = (prt_digit, ICP_{ent}, acct) => looks for reverse in journal
# But ICP codes in ICM use ICP_E... while journal ICP also uses ICP_E... for E-prefixed 
# and ICP_... for non-E-prefixed

for ent, prt in sorted(set(icm_pairs)):
    f.write(f"\nICM pair: ({ent}, {prt})\n")
    
    # Entity-side direct: (ent, prt) in journal?
    ent_match = (ent, prt) in journal_keys
    f.write(f"  Entity-side  ({ent}, {prt}) -> {'FOUND' if ent_match else 'NOT FOUND'}\n")
    
    # Partner-side reverse: strip ICP_ from prt to get partner entity code
    prt_raw = prt[4:] if prt.startswith('ICP_') else prt
    # Normalize: strip E prefix if present
    if prt_raw.startswith('E') and len(prt_raw) > 1 and prt_raw[1:].replace('_', '').isdigit():
        prt_entity = prt_raw[1:]
    else:
        prt_entity = prt_raw
    
    rev_plain = f"ICP_{ent}"
    rev_e = f"ICP_E{ent}"
    
    par_match_plain = (prt_entity, rev_plain) in journal_keys
    par_match_e = (prt_entity, rev_e) in journal_keys
    
    f.write(f"  Partner-side ({prt_entity}, {rev_plain}) -> {'FOUND' if par_match_plain else 'NOT FOUND'}\n")
    f.write(f"  Partner-side ({prt_entity}, {rev_e}) -> {'FOUND' if par_match_e else 'NOT FOUND'}\n")
    
    # Also try with E prefix on prt entity (not normalized)
    if prt_raw != prt_entity:
        par_match_raw_p = (prt_raw, rev_plain) in journal_keys  
        par_match_raw_e = (prt_raw, rev_e) in journal_keys
        f.write(f"  Partner-side ({prt_raw}, {rev_plain}) -> {'FOUND' if par_match_raw_p else 'NOT FOUND'}\n")
        f.write(f"  Partner-side ({prt_raw}, {rev_e}) -> {'FOUND' if par_match_raw_e else 'NOT FOUND'}\n")

f.close()
print("Done - see tmp_compare_out.txt")
