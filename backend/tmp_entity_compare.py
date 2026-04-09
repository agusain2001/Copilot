import openpyxl, re

# ICM source entities
wb = openpyxl.load_workbook(r'g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx', data_only=True)
ws = wb['Sheet1']

# Find header
for r in range(1, 50):
    v1 = str(ws.cell(r, 1).value or '').strip().lower()
    v2 = str(ws.cell(r, 2).value or '').strip().lower()
    if v1 == 'entity' and v2 == 'partner':
        print(f"Header at row {r}")
        break

# Print ALL entity values
icm_entities = set()
for dr in range(r+1, ws.max_row + 1):
    e = str(ws.cell(dr, 1).value or '').strip()
    p = str(ws.cell(dr, 2).value or '').strip()
    if not e and not p:
        continue
    m = re.match(r'(\d{6})', e)
    if m:
        icm_entities.add(m.group(1))
    elif e.startswith('E'):
        m2 = re.match(r'(E\d+)', e)
        if m2:
            icm_entities.add(m2.group(1))

print(f"ICM source unique entity codes: {sorted(icm_entities)}")

# Journal entities
wb2 = openpyxl.load_workbook(r'g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx', data_only=True)
ws2 = wb2.active
headers = [str(ws2.cell(30, c).value or '').strip().lower() for c in range(1, ws2.max_column + 1)]
col_map = {h: i for i, h in enumerate(headers) if h}
eidx = col_map.get('entity', 2)

journal_entities_raw = set()
journal_entities_norm = set()
for r2 in range(31, ws2.max_row + 1):
    vals = [ws2.cell(r2, c).value for c in range(1, ws2.max_column + 1)]
    label = str(vals[0] or '').strip()
    if label == 'Grand Total':
        break
    try:
        e = str(vals[eidx] or '').strip()
    except IndexError:
        continue
    m = re.match(r'(E?\d+\w*)', e)
    if m:
        code = m.group(1)
        journal_entities_raw.add(code)
        if code.startswith('E') and len(code) > 1 and code[1:].replace('_', '').isdigit():
            journal_entities_norm.add(code[1:])
        else:
            journal_entities_norm.add(code)

print(f"\nJournal entity codes (raw): {sorted(journal_entities_raw)}")
print(f"Journal entity codes (normalized): {sorted(journal_entities_norm)}")
print(f"\nICM entities NOT in journal: {sorted(icm_entities - journal_entities_norm)}")
print(f"Journal entities NOT in ICM: {sorted(journal_entities_norm - icm_entities)}")
print(f"Intersection: {sorted(icm_entities & journal_entities_norm)}")
