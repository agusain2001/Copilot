"""
Count how many times each Contribution journal transaction appears in the ICM Output.
"""
import sys, re
sys.path.insert(0, '.')
import openpyxl
from collections import defaultdict
from app.ic_processor import (
    extract_entity_code_journal, normalize_entity_code, extract_icp_code,
    extract_account_code, apply_sign, to_float, get_journal_indices,
    JOURNAL_DATA_START, is_detail_row
)

f = open("tmp_contrib_count.txt", "w", encoding="utf-8")

# Read Contribution journal
wb_j = openpyxl.load_workbook(r"uploads\reports\31\inputs\Contribution report.xlsx", data_only=True)
ws_j = wb_j.active
indices = get_journal_indices(ws_j)

transactions = []
for row_num, row in enumerate(ws_j.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws_j.max_row), start=JOURNAL_DATA_START):
    vals = [cell.value for cell in row]
    if not vals: continue
    label = str(vals[0] or "").strip()
    if label == "Grand Total": break
    if not is_detail_row(vals, indices): continue
    try:
        entity_raw = str(vals[indices["entity"]] or "").strip()
        account_raw = str(vals[indices["acct"]] or "").strip()
        icp_raw = str(vals[indices["icp"]] or "").strip()
        debit = to_float(vals[indices["debit"]])
        credit = to_float(vals[indices["credit"]])
    except IndexError:
        continue
    ent = normalize_entity_code(extract_entity_code_journal(entity_raw))
    icp = extract_icp_code(icp_raw)
    acct = extract_account_code(account_raw)
    is_group = bool(re.match(r"^E\d+", extract_entity_code_journal(entity_raw)))
    signed = apply_sign(debit, credit, acct)
    if ent and icp and acct and signed != 0:
        transactions.append({
            "row": row_num, "entity": ent, "icp": icp, "account": acct,
            "debit": debit, "credit": credit, "signed": signed,
            "entity_raw": entity_raw, "is_group": is_group,
        })

f.write(f"Contribution journal: {len(transactions)} transactions\n\n")
f.write("ALL TRANSACTIONS:\n")
for t in transactions:
    grp = " (GROUP/E-prefix)" if t["is_group"] else ""
    f.write(f"  JnlRow {t['row']:3d}: ({t['entity']}, {t['icp']}, {t['account']}) "
            f"debit={t['debit']:.2f} credit={t['credit']:.2f} signed={t['signed']:.2f}{grp}\n")

# Check which accounts are in the ICM valid accounts list
# Read the ICM source to get valid accounts
wb_icm = openpyxl.load_workbook(
    r"uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx",
    data_only=True)
ws_icm = wb_icm.active

# Find header row
from app.ic_processor import read_icm_headers, ICM_HEADER_ROW
icm_header_map = read_icm_headers(ws_icm)
valid_accounts = set()
for (code, tag) in icm_header_map.keys():
    valid_accounts.add(code)

f.write(f"\nICM Valid accounts: {sorted(valid_accounts)}\n\n")

# Check each transaction against valid accounts
valid_txns = []
invalid_txns = []
for t in transactions:
    if t["account"] in valid_accounts:
        valid_txns.append(t)
    else:
        invalid_txns.append(t)

f.write(f"Transactions with valid ICM accounts: {len(valid_txns)}\n")
f.write(f"Transactions with NON-ICM accounts (skipped): {len(invalid_txns)}\n\n")

if invalid_txns:
    f.write("SKIPPED (account not in ICM):\n")
    for t in invalid_txns:
        f.write(f"  JnlRow {t['row']:3d}: ({t['entity']}, {t['icp']}, {t['account']}) signed={t['signed']:.2f}\n")
    f.write("\n")

# Now check the output for Contribution block
wb_o = openpyxl.load_workbook(r"uploads\reports\31\outputs\ICM_Output_31_v3.xlsx", data_only=True)
ws_o = wb_o.active

# Contribution block columns: entity-side 32-36, partner-side 38-42
contrib_entity_cols = {32: "534018", 33: "433002", 34: "111006", 35: "120030", 36: "121015"}
contrib_partner_cols = {38: "433002", 39: "534018", 40: "111006", 41: "120030", 42: "121015"}

output_contrib_values = []
for r in range(33, ws_o.max_row + 1):
    ent_disp = str(ws_o.cell(r, 1).value or "").strip()
    prt_disp = str(ws_o.cell(r, 2).value or "").strip()
    if not ent_disp or not prt_disp: continue
    
    m = re.match(r"(\d{6})", ent_disp)
    if m:
        ent_code = m.group(1)
    else:
        m = re.match(r"E(\d+)", ent_disp)
        ent_code = m.group(1) if m else ""
    m = re.match(r"(ICP_\w+)", prt_disp)
    prt_code = m.group(1) if m else ""
    if not ent_code or not prt_code: continue
    
    prt_digit = prt_code[4:] if prt_code.startswith("ICP_") else prt_code
    if prt_digit.startswith("E") and len(prt_digit) > 1 and prt_digit[1:].isdigit():
        prt_entity = prt_digit[1:]
    else:
        prt_entity = prt_digit
    
    for c, acct in contrib_entity_cols.items():
        v = ws_o.cell(r, c).value
        if v is not None and v != "" and v != 0:
            output_contrib_values.append((r, ent_code, prt_code, "entity", acct, to_float(v)))
    for c, acct in contrib_partner_cols.items():
        v = ws_o.cell(r, c).value
        if v is not None and v != "" and v != 0:
            output_contrib_values.append((r, prt_entity, f"ICP_{ent_code}", "partner", acct, to_float(v)))

f.write(f"Total Contribution block non-zero cells in output: {len(output_contrib_values)}\n")
f.write(f"Expected valid transactions: {len(valid_txns)}\n\n")

if output_contrib_values:
    f.write("OUTPUT VALUES:\n")
    for ov in sorted(output_contrib_values, key=lambda x: (x[0], x[3], x[4])):
        r, ent, icp, side, acct, val = ov
        matches = [t for t in valid_txns if t["entity"] == ent and t["account"] == acct]
        matches = [t for t in matches if t["icp"] == icp or t["icp"] == icp.replace("ICP_", "ICP_E")]
        status = "OK" if matches else "NO MATCH"
        f.write(f"  OutRow={r:3d} side={side:7s} ({ent}, {icp}, {acct}) val={val:>15.2f} [{status}]\n")

# Transaction usage count
f.write(f"\n--- Transaction usage count (valid accounts only) ---\n")
txn_usage = defaultdict(int)
for ov in output_contrib_values:
    r, ent, icp, side, acct, val = ov
    for t in valid_txns:
        if t["entity"] == ent and t["account"] == acct:
            if t["icp"] == icp or t["icp"] == icp.replace("ICP_", "ICP_E"):
                txn_usage[t["row"]] += 1

for t in valid_txns:
    count = txn_usage.get(t["row"], 0)
    status = "OK" if count == 1 else ("MISSING" if count == 0 else f"DUPLICATE x{count}")
    f.write(f"  JnlRow {t['row']:3d}: ({t['entity']}, {t['icp']}, {t['account']}) = {t['signed']:>15.2f} -> used {count} times [{status}]\n")

total_valid_uses = sum(txn_usage.get(t["row"], 0) for t in valid_txns)
f.write(f"\nTotal valid transaction uses: {total_valid_uses} (should be {len(valid_txns)})\n")
f.write(f"Duplicated: {sum(1 for t in valid_txns if txn_usage.get(t['row'], 0) > 1)}\n")
f.write(f"Missing: {sum(1 for t in valid_txns if txn_usage.get(t['row'], 0) == 0)}\n")

f.close()
print("Done - see tmp_contrib_count.txt")
