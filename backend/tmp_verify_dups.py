"""Verify no journal value appears more than once in v9 output."""
import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from collections import defaultdict

path = r'G:\FCCS\Update files\ICM_Output_ENHANCED_v9.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

# Parent block starts at col 38, Contribution at col 74
# Check entity 001001 with partner ICP_022001 AND entity 022001 with partner ICP_001001
# These are the mirror rows that could cause duplication

print("Checking for value duplication in Parent Input block...")
print("Looking for rows with entity 001001/022001 pair:\n")

for r in range(33, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or '').strip()
    par = str(ws.cell(r, 2).value or '').strip()
    
    if ('001001' in ent and '022001' in par) or ('022001' in ent and '001001' in par):
        # Parent block: cols 38-73 (entity accounts start at 38)
        parent_vals = []
        for c in range(38, 74):
            v = ws.cell(r, c).value
            if v is not None and v != 0:
                hdr = str(ws.cell(32, c).value or '')[:40]
                parent_vals.append(f"  Col {c} [{hdr}]: {v}")
        
        # Contrib block: cols 74-109
        contrib_vals = []
        for c in range(74, 110):
            v = ws.cell(r, c).value
            if v is not None and v != 0:
                hdr = str(ws.cell(32, c).value or '')[:40]
                contrib_vals.append(f"  Col {c} [{hdr}]: {v}")
        
        print(f"Row {r}: Entity={ent[:30]}  Partner={par[:30]}")
        if parent_vals:
            print(f"  Parent Input values:")
            for pv in parent_vals:
                print(f"    {pv}")
        else:
            print(f"  Parent Input: (empty)")
        if contrib_vals:
            print(f"  Contrib Input values:")
            for cv in contrib_vals:
                print(f"    {cv}")
        else:
            print(f"  Contrib Input: (empty)")
        print()

# Global check: count how many times each non-zero value appears across parent+contrib blocks
print("\n--- Global duplication check ---")
value_locations = defaultdict(list)
for r in range(33, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or '').strip()[:15]
    par = str(ws.cell(r, 2).value or '').strip()[:15]
    # Parent block data columns (skip Variance/Total)
    for c in range(38, 70):  # account columns only
        v = ws.cell(r, c).value
        if v is not None and v != 0 and isinstance(v, (int, float)):
            hdr = str(ws.cell(32, c).value or '')[:20]
            value_locations[(v, hdr)].append(f"R{r} {ent}/{par}")
    # Contrib block
    for c in range(74, 106):
        v = ws.cell(r, c).value
        if v is not None and v != 0 and isinstance(v, (int, float)):
            hdr = str(ws.cell(32, c).value or '')[:20]
            value_locations[(v, hdr)].append(f"R{r} {ent}/{par}")

dupes = {k: v for k, v in value_locations.items() if len(v) > 1}
if dupes:
    print(f"Found {len(dupes)} values appearing more than once:")
    for (val, hdr), locs in sorted(dupes.items(), key=lambda x: -len(x[1]))[:10]:
        print(f"  Value={val} Account={hdr}: appears {len(locs)} times")
        for loc in locs[:3]:
            print(f"    at {loc}")
else:
    print("No duplications found! Each value appears exactly once.")
