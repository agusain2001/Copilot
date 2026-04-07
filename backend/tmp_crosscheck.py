"""
Cross-check ICM Output values against source Parent and Contribution journal files.
Validates every non-empty cell in the Parent and Contribution blocks.
"""
import sys, re, os
sys.path.insert(0, '.')
import openpyxl
from collections import defaultdict
from app.ic_processor import (
    extract_entity_code_journal, normalize_entity_code, extract_icp_code,
    extract_account_code, apply_sign, to_float, get_journal_indices,
    JOURNAL_DATA_START, is_detail_row
)

def read_journal_full(filepath, name):
    """Read all journal entries with full detail."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    indices = get_journal_indices(ws)
    
    entries = []
    for row_num, row in enumerate(ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row), start=JOURNAL_DATA_START):
        vals = [cell.value for cell in row]
        if not vals: continue
        label = str(vals[0] or "").strip()
        if label == "Grand Total": break
        if not is_detail_row(vals, indices): continue
        
        try:
            entity_raw = str(vals[indices["entity"]] or "").strip()
            account_raw = str(vals[indices["acct"]] or "").strip()
            icp_raw = str(vals[indices["icp"]] or "").strip()
            debit = to_float(vals[indices["debit"]])
            credit = to_float(vals[indices["credit"]])
        except IndexError:
            continue
        
        ent_code_raw = extract_entity_code_journal(entity_raw)
        ent_code = normalize_entity_code(ent_code_raw)
        acct_code = extract_account_code(account_raw)
        icp_code = extract_icp_code(icp_raw)
        is_group = bool(re.match(r"^E\d+", ent_code_raw))
        
        if not ent_code or not icp_code or not acct_code:
            continue
        
        signed_val = apply_sign(debit, credit, acct_code)
        
        entries.append({
            "entity_code": ent_code,
            "icp_code": icp_code,
            "account_code": acct_code,
            "debit": debit,
            "credit": credit,
            "signed_value": signed_val,
            "is_group": is_group,
            "entity_raw": entity_raw,
            "row": row_num,
        })
    
    # Build lookups: (entity, icp, acct) -> list of entries
    primary = defaultdict(list)  # non-group
    fallback = defaultdict(list)  # group (E-prefix)
    for e in entries:
        key = (e["entity_code"], e["icp_code"], e["account_code"])
        if not e["is_group"]:
            primary[key].append(e)
        else:
            fallback[key].append(e)
    
    return entries, primary, fallback

def read_icm_output(filepath):
    """Read ICM output structured data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Find header row
    header_row = None
    for r in range(1, 40):
        v1 = str(ws.cell(r, 1).value or "").strip().lower()
        v2 = str(ws.cell(r, 2).value or "").strip().lower()
        if v1 == "entity" and v2 == "partner":
            header_row = r
            break
    
    if not header_row:
        raise ValueError("Could not find header row")
    
    # Read headers
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(header_row, c).value or "").strip()
        if v:
            headers[c] = v
    
    # Identify block boundaries by looking at section labels (row 29)
    # Read all data rows
    data_rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        ent = str(ws.cell(r, 1).value or "").strip()
        prt = str(ws.cell(r, 2).value or "").strip()
        if not ent and not prt:
            continue
        
        row_data = {"row": r, "entity": ent, "partner": prt, "values": {}}
        for c in range(3, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and v != "" and v != 0:
                row_data["values"][c] = v
        data_rows.append(row_data)
    
    return ws, headers, header_row, data_rows

# ═══════════════════════════════════════════════════════════════════
# MAIN ANALYSIS
# ═══════════════════════════════════════════════════════════════════

f = open("tmp_crosscheck_result.txt", "w", encoding="utf-8")

f.write("=" * 80 + "\n")
f.write("  CROSS-CHECK: ICM Output vs Source Journals\n")
f.write("=" * 80 + "\n\n")

# Read journals
parent_entries, parent_primary, parent_fallback = read_journal_full(
    r"uploads\reports\31\inputs\Parent report.xlsx", "Parent")
contrib_entries, contrib_primary, contrib_fallback = read_journal_full(
    r"uploads\reports\31\inputs\Contribution report.xlsx", "Contribution")

f.write(f"Parent journal: {len(parent_entries)} entries, {len(parent_primary)} primary keys, {len(parent_fallback)} fallback keys\n")
f.write(f"Contribution journal: {len(contrib_entries)} entries, {len(contrib_primary)} primary keys, {len(contrib_fallback)} fallback keys\n\n")

# Aggregate by (entity, icp, account) -> net signed value
def aggregate_journal(primary, fallback):
    """Compute net signed value per (entity, icp, account) key."""
    result = {}
    for key, entries in primary.items():
        net = sum(e["signed_value"] for e in entries)
        if net != 0:
            result[key] = {"net": net, "source": "primary", "entries": entries}
    for key, entries in fallback.items():
        net = sum(e["signed_value"] for e in entries)
        if net != 0:
            if key not in result:
                result[key] = {"net": net, "source": "fallback", "entries": entries}
    return result

parent_agg = aggregate_journal(parent_primary, parent_fallback)
contrib_agg = aggregate_journal(contrib_primary, contrib_fallback)

f.write("--- PARENT JOURNAL: All (entity, icp, account) -> net value ---\n")
for key in sorted(parent_agg.keys()):
    info = parent_agg[key]
    f.write(f"  {key} -> net={info['net']:.2f} ({info['source']}, {len(info['entries'])} entries)\n")
    for e in info['entries']:
        f.write(f"    Row {e['row']}: debit={e['debit']:.2f}, credit={e['credit']:.2f}, signed={e['signed_value']:.2f}\n")

f.write(f"\n--- CONTRIBUTION JOURNAL: All (entity, icp, account) -> net value ---\n")
for key in sorted(contrib_agg.keys()):
    info = contrib_agg[key]
    f.write(f"  {key} -> net={info['net']:.2f} ({info['source']}, {len(info['entries'])} entries)\n")
    for e in info['entries']:
        f.write(f"    Row {e['row']}: debit={e['debit']:.2f}, credit={e['credit']:.2f}, signed={e['signed_value']:.2f}\n")

# Read ICM output
ws_out, headers_out, hdr_row, output_rows = read_icm_output(
    r"uploads\reports\31\outputs\ICM_Output_31.xlsx")

f.write(f"\n\nICM Output: {len(output_rows)} data rows, header at row {hdr_row}\n")
f.write(f"Headers:\n")
for c in sorted(headers_out.keys()):
    f.write(f"  Col {c:3d}: {headers_out[c][:60]}\n")

# Identify column blocks
# Base block: cols 3-15 (entity-side accounts, variance, partner-side accounts, variance, total)
# Parent block: cols 17-29 (same structure)  -> col 30 = plug
# Contribution block: cols 32-44 (same structure) -> col 45 = plug
# Plug section: col 47-48
# Final total: col 50

# Find account codes from headers
def extract_acct_from_header(hdr_text):
    """Extract account code from header like '534018 - 534018:Interest...'"""
    m = re.match(r"(\d{6})", hdr_text)
    return m.group(1) if m else None

# Map columns to (block, side, account_code)
col_mapping = {}
for c, hdr in headers_out.items():
    acct = extract_acct_from_header(hdr)
    if not acct:
        continue
    
    # Determine block and side
    if 3 <= c <= 7:
        col_mapping[c] = ("base", "entity", acct)
    elif 9 <= c <= 13:
        col_mapping[c] = ("base", "partner", acct)
    elif 17 <= c <= 21:
        col_mapping[c] = ("parent", "entity", acct)
    elif 23 <= c <= 27:
        col_mapping[c] = ("parent", "partner", acct)
    elif 32 <= c <= 36:
        col_mapping[c] = ("contrib", "entity", acct)
    elif 38 <= c <= 42:
        col_mapping[c] = ("contrib", "partner", acct)

f.write(f"\nColumn mapping:\n")
for c in sorted(col_mapping.keys()):
    block, side, acct = col_mapping[c]
    f.write(f"  Col {c:3d}: block={block}, side={side}, account={acct}\n")

# ═══════════════════════════════════════════════════════════════════
# CROSS-CHECK each non-empty Parent/Contribution cell
# ═══════════════════════════════════════════════════════════════════

f.write(f"\n{'='*80}\n")
f.write(f"  DETAILED CROSS-CHECK\n")
f.write(f"{'='*80}\n\n")

def extract_entity_code_from_display(display):
    """Extract normalized entity code from display text like '001001 - ...' or 'E101000:...'"""
    display = str(display or "").strip()
    m = re.match(r"(\d{6})", display)
    if m: return m.group(1)
    m = re.match(r"E(\d+)", display)
    if m: return m.group(1)
    return ""

def extract_icp_from_display(display):
    """Extract ICP code from display text."""
    display = str(display or "").strip()
    m = re.match(r"(ICP_\w+)", display)
    return m.group(1) if m else ""

mismatches = []
correct = 0
total_checked = 0

for row_data in output_rows:
    ent_display = row_data["entity"]
    prt_display = row_data["partner"]
    ent_code = extract_entity_code_from_display(ent_display)
    prt_code = extract_icp_from_display(prt_display)
    
    if not ent_code or not prt_code:
        continue
    
    # Compute reverse codes
    prt_digit = prt_code[4:] if prt_code.startswith("ICP_") else prt_code
    if prt_digit.startswith("E") and len(prt_digit) > 1 and prt_digit[1:].replace("_","").isdigit():
        prt_entity = prt_digit[1:]
    else:
        prt_entity = prt_digit
    reverse_icp_plain = f"ICP_{ent_code}"
    reverse_icp_e = f"ICP_E{ent_code}"
    
    for c, val in row_data["values"].items():
        if c not in col_mapping:
            continue
        
        block, side, acct = col_mapping[c]
        
        # Determine which journal and which key to check
        if block == "parent":
            journal_agg = parent_agg
            journal_name = "Parent"
        elif block == "contrib":
            journal_agg = contrib_agg
            journal_name = "Contribution"
        else:
            continue  # Base block comes from ICM source, not journals
        
        # Determine the lookup key based on side
        if side == "entity":
            # Entity-side: direct key (ent, prt, acct)
            lookup_key = (ent_code, prt_code, acct)
        else:
            # Partner-side: reverse key (prt_entity, ICP_{ent}, acct)
            lookup_key = None
            for rk in [(prt_entity, reverse_icp_plain, acct),
                       (prt_entity, reverse_icp_e, acct)]:
                if rk in journal_agg:
                    lookup_key = rk
                    break
            if lookup_key is None:
                lookup_key = (prt_entity, reverse_icp_plain, acct)
        
        # Check the value
        total_checked += 1
        expected_info = journal_agg.get(lookup_key)
        expected_val = expected_info["net"] if expected_info else None
        
        actual_val = to_float(val)
        
        if expected_val is not None:
            diff = abs(actual_val - expected_val)
            if diff < 0.01:
                correct += 1
            else:
                mismatches.append({
                    "row": row_data["row"],
                    "col": c,
                    "entity": ent_display[:40],
                    "partner": prt_display[:40],
                    "block": block,
                    "side": side,
                    "account": acct,
                    "lookup_key": lookup_key,
                    "expected": expected_val,
                    "actual": actual_val,
                    "diff": actual_val - expected_val,
                })
                f.write(f"MISMATCH Row {row_data['row']}, Col {c} ({block}/{side}/{acct}):\n")
                f.write(f"  Entity: {ent_display[:40]}\n")
                f.write(f"  Partner: {prt_display[:40]}\n")
                f.write(f"  Lookup key: {lookup_key}\n")
                f.write(f"  Expected: {expected_val:.2f}\n")
                f.write(f"  Actual:   {actual_val:.2f}\n")
                f.write(f"  Diff:     {actual_val - expected_val:.2f}\n")
                if expected_info:
                    for e in expected_info["entries"]:
                        f.write(f"    Journal row {e['row']}: D={e['debit']:.2f}, C={e['credit']:.2f}, signed={e['signed_value']:.2f}\n")
                f.write("\n")
        else:
            # Value in output but NOT in journal - unexpected
            if abs(actual_val) > 0.01:
                mismatches.append({
                    "row": row_data["row"],
                    "col": c,
                    "entity": ent_display[:40],
                    "partner": prt_display[:40],
                    "block": block,
                    "side": side,
                    "account": acct,
                    "lookup_key": lookup_key,
                    "expected": None,
                    "actual": actual_val,
                    "diff": None,
                })
                f.write(f"UNEXPECTED VALUE Row {row_data['row']}, Col {c} ({block}/{side}/{acct}):\n")
                f.write(f"  Entity: {ent_display[:40]}\n")
                f.write(f"  Partner: {prt_display[:40]}\n")
                f.write(f"  Lookup key: {lookup_key}\n")
                f.write(f"  Expected: NOT IN JOURNAL\n")
                f.write(f"  Actual:   {actual_val:.2f}\n\n")

# Also check: journal values that appear NOWHERE in the output
f.write(f"\n{'='*80}\n")
f.write(f"  JOURNAL VALUES NOT FOUND IN OUTPUT\n")
f.write(f"{'='*80}\n\n")

# Collect all (entity, partner, account, block, side) -> value from output
output_values = {}
for row_data in output_rows:
    ent_code = extract_entity_code_from_display(row_data["entity"])
    prt_code = extract_icp_from_display(row_data["partner"])
    if not ent_code or not prt_code:
        continue
    
    prt_digit = prt_code[4:] if prt_code.startswith("ICP_") else prt_code
    if prt_digit.startswith("E") and len(prt_digit) > 1 and prt_digit[1:].replace("_","").isdigit():
        prt_entity = prt_digit[1:]
    else:
        prt_entity = prt_digit
    
    for c, val in row_data["values"].items():
        if c not in col_mapping:
            continue
        block, side, acct = col_mapping[c]
        if block == "parent":
            if side == "entity":
                key = (ent_code, prt_code, acct)
            else:
                key = (prt_entity, f"ICP_{ent_code}", acct)
                key2 = (prt_entity, f"ICP_E{ent_code}", acct)
                output_values[("parent_partner", key2)] = to_float(val)
            output_values[("parent_" + side, key)] = to_float(val)
        elif block == "contrib":
            if side == "entity":
                key = (ent_code, prt_code, acct)
            else:
                key = (prt_entity, f"ICP_{ent_code}", acct)
                key2 = (prt_entity, f"ICP_E{ent_code}", acct)
                output_values[("contrib_partner", key2)] = to_float(val)
            output_values[("contrib_" + side, key)] = to_float(val)

missing_parent = 0
for key, info in sorted(parent_agg.items()):
    found_entity = ("parent_entity", key) in output_values
    found_partner = ("parent_partner", key) in output_values
    if not found_entity and not found_partner:
        f.write(f"  Parent {key} -> net={info['net']:.2f} NOT FOUND in any output row\n")
        missing_parent += 1

missing_contrib = 0
for key, info in sorted(contrib_agg.items()):
    found_entity = ("contrib_entity", key) in output_values
    found_partner = ("contrib_partner", key) in output_values
    if not found_entity and not found_partner:
        f.write(f"  Contrib {key} -> net={info['net']:.2f} NOT FOUND in any output row\n")
        missing_contrib += 1

# Summary
f.write(f"\n{'='*80}\n")
f.write(f"  SUMMARY\n")
f.write(f"{'='*80}\n\n")
f.write(f"Total cells checked: {total_checked}\n")
f.write(f"Correct matches: {correct}\n")
f.write(f"Mismatches: {len(mismatches)}\n")
f.write(f"Parent journal keys not in output: {missing_parent}\n")
f.write(f"Contribution journal keys not in output: {missing_contrib}\n\n")

if mismatches:
    f.write("MISMATCH SUMMARY:\n")
    for m in mismatches:
        exp = f"{m['expected']:.2f}" if m['expected'] is not None else "N/A"
        f.write(f"  Row {m['row']:3d} Col {m['col']:2d} [{m['block']:7s}/{m['side']:7s}/{m['account']}] "
                f"Expected={exp:>15s}  Actual={m['actual']:>15.2f}  "
                f"Entity={m['entity'][:25]}  Partner={m['partner'][:25]}\n")

f.close()
print("Done - see tmp_crosscheck_result.txt")
