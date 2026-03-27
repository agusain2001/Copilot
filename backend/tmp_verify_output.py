"""Verify structure of the generated ICM output file."""
import openpyxl

path = r"g:\FCCS\backend\uploads\reports\9\outputs\ICM_Output_TEST.xlsx"
wb   = openpyxl.load_workbook(path, data_only=True)
ws   = wb.active

print(f"Sheet: {ws.title}")
print(f"Max col: {ws.max_column}  Max row: {ws.max_row}")

# Rows 1-31 should be empty
non_empty = []
for r in range(1, 32):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v is not None:
            non_empty.append((r, c, v))
print(f"\nNon-empty cells in rows 1-31: {len(non_empty)}")
if non_empty:
    for r, c, v in non_empty[:5]:
        print(f"  Row {r} Col {c}: {v}")

# Row 29: section labels
print("\nRow 29 (section labels):")
for c in range(1, ws.max_column + 1):
    v = ws.cell(29, c).value
    if v:
        print(f"  Col {c}: {v}")

# Row 32: headers
print("\nRow 32 (headers):")
for c in range(1, ws.max_column + 1):
    v = ws.cell(32, c).value
    if v:
        print(f"  Col {c}: {v}")

# First 2 data rows
print("\nRow 33 (first data row), cols 1-10:")
for c in range(1, 11):
    print(f"  Col {c}: {ws.cell(33, c).value}")
print("\nRow 34 (second data row), cols 1-5:")
for c in range(1, 6):
    print(f"  Col {c}: {ws.cell(34, c).value}")
