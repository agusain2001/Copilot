"""Analyze the Update files to understand what accounts/records exist."""
import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import (
    extract_entity_code_journal, extract_icp_code, extract_account_code,
    is_detail_row, J_ENTITY, J_ACCT, J_ICP, J_DEBIT, J_CREDIT,
    JOURNAL_DATA_START, ICM_INPUT_HEADER_ROW, ICM_INPUT_DATA_START,
    read_icm_headers, read_icm_data, extract_entity_code_icm,
    FIXED_ENT_COLS, to_float
)

print("=" * 80)
print("1. ICM_Output_NEW_RUN.xlsx — Analyzing existing output")
print("=" * 80)
icm_out = r'G:\FCCS\Update files\ICM_Output_NEW_RUN.xlsx'
wb = openpyxl.load_workbook(icm_out, data_only=True)
ws = wb.active

# Find header row — check row 32 first, then scan
header_row = None
for r in [32, 4, 1]:
    v = str(ws.cell(r, 1).value or "").strip().lower()
    if v in ("entity", ""):
        v2 = str(ws.cell(r, 2).value or "").strip().lower()
        if v2 == "partner" or v == "entity":
            header_row = r
            break

if not header_row:
    # scan for entity/partner header
    for r in range(1, 50):
        v = str(ws.cell(r, 1).value or "").strip().lower()
        if v == "entity":
            header_row = r
            break

print(f"Header row: {header_row}")

# Show all headers
if header_row:
    all_accts_in_output = set()
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(header_row, c).value or "").strip()
        if v:
            acct = extract_account_code(v)
            if acct:
                all_accts_in_output.add(acct)
            if c <= 5 or c > ws.max_column - 5:
                print(f"  Col {c}: {v[:70]}")
    print(f"\n  Unique account codes in output headers: {sorted(all_accts_in_output)}")

# Count data rows
data_start = header_row + 1 if header_row else 33
row_count = 0
existing_pairs = set()
for r in range(data_start, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or "").strip()
    par = str(ws.cell(r, 2).value or "").strip()
    if ent or par:
        row_count += 1
        ent_code = extract_entity_code_icm(ent)
        icp_code = extract_icp_code(par)
        if ent_code and icp_code:
            existing_pairs.add((ent_code, icp_code))
print(f"\n  Total data rows: {row_count}")
print(f"  Unique (entity, partner) pairs: {len(existing_pairs)}")

print("\n" + "=" * 80)
print("2. Journals — Analyzing all valid records")
print("=" * 80)

journals = {
    "Parent (J1)":       r'G:\FCCS\Update files\Journal Report.xlsx',
    "Contribution (J2)": r'G:\FCCS\Update files\Journal Report (2).xlsx',
    "Plug (J4)":         r'G:\FCCS\Update files\Journal Report (4).xlsx',
}

all_journal_pairs = set()
all_journal_accounts = set()
all_journal_combos = set()
missing_pairs = set()
missing_combos_by_account = {}

for jname, jpath in journals.items():
    wb_j = openpyxl.load_workbook(jpath, data_only=True)
    ws_j = wb_j.active
    
    pairs = set()
    accounts = set()
    combos = set()
    skipped = 0
    
    for row in ws_j.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws_j.max_row):
        vals = [cell.value for cell in row]
        if str(vals[0] or "").strip() == "Grand Total":
            break
        if not is_detail_row(vals):
            continue
        
        entity_code = extract_entity_code_journal(str(vals[J_ENTITY - 1] or "").strip())
        icp_code = extract_icp_code(str(vals[J_ICP - 1] or "").strip())
        acct_raw = str(vals[J_ACCT - 1] or "").strip()
        acct_code = extract_account_code(acct_raw)
        
        # Validation: all three must be present
        if not entity_code or not icp_code or not acct_code:
            skipped += 1
            continue
        
        pairs.add((entity_code, icp_code))
        accounts.add(acct_code)
        combos.add((entity_code, icp_code, acct_code))
        
        all_journal_pairs.add((entity_code, icp_code))
        all_journal_accounts.add(acct_code)
        all_journal_combos.add((entity_code, icp_code, acct_code))
        
        if (entity_code, icp_code) not in existing_pairs:
            missing_pairs.add((entity_code, icp_code))
            if acct_code not in missing_combos_by_account:
                missing_combos_by_account[acct_code] = []
            missing_combos_by_account[acct_code].append((entity_code, icp_code, jname))
    
    print(f"\n  {jname}:")
    print(f"    Valid entity/ICP pairs: {len(pairs)}")
    print(f"    Unique account codes: {sorted(accounts)}")
    print(f"    Unique (ent, icp, acct) combos: {len(combos)}")
    print(f"    Skipped (missing field): {skipped}")

# Fixed account codes
fixed_codes = {c[0] for c in FIXED_ENT_COLS}
extra_accounts = all_journal_accounts - fixed_codes
print(f"\n{'=' * 80}")
print(f"3. Summary")
print(f"{'=' * 80}")
print(f"  All journal account codes: {sorted(all_journal_accounts)}")
print(f"  Fixed (known 15) codes:    {sorted(fixed_codes)}")
print(f"  EXTRA accounts to add:     {sorted(extra_accounts)}")
print(f"  Missing entity/partner pairs: {len(missing_pairs)}")

if missing_pairs:
    print(f"\n  Sample missing pairs:")
    for p in sorted(missing_pairs)[:15]:
        print(f"    Entity={p[0]}, ICP={p[1]}")
    if len(missing_pairs) > 15:
        print(f"    ... and {len(missing_pairs) - 15} more")

if missing_combos_by_account:
    print(f"\n  Missing combos by account:")
    for acct in sorted(missing_combos_by_account.keys()):
        items = missing_combos_by_account[acct]
        print(f"    Account {acct}: {len(items)} missing combos")
