import openpyxl, sys

try:
    with open('output_check.txt', 'w', encoding='utf-8') as out:
        wb = openpyxl.load_workbook(r'G:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx', data_only=True)
        ws = wb.active

        found = False
        for r in range(33, ws.max_row+1):
            ent = str(ws.cell(r, 1).value or '')
            prt = str(ws.cell(r, 2).value or '')
            if '117100' in ent and '007009' in prt:
                out.write(f'Found E117100 / 007009 at row {r}\n')
                found = True
            elif '007009' in ent and '117100' in prt:
                out.write(f'Found 007009 / E117100 at row {r}\n')
                found = True

        out.write(f'Was it in base ICM? {found}\n')

        # Check the generated output
        wb2 = openpyxl.load_workbook(r'G:\FCCS\backend\uploads\reports\31\output\ICM_Report31_Output.xlsx', data_only=True)
        ws2 = wb2.active
        found_out = False
        for r in range(33, ws2.max_row+1):
            ent = str(ws2.cell(r, 1).value or '')
            prt = str(ws2.cell(r, 2).value or '')
            if '117100' in ent and '007009' in prt:
                out.write(f'Found E117100 / 007009 in OUTPUT at row {r}\n')
                found_out = True
                for c in range(17, 31):
                    val = ws2.cell(r, c).value
                    if val:
                        hdr = str(ws2.cell(32, c).value or '')
                        out.write(f'  Col {c} ({hdr[:40]}): {val}\n')
            elif '007009' in ent and '117100' in prt:
                out.write(f'Found 007009 / E117100 in OUTPUT at row {r}\n')
                found_out = True
                for c in range(17, 31):
                    val = ws2.cell(r, c).value
                    if val:
                        hdr = str(ws2.cell(32, c).value or '')
                        out.write(f'  Col {c} ({hdr[:40]}): {val}\n')

        out.write(f'Was it in output? {found_out}\n')
except Exception as e:
    with open('output_check.txt', 'w', encoding='utf-8') as out:
        out.write(f'Error: {e}\n')
