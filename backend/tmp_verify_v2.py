import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import extract_account_code

path = r'G:\FCCS\Update files\ICM_Output_ENHANCED_v2.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Find all unique account codes and check for Plug_InvSh
target_codes = {'155020', '310001', '315005', 'Plug_InvSh', '189501'}
print("\nSearching for key accounts in header row 32:")
for c in range(1, ws.max_column + 1):
    v = str(ws.cell(32, c).value or '').strip()
    code = extract_account_code(v)
    if code in target_codes:
        print(f"  Col {c}: {code} -> '{v[:80]}'")

# Count all unique account codes  
all_codes = set()
for c in range(1, ws.max_column + 1):
    v = str(ws.cell(32, c).value or '').strip()
    code = extract_account_code(v)
    if code:
        all_codes.add(code)
print(f"\nAll account codes in output: {sorted(all_codes)}")
print(f"Total: {len(all_codes)} unique accounts")
