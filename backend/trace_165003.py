"""Trace all journal entries for account 165003 and how they land in the output."""
import sys, os, re
sys.path.insert(0, '.')
from collections import defaultdict
from app.ic_processor import (
    extract_entity_code_journal, extract_account_code, extract_icp_code,
    to_float, is_detail_row, apply_sign, extract_entity_code_icm,
    JOURNAL_DATA_START, J_ENTITY, J_ACCT, J_ICP, J_DEBIT, J_CREDIT,
    ICM_INPUT_HEADER_ROW, ICM_INPUT_DATA_START,
    FIXED_ENT_COLS, FIXED_PAR_COLS,
    BLK_PAR, BLK_CONT, NUM_ACCTS
)
import openpyxl

BASE = r"G:\FCCS\Update files"

# ── 1. Read ALL journal rows with account 165003 from Parent Journal ──────
print("=" * 100)
print("  ALL JOURNAL ROWS WITH ACCOUNT 165003 IN PARENT JOURNAL (Journal Report.xlsx)")
print("=" * 100)
wb = openpyxl.load_workbook(os.path.join(BASE, "Journal Report.xlsx"), data_only=True)
ws = wb.active

rows_165003 = []
all_rows = []
for row_num, row in enumerate(ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row), start=JOURNAL_DATA_START):
    vals = [cell.value for cell in row]
    label = str(vals[0] or "").strip()
    if label == "Grand Total": break
    if not is_detail_row(vals): continue

    entity_raw = str(vals[J_ENTITY - 1] or "").strip()
    account_raw = str(vals[J_ACCT - 1] or "").strip()
    icp_raw = str(vals[J_ICP - 1] or "").strip()
    entity_code = extract_entity_code_journal(entity_raw)
    account_code = extract_account_code(account_raw)
    icp_code = extract_icp_code(icp_raw)
    debit = to_float(vals[J_DEBIT - 1])
    credit = to_float(vals[J_CREDIT - 1])
    is_group = bool(re.match(r"^E\d+", entity_code))
    net = apply_sign(debit, credit, account_code)

    entry = {
        "row": row_num,
        "entity_raw": entity_raw[:70],
        "entity_code": entity_code,
        "account_raw": account_raw[:70],
        "account_code": account_code,
        "icp_raw": icp_raw[:70],
        "icp_code": icp_code,
        "debit": debit,
        "credit": credit,
        "net": net,
        "is_group": is_group,
    }
    all_rows.append(entry)
    if account_code == "165003":
        rows_165003.append(entry)

print("  Found {} rows with account 165003 out of {} total detail rows".format(len(rows_165003), len(all_rows)))
print()

for r in rows_165003:
    grp = " [GROUP]" if r["is_group"] else ""
    print("  Row {}: Entity={} ({}){}".format(r["row"], r["entity_code"], r["entity_raw"][:50], grp))
    print("          Account={} ({})".format(r["account_code"], r["account_raw"][:50]))
    print("          ICP={} ({})".format(r["icp_code"], r["icp_raw"][:50]))
    print("          Debit={:>20,.2f}  Credit={:>20,.2f}  Net={:>20,.2f}".format(r["debit"], r["credit"], r["net"]))
    print("          Key: ({}, {}, 165003)".format(r["entity_code"], r["icp_code"]))
    print()

# Group by (entity, icp) key
key_groups = defaultdict(list)
for r in rows_165003:
    key_groups[(r["entity_code"], r["icp_code"])].append(r)

print("  GROUPED BY (entity, icp) KEY:")
print("  " + "-" * 80)
for (ent, icp), rows in sorted(key_groups.items()):
    total_net = sum(r["net"] for r in rows)
    is_group = rows[0]["is_group"]
    grp = " [GROUP/FALLBACK]" if is_group else " [PRIMARY]"
    print("  ({}, {}, 165003){} = {:>20,.2f}  ({} rows merged)".format(
        ent, icp, grp, total_net, len(rows)))
    for r in rows:
        print("    Row {}: D={:>15,.2f} C={:>15,.2f} Net={:>15,.2f}".format(
            r["row"], r["debit"], r["credit"], r["net"]))
    print()

# ── 2. Read ICM rows for relevant entities ──────
print("=" * 100)
print("  ICM ROWS RELEVANT TO 001033 / E118000 / 001051")
print("=" * 100)
wb_icm = openpyxl.load_workbook(os.path.join(BASE, "Intercompany Balances IC Matching Report (1).xlsx"), data_only=True)
candidates = []
for sheet in wb_icm.worksheets:
    c1 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 1).value or "").strip()
    c2 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 2).value or "").strip()
    if c1.lower() == "entity" and c2.lower() == "partner":
        candidates.append(sheet)
ws_icm = min(candidates, key=lambda s: s.max_column) if candidates else wb_icm.active

for row_num, row in enumerate(
    ws_icm.iter_rows(min_row=ICM_INPUT_DATA_START, max_row=ws_icm.max_row), start=ICM_INPUT_DATA_START):
    vals = [cell.value for cell in row]
    entity_raw = str(vals[0] or "").strip()
    partner_raw = str(vals[1] or "").strip() if len(vals) > 1 else ""
    entity_code = extract_entity_code_icm(entity_raw)

    # Show rows where entity or partner involves 001033
    if "001033" in entity_raw or "001033" in partner_raw or \
       "001051" in entity_raw or "Cube" in entity_raw or \
       entity_raw.startswith("001033") or entity_code == "001033":
        print("  ICM Row {}: Entity='{}' | Partner='{}'".format(
            row_num, entity_raw[:60], partner_raw[:60]))

print()

# ── 3. Trace matching logic ──────
print("=" * 100)
print("  MATCHING TRACE")
print("=" * 100)

# For the Parent block, the processor does:
# Direct match: primary_key = (icm_entity, icm_partner, account)
# Vice-versa: primary_key = (icm_partner_stripped, ICP_icm_entity, account)
# Fallback: any group entry where fb_key[1] == icm_partner

print()
print("  How journal entries for 165003 SHOULD be placed:")
print("  " + "-" * 80)

for (ent, icp), rows in sorted(key_groups.items()):
    total_net = sum(r["net"] for r in rows)
    if total_net == 0:
        print("  ({}, {}) net=0 -> CANCELLED, not placed".format(ent, icp))
        continue
    is_group = rows[0]["is_group"]

    if not is_group:
        # PRIMARY: placed in ICM row where (icm_entity=ent AND icm_partner=icp)
        # OR vice-versa: (icm_entity=icp_stripped AND icm_partner=ICP_ent)
        icp_stripped = icp[4:] if icp.startswith("ICP_") else icp
        ent_as_icp = "ICP_{}".format(ent)

        print("  PRIMARY ({}, {}, 165003) net={:,.2f}".format(ent, icp, total_net))
        print("    -> Direct match: ICM row with entity={} partner={}".format(ent, icp))
        print("    -> Vice-versa:   ICM row with entity={} partner={}".format(icp_stripped, ent_as_icp))

        if icp == "":
            print("    -> ICP is EMPTY! Direct match requires ICM row (entity={}, partner='')".format(ent))
            print("       This matches ICM rows where 001033 has NO partner")
    else:
        # FALLBACK: placed in FIRST ICM row where icm_partner == icp (consumed once)
        print("  FALLBACK/GROUP ({}, {}, 165003) net={:,.2f}".format(ent, icp, total_net))
        print("    -> Matches first ICM row where partner={}".format(icp))

print()

# ── 4. Check what entity code "001033" rows in journal have ──────
print("=" * 100)
print("  ALL JOURNAL ROWS WHERE ENTITY CONTAINS '001033' (any account)")
print("=" * 100)
rows_001033 = [r for r in all_rows if "001033" in r["entity_code"]]
print("  Found {} rows for entity 001033".format(len(rows_001033)))

# Group by (entity, icp, account)
by_key = defaultdict(list)
for r in rows_001033:
    by_key[(r["entity_code"], r["icp_code"], r["account_code"])].append(r)

for (e, i, a), rows in sorted(by_key.items()):
    total = sum(r["net"] for r in rows)
    in_cols = "YES" if a in set(c[0] for c in FIXED_ENT_COLS) else "no"
    print("  ({}, {}, {}) -> {} rows, net={:>15,.2f}  in_columns={}".format(
        e, i, a, len(rows), total, in_cols))
