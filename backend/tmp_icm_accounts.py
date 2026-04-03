"""Analyze ICM source to get all account columns — these should drive the output."""
import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl, re
from app.ic_processor import extract_account_code, detect_icm_header_row

icm_path = r'G:\FCCS\Update files\Intercompany Balances IC Matching Report (1).xlsx'
wb = openpyxl.load_workbook(icm_path, data_only=True)

for sheet in wb.worksheets:
    hdr_row, data_start = detect_icm_header_row(sheet)
    if hdr_row == 4 and data_start == 5:
        # fallback defaults = not found
        continue
    
    print(f"\nSheet: '{sheet.title}' (max_col={sheet.max_column})")
    print(f"Header row: {hdr_row}")
    
    ent_accts = []  # Entity-tagged accounts
    par_accts = []  # Partner-tagged accounts
    
    for c in range(1, sheet.max_column + 1):
        v = str(sheet.cell(hdr_row, c).value or "").strip()
        if not v:
            continue
        code = extract_account_code(v)
        if not code:
            continue
        is_partner = bool(re.search(r'\bPartner\b', v))
        tag = "Partner" if is_partner else "Entity"
        
        if tag == "Entity":
            ent_accts.append((code, v, c))
        else:
            par_accts.append((code, v, c))
    
    print(f"\nEntity-tagged accounts ({len(ent_accts)}):")
    for code, desc, col in ent_accts:
        print(f"  Col {col:3d}: {code:15s} | {desc[:70]}")
    
    print(f"\nPartner-tagged accounts ({len(par_accts)}):")
    for code, desc, col in par_accts:
        print(f"  Col {col:3d}: {code:15s} | {desc[:70]}")
    
    all_codes = sorted(set(c[0] for c in ent_accts) | set(c[0] for c in par_accts))
    print(f"\nAll unique account codes in ICM: {all_codes}")
    print(f"Total unique: {len(all_codes)}")

# Also check Sheet1 (2) which may have more accounts
print("\n" + "=" * 80)
ws2 = wb['Sheet1 (2)']
hdr_row2, _ = detect_icm_header_row(ws2)
print(f"\nSheet1 (2): max_col={ws2.max_column}, header_row={hdr_row2}")
all_accts_2 = []
for c in range(1, ws2.max_column + 1):
    v = str(ws2.cell(hdr_row2, c).value or "").strip()
    if not v:
        continue
    code = extract_account_code(v)
    if code:
        is_partner = bool(re.search(r'\bPartner\b', v))
        tag = "Partner" if is_partner else "Entity"
        all_accts_2.append((code, tag, v[:70], c))

print(f"\nAll accounts in Sheet1 (2) ({len(all_accts_2)}):")
for code, tag, desc, col in all_accts_2:
    print(f"  Col {col:3d}: [{tag:7s}] {code:15s} | {desc}")

all_codes_2 = sorted(set(c[0] for c in all_accts_2))
print(f"\nUnique account codes: {all_codes_2}")
print(f"Total unique: {len(all_codes_2)}")
