"""
Run the full processor and check the output for entity-side vs partner-side values.
"""
import sys, os
sys.path.insert(0, '.')
import logging
logging.basicConfig(level=logging.WARNING)

from app.ic_processor import process_icm_report

icm_path = r'uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx'
journal_paths = {
    'parent_journal': r'uploads\reports\31\inputs\Parent report.xlsx',
    'contribution_journal': r'uploads\reports\31\inputs\Contribution report.xlsx',
}
report_inputs = r'uploads\reports\31\inputs\report Inputs.xlsx'
output_path = r'uploads\reports\31\outputs\ICM_Output_31_test.xlsx'

result = process_icm_report(icm_path, journal_paths, output_path, report_inputs)

# Analyze the output
import openpyxl
wb = openpyxl.load_workbook(output_path, data_only=True)
ws = wb.active

with open('tmp_output_analysis.txt', 'w', encoding='utf-8') as f:
    f.write("=== OUTPUT ANALYSIS ===\n")
    f.write(f"Output: {output_path}\n")
    f.write(f"Max rows: {ws.max_row}, Max cols: {ws.max_column}\n\n")
    
    # Header row
    f.write("Header row 32:\n")
    for c in range(1, min(ws.max_column + 1, 50)):
        v = str(ws.cell(32, c).value or '').strip()
        if v:
            f.write(f"  Col {c}: {v[:60]}\n")
    
    f.write("\n=== ROWS WITH DATA ===\n")
    for r in range(33, ws.max_row + 1):
        ent = str(ws.cell(r, 1).value or '').strip()
        prt = str(ws.cell(r, 2).value or '').strip()
        if not ent and not prt:
            continue
        
        has_data = False
        for c in range(3, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and v != 0 and v != '':
                has_data = True
                break
        
        # Count entity-side values (cols 3 to first Variance)
        # Count partner-side values (cols after first Variance)
        ent_vals = []
        par_vals = []
        all_vals = []
        for c in range(3, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and v != 0 and v != '':
                all_vals.append((c, v))
        
        status = "HAS DATA" if has_data else "EMPTY"
        f.write(f"\nRow {r}: [{status}]\n")
        f.write(f"  Entity: {ent[:40]}\n")
        f.write(f"  Partner: {prt[:40]}\n")
        if all_vals:
            for c, v in all_vals[:15]:
                hdr = str(ws.cell(32, c).value or '').strip()[:40]
                f.write(f"  Col {c:3d} ({hdr}): {v}\n")
            if len(all_vals) > 15:
                f.write(f"  ... and {len(all_vals) - 15} more values\n")

print("Done - see tmp_output_analysis.txt")
