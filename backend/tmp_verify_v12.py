"""Verify the fix: check rows with Entity=001001/001033 and Partner=ICP_001033."""
import openpyxl

path = r'G:\FCCS\Update files\ICM_Output_ENHANCED_v12.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

print("=== Checking rows with 001001/001033 pair ===\n")

for r in range(33, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or '').strip()
    par = str(ws.cell(r, 2).value or '').strip()
    
    if ('001001' in ent and '001033' in par) or \
       ('001033' in ent and '001001' in par) or \
       ('001032' in ent and '001033' in par) or \
       ('001033' in ent and '001032' in par):
        # Parent block: cols 38-73
        parent_vals = []
        for c in range(38, 74):
            v = ws.cell(r, c).value
            if v is not None and v != 0:
                hdr = str(ws.cell(32, c).value or '')[:50]
                parent_vals.append(f"  Col {c} [{hdr}]: {v}")
        
        print(f"Row {r}: Entity={ent[:35]}  Partner={par[:35]}")
        if parent_vals:
            print(f"  Parent Input values:")
            for pv in parent_vals:
                print(f"    {pv}")
        else:
            print(f"  Parent Input: (empty)")
        print()
