"""Analyze IC processor output - detailed breakdown per journal."""
import sys, os, re
sys.path.insert(0, '.')
from collections import defaultdict
from app.ic_processor import (
    read_journal_report, read_icm_data, read_icm_headers, parse_report_inputs,
    match_journal_to_icm, extract_entity_code_journal, extract_account_code,
    extract_icp_code, to_float, is_detail_row, apply_sign, classify_account,
    JOURNAL_DATA_START, J_ENTITY, J_ACCT, J_ICP, J_DEBIT, J_CREDIT,
    ICM_INPUT_HEADER_ROW, FIXED_ENT_COLS, FIXED_PAR_COLS,
    BLK_BASE, BLK_PAR, BLK_CONT, COL_PLUG, NUM_ACCTS
)
import openpyxl

BASE = r"G:\FCCS\Update files"
icm_path = os.path.join(BASE, "Intercompany Balances IC Matching Report (1).xlsx")
report_inputs_path = r"G:\FCCS\AI\report_inputs.xlsx"

# Parse report_inputs for plug mapping
plug_mapping = parse_report_inputs(report_inputs_path)
plug_code = plug_mapping.get("plug_code") if plug_mapping else None
elim_codes_ri = plug_mapping.get("elim_codes", set()) if plug_mapping else set()

# Known elimination account codes (from FIXED_ENT_COLS)
fixed_acct_codes = set(c[0] for c in FIXED_ENT_COLS)

# Read ICM data
wb_icm = openpyxl.load_workbook(icm_path, data_only=True)
candidates = []
for sheet in wb_icm.worksheets:
    c1 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 1).value or "").strip()
    c2 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 2).value or "").strip()
    if c1.lower() == "entity" and c2.lower() == "partner":
        candidates.append(sheet)
ws_icm = min(candidates, key=lambda s: s.max_column) if candidates else wb_icm.active
data_rows = read_icm_data(ws_icm)
icm_header_map = read_icm_headers(ws_icm)

# Build ICM entity/partner lookup
icm_entities = set()
icm_partners = set()
icm_pairs = set()
for dr in data_rows:
    icm_entities.add(dr["entity_code"])
    icm_partners.add(dr["partner_code"])
    icm_pairs.add((dr["entity_code"], dr["partner_code"]))

print("ICM Data: {} entity-partner rows".format(len(data_rows)))
print("Plug code: {}".format(plug_code))
print("Elimination codes from report_inputs: {}".format(sorted(elim_codes_ri)))
print("Fixed account codes (15 columns): {}".format(sorted(fixed_acct_codes)))
print()

# ============================================================
# Analyze each journal
# ============================================================
journals = [
    ("Journal 1 (Parent)", os.path.join(BASE, "Journal Report.xlsx"), None),
    ("Journal 2 (Contribution)", os.path.join(BASE, "Journal Report (2).xlsx"), None),
    ("Journal 4 (Plug Account)", os.path.join(BASE, "Journal Report (4).xlsx"), plug_mapping),
]

for jname, jpath, jmap in journals:
    print("=" * 90)
    print("  {} -- {}".format(jname, os.path.basename(jpath)))
    print("=" * 90)

    wb = openpyxl.load_workbook(jpath, data_only=True)
    ws = wb.active

    detail_rows_j = []
    for row_num, row in enumerate(
        ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row),
        start=JOURNAL_DATA_START,
    ):
        vals = [cell.value for cell in row]
        label = str(vals[0] or "").strip()
        if label == "Grand Total":
            break
        if not is_detail_row(vals):
            continue

        entity_raw = str(vals[J_ENTITY - 1] or "").strip()
        account_raw = str(vals[J_ACCT - 1] or "").strip()
        icp_raw = str(vals[J_ICP - 1] or "").strip()
        entity_code = extract_entity_code_journal(entity_raw)
        account_code = extract_account_code(account_raw)
        icp_code = extract_icp_code(icp_raw)
        debit = to_float(vals[J_DEBIT - 1])
        credit = to_float(vals[J_CREDIT - 1])
        is_group = bool(re.match(r"^E\d+", entity_code))

        # What the account code becomes after plug mapping
        mapped_code = account_code
        if jmap:
            pc = jmap.get("plug_code")
            ec = jmap.get("elim_codes", set())
            if account_code in ec or not account_code[:1].isdigit():
                mapped_code = pc

        is_elim = account_code in fixed_acct_codes

        detail_rows_j.append({
            "row": row_num,
            "entity_code": entity_code,
            "entity_raw": entity_raw[:60],
            "account_code": account_code,
            "mapped_code": mapped_code,
            "icp_code": icp_code,
            "icp_raw": icp_raw[:60],
            "debit": debit,
            "credit": credit,
            "is_group": is_group,
            "is_elim": is_elim,
            "net": apply_sign(debit, credit, mapped_code if jmap else account_code),
        })

    total_detail = len(detail_rows_j)
    elim_rows = [r for r in detail_rows_j if r["is_elim"]]
    non_elim_rows = [r for r in detail_rows_j if not r["is_elim"]]

    print("  Total detail rows:        {}".format(total_detail))

    if jmap:
        # For plug journal, show mapping
        non_numeric_rows = [r for r in detail_rows_j if not r["account_code"][:1].isdigit()]
        plug_mapped = [r for r in detail_rows_j if r["mapped_code"] == plug_code]
        acct_counts = defaultdict(int)
        for r in detail_rows_j:
            acct_counts[r["account_code"]] += 1
        elim_in_ri = [r for r in detail_rows_j if r["account_code"] in elim_codes_ri]
        print("  Elimination account rows: {}   ALL are: {}".format(
            len(elim_in_ri),
            ", ".join(sorted(set(r["account_code"] for r in elim_in_ri))) if elim_in_ri else "none"
        ))
        non_col = {a: c for a, c in acct_counts.items() if a not in fixed_acct_codes}
        if non_col:
            parts = ["{} x{}".format(a, c) for a, c in sorted(non_col.items())]
            print("  Non-column accounts:      {}".format(", ".join(parts)))
        print("  All mapped to plug code:  {}   (via plug_mapping -> {})".format(len(plug_mapped), plug_code))
    else:
        elim_acct_codes_found = sorted(set(r["account_code"] for r in elim_rows))
        print("  Elimination account rows: {}    Accounts: {}".format(
            len(elim_rows), ", ".join(elim_acct_codes_found)
        ))
        non_elim_acct_counts = defaultdict(int)
        for r in non_elim_rows:
            non_elim_acct_counts[r["account_code"]] += 1
        if non_elim_acct_counts:
            parts = ["{} x{}".format(a, c) for a, c in sorted(non_elim_acct_counts.items())]
            print("  Non-elimination rows:     {}    {} -- not in the 15 output columns".format(
                len(non_elim_rows), ", ".join(parts)
            ))

    # Build unique keys and check matching
    primary_keys = defaultdict(float)
    fallback_keys = defaultdict(float)
    for r in detail_rows_j:
        acct = r["mapped_code"] if jmap else r["account_code"]
        # Skip if account not in fixed columns AND not plug code
        if acct not in fixed_acct_codes and acct != plug_code:
            continue
        net = r["net"]
        key = (r["entity_code"], r["icp_code"], acct)
        if r["is_group"]:
            fallback_keys[key] += net
        else:
            primary_keys[key] += net

    # Remove zero-net keys
    primary_keys = {k: v for k, v in primary_keys.items() if v != 0}
    fallback_keys = {k: v for k, v in fallback_keys.items() if v != 0}

    total_unique = len(primary_keys) + len(fallback_keys)

    # Show merging info
    all_raw_keys_pri = defaultdict(int)
    all_raw_keys_fb = defaultdict(int)
    for r in detail_rows_j:
        acct = r["mapped_code"] if jmap else r["account_code"]
        if acct not in fixed_acct_codes and acct != plug_code:
            continue
        key = (r["entity_code"], r["icp_code"], acct)
        if r["is_group"]:
            all_raw_keys_fb[key] += 1
        else:
            all_raw_keys_pri[key] += 1

    merged_keys = {k: v for k, v in all_raw_keys_pri.items() if v > 1}
    merged_keys.update({k: v for k, v in all_raw_keys_fb.items() if v > 1})
    cancelled = []
    for k in list(all_raw_keys_pri.keys()) + list(all_raw_keys_fb.keys()):
        if k in primary_keys or k in fallback_keys:
            continue
        # This key cancelled out to zero
        cancelled.append(k)

    merge_note = ""
    if merged_keys:
        for mk, mc in merged_keys.items():
            merge_note += "    {} rows share key ({}) -> merged to 1\n".format(mc, mk)
    if cancelled:
        merge_note += "    {} keys cancelled out (D=C, net=0)\n".format(len(cancelled))
        for ck in cancelled[:5]:
            merge_note += "      cancelled: {}\n".format(ck)

    print("  Unique matched entries:   {}   (primary: {}, group/fallback: {})".format(
        total_unique, len(primary_keys), len(fallback_keys)
    ))
    if merge_note:
        print(merge_note.rstrip())

    # Check which ones actually land in the output
    matched_in_output = 0
    unmatched = []

    for (e, p, a), net in primary_keys.items():
        found = False
        # Direct entity match
        for dr in data_rows:
            if dr["entity_code"] == e and dr["partner_code"] == p:
                found = True
                break
        # Vice-versa partner match
        if not found:
            p_stripped = p[4:] if p.startswith("ICP_") else p
            e_as_icp = "ICP_{}".format(e)
            for dr in data_rows:
                if dr["entity_code"] == p_stripped and dr["partner_code"] == e_as_icp:
                    found = True
                    break
        # Fallback: partner match only (for plug)
        if not found and a == plug_code:
            for dr in data_rows:
                if dr["partner_code"] == p:
                    found = True
                    break
        if found:
            matched_in_output += 1
        else:
            unmatched.append((e, p, a, net))

    for (ge, p, a), net in fallback_keys.items():
        found = False
        for dr in data_rows:
            if dr["partner_code"] == p:
                found = True
                break
        if found:
            matched_in_output += 1
        else:
            unmatched.append((ge, p, a, net))

    print("  In output:                {}".format(matched_in_output))

    if unmatched:
        print("  MISSING from output:      {}".format(len(unmatched)))
        for (e, p, a, net) in unmatched:
            print("    {} / {} / {} = {:,.2f}".format(e, p, a, net))
            e_in_icm = e in icm_entities
            p_in_icm = p in icm_partners
            if not e_in_icm and not p_in_icm:
                print("      -> Neither entity {} nor partner {} found in ICM".format(e, p))
            elif not e_in_icm:
                close = [ie for ie in icm_entities if e in ie or ie in e]
                print("      -> Entity {} NOT in ICM entities. Close: {}".format(e, close[:5]))
            elif not p_in_icm:
                print("      -> Partner {} NOT in ICM partners".format(p))
            elif (e, p) not in icm_pairs:
                print("      -> Pair ({}, {}) not in ICM -- entity and partner exist separately".format(e, p))

    print()

# ============================================================
# Output file verification
# ============================================================
print("=" * 90)
print("  OUTPUT FILE VERIFICATION -- ICM_Output_NEW.xlsx")
print("=" * 90)
wb_out = openpyxl.load_workbook(r"G:\FCCS\Update files\ICM_Output_NEW.xlsx", data_only=True)
ws_out = wb_out.active

parent_ent = 0; parent_par = 0
contrib_ent = 0; contrib_par = 0
plug_count = 0
base_ent = 0; base_par = 0

output_rows = 0
for rn in range(33, ws_out.max_row + 1):
    entity = ws_out.cell(rn, 1).value
    if not entity:
        break
    output_rows += 1

    # Base block
    for i in range(NUM_ACCTS):
        v = ws_out.cell(rn, BLK_BASE[0] + i).value
        if v is not None and v != 0:
            base_ent += 1
    for i in range(NUM_ACCTS):
        v = ws_out.cell(rn, BLK_BASE[2] + i).value
        if v is not None and v != 0:
            base_par += 1

    # Parent block
    for i in range(NUM_ACCTS):
        v = ws_out.cell(rn, BLK_PAR[0] + i).value
        if v is not None and v != 0:
            parent_ent += 1
    for i in range(NUM_ACCTS):
        v = ws_out.cell(rn, BLK_PAR[2] + i).value
        if v is not None and v != 0:
            parent_par += 1

    # Contribution block
    for i in range(NUM_ACCTS):
        v = ws_out.cell(rn, BLK_CONT[0] + i).value
        if v is not None and v != 0:
            contrib_ent += 1
    for i in range(NUM_ACCTS):
        v = ws_out.cell(rn, BLK_CONT[2] + i).value
        if v is not None and v != 0:
            contrib_par += 1

    # Plug
    v = ws_out.cell(rn, COL_PLUG).value
    if v is not None and v != 0:
        plug_count += 1

print("  Output rows:               {}".format(output_rows))
print("  Base Input values:         {} (entity-side: {}, partner-side: {})".format(
    base_ent + base_par, base_ent, base_par))
print("  Parent Input (J1) values:  {} (entity-side: {}, partner-side: {})".format(
    parent_ent + parent_par, parent_ent, parent_par))
print("  Contrib Input (J2) values: {} (entity-side: {}, partner-side: {})".format(
    contrib_ent + contrib_par, contrib_ent, contrib_par))
print("  Plug Account (J4) values:  {}".format(plug_count))
print("  ----------------------------------------")
print("  TOTAL JOURNAL VALUES:      {}".format(parent_ent + parent_par + contrib_ent + contrib_par + plug_count))
