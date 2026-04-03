import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl

# Check report 22's source ICM header position
print("Report 22 source ICM:")
wb22 = openpyxl.load_workbook(r'G:\FCCS\backend\uploads\reports\22\inputs\Intercompany Balances IC Matching Report (1).xlsx', data_only=True)
for sheet in wb22.worksheets:
    for r in [4, 32]:
        v1 = str(sheet.cell(r, 1).value or '').strip()
        v2 = str(sheet.cell(r, 2).value or '').strip()
        if v1.lower() == 'entity':
            print(f'  Sheet "{sheet.title}" - Entity/Partner at row {r}')
            for c in range(1, min(6, sheet.max_column + 1)):
                v = str(sheet.cell(r, c).value or '').strip()
                if v: print(f'    Col {c}: {v[:60]}')
            break

# Check Update files source ICM
print("\nUpdate files source ICM:")
wb_upd = openpyxl.load_workbook(r'G:\FCCS\Update files\Intercompany Balances IC Matching Report (1).xlsx', data_only=True)
for sheet in wb_upd.worksheets:
    for r in [4, 32]:
        v1 = str(sheet.cell(r, 1).value or '').strip()
        v2 = str(sheet.cell(r, 2).value or '').strip()
        if v1.lower() == 'entity':
            print(f'  Sheet "{sheet.title}" (max_col={sheet.max_column}) - Entity/Partner at row {r}')
            for c in range(1, min(8, sheet.max_column + 1)):
                v = str(sheet.cell(r, c).value or '').strip()
                if v: print(f'    Col {c}: {v[:60]}')
            # Count data rows
            cnt = 0
            for r2 in range(r + 1, sheet.max_row + 1):
                e = str(sheet.cell(r2, 1).value or '').strip()
                p = str(sheet.cell(r2, 2).value or '').strip()
                if e and p: cnt += 1
            print(f'    Data rows: {cnt}')
            break
