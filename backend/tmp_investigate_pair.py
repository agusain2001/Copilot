"""Investigate the 001033/001032 pair mismatch."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
from app.ic_processor import (
    read_journal_report, normalize_to_numeric, extract_account_code,
    extract_entity_code_journal, extract_icp_code, get_journal_indices,
    JOURNAL_DATA_START, is_detail_row, apply_sign, to_float
)

# Check what journal entries exist for 001032/001033
parent = r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx"
contrib = r"g:\FCCS\backend\uploads\reports\31\inputs\Contribution report.xlsx"

for jname, jpath in [("PARENT", parent), ("CONTRIB", contrib)]:
    print(f"\n=== {jname} JOURNAL ===")
    wb = openpyxl.load_workbook(jpath, data_only=True)
    ws = wb.active
    indices = get_journal_indices(ws)
    
    for row in ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row):
        vals = [cell.value for cell in row]
        if not vals: continue
        if str(vals[0] or "").strip() == "Grand Total": break
        if not is_detail_row(vals, indices): continue
        
        entity_raw = str(vals[indices["entity"]] or "").strip()
        icp_raw = str(vals[indices["icp"]] or "").strip()
        acct_raw = str(vals[indices["acct"]] or "").strip()
        
        ent_code = extract_entity_code_journal(entity_raw)
        ent_num = normalize_to_numeric(ent_code)
        icp_code = extract_icp_code(icp_raw)
        icp_num = normalize_to_numeric(icp_code)
        acct = extract_account_code(acct_raw)
        
        if ent_num in ("001032", "001033") or icp_num in ("001032", "001033"):
            debit = to_float(vals[indices["debit"]])
            credit = to_float(vals[indices["credit"]])
            net = apply_sign(debit, credit, acct)
            print(f"  Entity={ent_num} ICP={icp_num} Acct={acct} D={debit} C={credit} Net={net}")
            print(f"    Raw: entity='{entity_raw[:50]}' icp='{icp_raw[:50]}' acct='{acct_raw[:50]}'")

# Check what the lookup produces
print("\n=== LOOKUP KEYS FOR 001033 ===")
lookup = read_journal_report(parent)
for (e, p, a), lines in lookup.items():
    if e in ("001032", "001033") or p in ("001032", "001033"):
        net = sum(apply_sign(l["debit"], l["credit"], a) for l in lines)
        print(f"  ({e}, {p}, {a}) -> net={net}, lines={len(lines)}")
