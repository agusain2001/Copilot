"""Check what 430015 entries are in Journal Report (5) and whether they match ICM rows."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from app.ic_processor import read_journal_report, get_journal_indices, JOURNAL_DATA_START
import openpyxl

# Parent journal = Journal Report (5)
jpath = r"g:\FCCS\backend\uploads\reports\26\inputs\Journal Report (5).xlsx"
wb = openpyxl.load_workbook(jpath, data_only=True)
ws = wb.active

# Check the header row for column structure
indices = get_journal_indices(ws)
print(f"Journal indices: {indices}")
print(f"Header row ({JOURNAL_DATA_START - 1}):")
header_row = JOURNAL_DATA_START - 1
for c in range(ws.max_column):
    v = ws.cell(header_row, c + 1).value
    if v:
        print(f"  Col {c}: '{str(v).strip()}'")

# Read with no plug mapping
primary, fallback = read_journal_report(jpath)
print(f"\nTotal primary keys: {len(primary)}")
print(f"Total fallback keys: {len(fallback)}")

# Find 430015 entries
print(f"\n=== 430015 journal entries ===")
for (e, icp, acct), lines in primary.items():
    if acct == "430015":
        for l in lines:
            print(f"  Entity={e}, ICP={icp}, Debit={l['debit']}, Credit={l['credit']}")

for (e, icp, acct), lines in fallback.items():
    if acct == "430015":
        for l in lines:
            print(f"  FALLBACK: Entity={e}, ICP={icp}, Debit={l['debit']}, Credit={l['credit']}")

# Check if ICM has entity 013011 / partner matching any of these
icm_path = r"g:\FCCS\backend\uploads\reports\26\inputs\Intercompany Report_1150_Intercompany Report.xlsx"
from app.ic_processor import detect_icm_header_row, read_icm_data
wb_icm = openpyxl.load_workbook(icm_path, data_only=True)
ws_icm = wb_icm.active
hdr_row, data_start = detect_icm_header_row(ws_icm)

# Get the narrower sheet
candidates = []
for sheet in wb_icm.worksheets:
    h, d = detect_icm_header_row(sheet)
    c1 = str(sheet.cell(h, 1).value or "").strip().lower()
    c2 = str(sheet.cell(h, 2).value or "").strip().lower()
    if c1 == "entity" and c2 == "partner":
        candidates.append((sheet, h, d))

ws_icm, hdr_row, data_start = min(candidates, key=lambda t: t[0].max_column) if candidates else (wb_icm.active, hdr_row, data_start)

data_rows = read_icm_data(ws_icm, data_start)
print(f"\nICM rows with entity 013011:")
for r in data_rows:
    if r["entity_code"] == "013011":
        print(f"  Entity={r['entity']}, Partner={r['partner']}, eCode={r['entity_code']}, pCode={r['partner_code']}")
