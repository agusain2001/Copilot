"""Clean final verification — output to file."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl, logging
logging.basicConfig(level=logging.WARNING)

ref_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_FIXED.xlsx"
output_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_VERIFY2.xlsx"
icm_path = r"g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx"
journal_paths = {
    "parent_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx",
    "contribution_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Contribution report.xlsx",
    "plugaccount_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Journal Report (4).xlsx",
}
report_inputs_path = r"g:\FCCS\backend\uploads\reports\31\inputs\report Inputs.xlsx"

from app.ic_processor import process_icm_report
process_icm_report(icm_path, journal_paths, output_path, report_inputs_path)

wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
ws_ref = wb_ref.active
wb_new = openpyxl.load_workbook(output_path, data_only=True)
ws_new = wb_new.active

def to_num(v):
    if v is None or str(v).strip() in ("", " "): return 0.0
    try: return round(float(v), 2)
    except: return v

def pair_map(ws):
    m = {}
    for r in range(33, ws.max_row + 1):
        e = str(ws.cell(r, 1).value or "").strip()
        p = str(ws.cell(r, 2).value or "").strip()
        if e or p: m[(e, p)] = r
    return m

ref_pm = pair_map(ws_ref)
new_pm = pair_map(ws_new)
common = set(ref_pm) & set(new_pm)
only_ref = set(ref_pm) - set(new_pm)
only_new = set(new_pm) - set(ref_pm)

out = open(r"g:\FCCS\backend\tmp_final_verify_clean.txt", "w", encoding="utf-8")

out.write(f"REF: {ws_ref.max_row} rows x {ws_ref.max_column} cols\n")
out.write(f"NEW: {ws_new.max_row} rows x {ws_new.max_column} cols\n")
out.write(f"REF data rows: {len(ref_pm)}\n")
out.write(f"NEW data rows: {len(new_pm)}\n")
out.write(f"Common: {len(common)}, Only REF: {len(only_ref)}, Only NEW: {len(only_new)}\n\n")

if only_ref:
    out.write(f"--- ONLY IN REF ({len(only_ref)}) ---\n")
    for p in sorted(only_ref):
        out.write(f"  {p}\n")

if only_new:
    out.write(f"\n--- ONLY IN NEW ({len(only_new)}) ---\n")
    for p in sorted(only_new):
        out.write(f"  {p}\n")

# Value comparison
out.write(f"\n--- VALUE COMPARISON (common={len(common)}) ---\n")
mismatches = 0
for pair in sorted(common):
    rr = ref_pm[pair]
    nr = new_pm[pair]
    for c in range(3, max(ws_ref.max_column, ws_new.max_column) + 1):
        rv = to_num(ws_ref.cell(rr, c).value)
        nv = to_num(ws_new.cell(nr, c).value)
        if rv != nv:
            mismatches += 1
            hdr = str(ws_ref.cell(32, c).value or f"Col{c}")[:60]
            out.write(f"  MISMATCH [{pair[0][:35]} | {pair[1][:35]}]\n")
            out.write(f"    Col {c} ({hdr}): REF={rv} NEW={nv}\n")

out.write(f"\nTotal mismatches: {mismatches}\n")
out.close()
print("Done. See tmp_final_verify_clean.txt")
