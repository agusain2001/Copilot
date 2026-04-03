import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl

path = r'G:\FCCS\Update files\ICM_Output_ENHANCED_v8.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

# Check headers (Condition 1 & 3)
headers_r32 = [str(ws.cell(32, c).value or '').strip() for c in range(1, ws.max_column + 1)]
headers_r29 = [str(ws.cell(29, c).value or '').strip() for c in range(1, ws.max_column + 1)]

print(f"Total Columns: {len(headers_r32)}")

# Find indices for Parent Input block
parent_start = headers_r29.index("Parent Input") + 1
contrib_start = headers_r29.index("Contribution Input") + 1
plug_start = headers_r29.index("Plug Account") + 1

print(f"\nParent Block from col {parent_start} to {contrib_start - 1}")
parent_cols = headers_r32[parent_start-1 : contrib_start-1]
print(f"Parent Block Ending columns:")
for i, h in enumerate(parent_cols[-5:]):
    print(f"  {len(parent_cols)-5+i}: {h[:50]}")

print(f"\nContrib Block from col {contrib_start} to {plug_start - 1}")
contrib_cols = headers_r32[contrib_start-1 : plug_start-1]
print(f"Contrib Block Ending columns:")
for i, h in enumerate(contrib_cols[-5:]):
    print(f"  {len(contrib_cols)-5+i}: {h[:50]}")

# Count missing rows added (Condition 2)
# Original ICM data ended at row ~1293 (Data started at row 5 -> 1289 records total)
added_rows = 0
for r in range(33, ws.max_row + 1):
    # Check if there is NO base data but there IS Parent or Contrib data
    base_data = any(ws.cell(r, c).value is not None for c in range(3, parent_start))
    if not base_data:
        added_rows += 1

print(f"\nTotal synthetic rows (no base data): {added_rows}")
