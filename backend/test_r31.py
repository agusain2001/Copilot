"""Count all parent journal values."""
import openpyxl, sys
sys.path.insert(0, r'G:\FCCS\backend')
from app.ic_processor import read_journal_report, apply_sign

# Count ALL individual account values in parent journal
lookup = read_journal_report(r'G:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx')
non_zero_count = 0
for (e, p, a), lines in lookup.items():
    net = sum(apply_sign(j['debit'], j['credit'], a) for j in lines)
    if net != 0:
        non_zero_count += 1
        print(f"  ({e}, {p}, {a}) = {net:.2f}")
        
print(f"\nTotal non-zero (ent, icp, acct) entries in Parent journal: {non_zero_count}")

# Count in output
wb = openpyxl.load_workbook(r'G:\FCCS\backend\uploads\reports\31\output\ICM_Report31_Output.xlsx', data_only=True)
ws = wb.active

# Count non-zero cells in parent block
ent_cells = 0
par_cells = 0
for r in range(33, ws.max_row+1):
    for c in range(17, 22):
        v = ws.cell(r, c).value
        if v is not None and v != 0:
            ent_cells += 1
    for c in range(23, 28):
        v = ws.cell(r, c).value
        if v is not None and v != 0:
            par_cells += 1

print(f"\nIn output Parent block:")
print(f"  Entity-side non-zero cells: {ent_cells}")
print(f"  Partner-side non-zero cells: {par_cells}")
print(f"  Total: {ent_cells + par_cells}")
