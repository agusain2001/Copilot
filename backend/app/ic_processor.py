"""
IC Matching Report — Importable Processing Module
===================================================
Refactored from AI/ic_matching_v3.py for use as a backend service.

Key design:
- Output columns are derived DYNAMICALLY from the ICM source file accounts
  and the elimination accounts in report_inputs.  NO journal-only accounts
  are added as columns.
- Missing Entity–Partner rows are added only when the journal account
  exists in the ICM-defined account set.
- Parent and Contribution blocks each have their own Plug Account column.
- Computes Variance and Total values in Python (not Excel formulas).
"""

import re
import os
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

logger = logging.getLogger(__name__)


class AssignmentAuditError(RuntimeError):
    """Raised when a normalized journal key is not assigned exactly once."""

# ── Row / column constants ─────────────────────────────────────────────────────
ICM_INPUT_HEADER_ROW  = 4
ICM_INPUT_DATA_START  = 5
ICM_OUTPUT_HEADER_ROW = 32
ICM_OUTPUT_DATA_START = 33
JOURNAL_DATA_START    = 31
SECTION_LABEL_ROW     = 29



# ── Styling ───────────────────────────────────────────────────────────────────
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=9)
HEADER_FILL    = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
ID_FILL        = PatternFill(start_color="C8DCF0", end_color="C8DCF0", fill_type="solid")
VARIANCE_FILL  = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
TOTAL_FILL     = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
MATCH_FILL     = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
GROUP_FILL     = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
SECTION_FILL   = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
NO_FILL        = PatternFill(fill_type=None)
ID_FONT        = Font(bold=True, size=11)
DATA_FONT      = Font(size=11)
SECTION_FONT   = Font(bold=True, color="FFFFFF", size=11)
NUM_FORMAT     = r"###,##0;\-###,##0"

DATA_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ═══════════════════════════════════════════════════════════════════════════
# UTILS
# ═══════════════════════════════════════════════════════════════════════════

def extract_icp_code(raw: str) -> str:
    m = re.match(r"(ICP_\w+)", str(raw or "").strip())
    return m.group(1) if m else ""

def extract_6digit_from_icp(icp_code: str) -> str:
    s = str(icp_code or "").strip()
    return s[4:] if s.startswith("ICP_") else s

def extract_account_code(raw: str) -> str:
    raw = str(raw or "").strip()
    m = re.search(r"]\.\[?(\w+)\]?:", raw)
    if m:
        val = m.group(1)
        if not val.isdigit():
            m2 = re.search(r"]:(\d{6}):", raw)
            if m2: return m2.group(1)
        return val
    # Pattern 2: "NNNNNN - NNNNNN:description" → extract second numeric (middle value)
    m = re.match(r"\d{6}\s*-\s*(\d{6}):", raw)
    if m:
        return m.group(1)
    # Pattern 3: plain 6-digit at start
    m = re.match(r"(\d{6})", raw)
    return m.group(1) if m else ""

def extract_entity_code_icm(raw: str) -> str:
    """Extract and normalize entity code from ICM source data.
    Handles both pure numeric (001001) and E-prefixed (E101000) entities."""
    raw = str(raw or "").strip()
    # Try pure 6-digit first
    m = re.match(r"(\d{6})", raw)
    if m:
        return m.group(1)
    # Try E-prefixed entity
    m = re.match(r"E(\d+)", raw)
    if m:
        return m.group(1)  # Strip the E prefix for normalized matching
    return ""

def extract_entity_code_journal(raw: str) -> str:
    raw = str(raw or "").strip()
    m = re.match(r"(E?\d+\w*)", raw)
    return m.group(1) if m else ""

def normalize_entity_code(code: str) -> str:
    """Normalize entity codes to numeric-only format for matching.
    Strips 'E' prefix so journal codes ('E101000') match ICM codes ('101000')."""
    code = str(code or "").strip()
    if code.startswith("E") and len(code) > 1 and code[1:].replace("_", "").isdigit():
        return code[1:]
    return code

def normalize_to_numeric(code: str) -> str:
    """Normalize any entity/ICP code to pure numeric.
    Strips ICP_ prefix and E prefix.
    E117100 -> 117100, ICP_007009 -> 007009, ICP_E117100 -> 117100, 007009 -> 007009"""
    s = str(code or "").strip()
    if s.startswith("ICP_"):
        s = s[4:]
    if s.startswith("E") and len(s) > 1 and s[1:].isdigit():
        s = s[1:]
    return s

def to_float(val) -> float:
    if val is None or str(val).strip() in ("", " "):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def normalize_output_label(raw: str) -> str:
    """Normalize journal labels to the workbook display format."""
    text = str(raw or "").strip()
    if text and ":" in text and " - " not in text:
        text = text.replace(":", " - ", 1)
    return text


def classify_pair_direction(raw_entity: str) -> str:
    """Classify direction using the business rule: entity codes starting with E are forward."""
    entity_code = extract_entity_code_journal(raw_entity)
    if str(entity_code or "").upper().startswith("E"):
        return "entity_to_partner"
    return "partner_to_entity"

def get_journal_indices(ws):
    header_row = JOURNAL_DATA_START - 1
    headers = [str(c.value or "").strip().lower() for c in ws[header_row]]
    col_map = {}
    for i, h in enumerate(headers):
        if h: col_map[h] = i
    return {
        "entity": col_map.get('entity', 2),
        "acct": col_map.get('account', 3),
        "icp": col_map.get('intercompany', 4),
        "debit": col_map.get('debit', 14),
        "credit": col_map.get('credit', 15)
    }

def is_detail_row(vals: list, indices: dict) -> bool:
    try:
        entity = str(vals[indices["entity"]] or "").strip()
        acct   = str(vals[indices["acct"]]   or "").strip()
        icp    = str(vals[indices["icp"]]    or "").strip()
        return bool(entity or acct or icp)
    except IndexError:
        return False

def classify_account(code: str) -> str:
    """Classify numeric account as S1 (asset) or S2 (liability)."""
    code = str(code or "").strip()
    if not code or not code[0].isdigit():
        return "EXTRA"
    s = code[0]
    if s in ("1", "5"): return "S1"
    elif s in ("2", "3", "4"): return "S2"
    return "EXTRA"

def apply_sign(debit: float, credit: float, account_code: str) -> float:
    code = str(account_code or "").strip()
    s = code[0] if code and code[0].isdigit() else "0"
    if s in ("1", "5"):
        return debit - credit
    elif s in ("2", "3", "4"):
        return credit - debit
    return debit - credit

def parse_report_inputs(filepath: str) -> dict:
    if not filepath or not os.path.exists(filepath): return None
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        plug_code = None
        for h in headers:
            if "Plug Account:" in h:
                m = re.search(r"(\d{6})", h)
                if m: plug_code = m.group(1)
                break
        if not plug_code: return None
        elim_codes = set()
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            val = str(row[0] or "").strip()
            if not val: continue
            code = extract_account_code(val)
            if code: elim_codes.add(code)
        return {"plug_code": plug_code, "elim_codes": elim_codes}
    except Exception as e:
        logger.error("Error parsing report inputs: %s", e)
        return None

# ═══════════════════════════════════════════════════════════════════════════
# ICM PARSING
# ═══════════════════════════════════════════════════════════════════════════

def detect_icm_header_row(ws):
    """Auto-detect the row containing 'Entity' col 1 / 'Partner' col 2.
    Scans first 50 rows.  Returns (header_row, data_start_row)."""
    for r in range(1, min(51, ws.max_row + 1)):
        v1 = str(ws.cell(r, 1).value or "").strip().lower()
        v2 = str(ws.cell(r, 2).value or "").strip().lower()
        if v1 == "entity" and v2 == "partner":
            return r, r + 1
    return ICM_INPUT_HEADER_ROW, ICM_INPUT_DATA_START

def read_icm_headers(ws, header_row=None):
    """Builds a map of (code, tag) -> col_index from the ICM file header row."""
    if header_row is None:
        header_row = ICM_INPUT_HEADER_ROW
    headers = [cell.value for cell in ws[header_row]]
    col_map = {}
    for i, h in enumerate(headers, start=1):
        hs = str(h or "").strip()
        code = extract_account_code(hs)
        if not code: continue
        tag = "Partner" if re.search(r"\bPartner\b", hs) else "Entity"
        col_map[(code, tag)] = i
    return col_map

def read_icm_account_columns(ws, header_row=None):
    """Extract the ordered list of (code, description, tag) from the FIRST
    entity-side + partner-side block in the ICM source headers.

    The ICM source repeats the same block of accounts 3 times (for Base /
    Parent / Contribution).  We only need the first block to define the
    output structure.

    Returns (ent_cols, par_cols) where each is a list of
    (code, description, series, tag) tuples — one per unique account,
    preserving order."""
    if header_row is None:
        header_row = ICM_INPUT_HEADER_ROW

    # Collect all (code, desc, tag, col) from the header row
    all_headers = []
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(header_row, c).value or "").strip()
        if not v:
            continue
        code = extract_account_code(v)
        if not code:
            continue
        tag = "Partner" if re.search(r"\bPartner\b", v) else "Entity"
        all_headers.append((code, v, tag, c))

    # De-duplicate: keep the FIRST occurrence of each (code, tag) pair
    seen = set()
    ent_cols = []
    par_cols = []
    for code, desc, tag, _col in all_headers:
        key = (code, tag)
        if key in seen:
            continue
        seen.add(key)
        series = classify_account(code)
        if tag == "Entity":
            ent_cols.append((code, desc, series, tag))
        else:
            par_cols.append((code, desc, series, tag))

    return ent_cols, par_cols

def read_icm_data(ws, data_start=None):
    if data_start is None:
        data_start = ICM_INPUT_DATA_START
    data_rows = []
    for row_num, row in enumerate(
            ws.iter_rows(min_row=data_start, max_row=ws.max_row), start=data_start):
        vals = [cell.value for cell in row]
        entity_raw  = str(vals[0] or "").strip()
        partner_raw = str(vals[1] or "").strip() if len(vals) > 1 else ""
        if not entity_raw and not partner_raw: continue
        partner_icp = extract_icp_code(partner_raw)
        data_rows.append({
            "row_num":      row_num,
            "entity":       entity_raw,
            "partner":      partner_raw,
            "entity_code":  extract_entity_code_icm(entity_raw),
            "partner_code": partner_icp,
            "partner_num":  normalize_to_numeric(partner_icp),
            "row_direction": classify_pair_direction(entity_raw),
            "is_synthetic": False,
        })
    return data_rows

# ═══════════════════════════════════════════════════════════════════════════
# JOURNAL READING & MATCHING
# ═══════════════════════════════════════════════════════════════════════════

def parse_journal_line(vals, indices: dict, row_num: int, plug_mapping: dict = None):
    """Normalize one journal detail row into a single numeric-key record."""
    try:
        label = str(vals[0] or "").strip()
        entity_raw = str(vals[indices["entity"]] or "").strip()
        account_raw = str(vals[indices["acct"]] or "").strip()
        icp_raw = str(vals[indices["icp"]] or "").strip()
        debit = to_float(vals[indices["debit"]])
        credit = to_float(vals[indices["credit"]])
    except IndexError:
        return None

    entity_num = normalize_to_numeric(extract_entity_code_journal(entity_raw))
    account_code = extract_account_code(account_raw)
    icp_num = normalize_to_numeric(extract_icp_code(icp_raw))

    if plug_mapping and account_raw:
        plug_code = plug_mapping.get("plug_code")
        elim_codes = plug_mapping.get("elim_codes", set())
        if account_code in elim_codes or (not account_code[:1].isdigit() and plug_code):
            account_code = plug_code

    if not entity_num or not icp_num or not account_code:
        return None

    return {
        "label": label,
        "source_row": row_num,
        "entity_raw": entity_raw,
        "entity_num": entity_num,
        "account_raw": account_raw,
        "account_code": account_code,
        "icp_raw": icp_raw,
        "icp_num": icp_num,
        "direction": classify_pair_direction(entity_raw),
        "debit": debit,
        "credit": credit,
        "net": apply_sign(debit, credit, account_code),
    }


def read_journal_lines(filepath: str, plug_mapping: dict = None):
    """Read normalized journal detail rows."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    indices = get_journal_indices(ws)
    lines = []

    for row_num, row in enumerate(ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row), start=JOURNAL_DATA_START):
        vals = [cell.value for cell in row]
        if not vals: continue
        label = str(vals[0] or "").strip()
        if label == "Grand Total": break
        if not is_detail_row(vals, indices): continue
        line = parse_journal_line(vals, indices, row_num, plug_mapping)
        if line is not None:
            lines.append(line)
    return lines


def read_journal_report(filepath: str, plug_mapping: dict = None):
    """Read journal entries and build a lookup keyed by pure numeric codes.
    Keys are (entity_num, icp_num, account_code) where entity_num and icp_num
    are normalized to pure numeric (E prefix and ICP_ prefix stripped).

    Matching rules (from user spec):
      Entity col:  E117100:… → 117100,  007009:… → 007009
      ICP col:     ICP_007009:… → 007009,  ICP_E117100:… → 117100
      Account col: 534018 - 534018:Interest expense → 534018
    """
    lines = read_journal_lines(filepath, plug_mapping)


    # ── Deduplication: prevent double-counting identical rows (§4) ────
    # Single unified lookup — all codes are pure numeric
    lookup = defaultdict(list)
    for line in lines:
        lookup[(line["entity_num"], line["icp_num"], line["account_code"])].append(line)

    return lookup


def read_journal_labels(filepath: str) -> tuple:
    """Returns (entity_labels, icp_labels) — two separate maps.
    entity_labels: numeric_code → entity raw string  (e.g. '117100' → 'E117100:QD UK...')
    icp_labels:    numeric_code → ICP raw string      (e.g. '117100' → 'ICP_E117100:QD UK... ICP')
    Keeping them separate prevents entity labels from overwriting ICP labels."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    indices = get_journal_indices(ws)
    entity_labels = {}
    icp_labels = {}
    for row in ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row):
        vals = [cell.value for cell in row]
        if not vals: continue
        if str(vals[0] or "").strip() == "Grand Total":
            break
        if not is_detail_row(vals, indices):
            continue
        try:
            entity_raw = str(vals[indices["entity"]] or "").strip()
            icp_raw    = str(vals[indices["icp"]] or "").strip()
        except IndexError:
            continue
        ent_code = extract_entity_code_journal(entity_raw)
        ent_num  = normalize_to_numeric(ent_code)
        icp_code = extract_icp_code(icp_raw)
        icp_num  = normalize_to_numeric(icp_code)
        if ent_code and entity_raw:
            entity_labels[ent_code] = entity_raw
            if ent_num != ent_code:
                entity_labels[ent_num] = entity_raw
        if icp_code and icp_raw:
            icp_labels[icp_code] = icp_raw
            if icp_num != icp_code:
                icp_labels[icp_num] = icp_raw
    return entity_labels, icp_labels


def read_journal_pair_labels(filepath: str) -> dict:
    """Returns {(entity_num, icp_num): {'entity': raw_entity, 'icp': raw_icp}}."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    indices = get_journal_indices(ws)
    pair_labels = {}
    for row_num, row in enumerate(
            ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row),
            start=JOURNAL_DATA_START):
        vals = [cell.value for cell in row]
        if not vals:
            continue
        if str(vals[0] or "").strip() == "Grand Total":
            break
        if not is_detail_row(vals, indices):
            continue
        line = parse_journal_line(vals, indices, row_num)
        if line is None:
            continue
        key = (line["entity_num"], line["icp_num"])
        if key not in pair_labels:
            pair_labels[key] = {
                "entity": str(vals[indices["entity"]] or "").strip(),
                "icp": str(vals[indices["icp"]] or "").strip(),
            }
    return pair_labels


def read_ic_elimination_report(filepath: str, plug_code: str = None):
    """Parse summarized IC elimination reports with Entity/Partner rows and account columns."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    candidates = []
    for sheet in wb.worksheets:
        hdr_row, data_start = detect_icm_header_row(sheet)
        c1 = str(sheet.cell(hdr_row, 1).value or "").strip().lower()
        c2 = str(sheet.cell(hdr_row, 2).value or "").strip().lower()
        if c1 == "entity" and c2 == "partner":
            candidates.append((sheet, hdr_row, data_start))
    if not candidates:
        return {}, set(), {}, {}, {}, []

    ws, hdr_row, data_start = max(candidates, key=lambda item: (item[0].max_row, item[0].max_column))

    account_headers = []
    total_col = None
    for col_idx in range(3, ws.max_column + 1):
        header_text = str(ws.cell(hdr_row, col_idx).value or "").strip()
        if not header_text:
            continue
        if header_text.lower() == "total":
            total_col = col_idx
            continue
        account_code = extract_account_code(header_text)
        if not account_code:
            continue
        tag = "Partner" if re.search(r"\bPartner\b", header_text) else "Entity"
        account_headers.append((col_idx, header_text, account_code, tag))

    updates = defaultdict(float)
    pair_keys = set()
    entity_labels = {}
    icp_labels = {}
    pair_labels = {}
    raw_lines = []

    for row_num in range(data_start, ws.max_row + 1):
        entity_raw = str(ws.cell(row_num, 1).value or "").strip()
        icp_raw = str(ws.cell(row_num, 2).value or "").strip()
        if not entity_raw and not icp_raw:
            continue
        if not entity_raw or not icp_raw:
            continue

        entity_code = extract_entity_code_icm(entity_raw)
        icp_code_raw = extract_icp_code(icp_raw)
        icp_num = normalize_to_numeric(icp_code_raw)
        if not entity_code or not icp_num:
            continue

        pair_key = (entity_code, icp_num)
        pair_keys.add(pair_key)
        entity_labels.setdefault(entity_code, entity_raw)
        if icp_code_raw:
            icp_labels.setdefault(icp_num, icp_raw)
        pair_labels.setdefault(pair_key, {"entity": entity_raw, "icp": icp_raw})

        for col_idx, header_text, account_code, tag in account_headers:
            cell_val = ws.cell(row_num, col_idx).value
            if cell_val in (None, "", " "):
                continue
            amount = to_float(cell_val)
            if amount == 0:
                continue

            if tag == "Entity":
                key = (entity_code, icp_num, account_code)
                direction = "entity_to_partner"
            else:
                key = (icp_num, entity_code, account_code)
                direction = "partner_to_entity"
            raw_lines.append({
                "label": f"IC Elim Row {row_num}",
                "source_row": row_num,
                "direction": direction,
                "entity_raw": entity_raw,
                "entity_num": entity_code,
                "account_raw": header_text,
                "account_code": account_code,
                "icp_raw": icp_raw,
                "icp_num": icp_num,
                "debit": amount if amount > 0 else 0.0,
                "credit": -amount if amount < 0 else 0.0,
                "net": amount,
            })

        if plug_code and total_col is not None:
            total_val = ws.cell(row_num, total_col).value
            if total_val not in (None, "", " "):
                total_amount = to_float(total_val)
                if total_amount != 0:
                    updates[(entity_code, icp_num, plug_code)] += total_amount
                    raw_lines.append({
                        "label": f"IC Elim Total {row_num}",
                        "source_row": row_num,
                        "direction": "entity_to_partner",
                        "entity_raw": entity_raw,
                        "entity_num": entity_code,
                        "account_raw": f"{plug_code}:Intercompany Balances Plug A/c Total",
                        "account_code": plug_code,
                        "icp_raw": icp_raw,
                        "icp_num": icp_num,
                        "debit": total_amount if total_amount > 0 else 0.0,
                        "credit": -total_amount if total_amount < 0 else 0.0,
                        "net": total_amount,
                    })

    return dict(updates), pair_keys, entity_labels, icp_labels, pair_labels, raw_lines


def merge_reciprocal_rows(data_rows):
    """Collapse reciprocal rows into one canonical entity-to-partner row when possible."""
    pair_to_index = {}
    for idx, row in enumerate(data_rows):
        pair_key = (row.get("entity_code", ""), row.get("partner_num", ""))
        if pair_key[0] and pair_key[1]:
            pair_to_index[pair_key] = idx

    removed = set()
    for idx, row in enumerate(data_rows):
        if idx in removed:
            continue
        ent = row.get("entity_code", "")
        prt = row.get("partner_num", "")
        if not ent or not prt:
            continue
        rev_idx = pair_to_index.get((prt, ent))
        if rev_idx is None or rev_idx == idx or rev_idx in removed:
            continue

        rev_row = data_rows[rev_idx]
        candidates = [
            (idx, row),
            (rev_idx, rev_row),
        ]
        forward_candidates = [
            (cand_idx, cand_row)
            for cand_idx, cand_row in candidates
            if cand_row.get("row_direction") == "entity_to_partner"
        ]
        if not forward_candidates:
            continue

        keep_idx, keep_row = sorted(
            forward_candidates,
            key=lambda item: (
                item[1].get("is_synthetic", False),
                item[1].get("row_num") is None,
                item[0],
            ),
        )[0]
        drop_idx = rev_idx if keep_idx == idx else idx
        keep_row["merged_reciprocal"] = True
        removed.add(drop_idx)

    if not removed:
        return data_rows
    return [row for idx, row in enumerate(data_rows) if idx not in removed]


def build_covered_pairs(data_rows):
    """Pairs considered already represented in the consolidated report."""
    covered = set()
    for row in data_rows:
        ent = row.get("entity_code", "")
        prt = row.get("partner_num", "")
        if not ent or not prt:
            continue
        covered.add((ent, prt))
        if ent != prt:
            covered.add((prt, ent))
    return covered


def _row_choice_rank(row):
    return (
        row.get("is_synthetic", False),
        row.get("row_num") is None,
        row.get("row_num") if row.get("row_num") is not None else float("inf"),
        0 if row.get("row_direction") == "entity_to_partner" else 1,
        str(row.get("entity", "")),
        str(row.get("partner", "")),
    )


def _choose_canonical_pair(pair_key, all_pairs, icm_rows_by_pair, all_pair_labels, all_entity_labels):
    ent, prt = pair_key
    rev_key = (prt, ent)
    candidates = [key for key in (pair_key, rev_key) if key in all_pairs]
    if not candidates:
        candidates = [pair_key]

    for key in candidates:
        row = icm_rows_by_pair.get(key)
        if row and row.get("row_direction") == "entity_to_partner":
            return key

    for key in candidates:
        pair_labels = all_pair_labels.get(key)
        if pair_labels and classify_pair_direction(pair_labels["entity"]) == "entity_to_partner":
            return key

    for key in candidates:
        ent_label = all_entity_labels.get(key[0], "")
        if ent_label and classify_pair_direction(ent_label) == "entity_to_partner":
            return key

    for key in candidates:
        if key in icm_rows_by_pair:
            return key

    return min(candidates)


def _resolve_consolidated_labels(pair_key, icm_rows_by_pair, all_pair_labels,
                                 all_entity_labels, all_icp_labels):
    row = icm_rows_by_pair.get(pair_key)
    pair_labels = all_pair_labels.get(pair_key)
    if pair_labels:
        ent_label = normalize_output_label(pair_labels["entity"])
        prt_label = normalize_output_label(pair_labels["icp"])
        return (
            ent_label,
            prt_label,
            row["row_num"] if row else None,
            row["row_direction"] if row else classify_pair_direction(pair_labels["entity"]),
            row["is_synthetic"] if row else True,
        )

    if row:
        return row["entity"], row["partner"], row["row_num"], row["row_direction"], row["is_synthetic"]

    ent_code, prt_code = pair_key
    ent_raw = all_entity_labels.get(ent_code, ent_code)
    prt_raw = all_icp_labels.get(prt_code, "")
    if not prt_raw:
        fallback = all_entity_labels.get(prt_code, prt_code)
        if ":" in str(fallback):
            suffix = str(fallback).split(":", 1)[1]
            raw_code = extract_entity_code_journal(fallback) or prt_code
            prt_raw = f"ICP_{raw_code}:{suffix}"
        else:
            prt_raw = f"ICP_{prt_code}"
    ent_label = normalize_output_label(ent_raw)
    prt_label = normalize_output_label(prt_raw)
    return ent_label, prt_label, None, classify_pair_direction(ent_raw), True


def build_consolidated_rows(icm_rows, journal_pair_keys, all_pair_labels,
                            all_entity_labels, all_icp_labels):
    """Build one canonical row per unique normalized entity/partner pair."""
    icm_rows_by_pair = {}
    for row in icm_rows:
        pair_key = (row.get("entity_code", ""), row.get("partner_num", ""))
        if not pair_key[0] or not pair_key[1]:
            continue
        current = icm_rows_by_pair.get(pair_key)
        if current is None or _row_choice_rank(row) < _row_choice_rank(current):
            icm_rows_by_pair[pair_key] = dict(row)

    all_pairs = set(icm_rows_by_pair) | {
        (ent, prt) for ent, prt in journal_pair_keys if ent and prt
    }

    consolidated_rows = []
    seen_groups = set()
    for pair_key in sorted(all_pairs):
        ent, prt = pair_key
        group_key = (ent, prt) if ent == prt else tuple(sorted((ent, prt)))
        if group_key in seen_groups:
            continue
        seen_groups.add(group_key)

        canonical_pair = _choose_canonical_pair(
            pair_key, all_pairs, icm_rows_by_pair, all_pair_labels, all_entity_labels
        )
        ent_label, prt_label, row_num, row_direction, is_synthetic = _resolve_consolidated_labels(
            canonical_pair, icm_rows_by_pair, all_pair_labels, all_entity_labels, all_icp_labels
        )
        consolidated_rows.append({
            "row_num": row_num,
            "entity": ent_label,
            "partner": prt_label,
            "entity_code": canonical_pair[0],
            "partner_code": f"ICP_{canonical_pair[1]}",
            "partner_num": canonical_pair[1],
            "row_direction": row_direction,
            "is_synthetic": is_synthetic,
        })

    return consolidated_rows


def canonicalize_updates_to_rows(updates, data_rows):
    """Fold directional updates onto the canonical consolidated row pair when needed."""
    row_pairs = {
        (row.get("entity_code", ""), row.get("partner_num", ""))
        for row in data_rows
        if row.get("entity_code") and row.get("partner_num")
    }
    canonical = defaultdict(float)
    for (ent, prt, acct), value in updates.items():
        if (ent, prt) in row_pairs:
            canonical[(ent, prt, acct)] += value
        elif (prt, ent) in row_pairs:
            canonical[(prt, ent, acct)] += value
        else:
            canonical[(ent, prt, acct)] += value
    return dict(canonical)


def match_journal_to_icm(data_rows, lookup):
    """Returns updates dict of (entity_num, icp_num, acct) → net.
    All keys are pure numeric (no ICP_ or E prefixes)."""
    updates = {}
    for (e, p, a), jlines in lookup.items():
        net = sum(j["net"] for j in jlines)
        if net != 0:
            updates[(e, p, a)] = net
    return updates

# ═══════════════════════════════════════════════════════════════════════════
# OUTPUT WRITER
# ═══════════════════════════════════════════════════════════════════════════

def _style_header_cell(cell, text):
    cell.value = text; cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _style_id_cell(cell, text):
    cell.value = text; cell.font = ID_FONT; cell.fill = ID_FILL; cell.border = DATA_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

def _style_data_cell(cell, value, fill=None):
    cell.value = value; cell.font = DATA_FONT; cell.fill = fill if fill is not None else NO_FILL
    cell.border = DATA_BORDER; cell.alignment = Alignment(vertical="top", wrap_text=True)
    if isinstance(value, (int, float)): cell.number_format = NUM_FORMAT


def _block_positions(start, num_accts_ent, num_accts_par, has_plug=False):
    """Compute column positions for one block.
    Returns dict with keys: ent_start, var1, par_start, var2, total, plug (or None), spacer."""
    ent_start = start
    var1      = start + num_accts_ent
    par_start = var1 + 1
    var2      = par_start + num_accts_par
    total     = var2 + 1
    if has_plug:
        plug   = total + 1
        spacer = plug + 1
    else:
        plug   = None
        spacer = total + 1
    return {
        "ent_start": ent_start, "var1": var1,
        "par_start": par_start, "var2": var2,
        "total": total, "plug": plug, "spacer": spacer,
    }


def write_output(ws_icm_source, data_rows, icm_header_map,
                 updates_list, output_path, plug_code=None,
                 ent_cols=None, par_cols=None, raw_plug_lines=None):
    """Write the ICM output workbook.

    Parameters
    ----------
    ent_cols, par_cols : list of (code, desc, series, tag) tuples
        Defines the account columns for every block — driven by the ICM source.
    """
    if not ent_cols or not par_cols:
        raise ValueError("ent_cols and par_cols must be provided (derived from ICM source)")

    n_ent = len(ent_cols)
    n_par = len(par_cols)

    # ── Compute block positions ──────────────────────────────────────────
    blk_base = _block_positions(3, n_ent, n_par, has_plug=False)
    blk_par  = _block_positions(blk_base["spacer"] + 1, n_ent, n_par, has_plug=True)
    blk_cont = _block_positions(blk_par["spacer"] + 1, n_ent, n_par, has_plug=True)

    # Plug Account section: plug_value | Total
    plug_section_start = blk_cont["spacer"] + 1
    plug_section = {
        "plug_col": plug_section_start,
        "total": plug_section_start + 1,
        "spacer": plug_section_start + 2,
    }
    col_final = plug_section["spacer"] + 1
    total_cols = col_final

    spacer_cols = {blk_base["spacer"], blk_par["spacer"], blk_cont["spacer"], plug_section["spacer"]}

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "ICM Matched"

    for i in range(total_cols):
        col_num = i + 1
        ws.column_dimensions[get_column_letter(col_num)].width = 4 if col_num in spacer_cols else 17.5714

    ws.row_dimensions[1].height = 27.0
    for r in list(range(10, 20)) + [23, 28]:
        ws.row_dimensions[r].height = 14.45
    ws.row_dimensions[ICM_OUTPUT_HEADER_ROW].height = 78.75

    # ── Section Labels ───────────────────────────────────────────────────
    def _section_label(start, end, text):
        ws.merge_cells(start_row=SECTION_LABEL_ROW, start_column=start,
                       end_row=SECTION_LABEL_ROW, end_column=end)
        c = ws.cell(SECTION_LABEL_ROW, start)
        c.value = text; c.fill = SECTION_FILL; c.font = SECTION_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

    _section_label(blk_par["ent_start"], blk_par["plug"] or blk_par["total"],  "Parent Input")
    _section_label(blk_cont["ent_start"], blk_cont["plug"] or blk_cont["total"], "Contribution Input")
    _section_label(plug_section["plug_col"], plug_section["total"], "Plug Account")

    # ── Header Row ───────────────────────────────────────────────────────
    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, 1), "Entity")
    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, 2), "Partner")

    plug_label_base = f"{plug_code}:Intercompany Balances Plug A/c" if plug_code else "188800:Intercompany Balances Plug A/c"
    plug_label_parent = f"{plug_label_base}\n(Entity \u2192 Parent)"
    plug_label_contrib = f"{plug_label_base}\n(Parent \u2192 Entity)"

    def _write_block_headers(blk, has_plug_hdr=False, plug_hdr_text=None):
        for i, (code, desc, _, tag) in enumerate(ent_cols):
            _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["ent_start"] + i),
                               f"{code} - {desc} {tag}" if tag not in desc else desc)
        _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["var1"]), "Variance")

        for i, (code, desc, _, tag) in enumerate(par_cols):
            _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["par_start"] + i),
                               f"{code} - {desc} {tag}" if tag not in desc else desc)
        _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["var2"]), "Variance")
        _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["total"]), "Total")

        if has_plug_hdr and blk["plug"] is not None:
            _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["plug"]),
                               plug_hdr_text or plug_label_base)

    _write_block_headers(blk_base, has_plug_hdr=False)
    _write_block_headers(blk_par, has_plug_hdr=True, plug_hdr_text=plug_label_parent)
    _write_block_headers(blk_cont, has_plug_hdr=True, plug_hdr_text=plug_label_contrib)

    # Plug Account section headers
    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, plug_section["plug_col"]), plug_label_base)
    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, plug_section["total"]), "Total")

    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, col_final), "Final Total")

    # ── Scale factor for output values ────────────────────────────────────
    SCALE = 1.0

    # ── Helper: read base value from ICM source ──────────────────────────
    def _icm_num(src_r, code, tag):
        if src_r is None: return None
        col_idx = icm_header_map.get((code, tag))
        if not col_idx: return None
        v = ws_icm_source.cell(src_r, col_idx).value
        try:
            val = float(v) if v not in (None, "", " ") else None
            return val * SCALE if val is not None else None
        except: return None

    def _extract_val(raw):
        if raw is None:
            return None, NO_FILL
        return raw * SCALE, MATCH_FILL

    # ── Unpack updates (single dict per journal, no more primary/fallback) ──
    def _unpack(idx):
        return updates_list[idx] if len(updates_list) > idx else {}

    parent_updates  = _unpack(0)
    contrib_updates = _unpack(1)
    plug_updates    = _unpack(2)
    pair_owner_index = {}
    for row_idx, row in enumerate(data_rows):
        pair_key = (row["entity_code"], row.get("partner_num", ""))
        if not pair_key[0] or not pair_key[1]:
            continue
        current_owner = pair_owner_index.get(pair_key)
        if current_owner is None:
            pair_owner_index[pair_key] = row_idx
            continue
        if data_rows[current_owner].get("is_synthetic") and not row.get("is_synthetic"):
            pair_owner_index[pair_key] = row_idx

    # S1/S2 groups for variance
    s1_ent_codes = [c[0] for c in ent_cols if c[2] == "S1"]
    s2_ent_codes = [c[0] for c in ent_cols if c[2] == "S2"]
    s1_par_codes = [c[0] for c in par_cols if c[2] == "S1"]
    s2_par_codes = [c[0] for c in par_cols if c[2] == "S2"]

    # §4 Transaction Uniqueness: Each journal key (entity, icp, acct) must
    # map to exactly ONE output cell.  Consumed sets track which keys have
    # already been written so paired rows don't double-count.
    parent_consumed  = set()
    contrib_consumed = set()
    plug_consumed    = set()
    assignment_audit = {
        "parent": defaultdict(list),
        "contrib": defaultdict(list),
        "plug": defaultdict(list),
    }

    def _claim_update(family, key, updates, consumed, row_num, slot_name):
        if key not in updates or key in consumed:
            return None
        consumed.add(key)
        assignment_audit[family][key].append({
            "row_num": row_num,
            "slot": slot_name,
        })
        return updates[key]

    def _validate_assignments(family, updates, consumed):
        missing = sorted(k for k in updates if k not in consumed)
        duplicates = {
            key: claims for key, claims in assignment_audit[family].items()
            if len(claims) != 1
        }
        if missing or duplicates:
            pieces = []
            if missing:
                pieces.append(
                    f"unassigned={missing[:5]}"
                    + ("..." if len(missing) > 5 else "")
                )
            if duplicates:
                pieces.append(f"duplicate_claims={list(duplicates.items())[:5]}")
            msg = f"{family} journal assignment audit failed: " + "; ".join(pieces)
            logger.error(msg)
            raise AssignmentAuditError(msg)
        logger.info("Assignment audit passed for %s journal: %d keys", family, len(updates))

    def _is_exact_owner(row_idx, pair_key):
        return pair_owner_index.get(pair_key) == row_idx


    # ── Write data rows ──────────────────────────────────────────────────
    for out_idx, icm_row in enumerate(data_rows):
        src_r = icm_row["row_num"]
        out_r = ICM_OUTPUT_DATA_START + out_idx
        ent   = icm_row["entity_code"]          # pure numeric (e.g. "117100")
        prt_num = icm_row.get("partner_num", "")  # pure numeric (e.g. "007009")
        pair_key = (ent, prt_num)
        owns_pair = _is_exact_owner(out_idx, pair_key)

        _style_id_cell(ws.cell(out_r, 1), icm_row["entity"])
        _style_id_cell(ws.cell(out_r, 2), icm_row["partner"])

        # All matching uses pure numeric codes
        can_match = bool(ent and prt_num)

        def _write_block(blk, is_base, updates=None, consumed=None, audit_family=None):
            """Write entity-side, partner-side, variances, total for one block.
            Entity-side:  key (ent, prt_num, acct)  — entity-to-parent direction.
            Partner-side: key (prt_num, ent, acct)  — parent-to-entity direction.
            consumed: set tracking used keys — ensures §4 uniqueness."""
            ent_calc_vals = {}
            par_calc_vals = {}

            # ── Entity-side (entity-to-parent: ent → prt_num) ────────────
            for i, (code, _, series, tag) in enumerate(ent_cols):
                if is_base:
                    # Mandatory: only assign values when BOTH entity and partner are present
                    v = _icm_num(src_r, code, tag) if can_match else None
                    _style_data_cell(ws.cell(out_r, blk["ent_start"] + i), v)
                    ent_calc_vals[code] = v
                else:
                    ent_key = (ent, prt_num, code)
                    raw = None
                    if can_match and owns_pair and updates and audit_family and consumed is not None:
                        raw = _claim_update(audit_family, ent_key, updates, consumed, out_r, "entity_side")
                    v, fill = _extract_val(raw)
                    _style_data_cell(ws.cell(out_r, blk["ent_start"] + i), v,
                                     fill if v is not None else None)
                    ent_calc_vals[code] = v

            # ── Variance 1 ──────────────────────────────────────────────
            sum_s1 = sum(to_float(ent_calc_vals.get(c)) for c in s1_ent_codes)
            sum_s2 = sum(to_float(ent_calc_vals.get(c)) for c in s2_ent_codes)
            var1 = sum_s1 - sum_s2
            _style_data_cell(ws.cell(out_r, blk["var1"]), var1, VARIANCE_FILL)
            ws.cell(out_r, blk["var1"]).number_format = NUM_FORMAT

            # ── Partner-side (parent-to-entity: prt_num → ent, reversed) ──
            for i, (code, _, series, tag) in enumerate(par_cols):
                if is_base:
                    # Mandatory: only assign values when BOTH entity and partner are present
                    v = _icm_num(src_r, code, tag) if can_match else None
                    _style_data_cell(ws.cell(out_r, blk["par_start"] + i), v)
                    par_calc_vals[code] = v
                else:
                    rev_key = (prt_num, ent, code)
                    rev_pair_key = (prt_num, ent)
                    raw_claim = None
                    if (
                        can_match and owns_pair and updates and audit_family and consumed is not None
                        and rev_pair_key not in pair_owner_index
                    ):
                        raw_claim = _claim_update(audit_family, rev_key, updates, consumed, out_r, "partner_side")
                    raw_display = raw_claim
                    if (
                        raw_display is None
                        and can_match
                        and updates
                        and icm_row.get("row_direction") == "entity_to_partner"
                        and rev_key in updates
                    ):
                        raw_display = updates[rev_key]
                    v, fill = _extract_val(raw_display)
                    _style_data_cell(ws.cell(out_r, blk["par_start"] + i), v,
                                     fill if v is not None else None)
                    calc_v, _ = _extract_val(raw_claim)
                    par_calc_vals[code] = calc_v

            # ── Variance 2 ──────────────────────────────────────────────
            sum_s1p = sum(to_float(par_calc_vals.get(c)) for c in s1_par_codes)
            sum_s2p = sum(to_float(par_calc_vals.get(c)) for c in s2_par_codes)
            var2 = sum_s1p - sum_s2p
            _style_data_cell(ws.cell(out_r, blk["var2"]), var2, VARIANCE_FILL)
            ws.cell(out_r, blk["var2"]).number_format = NUM_FORMAT

            # ── Total ───────────────────────────────────────────────────
            total = var1 + var2
            _style_data_cell(ws.cell(out_r, blk["total"]), total, TOTAL_FILL)
            ws.cell(out_r, blk["total"]).number_format = NUM_FORMAT

            return total

        # ── Write blocks ─────────────────────────────────────────────────
        base_total = _write_block(blk_base, is_base=True)

        parent_total = _write_block(blk_par, is_base=False,
                                    updates=parent_updates,
                                    consumed=parent_consumed,
                                    audit_family="parent")

        contrib_total = _write_block(blk_cont, is_base=False,
                                     updates=contrib_updates,
                                     consumed=contrib_consumed,
                                     audit_family="contrib")

        # ── Plug Account — Parent block (entity-to-parent direction) ─────
        plug_par_val = 0.0
        if plug_code and can_match and blk_par["plug"]:
            plug_key = (ent, prt_num, plug_code)
            raw = _claim_update("parent", plug_key, parent_updates, parent_consumed, out_r, "plug_parent")
            v, fill = _extract_val(raw)
            if v is not None:
                plug_par_val = to_float(v)
                _style_data_cell(ws.cell(out_r, blk_par["plug"]), v, fill)
            else:
                _style_data_cell(ws.cell(out_r, blk_par["plug"]), None)

        # ── Plug Account — Contribution block (parent-to-entity, reversed) ──
        plug_cont_val = 0.0
        if plug_code and can_match and blk_cont["plug"]:
            rev_plug_key = (prt_num, ent, plug_code)
            rev_pair_key = (prt_num, ent)
            raw_claim = None
            if owns_pair and rev_pair_key not in pair_owner_index:
                raw_claim = _claim_update("contrib", rev_plug_key, contrib_updates, contrib_consumed, out_r, "plug_contrib")
            raw_display = raw_claim
            if (
                raw_display is None
                and icm_row.get("row_direction") == "entity_to_partner"
                and rev_plug_key in contrib_updates
            ):
                raw_display = contrib_updates[rev_plug_key]
            v, fill = _extract_val(raw_display)
            if v is not None:
                if raw_claim is not None:
                    plug_cont_val = to_float(_extract_val(raw_claim)[0])
                _style_data_cell(ws.cell(out_r, blk_cont["plug"]), v, fill)
            else:
                _style_data_cell(ws.cell(out_r, blk_cont["plug"]), None)

        # ── Plug Account Section (Standalone) ────────────────────────────
        plug_val = 0.0
        if plug_code and can_match:
            plug_key = (ent, prt_num, plug_code)
            rev_plug_key = (prt_num, ent, plug_code)
            raw = _claim_update("plug", plug_key, plug_updates, plug_consumed, out_r, "plug_section")
            if raw is None:
                raw = _claim_update("plug", rev_plug_key, plug_updates, plug_consumed, out_r, "plug_section")
            v, fill = _extract_val(raw)
            if v is not None:
                plug_val = to_float(v)
                _style_data_cell(ws.cell(out_r, plug_section["plug_col"]), v, fill)
            else:
                _style_data_cell(ws.cell(out_r, plug_section["plug_col"]), None)
        else:
            _style_data_cell(ws.cell(out_r, plug_section["plug_col"]), None)

        # Plug Section Total
        _style_data_cell(ws.cell(out_r, plug_section["total"]),
                         plug_val if plug_val else None, TOTAL_FILL)
        ws.cell(out_r, plug_section["total"]).number_format = NUM_FORMAT

        # ── Final Total ──────────────────────────────────────────────────
        final_total = base_total + parent_total + contrib_total + plug_par_val + plug_cont_val + plug_val
        _style_data_cell(ws.cell(out_r, col_final), final_total, TOTAL_FILL)
        ws.cell(out_r, col_final).number_format = NUM_FORMAT

    _validate_assignments("parent", parent_updates, parent_consumed)
    _validate_assignments("contrib", contrib_updates, contrib_consumed)
    _validate_assignments("plug", plug_updates, plug_consumed)

    if raw_plug_lines is not None:
        ws_raw = out_wb.create_sheet("IC Elim Detail")
        headers = [
            "Label", "Source Row", "Direction",
            "Entity", "Entity Code",
            "Account", "Account Code",
            "Intercompany", "Intercompany Code",
            "Debit", "Credit", "Net",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_raw.cell(1, col_idx)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws_raw.column_dimensions[get_column_letter(col_idx)].width = 22
        for row_idx, line in enumerate(raw_plug_lines, start=2):
            values = [
                line.get("label"),
                line.get("source_row"),
                line.get("direction"),
                line.get("entity_raw"),
                line.get("entity_num"),
                line.get("account_raw"),
                line.get("account_code"),
                line.get("icp_raw"),
                line.get("icp_num"),
                line.get("debit"),
                line.get("credit"),
                line.get("net"),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws_raw.cell(row_idx, col_idx)
                cell.value = value
                cell.border = DATA_BORDER
                if col_idx >= 10:
                    cell.number_format = NUM_FORMAT

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    out_wb.save(output_path)
    logger.info("Saved output to %s  (%d rows, %d cols)", output_path, len(data_rows), total_cols)


# ═══════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

def process_icm_report(icm_path, journal_paths, output_path, report_inputs_path=None):
    logger.info("=" * 65)
    logger.info("  IC Matching — Dynamic ICM-Driven Columns")
    logger.info("=" * 65)

    wb_icm = openpyxl.load_workbook(icm_path, data_only=True)

    # ── Auto-detect sheet and header row ─────────────────────────────────
    candidates = []
    for sheet in wb_icm.worksheets:
        hdr_row, data_start = detect_icm_header_row(sheet)
        c1 = str(sheet.cell(hdr_row, 1).value or "").strip().lower()
        c2 = str(sheet.cell(hdr_row, 2).value or "").strip().lower()
        if c1 == "entity" and c2 == "partner":
            candidates.append((sheet, hdr_row, data_start))

    if candidates:
        ws_icm, icm_hdr_row, icm_data_start = min(candidates, key=lambda t: t[0].max_column)
    else:
        ws_icm = wb_icm.active
        icm_hdr_row, icm_data_start = detect_icm_header_row(ws_icm)

    logger.info("Using sheet '%s' header=%d data=%d", ws_icm.title, icm_hdr_row, icm_data_start)

    # ── Read ICM account columns (dynamic) ───────────────────────────────
    icm_header_map = read_icm_headers(ws_icm, header_row=icm_hdr_row)
    ent_cols_from_icm, par_cols_from_icm = read_icm_account_columns(ws_icm, header_row=icm_hdr_row)

    # Also read from the wider sheet if available — to capture all accounts
    if len(candidates) > 1:
        ws_wide = max(candidates, key=lambda t: t[0].max_column)[0]
        wide_hdr = max(candidates, key=lambda t: t[0].max_column)[1]
        ent_wide, par_wide = read_icm_account_columns(ws_wide, header_row=wide_hdr)
        icm_header_map_wide = read_icm_headers(ws_wide, header_row=wide_hdr)
        # Merge: add any accounts from wider sheet not already present
        existing_ent = {c[0] for c in ent_cols_from_icm}
        existing_par = {c[0] for c in par_cols_from_icm}
        for col in ent_wide:
            if col[0] not in existing_ent:
                ent_cols_from_icm.append(col)
                existing_ent.add(col[0])
        for col in par_wide:
            if col[0] not in existing_par:
                par_cols_from_icm.append(col)
                existing_par.add(col[0])
        icm_header_map.update(icm_header_map_wide)

    # ── Read elimination accounts from report_inputs ─────────────────────
    plug_mapping = None
    plug_code = None
    if report_inputs_path:
        plug_mapping = parse_report_inputs(report_inputs_path)
        if plug_mapping:
            plug_code = plug_mapping.get("plug_code")

    # Merge elimination accounts into the column lists if not already present
    if plug_mapping:
        elim_codes = plug_mapping.get("elim_codes", set())
        existing_ent_codes = {c[0] for c in ent_cols_from_icm}
        existing_par_codes = {c[0] for c in par_cols_from_icm}
        for ec in sorted(elim_codes):
            if not ec or not ec[0].isdigit():
                continue
            series = classify_account(ec)
            if series == "EXTRA":
                series = "S1"  # Default for accounts starting with 6-9
            if ec not in existing_ent_codes:
                # Entity-side: S1 → Entity tag, S2 → Partner tag
                ent_tag = "Entity" if series == "S1" else "Partner"
                ent_cols_from_icm.append((ec, f"{ec}:{ec}", series, ent_tag))
                existing_ent_codes.add(ec)
            if ec not in existing_par_codes:
                # Partner-side: reversed tags
                par_tag = "Partner" if series == "S1" else "Entity"
                par_cols_from_icm.append((ec, f"{ec}:{ec}", series, par_tag))
                existing_par_codes.add(ec)

    # Build the set of valid account codes (ICM + elimination) for filtering
    valid_accounts = {c[0] for c in ent_cols_from_icm} | {c[0] for c in par_cols_from_icm}
    if plug_code:
        valid_accounts.add(plug_code)

    logger.info("Output columns: %d entity-side, %d partner-side",
                len(ent_cols_from_icm), len(par_cols_from_icm))
    logger.info("Valid account codes: %s", sorted(valid_accounts))

    # ── Read ICM data rows ───────────────────────────────────────────────
    icm_rows = read_icm_data(ws_icm, data_start=icm_data_start)

    # ── Read journals and build the normalized pair universe ─────────────
    journal_order = ["parent_journal", "contribution_journal", "plugaccount_journal"]
    updates_list_final = []
    all_entity_labels = {}
    all_icp_labels = {}
    all_pair_labels = {}
    journal_pair_keys = set()
    raw_plug_lines = None

    for jkey in journal_order:
        jpath = journal_paths.get(jkey)
        if not jpath:
            updates_list_final.append({})
            continue

        if jkey == "plugaccount_journal":
            updates, pair_keys, ent_lbls, icp_lbls, pair_lbls, raw_lines = read_ic_elimination_report(
                jpath, plug_code=plug_code
            )
            updates = {k: v for k, v in updates.items() if k[2] in valid_accounts}
            raw_plug_lines = raw_lines
            updates_list_final.append(updates)
            journal_pair_keys.update(pair_keys)
            all_entity_labels.update(ent_lbls)
            all_icp_labels.update(icp_lbls)
            all_pair_labels.update(pair_lbls)
            continue

        jmap = plug_mapping if jkey == "plugaccount_journal" else None
        lines = read_journal_lines(jpath, jmap)
        lookup_filtered = defaultdict(list)
        for line in lines:
            if line["account_code"] in valid_accounts:
                key = (line["entity_num"], line["icp_num"], line["account_code"])
                lookup_filtered[key].append(line)

        updates = match_journal_to_icm([], lookup_filtered)
        updates_list_final.append(updates)
        journal_pair_keys.update((ent, icp) for (ent, icp, _acct) in updates)

        ent_lbls, icp_lbls = read_journal_labels(jpath)
        all_entity_labels.update(ent_lbls)
        all_icp_labels.update(icp_lbls)
        all_pair_labels.update(read_journal_pair_labels(jpath))

    data_rows = build_consolidated_rows(
        icm_rows,
        journal_pair_keys,
        all_pair_labels,
        all_entity_labels,
        all_icp_labels,
    )
    if len(updates_list_final) > 2:
        updates_list_final[2] = canonicalize_updates_to_rows(updates_list_final[2], data_rows)
    synthetic_count = sum(1 for row in data_rows if row.get("is_synthetic"))
    if synthetic_count:
        logger.info("Built %d synthetic consolidated rows", synthetic_count)

    # ── Final validation: log unmatched journal keys ──────────────────────
    total_journal_keys = sum(len(u) for u in updates_list_final)
    matched_keys = 0
    covered_pairs_final = build_covered_pairs(data_rows)
    for upd in updates_list_final:
        for (e, p, a) in upd:
            if (e, p) in covered_pairs_final:
                matched_keys += 1
    unmatched = total_journal_keys - matched_keys
    if unmatched > 0:
        logger.warning("FINAL: %d journal keys unmatched (no ICM row)", unmatched)
    else:
        logger.info("FINAL: All %d journal keys matched to output rows", total_journal_keys)

    # ── Write output ─────────────────────────────────────────────────────
    write_output(ws_icm, data_rows, icm_header_map, updates_list_final, output_path,
                 plug_code, ent_cols=ent_cols_from_icm, par_cols=par_cols_from_icm,
                 raw_plug_lines=raw_plug_lines)
    return output_path


process_icm_report_v1 = process_icm_report

from app.ic_refactor import (  # noqa: E402
    build_all_facts,
    build_cell_ledger,
    build_pair_registry,
    build_plug_section_facts,
    compare_v1_v2_report31,
    derive_plug_facts,
    process_icm_report_v2,
    write_diagnostics_sheets,
    write_output_v2,
)


def process_icm_report(icm_path, journal_paths, output_path, report_inputs_path=None):
    return process_icm_report_v2(
        icm_path=icm_path,
        journal_paths=journal_paths,
        output_path=output_path,
        report_inputs_path=report_inputs_path,
    )
