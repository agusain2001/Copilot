from __future__ import annotations

import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


def _count_non_zero_numeric_cells(ws):
    total = 0
    for row in ws.iter_rows(min_row=33, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value != 0:
                total += 1
    return total


def _sheet_metrics(filepath: str) -> dict[str, int]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["ICM Matched"]
    pairs = set()
    row_count = 0
    for row_num in range(33, ws.max_row + 1):
        entity = str(ws.cell(row_num, 1).value or "").strip()
        partner = str(ws.cell(row_num, 2).value or "").strip()
        if not entity and not partner:
            continue
        row_count += 1
        pairs.add((entity, partner))
    return {
        "data_row_count": row_count,
        "canonical_pair_count": len(pairs),
        "non_zero_numeric_cells": _count_non_zero_numeric_cells(ws),
    }


def _collect_diagnostics_counts(filepath: str) -> dict[str, int]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    counts = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Diagnostics_"):
            continue
        ws = wb[sheet_name]
        counts[sheet_name] = max(ws.max_row - 1, 0)
    return counts


def _collect_sample_diffs(current_path: str, refactor_path: str, limit: int = 200) -> list[dict]:
    current_wb = openpyxl.load_workbook(current_path, data_only=True)
    refactor_wb = openpyxl.load_workbook(refactor_path, data_only=True)
    current_ws = current_wb["ICM Matched"]
    refactor_ws = refactor_wb["ICM Matched"]
    max_row = max(current_ws.max_row, refactor_ws.max_row)
    max_col = max(current_ws.max_column, refactor_ws.max_column)
    diffs = []
    for row_num in range(1, max_row + 1):
        for col_num in range(1, max_col + 1):
            current_value = current_ws.cell(row_num, col_num).value
            refactor_value = refactor_ws.cell(row_num, col_num).value
            if current_value == refactor_value:
                continue
            diffs.append(
                {
                    "row": row_num,
                    "column": col_num,
                    "current_value": current_value,
                    "refactor_value": refactor_value,
                }
            )
            if len(diffs) >= limit:
                return diffs
    return diffs


def _write_sheet(ws, headers, rows):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = row.get(header)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _available_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    try:
        with open(path, "ab"):
            return path
    except OSError:
        stem = path.stem
        suffix = path.suffix
        for idx in range(1, 100):
            candidate = path.with_name(f"{stem}_{idx}{suffix}")
            if not candidate.exists():
                return candidate
            try:
                with open(candidate, "ab"):
                    return candidate
            except OSError:
                continue
        return path.with_name(f"{stem}_latest{suffix}")


def compare_v1_v2_report31():
    base_dir = Path(r"g:\FCCS\backend\uploads\reports\31")
    inputs = base_dir / "inputs"
    output_dir = base_dir / "outputs" / "refactor_compare"
    output_dir.mkdir(parents=True, exist_ok=True)

    icm_path = str(inputs / "IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx")
    journal_paths = {
        "parent_journal": str(inputs / "Parent report.xlsx"),
        "contribution_journal": str(inputs / "Contribution report.xlsx"),
        "plugaccount_journal": str(inputs / "Journal Report (4).xlsx"),
    }
    report_inputs_path = str(inputs / "report Inputs.xlsx")
    current_output = str(_available_output_path(output_dir / "ICM_Output_current.xlsx"))
    refactor_output = str(_available_output_path(output_dir / "ICM_Output_refactor.xlsx"))
    comparison_output = str(_available_output_path(output_dir / "Comparison_Notes.xlsx"))

    from app.ic_processor import process_icm_report_v1
    from .pipeline import process_icm_report_v2

    process_icm_report_v1(icm_path, journal_paths, current_output, report_inputs_path=report_inputs_path)
    process_icm_report_v2(icm_path, journal_paths, refactor_output, report_inputs_path=report_inputs_path)

    current_metrics = _sheet_metrics(current_output)
    refactor_metrics = _sheet_metrics(refactor_output)
    refactor_diagnostics = _collect_diagnostics_counts(refactor_output)
    diffs = _collect_sample_diffs(current_output, refactor_output)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        {"metric": "data_row_count", "current": current_metrics["data_row_count"], "refactor": refactor_metrics["data_row_count"]},
        {"metric": "canonical_pair_count", "current": current_metrics["canonical_pair_count"], "refactor": refactor_metrics["canonical_pair_count"]},
        {"metric": "non_zero_numeric_cells", "current": current_metrics["non_zero_numeric_cells"], "refactor": refactor_metrics["non_zero_numeric_cells"]},
    ]
    _write_sheet(ws, ["metric", "current", "refactor"], summary_rows)

    diag_ws = wb.create_sheet("Diagnostics")
    diag_rows = [{"metric": key, "current": 0, "refactor": value} for key, value in sorted(refactor_diagnostics.items())]
    _write_sheet(diag_ws, ["metric", "current", "refactor"], diag_rows)

    diff_ws = wb.create_sheet("CellDiffs")
    _write_sheet(diff_ws, ["row", "column", "current_value", "refactor_value"], diffs)

    wb.save(comparison_output)
    return {
        "current_output": current_output,
        "refactor_output": refactor_output,
        "comparison_output": comparison_output,
    }
