from __future__ import annotations

import os

import openpyxl

from .models import WorkbookSource


def _normalize_text(value) -> str:
    return str(value or "").strip().lower()


def _extract_accountish_headers(ws, row_num: int) -> tuple[int, bool, int | None]:
    acct_headers = 0
    total_col = None
    for col_idx in range(3, ws.max_column + 1):
        text = str(ws.cell(row_num, col_idx).value or "").strip()
        if not text:
            continue
        if text.lower() == "total":
            total_col = col_idx
            continue
        if any(ch.isdigit() for ch in text):
            acct_headers += 1
    return acct_headers, total_col is not None, total_col


def detect_grid_source(filepath: str) -> WorkbookSource | None:
    if not filepath or not os.path.exists(filepath):
        return None
    wb = openpyxl.load_workbook(filepath, data_only=True)
    candidates: list[WorkbookSource] = []
    for ws in wb.worksheets:
        for row_num in range(1, min(ws.max_row, 60) + 1):
            if _normalize_text(ws.cell(row_num, 1).value) != "entity":
                continue
            if _normalize_text(ws.cell(row_num, 2).value) != "partner":
                continue
            acct_headers, has_total, total_col = _extract_accountish_headers(ws, row_num)
            if acct_headers == 0:
                continue
            kind = "ic_elim_grid" if has_total else "icm_grid"
            candidates.append(
                WorkbookSource(
                    filepath=filepath,
                    kind=kind,
                    sheet_name=ws.title,
                    header_row=row_num,
                    data_start=row_num + 1,
                    has_total_column=has_total,
                    total_column=total_col,
                )
            )
            break
    if not candidates:
        return None
    return max(
        candidates,
        key=lambda item: (
            1 if item.has_total_column else 0,
            wb[item.sheet_name].max_column,
            wb[item.sheet_name].max_row,
        ),
    )


def detect_journal_source(filepath: str) -> WorkbookSource | None:
    if not filepath or not os.path.exists(filepath):
        return None
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for ws in wb.worksheets:
        for row_num in range(1, min(ws.max_row, 60) + 1):
            headers = [_normalize_text(ws.cell(row_num, col).value) for col in range(1, min(ws.max_column, 8) + 1)]
            if "entity" in headers and "account" in headers and "intercompany" in headers:
                return WorkbookSource(
                    filepath=filepath,
                    kind="journal",
                    sheet_name=ws.title,
                    header_row=row_num,
                    data_start=row_num + 1,
                )
    return None


def detect_sources(icm_path: str, journal_paths: dict[str, str]) -> dict[str, WorkbookSource | None]:
    sources: dict[str, WorkbookSource | None] = {
        "base_grid": detect_grid_source(icm_path),
        "ic_elim_grid": None,
        "parent_journal": detect_journal_source(journal_paths.get("parent_journal", "")) if journal_paths.get("parent_journal") else None,
        "contribution_journal": detect_journal_source(journal_paths.get("contribution_journal", "")) if journal_paths.get("contribution_journal") else None,
        "plugaccount_journal": detect_journal_source(journal_paths.get("plugaccount_journal", "")) if journal_paths.get("plugaccount_journal") else None,
    }

    grid_candidates: list[WorkbookSource] = []
    if sources["base_grid"] is not None:
        grid_candidates.append(sources["base_grid"])
    plug_path = journal_paths.get("plugaccount_journal")
    if plug_path:
        plug_grid = detect_grid_source(plug_path)
        if plug_grid is not None:
            grid_candidates.append(plug_grid)

    ic_elim_candidates = [item for item in grid_candidates if item.has_total_column]
    if ic_elim_candidates:
        sources["ic_elim_grid"] = max(ic_elim_candidates, key=lambda item: (item.header_row, item.total_column or 0))
    elif sources["base_grid"] is not None and sources["base_grid"].has_total_column:
        sources["ic_elim_grid"] = sources["base_grid"]

    return sources
