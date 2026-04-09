from __future__ import annotations

import csv
import os
import re
from collections import defaultdict
from typing import Any

import openpyxl

from .config import PARTY_CODE_ALIASES
from .diagnostics import new_diagnostics, record_diagnostic
from .models import BaseRow, FactBuildResult, LayoutColumn, LayoutSpec, NormalizedFact
from .shape_detection import detect_sources


def to_float(value) -> float:
    if value in (None, "", " "):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def normalize_output_label(raw: str) -> str:
    text = str(raw or "").strip()
    if text and ":" in text and " - " not in text:
        text = text.replace(":", " - ", 1)
    return text


def normalize_account_code(raw: str) -> str | None:
    text = str(raw or "").strip()
    if not text:
        return None
    direct = re.match(r"(\d{6})(?![A-Za-z0-9])", text)
    if direct:
        return direct.group(1)
    bracket = re.search(r"]\.\[?([A-Za-z0-9_]+)\]?:", text)
    if bracket:
        value = bracket.group(1)
        if re.fullmatch(r"\d{6}", value):
            return value
    alt = re.search(r"]:(\d{6}):", text)
    if alt:
        return alt.group(1)
    middle = re.match(r"\d{6}\s*-\s*(\d{6})(?![A-Za-z0-9])", text)
    if middle:
        return middle.group(1)
    embedded = re.search(r"(?<![A-Za-z0-9])(\d{6})(?![A-Za-z0-9])", text)
    if embedded:
        return embedded.group(1)
    return None


def normalize_journal_account_code(raw: str) -> str | None:
    text = str(raw or "").strip()
    if not text:
        return None
    if not re.match(r"^\d{6}(?![A-Za-z0-9])", text):
        return None
    return normalize_account_code(text)


def normalize_party_code(
    raw: str,
    *,
    alias_map: dict[str, str] | None = None,
    diagnostics: dict[str, list[dict]] | None = None,
    source_file: str = "",
    source_row: int | None = None,
    field_name: str = "party",
) -> str | None:
    text = str(raw or "").strip()
    if not text:
        return None

    alias_map = alias_map or PARTY_CODE_ALIASES
    alias_key = text.upper()
    if alias_key in alias_map:
        value = alias_map[alias_key]
        if str(value).isdigit():
            return str(value).zfill(6) if len(str(value)) <= 6 else str(value)
        return None

    upper = alias_key
    if upper.startswith("FCCS_NO INTERCOMPANY") or upper.startswith("NO INTERCOMPANY"):
        return None

    match = re.match(r"^(?:ICP_)?E?(\d{5,6})(?![A-Za-z0-9])", upper)
    if match:
        return match.group(1).zfill(6)

    embedded = re.search(r"(?<![A-Za-z0-9])(?:ICP_)?E?(\d{5,6})(?![A-Za-z0-9])", upper)
    if embedded:
        return embedded.group(1).zfill(6)

    if any(ch.isdigit() for ch in upper):
        if diagnostics is not None:
            record_diagnostic(
                diagnostics,
                "bad_codes",
                f"unresolved_{field_name}_code",
                source_file=source_file,
                source_row=source_row,
                raw_entity=text if field_name == "entity" else "",
                raw_partner=text if field_name == "partner" else "",
                extra={"raw_value": text, "field_name": field_name},
            )
    return None


def classify_pair_direction(raw_entity: str) -> str:
    entity = str(raw_entity or "").strip()
    if entity.upper().startswith("E"):
        return "entity_to_partner"
    return "partner_to_entity"


def classify_account(code: str) -> str:
    text = str(code or "").strip()
    if not text or not text[0].isdigit():
        return "EXTRA"
    if text[0] in ("1", "5"):
        return "S1"
    if text[0] in ("2", "3", "4"):
        return "S2"
    return "EXTRA"


def apply_sign(debit: float, credit: float, account_code: str) -> float:
    code = str(account_code or "").strip()
    if not code:
        return debit - credit
    if code[0] in ("1", "5"):
        return debit - credit
    if code[0] in ("2", "3", "4"):
        return credit - debit
    return debit - credit


def parse_report_inputs(filepath: str | None) -> dict[str, Any]:
    if not filepath or not os.path.exists(filepath):
        return {"plug_code": None, "elim_codes": set()}
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    plug_code = None
    for cell in ws[1]:
        text = str(cell.value or "").strip()
        if "Plug Account:" not in text:
            continue
        plug_code = normalize_account_code(text)
        break

    elim_codes = set()
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        code = normalize_account_code(row[0])
        if code:
            elim_codes.add(code)

    return {"plug_code": plug_code, "elim_codes": elim_codes}


def _normalize_rate_type(raw: str) -> str | None:
    text = str(raw or "").strip().lower()
    if not text:
        return None
    if "avg" in text or text == "average":
        return "average"
    if "end" in text or text == "ending":
        return "ending"
    return None


def _fx_rate_type_for_account(account_code: str) -> str | None:
    code = str(account_code or "").strip()
    if not code or not code[0].isdigit():
        return None
    if code[0] in {"1", "2", "3"}:
        return "ending"
    if code[0] in {"4", "5"}:
        return "average"
    return None


def _parse_numeric_or_none(raw) -> float | None:
    if raw in (None, "", " "):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1].strip()
        try:
            return float(text) / 100.0
        except ValueError:
            return None
    try:
        return float(text)
    except ValueError:
        return None


def _find_header_row(rows: list[list], required_tokens: set[str]) -> tuple[int, dict[str, int]] | None:
    for idx, row in enumerate(rows):
        tokens = {str(cell or "").strip().lower() for cell in row}
        if not required_tokens.issubset(tokens):
            continue
        index_map = {}
        for col_idx, cell in enumerate(row):
            key = str(cell or "").strip().lower()
            if key:
                index_map[key] = col_idx
        return idx, index_map
    return None


def parse_entity_currency_lookup(
    filepath: str | None,
    alias_map: dict[str, str],
    diagnostics: dict[str, list[dict]],
) -> dict[str, str]:
    mapping: dict[str, str] = {}
    if not filepath or not os.path.exists(filepath):
        return mapping

    ext = os.path.splitext(filepath)[1].lower()
    rows: list[list] = []

    if ext == ".csv":
        with open(filepath, "r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
    else:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]

    hdr = _find_header_row(rows, {"entity"})
    if hdr is None:
        return mapping

    header_row, columns = hdr
    entity_col = columns.get("entity", 0)
    currency_col = None
    for key, idx in columns.items():
        if "currency" in key:
            currency_col = idx
            break
    if currency_col is None:
        return mapping

    for row_idx, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        entity_raw = row[entity_col] if entity_col < len(row) else None
        currency_raw = row[currency_col] if currency_col < len(row) else None
        entity_num = normalize_party_code(
            str(entity_raw or ""),
            alias_map=alias_map,
            diagnostics=diagnostics,
            source_file=filepath,
            source_row=row_idx,
            field_name="entity",
        )
        currency = str(currency_raw or "").strip().upper()
        if not entity_num or not currency:
            continue
        mapping[entity_num] = currency

    return mapping


def parse_ownership_lookup(
    filepath: str | None,
    alias_map: dict[str, str],
    diagnostics: dict[str, list[dict]],
) -> dict[str, float]:
    mapping: dict[str, float] = {}
    if not filepath or not os.path.exists(filepath):
        return mapping

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]

    pct_header = None
    header_row = 0
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            if "percent ownership" in str(cell or "").strip().lower():
                pct_header = col_idx
                header_row = row_idx
                break
        if pct_header is not None:
            break

    if pct_header is None:
        return mapping

    entity_col = 0
    header_values = [str(cell or "").strip().lower() for cell in rows[header_row]]
    for idx, value in enumerate(header_values):
        if value == "entity":
            entity_col = idx
            break

    for row_idx, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        entity_raw = row[entity_col] if entity_col < len(row) else None
        pct_raw = row[pct_header] if pct_header < len(row) else None
        pct = _parse_numeric_or_none(pct_raw)
        if pct is None:
            continue
        if pct > 1.0 and pct <= 100.0:
            pct = pct / 100.0
        entity_num = normalize_party_code(
            str(entity_raw or ""),
            alias_map=alias_map,
            diagnostics=diagnostics,
            source_file=filepath,
            source_row=row_idx,
            field_name="entity",
        )
        if not entity_num:
            continue
        mapping[entity_num] = pct

    return mapping


def parse_exchange_rates_lookup(filepath: str | None) -> dict[tuple[str, str], float]:
    rates: dict[tuple[str, str], float] = {}
    if not filepath or not os.path.exists(filepath):
        return rates

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]

    header_info = _find_header_row(rows, {"currency", "type"})
    if header_info is None:
        return rates

    header_row, columns = header_info
    currency_col = columns.get("currency")
    type_col = columns.get("type")
    if currency_col is None or type_col is None:
        return rates

    for row in rows[header_row + 1 :]:
        currency = str(row[currency_col] if currency_col < len(row) else "").strip().upper()
        rate_type = _normalize_rate_type(row[type_col] if type_col < len(row) else "")
        if not currency or rate_type is None:
            continue
        rate = None
        for idx, raw in enumerate(row):
            if idx in {currency_col, type_col}:
                continue
            rate = _parse_numeric_or_none(raw)
            if rate is not None:
                break
        if rate is None:
            continue
        rates[(currency, rate_type)] = rate

    return rates


def build_qar_facts_from_parent(
    parent_fact_ids: list[str],
    fact_registry: dict[str, NormalizedFact],
    entity_currency_map: dict[str, str],
    exchange_rates: dict[tuple[str, str], float],
    diagnostics: dict[str, list[dict]],
) -> list[NormalizedFact]:
    derived: list[NormalizedFact] = []
    for fact_id in parent_fact_ids:
        parent_fact = fact_registry[fact_id]
        currency = entity_currency_map.get(parent_fact.entity_num)
        if not currency:
            record_diagnostic(
                diagnostics,
                "lookup_fallbacks",
                "missing_entity_currency_default_rate_1_0",
                source_file=parent_fact.source_file,
                source_row=parent_fact.source_row,
                raw_entity=parent_fact.raw_entity,
                raw_partner=parent_fact.raw_partner,
                normalized_entity=parent_fact.entity_num,
                normalized_partner=parent_fact.partner_num,
                account=parent_fact.account_code,
                amount=parent_fact.amount,
            )
            currency = "UNKNOWN"

        rate_type = _fx_rate_type_for_account(parent_fact.account_code)
        if rate_type is None:
            record_diagnostic(
                diagnostics,
                "lookup_fallbacks",
                "unsupported_account_series_default_rate_1_0",
                source_file=parent_fact.source_file,
                source_row=parent_fact.source_row,
                raw_entity=parent_fact.raw_entity,
                raw_partner=parent_fact.raw_partner,
                normalized_entity=parent_fact.entity_num,
                normalized_partner=parent_fact.partner_num,
                account=parent_fact.account_code,
                amount=parent_fact.amount,
            )
            rate = 1.0
        else:
            rate = exchange_rates.get((currency, rate_type))
            if rate is None:
                record_diagnostic(
                    diagnostics,
                    "lookup_fallbacks",
                    "missing_exchange_rate_default_1_0",
                    source_file=parent_fact.source_file,
                    source_row=parent_fact.source_row,
                    raw_entity=parent_fact.raw_entity,
                    raw_partner=parent_fact.raw_partner,
                    normalized_entity=parent_fact.entity_num,
                    normalized_partner=parent_fact.partner_num,
                    account=parent_fact.account_code,
                    amount=parent_fact.amount,
                    extra={"currency": currency, "rate_type": rate_type},
                )
                rate = 1.0

        meta = dict(parent_fact.meta)
        meta.update(
            {
                "derived_from": "parent_fx",
                "currency": currency,
                "fx_rate_type": rate_type or "fallback_1_0",
                "fx_rate": rate,
                "source_fact_ids": [fact_id],
            }
        )
        derived.append(
            NormalizedFact(
                family="contrib",
                source_file=parent_fact.source_file,
                source_row=parent_fact.source_row,
                entity_num=parent_fact.entity_num,
                partner_num=parent_fact.partner_num,
                account_code=parent_fact.account_code,
                amount=parent_fact.amount * rate,
                direction=parent_fact.direction,
                raw_entity=parent_fact.raw_entity,
                raw_partner=parent_fact.raw_partner,
                raw_account=parent_fact.raw_account,
                meta=meta,
            )
        )
    return derived


def _get_sheet(source):
    wb = openpyxl.load_workbook(source.filepath, data_only=True)
    return wb[source.sheet_name]


def extract_layout(source, plug_code: str | None, elim_codes: set[str] | None = None) -> LayoutSpec:
    ws = _get_sheet(source)
    ent_cols: list[LayoutColumn] = []
    par_cols: list[LayoutColumn] = []
    seen: set[tuple[str, str]] = set()
    for col_idx in range(3, ws.max_column + 1):
        text = str(ws.cell(source.header_row, col_idx).value or "").strip()
        if not text or text.lower() == "total":
            continue
        code = normalize_account_code(text)
        if not code:
            continue
        tag = "Partner" if re.search(r"\bPartner\b", text) else "Entity"
        key = (code, tag)
        if key in seen:
            continue
        seen.add(key)
        column = LayoutColumn(code=code, description=text, series=classify_account(code), tag=tag)
        if tag == "Entity":
            ent_cols.append(column)
        else:
            par_cols.append(column)

    elim_codes = elim_codes or set()
    existing_ent = {column.code for column in ent_cols}
    existing_par = {column.code for column in par_cols}
    for code in sorted(elim_codes):
        series = classify_account(code)
        if series == "EXTRA":
            series = "S1"
        if code not in existing_ent:
            ent_tag = "Entity" if series == "S1" else "Partner"
            ent_cols.append(LayoutColumn(code=code, description=f"{code}:{code}", series=series, tag=ent_tag))
            existing_ent.add(code)
        if code not in existing_par:
            par_tag = "Partner" if series == "S1" else "Entity"
            par_cols.append(LayoutColumn(code=code, description=f"{code}:{code}", series=series, tag=par_tag))
            existing_par.add(code)

    plug_label_base = None
    if plug_code:
        plug_label_base = f"{plug_code}:Intercompany Balances Plug A/c"

    return LayoutSpec(ent_cols=ent_cols, par_cols=par_cols, plug_code=plug_code, plug_label_base=plug_label_base)


class FactCollector:
    def __init__(self):
        self.registry: dict[str, NormalizedFact] = {}
        self.counter = 0

    def add(self, fact: NormalizedFact) -> str:
        self.counter += 1
        fact_id = f"fact_{self.counter:06d}"
        self.registry[fact_id] = fact
        return fact_id


def _register_labels(label_maps: dict[str, Any], fact: NormalizedFact) -> None:
    entity_labels = label_maps["entity_labels"]
    partner_labels = label_maps["partner_labels"]
    pair_labels = label_maps["pair_labels"]
    journal_pair_votes = label_maps["journal_pair_votes"]

    if fact.entity_num and fact.raw_entity:
        entity_labels.setdefault(fact.entity_num, normalize_output_label(fact.raw_entity))
    if fact.partner_num:
        partner_value = normalize_output_label(fact.raw_partner) if fact.raw_partner else f"ICP_{fact.partner_num}"
        partner_labels.setdefault(fact.partner_num, partner_value)
    if fact.entity_num and fact.partner_num and fact.family in {"parent", "contrib", "plug_source"}:
        pair_labels.setdefault(
            (fact.entity_num, fact.partner_num),
            {
                "entity": normalize_output_label(fact.raw_entity) if fact.raw_entity else fact.entity_num,
                "partner": normalize_output_label(fact.raw_partner) if fact.raw_partner else f"ICP_{fact.partner_num}",
                "family": fact.family,
            },
        )
        journal_pair_votes[(fact.entity_num, fact.partner_num)] += 1


def _iter_account_headers(ws, header_row: int):
    headers = []
    for col_idx in range(3, ws.max_column + 1):
        text = str(ws.cell(header_row, col_idx).value or "").strip()
        if not text:
            continue
        if text.lower() == "total":
            headers.append((col_idx, text, None, "Total"))
            continue
        code = normalize_account_code(text)
        if not code:
            continue
        tag = "Partner" if re.search(r"\bPartner\b", text) else "Entity"
        headers.append((col_idx, text, code, tag))
    return headers


def build_icm_facts(
    source,
    collector: FactCollector,
    label_maps: dict[str, Any],
    diagnostics: dict[str, list[dict]],
    alias_map: dict[str, str],
) -> tuple[list[str], list[BaseRow]]:
    ws = _get_sheet(source)
    facts: list[str] = []
    base_rows: list[BaseRow] = []
    headers = [item for item in _iter_account_headers(ws, source.header_row) if item[3] != "Total"]

    for row_num in range(source.data_start, ws.max_row + 1):
        entity_raw = str(ws.cell(row_num, 1).value or "").strip()
        partner_raw = str(ws.cell(row_num, 2).value or "").strip()
        if not entity_raw and not partner_raw:
            continue

        entity_num = normalize_party_code(
            entity_raw,
            alias_map=alias_map,
            diagnostics=diagnostics,
            source_file=source.filepath,
            source_row=row_num,
            field_name="entity",
        )
        partner_num = normalize_party_code(
            partner_raw,
            alias_map=alias_map,
            diagnostics=diagnostics,
            source_file=source.filepath,
            source_row=row_num,
            field_name="partner",
        )
        if not entity_num or not partner_num:
            record_diagnostic(
                diagnostics,
                "unmatched_facts",
                "invalid_base_pair",
                source_file=source.filepath,
                source_row=row_num,
                raw_entity=entity_raw,
                raw_partner=partner_raw,
                normalized_entity=entity_num,
                normalized_partner=partner_num,
            )
            continue

        pair_group = tuple(sorted((entity_num, partner_num)))
        base_rows.append(
            BaseRow(
                pair_key=(entity_num, partner_num),
                pair_group=pair_group,
                row_num=row_num,
                row_direction=classify_pair_direction(entity_raw),
                entity_label=normalize_output_label(entity_raw),
                partner_label=normalize_output_label(partner_raw),
                source_file=source.filepath,
            )
        )

        for col_idx, header_text, account_code, tag in headers:
            amount = to_float(ws.cell(row_num, col_idx).value)
            if amount == 0 or not account_code:
                continue
            if tag == "Entity":
                fact = NormalizedFact(
                    family="base_icm",
                    source_file=source.filepath,
                    source_row=row_num,
                    entity_num=entity_num,
                    partner_num=partner_num,
                    account_code=account_code,
                    amount=amount,
                    direction="entity_to_partner",
                    raw_entity=entity_raw,
                    raw_partner=partner_raw,
                    raw_account=header_text,
                    meta={"column_tag": tag, "sheet_name": source.sheet_name},
                )
            else:
                fact = NormalizedFact(
                    family="base_icm",
                    source_file=source.filepath,
                    source_row=row_num,
                    entity_num=partner_num,
                    partner_num=entity_num,
                    account_code=account_code,
                    amount=amount,
                    direction="partner_to_entity",
                    raw_entity=entity_raw,
                    raw_partner=partner_raw,
                    raw_account=header_text,
                    meta={"column_tag": tag, "sheet_name": source.sheet_name},
                )
            fact_id = collector.add(fact)
            facts.append(fact_id)
            _register_labels(label_maps, fact)

    return facts, base_rows


def _journal_indices(ws, header_row: int) -> dict[str, int]:
    headers = [str(cell.value or "").strip().lower() for cell in ws[header_row]]
    mapping = {value: idx + 1 for idx, value in enumerate(headers)}
    return {
        "entity": mapping.get("entity", 3),
        "account": mapping.get("account", 4),
        "intercompany": mapping.get("intercompany", 5),
        "debit": mapping.get("debit", 15),
        "credit": mapping.get("credit", 16),
        "label": 1,
    }


def build_journal_facts(
    source,
    family: str,
    collector: FactCollector,
    label_maps: dict[str, Any],
    diagnostics: dict[str, list[dict]],
    alias_map: dict[str, str],
    *,
    plug_code: str | None = None,
    elim_codes: set[str] | None = None,
    plug_remap: bool = False,
    ownership_lookup: dict[str, float] | None = None,
    apply_ownership: bool = False,
    require_direct_account_format: bool = False,
) -> list[str]:
    ws = _get_sheet(source)
    indices = _journal_indices(ws, source.header_row)
    elim_codes = elim_codes or set()
    facts: list[str] = []

    for row_num in range(source.data_start, ws.max_row + 1):
        label = str(ws.cell(row_num, indices["label"]).value or "").strip()
        if label == "Grand Total":
            break

        entity_raw = str(ws.cell(row_num, indices["entity"]).value or "").strip()
        account_raw = str(ws.cell(row_num, indices["account"]).value or "").strip()
        partner_raw = str(ws.cell(row_num, indices["intercompany"]).value or "").strip()
        if not entity_raw and not account_raw and not partner_raw:
            continue

        raw_account_code = normalize_account_code(account_raw)

        entity_num = normalize_party_code(
            entity_raw,
            alias_map=alias_map,
            diagnostics=diagnostics,
            source_file=source.filepath,
            source_row=row_num,
            field_name="entity",
        )
        partner_num = normalize_party_code(
            partner_raw,
            alias_map=alias_map,
            diagnostics=diagnostics,
            source_file=source.filepath,
            source_row=row_num,
            field_name="partner",
        )

        if require_direct_account_format and account_raw and normalize_journal_account_code(account_raw) is None:
            debit = to_float(ws.cell(row_num, indices["debit"]).value)
            credit = to_float(ws.cell(row_num, indices["credit"]).value)
            amount = apply_sign(debit, credit, raw_account_code or "")
            record_diagnostic(
                diagnostics,
                "unmatched_facts",
                "unsupported_journal_account_format",
                source_file=source.filepath,
                source_row=row_num,
                raw_entity=entity_raw,
                raw_partner=partner_raw,
                normalized_entity=entity_num,
                normalized_partner=partner_num,
                account=account_raw,
                amount=amount,
                extra={"family": family, "label": label},
            )
            continue

        account_code = raw_account_code
        if plug_remap and plug_code and (account_code in elim_codes or account_code is None):
            account_code = plug_code

        debit = to_float(ws.cell(row_num, indices["debit"]).value)
        credit = to_float(ws.cell(row_num, indices["credit"]).value)
        amount = apply_sign(debit, credit, account_code or "")

        ownership_pct = 1.0
        if apply_ownership and entity_num:
            if ownership_lookup and entity_num in ownership_lookup:
                ownership_pct = ownership_lookup[entity_num]
            else:
                record_diagnostic(
                    diagnostics,
                    "lookup_fallbacks",
                    "missing_ownership_default_1_0",
                    source_file=source.filepath,
                    source_row=row_num,
                    raw_entity=entity_raw,
                    raw_partner=partner_raw,
                    normalized_entity=entity_num,
                    normalized_partner=partner_num,
                    account=account_code or account_raw,
                    amount=amount,
                )
            amount *= ownership_pct

        if amount == 0:
            continue
        if not entity_num or not partner_num or not account_code:
            record_diagnostic(
                diagnostics,
                "unmatched_facts",
                "invalid_journal_fact",
                source_file=source.filepath,
                source_row=row_num,
                raw_entity=entity_raw,
                raw_partner=partner_raw,
                normalized_entity=entity_num,
                normalized_partner=partner_num,
                account=account_raw,
                amount=amount,
                extra={"family": family, "label": label},
            )
            continue

        fact = NormalizedFact(
            family=family,
            source_file=source.filepath,
            source_row=row_num,
            entity_num=entity_num,
            partner_num=partner_num,
            account_code=account_code,
            amount=amount,
            direction=classify_pair_direction(entity_raw),
            raw_entity=entity_raw,
            raw_partner=partner_raw,
            raw_account=account_raw,
            meta={"label": label, "debit": debit, "credit": credit, "ownership_pct": ownership_pct},
        )
        fact_id = collector.add(fact)
        facts.append(fact_id)
        _register_labels(label_maps, fact)

    return facts


def build_ic_elim_facts(
    source,
    collector: FactCollector,
    label_maps: dict[str, Any],
    diagnostics: dict[str, list[dict]],
    alias_map: dict[str, str],
    *,
    plug_code: str | None = None,
) -> list[str]:
    ws = _get_sheet(source)
    facts: list[str] = []
    headers = _iter_account_headers(ws, source.header_row)

    for row_num in range(source.data_start, ws.max_row + 1):
        entity_raw = str(ws.cell(row_num, 1).value or "").strip()
        partner_raw = str(ws.cell(row_num, 2).value or "").strip()
        if not entity_raw and not partner_raw:
            continue

        entity_num = normalize_party_code(
            entity_raw,
            alias_map=alias_map,
            diagnostics=diagnostics,
            source_file=source.filepath,
            source_row=row_num,
            field_name="entity",
        )
        partner_num = normalize_party_code(
            partner_raw,
            alias_map=alias_map,
            diagnostics=diagnostics,
            source_file=source.filepath,
            source_row=row_num,
            field_name="partner",
        )
        if not entity_num or not partner_num:
            record_diagnostic(
                diagnostics,
                "unmatched_facts",
                "invalid_ic_elim_pair",
                source_file=source.filepath,
                source_row=row_num,
                raw_entity=entity_raw,
                raw_partner=partner_raw,
                normalized_entity=entity_num,
                normalized_partner=partner_num,
            )
            continue

        for col_idx, header_text, account_code, tag in headers:
            value = to_float(ws.cell(row_num, col_idx).value)
            if value == 0:
                continue

            if tag == "Total":
                if not plug_code:
                    continue
                fact = NormalizedFact(
                    family="ic_elim",
                    source_file=source.filepath,
                    source_row=row_num,
                    entity_num=entity_num,
                    partner_num=partner_num,
                    account_code=plug_code,
                    amount=value,
                    direction="entity_to_partner",
                    raw_entity=entity_raw,
                    raw_partner=partner_raw,
                    raw_account=header_text,
                    meta={"column_role": "total"},
                )
            elif tag == "Entity":
                fact = NormalizedFact(
                    family="ic_elim",
                    source_file=source.filepath,
                    source_row=row_num,
                    entity_num=entity_num,
                    partner_num=partner_num,
                    account_code=account_code,
                    amount=value,
                    direction="entity_to_partner",
                    raw_entity=entity_raw,
                    raw_partner=partner_raw,
                    raw_account=header_text,
                    meta={"column_role": "entity"},
                )
            else:
                fact = NormalizedFact(
                    family="ic_elim",
                    source_file=source.filepath,
                    source_row=row_num,
                    entity_num=partner_num,
                    partner_num=entity_num,
                    account_code=account_code,
                    amount=value,
                    direction="partner_to_entity",
                    raw_entity=entity_raw,
                    raw_partner=partner_raw,
                    raw_account=header_text,
                    meta={"column_role": "partner"},
                )
            fact_id = collector.add(fact)
            facts.append(fact_id)
            _register_labels(label_maps, fact)

    return facts


def build_all_facts(
    icm_path: str,
    journal_paths: dict[str, str],
    report_inputs_path: str | None = None,
    lookup_paths: dict[str, str] | None = None,
    alias_map: dict[str, str] | None = None,
) -> FactBuildResult:
    alias_map = alias_map or PARTY_CODE_ALIASES
    diagnostics = new_diagnostics()
    lookup_paths = lookup_paths or {}
    sources = detect_sources(icm_path, journal_paths)
    base_source = sources.get("base_grid")
    if base_source is None:
        raise ValueError(f"No base grid detected for {icm_path}")

    report_inputs = parse_report_inputs(report_inputs_path)
    plug_code = report_inputs.get("plug_code")
    elim_codes = set(report_inputs.get("elim_codes", set()))
    layout = extract_layout(base_source, plug_code, elim_codes)
    entity_currency_map = parse_entity_currency_lookup(
        lookup_paths.get("entity_with_currency"),
        alias_map,
        diagnostics,
    )
    ownership_lookup = parse_ownership_lookup(
        lookup_paths.get("ownership_structure"),
        alias_map,
        diagnostics,
    )
    exchange_rates = parse_exchange_rates_lookup(lookup_paths.get("exchange_rates"))

    collector = FactCollector()
    label_maps = {
        "entity_labels": {},
        "partner_labels": {},
        "pair_labels": {},
        "journal_pair_votes": defaultdict(int),
    }

    facts_icm, base_rows = build_icm_facts(base_source, collector, label_maps, diagnostics, alias_map)
    facts_parent = []
    if sources.get("parent_journal") is not None:
        facts_parent = build_journal_facts(
            sources["parent_journal"],
            "parent",
            collector,
            label_maps,
            diagnostics,
            alias_map,
            ownership_lookup=ownership_lookup,
            apply_ownership=True,
            require_direct_account_format=True,
        )

    if sources.get("contribution_journal") is not None:
        record_diagnostic(
            diagnostics,
            "lookup_fallbacks",
            "contribution_journal_ignored_qar_mode",
            source_file=sources["contribution_journal"].filepath,
            source_row=None,
            raw_entity="",
            raw_partner="",
            normalized_entity=None,
            normalized_partner=None,
            account="",
            amount=None,
        )

    facts_contrib: list[str] = []
    for fact in build_qar_facts_from_parent(
        facts_parent,
        collector.registry,
        entity_currency_map,
        exchange_rates,
        diagnostics,
    ):
        fact_id = collector.add(fact)
        facts_contrib.append(fact_id)
        _register_labels(label_maps, fact)

    facts_plug_source = []
    if sources.get("plugaccount_journal") is not None:
        facts_plug_source = build_journal_facts(
            sources["plugaccount_journal"],
            "plug_source",
            collector,
            label_maps,
            diagnostics,
            alias_map,
            plug_code=plug_code,
            elim_codes=elim_codes,
            plug_remap=True,
        )

    facts_ic_elim = []
    ic_elim_source = sources.get("ic_elim_grid")
    if ic_elim_source is not None:
        facts_ic_elim = build_ic_elim_facts(
            ic_elim_source,
            collector,
            label_maps,
            diagnostics,
            alias_map,
            plug_code=plug_code,
        )

    return FactBuildResult(
        fact_registry=collector.registry,
        facts_icm=facts_icm,
        facts_parent=facts_parent,
        facts_contrib=facts_contrib,
        facts_plug_source=facts_plug_source,
        facts_ic_elim=facts_ic_elim,
        base_rows=base_rows,
        layout=layout,
        label_maps=label_maps,
        diagnostics=diagnostics,
        plug_code=plug_code,
        elim_codes=elim_codes,
        sources=sources,
    )
