from __future__ import annotations

from collections import defaultdict

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
NUM_FORMAT = r"###,##0;\-###,##0"


def new_diagnostics() -> dict[str, list[dict]]:
    return defaultdict(list)


def record_diagnostic(
    diagnostics: dict[str, list[dict]],
    category: str,
    reason: str,
    *,
    source_file: str = "",
    source_row: int | None = None,
    raw_entity: str = "",
    raw_partner: str = "",
    normalized_entity: str | None = None,
    normalized_partner: str | None = None,
    account: str = "",
    amount: float | None = None,
    extra: dict | None = None,
) -> None:
    record = {
        "source_file": source_file,
        "source_row": source_row,
        "raw_entity": raw_entity,
        "raw_partner": raw_partner,
        "normalized_entity": normalized_entity,
        "normalized_partner": normalized_partner,
        "account": account,
        "amount": amount,
        "reason": reason,
    }
    if extra:
        record.update(extra)
    diagnostics[category].append(record)


def diagnostics_totals(diagnostics: dict[str, list[dict]]) -> dict[str, int]:
    return {category: len(records) for category, records in sorted(diagnostics.items())}


def _write_table_sheet(workbook, title: str, rows: list[dict]) -> None:
    ws = workbook.create_sheet(title)
    headers = [
        "source_file",
        "source_row",
        "raw_entity",
        "raw_partner",
        "normalized_entity",
        "normalized_partner",
        "account",
        "amount",
        "reason",
        "details",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    for row_idx, record in enumerate(rows, start=2):
        details = {
            key: value
            for key, value in record.items()
            if key not in headers[:-1]
        }
        values = [
            record.get("source_file"),
            record.get("source_row"),
            record.get("raw_entity"),
            record.get("raw_partner"),
            record.get("normalized_entity"),
            record.get("normalized_partner"),
            record.get("account"),
            record.get("amount"),
            record.get("reason"),
            str(details) if details else "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 8 and isinstance(value, (int, float)):
                cell.number_format = NUM_FORMAT


def write_diagnostics_sheets(
    workbook,
    diagnostics: dict[str, list[dict]],
    plug_reconciliation_log: list[dict],
    fact_assignment_log: dict[str, dict],
) -> None:
    _write_table_sheet(workbook, "Diagnostics_Pairs", diagnostics.get("pairs", []))
    _write_table_sheet(workbook, "Diagnostics_UnmatchedFacts", diagnostics.get("unmatched_facts", []))
    _write_table_sheet(workbook, "Diagnostics_Aliases", diagnostics.get("bad_codes", []))
    _write_table_sheet(workbook, "Diagnostics_Lookups", diagnostics.get("lookup_fallbacks", []))
    _write_table_sheet(workbook, "Diagnostics_PlugRecon", plug_reconciliation_log)

    assignment_rows = []
    for fact_id, payload in sorted(fact_assignment_log.items()):
        assignment_rows.append(
            {
                "source_file": payload.get("source_file"),
                "source_row": payload.get("source_row"),
                "raw_entity": payload.get("raw_entity"),
                "raw_partner": payload.get("raw_partner"),
                "normalized_entity": payload.get("entity_num"),
                "normalized_partner": payload.get("partner_num"),
                "account": payload.get("account_code"),
                "amount": payload.get("amount"),
                "reason": payload.get("reason", ""),
                "fact_id": fact_id,
                "destination": payload.get("destination"),
            }
        )
    _write_table_sheet(workbook, "Diagnostics_Assignment", assignment_rows)
