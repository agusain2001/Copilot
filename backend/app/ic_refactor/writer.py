from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .diagnostics import write_diagnostics_sheets


ICM_OUTPUT_HEADER_ROW = 32
ICM_OUTPUT_DATA_START = 33
SECTION_LABEL_ROW = 29

HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
ID_FILL = PatternFill(start_color="C8DCF0", end_color="C8DCF0", fill_type="solid")
VARIANCE_FILL = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
MATCH_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
SECTION_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
ID_FONT = Font(bold=True, size=11)
DATA_FONT = Font(size=11)
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
NUM_FORMAT = r"###,##0;\-###,##0"

DATA_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"))
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


def _style_header_cell(cell, text):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_id_cell(cell, text):
    cell.value = text
    cell.font = ID_FONT
    cell.fill = ID_FILL
    cell.border = DATA_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _style_data_cell(cell, value, fill=None):
    cell.value = value
    cell.font = DATA_FONT
    cell.fill = fill if fill is not None else NO_FILL
    cell.border = DATA_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if isinstance(value, (int, float)):
        cell.number_format = NUM_FORMAT


def _block_positions(start, num_accts_ent, num_accts_par, has_plug=False):
    ent_start = start
    var1 = start + num_accts_ent
    par_start = var1 + 1
    var2 = par_start + num_accts_par
    total = var2 + 1
    if has_plug:
        plug = total + 1
        spacer = plug + 1
    else:
        plug = None
        spacer = total + 1
    return {
        "ent_start": ent_start,
        "var1": var1,
        "par_start": par_start,
        "var2": var2,
        "total": total,
        "plug": plug,
        "spacer": spacer,
    }


def _get_map_value(mapping, source_row_num, side, account_code):
    if source_row_num is None:
        return None
    return mapping.get((source_row_num, side, account_code))


def _get_ledger_value(cell_ledger, row_key, block, side, account_code):
    routed = cell_ledger.get((row_key, block, side, account_code))
    return None if routed is None else routed.amount


def write_output_v2(
    row_registry,
    cell_ledger,
    base_value_map,
    output_path,
    layout,
    diagnostics,
    plug_reconciliation_log,
    fact_assignment_log,
):
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "ICM Matched"

    n_ent = len(layout.ent_cols)
    n_par = len(layout.par_cols)

    blk_base = _block_positions(3, n_ent, n_par, has_plug=False)
    blk_par = _block_positions(blk_base["spacer"] + 1, n_ent, n_par, has_plug=True)
    blk_cont = _block_positions(blk_par["spacer"] + 1, n_ent, n_par, has_plug=True)
    plug_section_start = blk_cont["spacer"] + 1
    plug_section = {"plug_col": plug_section_start, "total": plug_section_start + 1, "spacer": plug_section_start + 2}
    col_final = plug_section["spacer"] + 1

    spacer_cols = {blk_base["spacer"], blk_par["spacer"], blk_cont["spacer"], plug_section["spacer"]}
    total_cols = col_final
    for col_num in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 4 if col_num in spacer_cols else 17.5714

    ws.row_dimensions[ICM_OUTPUT_HEADER_ROW].height = 78.75

    def _section_label(start, end, text):
        ws.merge_cells(start_row=SECTION_LABEL_ROW, start_column=start, end_row=SECTION_LABEL_ROW, end_column=end)
        cell = ws.cell(SECTION_LABEL_ROW, start)
        cell.value = text
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    _section_label(blk_par["ent_start"], blk_par["plug"] or blk_par["total"], "Parent Input")
    _section_label(blk_cont["ent_start"], blk_cont["plug"] or blk_cont["total"], "Contribution Input")
    _section_label(plug_section["plug_col"], plug_section["total"], "Plug Account")

    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, 1), "Entity")
    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, 2), "Partner")

    plug_label_base = layout.plug_label_base or "Intercompany Balances Plug A/c"

    def _write_block_headers(blk, has_plug_hdr=False, plug_hdr_text=None):
        for idx, column in enumerate(layout.ent_cols):
            label = f"{column.code} - {column.description} {column.tag}" if column.tag not in column.description else column.description
            _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["ent_start"] + idx), label)
        _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["var1"]), "Variance")

        for idx, column in enumerate(layout.par_cols):
            label = f"{column.code} - {column.description} {column.tag}" if column.tag not in column.description else column.description
            _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["par_start"] + idx), label)
        _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["var2"]), "Variance")
        _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["total"]), "Total")
        if has_plug_hdr and blk["plug"] is not None:
            _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, blk["plug"]), plug_hdr_text or plug_label_base)

    _write_block_headers(blk_base)
    _write_block_headers(blk_par, has_plug_hdr=True, plug_hdr_text=f"{plug_label_base}\n(Entity \u2192 Parent)")
    _write_block_headers(blk_cont, has_plug_hdr=True, plug_hdr_text=f"{plug_label_base}\n(Parent \u2192 Entity)")
    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, plug_section["plug_col"]), plug_label_base)
    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, plug_section["total"]), "Total")
    _style_header_cell(ws.cell(ICM_OUTPUT_HEADER_ROW, col_final), "Final Total")

    s1_ent_codes = [column.code for column in layout.ent_cols if column.series == "S1"]
    s2_ent_codes = [column.code for column in layout.ent_cols if column.series == "S2"]
    s1_par_codes = [column.code for column in layout.par_cols if column.series == "S1"]
    s2_par_codes = [column.code for column in layout.par_cols if column.series == "S2"]

    def _sum_codes(values, codes):
        return sum(float(values.get(code, 0) or 0) for code in codes)

    for out_idx, row in enumerate(row_registry):
        out_row = ICM_OUTPUT_DATA_START + out_idx
        row_key = row.row_key
        _style_id_cell(ws.cell(out_row, 1), row.display_entity)
        _style_id_cell(ws.cell(out_row, 2), row.display_partner)

        def _write_block(blk, block_name, source):
            ent_values = {}
            par_values = {}
            for idx, column in enumerate(layout.ent_cols):
                value = _get_map_value(base_value_map, row.source_row_num, "entity_side", column.code) if source == "base" else _get_ledger_value(cell_ledger, row_key, block_name, "entity_side", column.code)
                _style_data_cell(ws.cell(out_row, blk["ent_start"] + idx), value, MATCH_FILL if value not in (None, 0) and source != "base" else None)
                ent_values[column.code] = value

            var1 = _sum_codes(ent_values, s1_ent_codes) - _sum_codes(ent_values, s2_ent_codes)
            _style_data_cell(ws.cell(out_row, blk["var1"]), var1, VARIANCE_FILL)

            for idx, column in enumerate(layout.par_cols):
                value = _get_map_value(base_value_map, row.source_row_num, "partner_side", column.code) if source == "base" else _get_ledger_value(cell_ledger, row_key, block_name, "partner_side", column.code)
                _style_data_cell(ws.cell(out_row, blk["par_start"] + idx), value, MATCH_FILL if value not in (None, 0) and source != "base" else None)
                par_values[column.code] = value

            var2 = _sum_codes(par_values, s1_par_codes) - _sum_codes(par_values, s2_par_codes)
            _style_data_cell(ws.cell(out_row, blk["var2"]), var2, VARIANCE_FILL)
            total = var1 + var2
            _style_data_cell(ws.cell(out_row, blk["total"]), total, TOTAL_FILL)
            return total

        base_total = _write_block(blk_base, "base", "base")
        parent_total = _write_block(blk_par, "parent", "ledger")
        contrib_total = _write_block(blk_cont, "contrib", "ledger")

        plug_parent = _get_ledger_value(cell_ledger, row_key, "plug", "plug_parent", layout.plug_code) if layout.plug_code else None
        _style_data_cell(ws.cell(out_row, blk_par["plug"]), plug_parent, MATCH_FILL if plug_parent not in (None, 0) else None)

        plug_contrib = _get_ledger_value(cell_ledger, row_key, "plug", "plug_contrib", layout.plug_code) if layout.plug_code else None
        _style_data_cell(ws.cell(out_row, blk_cont["plug"]), plug_contrib, MATCH_FILL if plug_contrib not in (None, 0) else None)

        plug_section_value = _get_ledger_value(cell_ledger, row_key, "plug", "plug_section", layout.plug_code) if layout.plug_code else None
        _style_data_cell(ws.cell(out_row, plug_section["plug_col"]), plug_section_value, MATCH_FILL if plug_section_value not in (None, 0) else None)
        _style_data_cell(ws.cell(out_row, plug_section["total"]), plug_section_value, TOTAL_FILL)

        final_total = (
            float(base_total or 0)
            + float(parent_total or 0)
            + float(contrib_total or 0)
            + float(plug_parent or 0)
            + float(plug_contrib or 0)
            + float(plug_section_value or 0)
        )
        _style_data_cell(ws.cell(out_row, col_final), final_total, TOTAL_FILL)

    write_diagnostics_sheets(out_wb, diagnostics, plug_reconciliation_log, fact_assignment_log)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    out_wb.save(output_path)
    return output_path
