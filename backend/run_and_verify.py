"""Run processor and verify ALL accounts are correctly matched."""
import sys, os
sys.path.insert(0, r'G:\FCCS\backend')
from app.ic_processor import (
    process_icm_report, read_journal_report, match_journal_to_icm,
    read_icm_data, read_icm_headers, ICM_INPUT_HEADER_ROW,
    BLK_PAR, BLK_CONT, FIXED_ENT_COLS, FIXED_PAR_COLS, ICM_OUTPUT_DATA_START
)
import openpyxl

# ── Correct file paths ──
icm = r'G:\FCCS\backend\uploads\reports\22\inputs\Intercompany Balances IC Matching Report (1).xlsx'
inputs = r'G:\FCCS\backend\uploads\reports\22\inputs\report_inputs.xlsx'
journals = {
    "parent_journal":       r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (1).xlsx',
    "contribution_journal": r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (2).xlsx',
    "plugaccount_journal":  r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (4).xlsx'
}
out = r'G:\FCCS\backend\uploads\reports\22\outputs\ICM_Output_22_final.xlsx'

# ── Run the processor ──
print("Running ICM processor...")
try:
    res = process_icm_report(icm, journals, out, inputs)
    print(f"Saved to: {res}")
except PermissionError:
    out = r'G:\FCCS\backend\uploads\reports\22\outputs\ICM_Output_22_v2.xlsx'
    res = process_icm_report(icm, journals, out, inputs)
    print(f"(Original locked) Saved to: {res}")

# ── Verify ALL accounts ──
print("\nVerifying output...")
wb = openpyxl.load_workbook(out, data_only=True)
ws = wb.active

with open(r'G:\FCCS\backend\verify_all_accounts.txt', 'w', encoding='utf-8') as f:
    f.write("COMPLETE VERIFICATION - ALL ACCOUNTS\n")
    f.write("=" * 100 + "\n\n")
    
    # First show ALL primary and fallback updates from all journals
    f.write("JOURNAL MATCHING RESULTS:\n")
    f.write("-" * 80 + "\n")
    
    wb_icm = openpyxl.load_workbook(icm, data_only=True)
    candidates = []
    for sheet in wb_icm.worksheets:
        c1 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 1).value or "").strip()
        c2 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 2).value or "").strip()
        if c1.lower() == "entity" and c2.lower() == "partner":
            candidates.append(sheet)
    ws_icm = min(candidates, key=lambda s: s.max_column) if candidates else wb_icm.active
    data_rows = read_icm_data(ws_icm)
    
    for jkey, jpath in journals.items():
        primary, fallback = read_journal_report(jpath)
        pri_up, fb_up = match_journal_to_icm(data_rows, primary, fallback)
        
        f.write(f"\n  {jkey} ({os.path.basename(jpath)}):\n")
        f.write(f"    Primary ({len(pri_up)} entries):\n")
        for k, v in sorted(pri_up.items()):
            f.write(f"      Entity={k[0]} Partner={k[1]} Account={k[2]} -> {v:,.2f}\n")
        f.write(f"    Fallback ({len(fb_up)} entries):\n")
        for k, v in sorted(fb_up.items()):
            f.write(f"      Entity={k[0]} Partner={k[1]} Account={k[2]} -> {v:,.2f}\n")
    
    # Now verify output rows - check ALL entities that have ANY journal values
    f.write(f"\n\n{'='*100}\n")
    f.write("OUTPUT FILE - ALL ROWS WITH JOURNAL VALUES:\n")
    f.write("=" * 100 + "\n")
    
    for r in range(ICM_OUTPUT_DATA_START, ws.max_row + 1):
        entity = str(ws.cell(r, 1).value or "").strip()
        partner = str(ws.cell(r, 2).value or "").strip()
        
        # Check parent entity-side
        par_ent_vals = []
        for i, (code, desc, _, tag) in enumerate(FIXED_ENT_COLS):
            v = ws.cell(r, BLK_PAR[0] + i).value
            if v is not None:
                par_ent_vals.append(f"{code}({tag})={v}")
        
        # Check parent partner-side
        par_par_vals = []
        for i, (code, desc, _, tag) in enumerate(FIXED_PAR_COLS):
            v = ws.cell(r, BLK_PAR[2] + i).value
            if v is not None:
                par_par_vals.append(f"{code}({tag})={v}")
        
        # Check contrib entity-side
        con_ent_vals = []
        for i, (code, desc, _, tag) in enumerate(FIXED_ENT_COLS):
            v = ws.cell(r, BLK_CONT[0] + i).value
            if v is not None:
                con_ent_vals.append(f"{code}({tag})={v}")
        
        # Check contrib partner-side
        con_par_vals = []
        for i, (code, desc, _, tag) in enumerate(FIXED_PAR_COLS):
            v = ws.cell(r, BLK_CONT[2] + i).value
            if v is not None:
                con_par_vals.append(f"{code}({tag})={v}")
        
        if par_ent_vals or par_par_vals or con_ent_vals or con_par_vals:
            f.write(f"\nRow {r}: {entity}\n")
            f.write(f"  Partner: {partner}\n")
            if par_ent_vals:
                f.write(f"  PARENT entity-side: {', '.join(par_ent_vals)}\n")
            if par_par_vals:
                f.write(f"  PARENT partner-side: {', '.join(par_par_vals)}\n")
            if con_ent_vals:
                f.write(f"  CONTRIB entity-side: {', '.join(con_ent_vals)}\n")
            if con_par_vals:
                f.write(f"  CONTRIB partner-side: {', '.join(con_par_vals)}\n")
    
    # Confirm 001001 has NO parent entity-side values
    f.write(f"\n\n{'='*100}\n")
    f.write("SANITY CHECK - Entity 001001 (should have NO journal values):\n")
    f.write("=" * 100 + "\n")
    found = False
    for r in range(ICM_OUTPUT_DATA_START, ws.max_row + 1):
        entity = str(ws.cell(r, 1).value or "").strip()
        if not entity.startswith('001001'):
            continue
        for i, (code, _, _, tag) in enumerate(FIXED_ENT_COLS):
            v = ws.cell(r, BLK_PAR[0] + i).value
            if v is not None:
                partner = str(ws.cell(r, 2).value or "").strip()
                f.write(f"  WRONG! Row {r}: parent ent {code}={v} (partner={partner[:40]})\n")
                found = True
        for i, (code, _, _, tag) in enumerate(FIXED_PAR_COLS):
            v = ws.cell(r, BLK_PAR[2] + i).value
            if v is not None:
                partner = str(ws.cell(r, 2).value or "").strip()
                f.write(f"  NOTE: Row {r}: parent par {code}={v} (partner={partner[:40]})\n")
                found = True
    if not found:
        f.write("  CORRECT - No journal values for entity 001001\n")

print("Verification written to verify_all_accounts.txt")
