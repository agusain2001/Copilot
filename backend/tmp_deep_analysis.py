"""
Deep structural analysis of both Excel files — output to a text file.
"""
import openpyxl
from openpyxl.utils import get_column_letter
import sys

out = open(r"g:\FCCS\backend\tmp_analysis_output.txt", "w", encoding="utf-8")

def p(msg=""):
    out.write(msg + "\n")

def analyze_sheet(path, label):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    p(f"\n{'='*80}")
    p(f"  {label}")
    p(f"  Sheet: '{ws.title}'  |  Rows: {ws.max_row}  |  Cols: {ws.max_column}")
    p(f"{'='*80}")

    # 1) Merged cells
    p(f"\n--- Merged Cells ({len(ws.merged_cells.ranges)}) ---")
    for mc in sorted(ws.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
        top_left = ws.cell(mc.min_row, mc.min_col)
        p(f"  {mc}  value='{top_left.value}'")

    # 2) Row heights
    p("\n--- Row Heights (non-default) ---")
    for r in range(1, min(ws.max_row + 1, 40)):
        rd = ws.row_dimensions.get(r)
        if rd and rd.height:
            p(f"  Row {r}: height={rd.height}")

    # 3) Column widths
    p("\n--- Column Widths ---")
    for c in range(1, min(ws.max_column + 1, 80)):
        ltr = get_column_letter(c)
        cd = ws.column_dimensions.get(ltr)
        w = cd.width if cd else None
        p(f"  Col {c} ({ltr}): width={w}")

    # 4) Row-by-row content for rows 1-35
    p("\n--- Row Content (rows 1-35) ---")
    for r in range(1, min(36, ws.max_row + 1)):
        vals = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is not None:
                vals.append(f"C{c}={repr(v)[:60]}")
        if vals:
            p(f"  Row {r}: {', '.join(vals)}")
        else:
            p(f"  Row {r}: (empty)")

    # 5) Header row 32 - all values
    p(f"\n--- Header Row 32 (all values) ---")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(32, c)
        if cell.value is not None:
            fg = "N/A"
            try:
                fg = cell.fill.start_color.rgb
            except:
                pass
            p(f"  Col {c}: val='{str(cell.value)[:80]}' bold={cell.font.bold} fill_fg={fg} font_color={cell.font.color} font_size={cell.font.size}")

    # 6) Section label row 29
    p(f"\n--- Section Label Row 29 (all values) ---")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(29, c)
        if cell.value is not None:
            fg = "N/A"
            try:
                fg = cell.fill.start_color.rgb
            except:
                pass
            p(f"  Col {c}: val='{cell.value}' bold={cell.font.bold} fill_fg={fg}")

    # 7) Data rows 33-35, first few cols
    p(f"\n--- Data Rows 33-35 (first 5 cols) ---")
    for r in range(33, min(36, ws.max_row + 1)):
        for c in range(1, min(6, ws.max_column + 1)):
            cell = ws.cell(r, c)
            fg = "N/A"
            try:
                fg = cell.fill.start_color.rgb
            except:
                pass
            p(f"  R{r}C{c}: val={repr(cell.value)[:60]} fill={fg} bold={cell.font.bold} size={cell.font.size} numfmt={cell.number_format}")

    # 8) AutoFilter
    p(f"\n--- AutoFilter ---")
    p(f"  ref={ws.auto_filter.ref}")

    # 9) Freeze panes
    p(f"\n--- Freeze Panes ---")
    p(f"  freeze={ws.freeze_panes}")

    wb.close()

analyze_sheet(r"g:\FCCS\backend\uploads\reports\9\outputs\ICM_Output_9.xlsx", "CURRENT OUTPUT (ICM_Output_9)")
analyze_sheet(r"g:\FCCS\AI\Intercompany Balances IC Matching Report 1.xlsx", "EXPECTED MASTER REPORT")

out.close()
print("Analysis written to tmp_analysis_output.txt")
