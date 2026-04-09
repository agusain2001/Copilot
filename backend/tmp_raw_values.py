"""Check actual raw debit/credit for E117100 entries and compare to output."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
from app.ic_processor import (
    read_journal_report, match_journal_to_icm, apply_sign,
    get_journal_indices, is_detail_row, extract_entity_code_journal,
    extract_account_code, extract_icp_code, normalize_to_numeric,
    to_float, JOURNAL_DATA_START
)

parent_path = r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx"
wb = openpyxl.load_workbook(parent_path, data_only=True)
ws = wb.active
indices = get_journal_indices(ws)

out = open("tmp_raw_journal_117100.txt", "w", encoding="utf-8")
out.write(f"Journal indices: {indices}\n\n")

out.write("=== RAW PARENT JOURNAL ENTRIES for E117100 ===\n")
for row_num, row in enumerate(ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row), start=JOURNAL_DATA_START):
    vals = [cell.value for cell in row]
    if not vals: continue
    label = str(vals[0] or "").strip()
    if label == "Grand Total": break
    if not is_detail_row(vals, indices): continue
    
    entity_raw = str(vals[indices["entity"]] or "").strip()
    entity_code = normalize_to_numeric(extract_entity_code_journal(entity_raw))
    
    if entity_code != "117100":
        continue
    
    acct_raw = str(vals[indices["acct"]] or "").strip()
    icp_raw = str(vals[indices["icp"]] or "").strip()
    debit = to_float(vals[indices["debit"]])
    credit = to_float(vals[indices["credit"]])
    acct_code = extract_account_code(acct_raw)
    icp_num = normalize_to_numeric(extract_icp_code(icp_raw))
    
    signed = apply_sign(debit, credit, acct_code)
    
    out.write(f"Row {row_num}:\n")
    out.write(f"  Entity: {entity_raw}\n")
    out.write(f"  Account: {acct_raw} -> code={acct_code}\n")
    out.write(f"  ICP: {icp_raw} -> num={icp_num}\n")
    out.write(f"  Debit={debit}, Credit={credit}\n")
    out.write(f"  apply_sign({debit}, {credit}, {acct_code}) = {signed}\n\n")

# Also check the output values
out.write("\n=== OUTPUT VALUES for E117100 rows ===\n")
wb2 = openpyxl.load_workbook(r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_FINAL_TEST.xlsx", data_only=True)
ws2 = wb2.active
for r in range(33, ws2.max_row + 1):
    ent = str(ws2.cell(r, 1).value or "").strip()
    prt = str(ws2.cell(r, 2).value or "").strip()
    if "117100" in ent or "E117100" in ent:
        out.write(f"\nRow {r}: Entity={ent[:50]} | Partner={prt[:50]}\n")
        for c in range(3, ws2.max_column + 1):
            v = ws2.cell(r, c).value
            hdr = str(ws2.cell(32, c).value or f"Col{c}")[:60]
            if v is not None and v != 0:
                out.write(f"  Col {c} ({hdr}): {v}\n")

out.close()
print("Done")
