"""Verify the new output - write results to file."""
import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import (
    BLK_PAR, BLK_CONT, FIXED_ENT_COLS, FIXED_PAR_COLS,
    ICM_OUTPUT_DATA_START
)

out_path = r'G:\FCCS\backend\uploads\reports\22\outputs\ICM_Output_22_v2.xlsx'
wb = openpyxl.load_workbook(out_path, data_only=True)
ws = wb.active

target_entities = ['022001', '001021', '001033', '013024']

with open(r'G:\FCCS\backend\verify_result.txt', 'w', encoding='utf-8') as f:
    f.write("PARENT BLOCK - Entity-side values for target entities:\n")
    f.write("=" * 80 + "\n")
    
    for r in range(ICM_OUTPUT_DATA_START, ws.max_row + 1):
        entity = str(ws.cell(r, 1).value or "").strip()
        partner = str(ws.cell(r, 2).value or "").strip()
        
        if not any(entity.startswith(t) for t in target_entities):
            continue
        
        parent_vals = []
        for i, (code, desc, _, tag) in enumerate(FIXED_ENT_COLS):
            v = ws.cell(r, BLK_PAR[0] + i).value
            if v is not None:
                parent_vals.append(f"{code}={v}")
        
        contrib_vals = []
        for i, (code, desc, _, tag) in enumerate(FIXED_ENT_COLS):
            v = ws.cell(r, BLK_CONT[0] + i).value
            if v is not None:
                contrib_vals.append(f"{code}={v}")
        
        if parent_vals or contrib_vals:
            f.write(f"\nRow {r}: {entity}\n")
            f.write(f"  Partner: {partner}\n")
            if parent_vals:
                f.write(f"  PARENT: {', '.join(parent_vals)}\n")
            if contrib_vals:
                f.write(f"  CONTRIB: {', '.join(contrib_vals)}\n")
    
    # Check 001001
    f.write(f"\n\n{'='*80}\n")
    f.write("001001 CHECK (should have NO parent entity-side values):\n")
    f.write("=" * 80 + "\n")
    found = False
    for r in range(ICM_OUTPUT_DATA_START, ws.max_row + 1):
        entity = str(ws.cell(r, 1).value or "").strip()
        if not entity.startswith('001001'):
            continue
        for i, (code, desc, _, tag) in enumerate(FIXED_ENT_COLS):
            v = ws.cell(r, BLK_PAR[0] + i).value
            if v is not None:
                partner = str(ws.cell(r, 2).value or "").strip()
                f.write(f"  Row {r}: {code}={v} (partner={partner})\n")
                found = True
    if not found:
        f.write("  CORRECT - No parent entity-side values for 001001\n")

print("Written to verify_result.txt")
