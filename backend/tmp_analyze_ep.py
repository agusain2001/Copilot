"""
Analyze: Entity-to-Partner vs Partner-to-Entity in IC Elimination report, 
Parent report, Contribution report, and ICM Output.
"""
import openpyxl
import re
import os

def extract_entity_code(raw):
    raw = str(raw or "").strip()
    m = re.match(r"(E?\d+\w*)", raw)
    return m.group(1) if m else ""

def extract_icp_code(raw):
    m = re.match(r"(ICP_\w+)", str(raw or "").strip())
    return m.group(1) if m else ""

def normalize_entity_code(code):
    code = str(code or "").strip()
    if code.startswith("E") and len(code) > 1 and code[1:].replace("_", "").isdigit():
        return code[1:]
    return code

def analyze_journal(filepath, name):
    """Analyze a journal file for Entity-to-Partner and Partner-to-Entity entries."""
    if not os.path.exists(filepath):
        print(f"\n{'='*60}")
        print(f"  FILE NOT FOUND: {name}")
        print(f"  Path: {filepath}")
        print(f"{'='*60}")
        return
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    print(f"\n{'='*60}")
    print(f"  FILE: {name}")
    print(f"  Path: {filepath}")
    print(f"  Sheet: {ws.title}")
    print(f"  Max rows: {ws.max_row}, Max cols: {ws.max_column}")
    print(f"{'='*60}")
    
    # Print first 35 rows to understand structure
    print(f"\n--- First 35 rows (raw) ---")
    for r in range(1, min(36, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 20))]
        non_empty = [(i, v) for i, v in enumerate(vals) if v is not None and str(v).strip()]
        if non_empty:
            print(f"  Row {r:3d}: {non_empty[:8]}")
    
    # Detect header row
    header_row = None
    for r in range(1, min(35, ws.max_row + 1)):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, min(ws.max_column + 1, 20))]
        if "entity" in vals and "account" in vals:
            header_row = r
            break
    
    if header_row is None:
        print(f"  *** Could not find header row with 'entity' and 'account'")
        return
    
    headers = [str(ws.cell(header_row, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
    print(f"\n  Header row: {header_row}")
    print(f"  Headers: {[(i, h) for i, h in enumerate(headers) if h]}")
    
    # Get column indices
    col_map = {}
    for i, h in enumerate(headers):
        if h: col_map[h] = i
    
    entity_idx = col_map.get('entity')
    acct_idx = col_map.get('account')
    icp_idx = col_map.get('intercompany')
    debit_idx = col_map.get('debit')
    credit_idx = col_map.get('credit')
    
    print(f"  Entity col: {entity_idx}, Account col: {acct_idx}, ICP col: {icp_idx}")
    print(f"  Debit col: {debit_idx}, Credit col: {credit_idx}")
    
    data_start = header_row + 1
    
    # Collect entity/partner pairs
    entity_to_partner = []  # Entity code is the main entity, ICP is the partner
    partner_to_entity = []  # Would need reverse lookup
    
    all_entries = []
    for r in range(data_start, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not vals: continue
        label = str(vals[0] or "").strip()
        if label == "Grand Total": break
        
        try:
            entity_raw = str(vals[entity_idx] or "").strip() if entity_idx is not None else ""
            acct_raw = str(vals[acct_idx] or "").strip() if acct_idx is not None else ""
            icp_raw = str(vals[icp_idx] or "").strip() if icp_idx is not None else ""
        except IndexError:
            continue
        
        if not entity_raw and not acct_raw and not icp_raw:
            continue
        
        entity_code = normalize_entity_code(extract_entity_code(entity_raw))
        icp_code = extract_icp_code(icp_raw)
        icp_digit = icp_code[4:] if icp_code.startswith("ICP_") else ""
        
        all_entries.append({
            "entity_raw": entity_raw,
            "entity_code": entity_code,
            "icp_raw": icp_raw,
            "icp_code": icp_code,
            "icp_digit": icp_digit,
            "acct_raw": acct_raw,
        })
    
    print(f"\n  Total detail entries: {len(all_entries)}")
    
    # Analyze unique Entity -> ICP pairs
    unique_pairs = set()
    for e in all_entries:
        if e["entity_code"] and e["icp_code"]:
            unique_pairs.add((e["entity_code"], e["icp_code"]))
    
    print(f"  Unique (Entity, ICP) pairs: {len(unique_pairs)}")
    for ent, icp in sorted(unique_pairs):
        print(f"    Entity={ent}  ->  Partner(ICP)={icp}")
    
    # Check: for each (E, P) pair, does the reverse (P_digit, ICP_E) also exist?
    print(f"\n  --- Entity-to-Partner vs Partner-to-Entity analysis ---")
    entity_to_partner_pairs = set()
    partner_to_entity_pairs = set()
    
    for ent, icp in sorted(unique_pairs):
        icp_digit = icp[4:] if icp.startswith("ICP_") else icp
        reverse_entity = icp_digit
        reverse_partner = f"ICP_{ent}"
        
        if (reverse_entity, reverse_partner) in unique_pairs:
            print(f"    BIDIRECTIONAL: ({ent}, {icp}) <-> ({reverse_entity}, {reverse_partner})")
        else:
            print(f"    ONE-WAY ONLY:  ({ent}, {icp})  -- reverse ({reverse_entity}, {reverse_partner}) NOT found")
    
    return unique_pairs

def analyze_icm_output(filepath, name):
    """Analyze ICM output to see Entity/Partner structure."""
    if not os.path.exists(filepath):
        print(f"\n{'='*60}")
        print(f"  FILE NOT FOUND: {name}")
        print(f"  Path: {filepath}")
        print(f"{'='*60}")
        return
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    print(f"\n{'='*60}")
    print(f"  FILE: {name}")
    print(f"  Path: {filepath}")
    print(f"  Sheet: {ws.title}")
    print(f"  Max rows: {ws.max_row}, Max cols: {ws.max_column}")
    print(f"{'='*60}")
    
    # Find the header row
    header_row = None
    for r in range(1, min(50, ws.max_row + 1)):
        v1 = str(ws.cell(r, 1).value or "").strip().lower()
        v2 = str(ws.cell(r, 2).value or "").strip().lower()
        if v1 == "entity" and v2 == "partner":
            header_row = r
            break
    
    if header_row is None:
        print("  *** Could not find header row")
        return
    
    print(f"  Header row: {header_row}")
    
    data_start = header_row + 1
    
    unique_pairs = set()
    for r in range(data_start, ws.max_row + 1):
        entity_raw = str(ws.cell(r, 1).value or "").strip()
        partner_raw = str(ws.cell(r, 2).value or "").strip()
        if not entity_raw and not partner_raw:
            continue
        
        m_ent = re.match(r"(\d{6})", entity_raw)
        ent_code = m_ent.group(1) if m_ent else ""
        m_prt = re.match(r"(ICP_\w+)", partner_raw)
        prt_code = m_prt.group(1) if m_prt else ""
        
        if ent_code and prt_code:
            unique_pairs.add((ent_code, prt_code))
    
    print(f"  Unique (Entity, Partner) pairs in ICM output: {len(unique_pairs)}")
    for ent, prt in sorted(unique_pairs)[:30]:
        print(f"    Entity={ent}  Partner={prt}")
    if len(unique_pairs) > 30:
        print(f"    ... ({len(unique_pairs) - 30} more)")
    
    # Check bidirectionality
    print(f"\n  --- Bidirectionality check ---")
    for ent, prt in sorted(unique_pairs):
        prt_digit = prt[4:] if prt.startswith("ICP_") else prt
        reverse = (prt_digit, f"ICP_{ent}")
        if reverse in unique_pairs:
            print(f"    BIDIRECTIONAL: ({ent}, {prt})")
        else:
            print(f"    ONE-WAY ONLY:  ({ent}, {prt})  -- reverse ({prt_digit}, ICP_{ent}) NOT in output")
    
    return unique_pairs

def analyze_ic_elimination(filepath, name):
    """Analyze IC Elimination report structure."""
    if not os.path.exists(filepath):
        print(f"\n{'='*60}")
        print(f"  FILE NOT FOUND: {name}")
        print(f"  Path: {filepath}")
        print(f"{'='*60}")
        return
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    print(f"\n{'='*60}")
    print(f"  FILE: {name}")
    print(f"  Path: {filepath}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"{'='*60}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  --- Sheet: {sheet_name} ---")
        print(f"  Max rows: {ws.max_row}, Max cols: {ws.max_column}")
        
        # Find header row
        header_row = None
        for r in range(1, min(50, ws.max_row + 1)):
            v1 = str(ws.cell(r, 1).value or "").strip().lower()
            v2 = str(ws.cell(r, 2).value or "").strip().lower()
            if v1 == "entity" and v2 == "partner":
                header_row = r
                break
        
        if header_row:
            print(f"  Header row: {header_row}")
            headers = []
            for c in range(1, min(ws.max_column + 1, 30)):
                v = str(ws.cell(header_row, c).value or "").strip()
                if v:
                    headers.append((c, v[:80]))
            print(f"  Headers: {headers}")
            
            data_start = header_row + 1
            unique_pairs = set()
            for r in range(data_start, ws.max_row + 1):
                entity_raw = str(ws.cell(r, 1).value or "").strip()
                partner_raw = str(ws.cell(r, 2).value or "").strip()
                if not entity_raw and not partner_raw:
                    continue
                m_ent = re.match(r"(\d{6})", entity_raw)
                ent_code = m_ent.group(1) if m_ent else entity_raw[:20]
                m_prt = re.match(r"(ICP_\w+)", partner_raw)
                prt_code = m_prt.group(1) if m_prt else partner_raw[:20]
                unique_pairs.add((ent_code, prt_code))
            
            print(f"  Unique (Entity, Partner) pairs: {len(unique_pairs)}")
            for ent, prt in sorted(unique_pairs):
                print(f"    Entity={ent}  Partner={prt}")
            
            # Bidirectionality
            print(f"\n  --- Bidirectionality in IC Elimination ---")
            for ent, prt in sorted(unique_pairs):
                prt_digit = prt[4:] if prt.startswith("ICP_") else prt
                reverse = (prt_digit, f"ICP_{ent}")
                if reverse in unique_pairs:
                    print(f"    BIDIRECTIONAL: ({ent}, {prt})")
                else:
                    print(f"    ONE-WAY ONLY:  ({ent}, {prt})")
        else:
            # Print first 20 rows
            print(f"  No 'Entity/Partner' header found. Dumping first 20 rows:")
            for r in range(1, min(21, ws.max_row + 1)):
                vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))]
                non_empty = [(i, v) for i, v in enumerate(vals) if v is not None]
                if non_empty:
                    print(f"    Row {r}: {non_empty[:8]}")


# ── Run analysis ──
print("=" * 70)
print("  COMPREHENSIVE ENTITY-TO-PARTNER / PARTNER-TO-ENTITY ANALYSIS")
print("=" * 70)

# IC Elimination Report
analyze_ic_elimination(
    r"g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx",
    "IC Elimination Report (inputs/31)"
)

# Parent Report
analyze_journal(
    r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx",
    "Parent Report (inputs/31)"
)

# Contribution Report
analyze_journal(
    r"g:\FCCS\backend\uploads\reports\31\inputs\Contribution report.xlsx",
    "Contribution Report (inputs/31)"
)

# Report Inputs
print(f"\n{'='*60}")
print(f"  REPORT INPUTS")
print(f"{'='*60}")
ri_path = r"g:\FCCS\backend\uploads\reports\31\inputs\report Inputs.xlsx"
if os.path.exists(ri_path):
    wb = openpyxl.load_workbook(ri_path, data_only=True)
    ws = wb.active
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 10))]
        non_empty = [(i, v) for i, v in enumerate(vals) if v is not None]
        if non_empty:
            print(f"  Row {r}: {non_empty}")

# ICM Output
analyze_icm_output(
    r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31.xlsx",
    "ICM Output 31"
)

# Also check New folder (4) files for comparison
print(f"\n\n{'#'*70}")
print(f"  NOW CHECKING 'New folder (4)' files for comparison")
print(f"{'#'*70}")

analyze_ic_elimination(
    r"g:\FCCS\New folder (4)\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx",
    "IC Elimination Report (New folder 4)"
)

analyze_journal(
    r"g:\FCCS\New folder (4)\Parent report.xlsx",
    "Parent Report (New folder 4)"
)

analyze_journal(
    r"g:\FCCS\New folder (4)\Contribution report.xlsx",
    "Contribution Report (New folder 4)"
)
