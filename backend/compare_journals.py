"""Check what values Journal Report (1).xlsx from report 22 inputs produces."""
import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import (
    read_journal_report, match_journal_to_icm, read_icm_data,
    ICM_INPUT_HEADER_ROW
)

# Read ICM
icm_path = r'G:\FCCS\backend\uploads\reports\22\inputs\Intercompany Balances IC Matching Report (1).xlsx'
wb_icm = openpyxl.load_workbook(icm_path, data_only=True)
candidates = []
for sheet in wb_icm.worksheets:
    c1 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 1).value or "").strip()
    c2 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 2).value or "").strip()
    if c1.lower() == "entity" and c2.lower() == "partner":
        candidates.append(sheet)
ws_icm = min(candidates, key=lambda s: s.max_column) if candidates else wb_icm.active
data_rows = read_icm_data(ws_icm)

# Check ALL journal files
journals = {
    "Report22 J(1)": r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (1).xlsx',
    "Report22 J(2)": r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (2).xlsx',
    "Report22 J(4)": r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (4).xlsx',
    "Update J": r'G:\FCCS\Update files\Journal Report.xlsx',
    "Update J(2)": r'G:\FCCS\Update files\Journal Report (2).xlsx',
    "Update J(4)": r'G:\FCCS\Update files\Journal Report (4).xlsx',
}

with open(r'G:\FCCS\backend\journal_compare.txt', 'w', encoding='utf-8') as f:
    for label, jpath in journals.items():
        try:
            primary, fallback = read_journal_report(jpath)
            primary_updates, fallback_updates = match_journal_to_icm(data_rows, primary, fallback)
            
            f.write(f"\n{'='*80}\n")
            f.write(f"{label}: {jpath}\n")
            f.write(f"  Primary updates: {len(primary_updates)}\n")
            for k, v in sorted(primary_updates.items()):
                f.write(f"    {k[0]}/{k[1]}/{k[2]} -> {v:,.2f}\n")
            f.write(f"  Fallback updates: {len(fallback_updates)}\n")
            for k, v in sorted(fallback_updates.items()):
                f.write(f"    {k[0]}/{k[1]}/{k[2]} -> {v:,.2f}\n")
        except Exception as e:
            f.write(f"\n{label}: ERROR - {e}\n")

print("Written to journal_compare.txt")
