"""Check where the '189501' entity row in the output comes from"""
import openpyxl, sys
sys.path.insert(0, '.')
from app.ic_processor import detect_icm_header_row, extract_entity_code_icm

BASE = "g:/FCCS/backend/uploads/reports/29"
icm_path = BASE + "/inputs/IC Elimination Report_188800_Intercompany Balances Plug A_c_1154_Intercompany Report.xlsx"
wb = openpyxl.load_workbook(icm_path, data_only=True)

out = open("tmp_189501_v4.txt", "w", encoding="utf-8")

# Check ALL rows in ICM source where column 1 value contains 189501
for sheet in wb.worksheets:
    hdr_row, ds = detect_icm_header_row(sheet)
    c1 = str(sheet.cell(hdr_row, 1).value or "").strip().lower()
    if c1 != "entity": continue
    out.write(f"Sheet: {sheet.title}, data_start={ds}, max_row={sheet.max_row}\n\n")
    
    # Print ALL rows from data_start to max_row
    for r in range(ds, sheet.max_row + 1):
        ent = str(sheet.cell(r, 1).value or "").strip()
        prt = str(sheet.cell(r, 2).value or "").strip()
        if not ent and not prt:
            continue
        ent_code = extract_entity_code_icm(ent)
        # Show ALL rows - just flag the 189501 ones
        if "189501" in ent or "189501" in prt:
            out.write(f"*** Row {r}: entity='{ent}' | partner='{prt}' | ent_code='{ent_code}'\n")
        # Also show nearby rows for context
        if r >= 430 and r <= 445:
            out.write(f"    Row {r}: entity='{ent}' | partner='{prt}' | ent_code='{ent_code}'\n")

# Count total ICM source data rows
out.write("\n\nTotal data rows in ICM source:\n")
for sheet in wb.worksheets:
    hdr_row, ds = detect_icm_header_row(sheet)
    c1 = str(sheet.cell(hdr_row, 1).value or "").strip().lower()
    if c1 != "entity": continue
    count = 0
    for r in range(ds, sheet.max_row + 1):
        ent = str(sheet.cell(r, 1).value or "").strip()
        prt = str(sheet.cell(r, 2).value or "").strip()
        if ent or prt:
            count += 1
    out.write(f"  Sheet '{sheet.title}': {count} data rows\n")

# Count output rows
out_path = BASE + "/outputs/ICM_Output_29.xlsx"
wb_out = openpyxl.load_workbook(out_path, data_only=True)
ws_out = wb_out.active
out.write(f"\nTotal output rows: {ws_out.max_row - 32}\n")

# Output row 438 is the 189501 entity. Where did it come from?
# Is it a missing pair synthetic row? Check the entities around it.
out.write("\nOutput rows 435-442:\n")
for r in range(435, 443):
    ent = str(ws_out.cell(r, 1).value or "").strip()
    prt = str(ws_out.cell(r, 2).value or "").strip()
    out.write(f"  Row {r}: entity='{ent[:60]}' | partner='{prt[:60]}'\n")

out.close()
print("Done -> tmp_189501_v4.txt")
