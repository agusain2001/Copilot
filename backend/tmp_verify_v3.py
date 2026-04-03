import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import extract_account_code

path = r'G:\FCCS\Update files\ICM_Output_ENHANCED_v3.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Header row 32
print("\nAll headers (row 32):")
all_codes = set()
for c in range(1, ws.max_column + 1):
    v = str(ws.cell(32, c).value or '').strip()
    if v:
        code = extract_account_code(v)
        if code: all_codes.add(code)
        print(f"  Col {c:3d}: {v[:80]}")

print(f"\nUnique account codes: {sorted(all_codes)}")
print(f"Total: {len(all_codes)}")

# Section labels (row 29)
print("\nSection labels (row 29):")
for c in range(1, ws.max_column + 1):
    v = ws.cell(29, c).value
    if v:
        print(f"  Col {c}: {v}")

# Verify NO journal-only accounts
journal_only = {'155020', '310001', '315005', 'Plug_InvSh'}
found_bad = [c for c in journal_only if c in all_codes]
print(f"\nJournal-only accounts in output: {found_bad if found_bad else 'NONE (correct!)'}")

# Count data rows
total = 0
for r in range(33, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or '').strip()
    if ent: total += 1
print(f"\nData rows: {total}")
