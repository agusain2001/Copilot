import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import extract_account_code

path = r'G:\FCCS\Update files\ICM_Output_ENHANCED.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

# Check section labels (row 29 - green cells)
print("Row 29 (Section labels):")
for c in range(1, ws.max_column + 1):
    v = ws.cell(29, c).value
    if v:
        print(f"  Col {c}: '{v}'")

# Check header row 32 for 189501 and 155020
print("\nRow 32 (Headers) - searching for 189501 and 155020:")
for c in range(1, ws.max_column + 1):
    v = str(ws.cell(32, c).value or '').strip()
    if '189501' in v or '155020' in v:
        # Also check cell fill color
        cell = ws.cell(32, c)
        fill_color = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'
        print(f"  Col {c}: '{v[:80]}' (fill={fill_color})")

# Check if any cell in the sheet contains [165000].[189501] format
print("\nSearching for '[165000].[189501]' pattern across key rows:")
for r in [29, 30, 31, 32]:
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(r, c).value or '')
        if '[165000]' in v or '[155000]' in v:
            cell = ws.cell(r, c)
            fill_color = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'
            print(f"  Row {r}, Col {c}: '{v[:80]}' (fill={fill_color})")

# Also check the ICM source headers (row 32 in source)
print("\nICM Source headers with 189501 or 155020:")
wb_src = openpyxl.load_workbook(r'G:\FCCS\Update files\Intercompany Balances IC Matching Report (1).xlsx', data_only=True)
for sheet in wb_src.worksheets:
    if sheet.max_column < 5:
        continue
    for c in range(1, sheet.max_column + 1):
        v = str(sheet.cell(32, c).value or '')
        if '189501' in v or '155020' in v:
            print(f"  Sheet '{sheet.title}', Col {c}: '{v[:80]}'")
