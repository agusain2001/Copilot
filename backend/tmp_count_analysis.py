"""Count Parent journal valid entries vs ICM output Parent Input values."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
from app.ic_processor import (
    read_journal_report, match_journal_to_icm, apply_sign,
    get_journal_indices, is_detail_row, extract_entity_code_journal,
    extract_account_code, extract_icp_code, normalize_to_numeric,
    to_float, JOURNAL_DATA_START
)

valid_accounts = {"111006", "120030", "121015", "188600", "433002", "534018"}

# === 1. Count valid Parent journal entries ===
parent_path = r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx"
wb = openpyxl.load_workbook(parent_path, data_only=True)
ws = wb.active
indices = get_journal_indices(ws)

out = open("tmp_count_analysis.txt", "w", encoding="utf-8")

valid_entries = []
for row_num, row in enumerate(ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row), start=JOURNAL_DATA_START):
    vals = [cell.value for cell in row]
    if not vals: continue
    label = str(vals[0] or "").strip()
    if label == "Grand Total": break
    if not is_detail_row(vals, indices): continue
    
    entity_raw = str(vals[indices["entity"]] or "").strip()
    acct_raw = str(vals[indices["acct"]] or "").strip()
    icp_raw = str(vals[indices["icp"]] or "").strip()
    
    entity_num = normalize_to_numeric(extract_entity_code_journal(entity_raw))
    acct_code = extract_account_code(acct_raw)
    icp_num = normalize_to_numeric(extract_icp_code(icp_raw))
    
    if entity_num and icp_num and acct_code in valid_accounts:
        debit = to_float(vals[indices["debit"]])
        credit = to_float(vals[indices["credit"]])
        valid_entries.append({
            "row": row_num,
            "entity": entity_num,
            "icp": icp_num,
            "acct": acct_code,
            "debit": debit,
            "credit": credit,
            "signed": apply_sign(debit, credit, acct_code),
        })

out.write(f"=== PARENT JOURNAL: {len(valid_entries)} valid entries ===\n")
for e in valid_entries:
    out.write(f"  Row {e['row']}: ({e['entity']}, {e['icp']}, {e['acct']}) D={e['debit']} C={e['credit']} signed={e['signed']}\n")

# Group by key and compute net
from collections import defaultdict
grouped = defaultdict(float)
for e in valid_entries:
    grouped[(e["entity"], e["icp"], e["acct"])] += e["signed"]

# Filter net != 0
non_zero = {k: v for k, v in grouped.items() if v != 0}
out.write(f"\n=== After grouping: {len(grouped)} unique keys, {len(non_zero)} with net != 0 ===\n")
for k, v in sorted(non_zero.items()):
    out.write(f"  {k} -> net={v}\n")

# === 2. Count non-zero values in ICM Output Parent Input block ===
out.write(f"\n=== ICM OUTPUT Parent Input block ===\n")
wb2 = openpyxl.load_workbook(r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_FINAL.xlsx", data_only=True)
ws2 = wb2.active

# The Parent Input block starts after the base block. Let me find column ranges.
# Base block: cols 3-15, Parent Input: cols 16-29 (approximately)
# Let me identify by headers
hdrs = {}
for c in range(1, ws2.max_column + 1):
    h = str(ws2.cell(32, c).value or "").strip()
    hdrs[c] = h

# Find section labels (row 29)
sections = {}
for c in range(1, ws2.max_column + 1):
    s = str(ws2.cell(29, c).value or "").strip()
    if s:
        sections[c] = s

out.write("Section labels (row 29):\n")
for c, s in sorted(sections.items()):
    out.write(f"  Col {c}: {s}\n")

# Count non-zero parent input cells
# Parent Input block is between "Parent Input" and "Contribution Input" sections
parent_start = None
parent_end = None
for c, s in sorted(sections.items()):
    if "Parent" in s and parent_start is None:
        parent_start = c
    elif "Contribution" in s and parent_start is not None:
        parent_end = c - 1
        break

if parent_start and parent_end:
    out.write(f"\nParent Input block: cols {parent_start} to {parent_end}\n")
    parent_nonzero_cells = 0
    parent_rows_with_values = set()
    for r in range(33, ws2.max_row + 1):
        for c in range(parent_start, parent_end + 1):
            v = ws2.cell(r, c).value
            hdr = hdrs.get(c, "")
            if "Variance" in hdr or "Total" in hdr:
                continue  # Skip computed columns
            if v is not None and v != 0:
                parent_nonzero_cells += 1
                parent_rows_with_values.add(r)
                ent = str(ws2.cell(r, 1).value or "")[:40]
                prt = str(ws2.cell(r, 2).value or "")[:40]
                out.write(f"  Row {r} Col {c} ({hdr[:30]}): {v}  [{ent} | {prt}]\n")
    
    out.write(f"\nTotal non-zero cells in Parent Input (excl Variance/Total): {parent_nonzero_cells}\n")
    out.write(f"Rows with at least one value: {len(parent_rows_with_values)}\n")

out.write(f"\n=== SUMMARY ===\n")
out.write(f"Parent journal valid entries: {len(valid_entries)}\n")
out.write(f"Unique journal keys: {len(grouped)}\n")
out.write(f"Keys with non-zero net: {len(non_zero)}\n")
out.write(f"Non-zero cells in Parent Input block: {parent_nonzero_cells}\n")
out.write(f"Expected (1 per key): {len(non_zero)}\n")
out.write(f"Excess: {parent_nonzero_cells - len(non_zero)}\n")

out.close()
print("Done")
