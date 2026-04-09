"""Clean comparison: run processor and compare against FIXED reference."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
import logging
logging.basicConfig(level=logging.WARNING)

# Paths
ref_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_FIXED.xlsx"
icm_path = r"g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx"
journal_paths = {
    "parent_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx",
    "contribution_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Contribution report.xlsx",
    "plugaccount_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Journal Report (4).xlsx",
}
report_inputs_path = r"g:\FCCS\backend\uploads\reports\31\inputs\report Inputs.xlsx"
output_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_VERIFY.xlsx"

# Run processor
from app.ic_processor import process_icm_report
process_icm_report(icm_path, journal_paths, output_path, report_inputs_path)

# Compare
wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
ws_ref = wb_ref.active
wb_new = openpyxl.load_workbook(output_path, data_only=True)
ws_new = wb_new.active

print(f"REF: {ws_ref.max_row} rows x {ws_ref.max_column} cols")
print(f"NEW: {ws_new.max_row} rows x {ws_new.max_column} cols")

# Count data rows
def count_data(ws):
    n = 0
    for r in range(33, ws.max_row + 1):
        if ws.cell(r, 1).value or ws.cell(r, 2).value:
            n += 1
    return n

print(f"REF data rows: {count_data(ws_ref)}")
print(f"NEW data rows: {count_data(ws_new)}")

# Header comparison
print("\n--- HEADER COMPARISON (row 32) ---")
ref_hdrs = {}
new_hdrs = {}
for c in range(1, max(ws_ref.max_column, ws_new.max_column) + 1):
    rv = str(ws_ref.cell(32, c).value or "").strip()
    nv = str(ws_new.cell(32, c).value or "").strip()
    if rv: ref_hdrs[c] = rv
    if nv: new_hdrs[c] = nv
    if rv != nv and (rv or nv):
        print(f"  Col {c}: REF='{rv[:60]}' vs NEW='{nv[:60]}'")

if ref_hdrs.keys() == new_hdrs.keys():
    print("  Headers MATCH (same columns)")
else:
    print(f"  REF has {len(ref_hdrs)} header cols, NEW has {len(new_hdrs)} header cols")

# Build pair maps
def pair_map(ws):
    m = {}
    for r in range(33, ws.max_row + 1):
        e = str(ws.cell(r, 1).value or "").strip()
        p = str(ws.cell(r, 2).value or "").strip()
        if e or p:
            m[(e, p)] = r
    return m

ref_pm = pair_map(ws_ref)
new_pm = pair_map(ws_new)
common = set(ref_pm) & set(new_pm)
only_ref = set(ref_pm) - set(new_pm)
only_new = set(new_pm) - set(ref_pm)

print(f"\nCommon pairs: {len(common)}")
if only_ref:
    print(f"Only in REF ({len(only_ref)}):")
    for p in sorted(only_ref)[:10]:
        print(f"  {p}")
if only_new:
    print(f"Only in NEW ({len(only_new)}):")
    for p in sorted(only_new)[:10]:
        print(f"  {p}")

# Value comparison for common pairs
print("\n--- VALUE COMPARISON ---")
mismatches = 0
match_count = 0

def to_num(v):
    if v is None or str(v).strip() in ("", " "): return 0.0
    try: return round(float(v), 2)
    except: return v

for pair in sorted(common):
    rr = ref_pm[pair]
    nr = new_pm[pair]
    max_c = max(ws_ref.max_column, ws_new.max_column)
    for c in range(3, max_c + 1):
        rv = to_num(ws_ref.cell(rr, c).value)
        nv = to_num(ws_new.cell(nr, c).value)
        if rv != nv:
            mismatches += 1
            if mismatches <= 20:
                hdr = str(ws_ref.cell(32, c).value or f"Col{c}")[:50]
                print(f"  [{pair[0][:30]} | {pair[1][:30]}] Col {c} ({hdr}): REF={rv} NEW={nv}")
        else:
            match_count += 1

print(f"\nMatching values: {match_count}")
print(f"Mismatched values: {mismatches}")
print(f"Match rate: {match_count/(match_count+mismatches)*100:.1f}%" if (match_count+mismatches) > 0 else "No data")
