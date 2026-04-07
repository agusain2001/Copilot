"""
Complete matching trace - which ICM rows match which journal keys, for both entity-side and partner-side.
"""
import openpyxl
import re
from collections import defaultdict

def extract_entity_code_journal(raw):
    raw = str(raw or "").strip()
    m = re.match(r"(E?\d+\w*)", raw)
    return m.group(1) if m else ""

def normalize_entity_code(code):
    code = str(code or "").strip()
    if code.startswith("E") and len(code) > 1 and code[1:].replace("_", "").isdigit():
        return code[1:]
    return code

def extract_icp_code(raw):
    m = re.match(r"(ICP_\w+)", str(raw or "").strip())
    return m.group(1) if m else ""

def extract_6digit_from_icp(icp_code):
    s = str(icp_code or "").strip()
    return s[4:] if s.startswith("ICP_") else s

# Read Parent journal keys
wb = openpyxl.load_workbook(r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx", data_only=True)
ws = wb.active
headers = [str(ws.cell(30, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
col_map = {h: i for i, h in enumerate(headers) if h}
eidx = col_map.get('entity', 2)
icpidx = col_map.get('intercompany', 4)

journal_keys = set()
for r in range(31, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    label = str(vals[0] or "").strip()
    if label == "Grand Total": break
    try:
        ent_raw = str(vals[eidx] or "").strip()
        icp_raw = str(vals[icpidx] or "").strip()
    except IndexError:
        continue
    ent_code_raw = extract_entity_code_journal(ent_raw)
    ent_code = normalize_entity_code(ent_code_raw)
    icp_code = extract_icp_code(icp_raw)
    if ent_code and icp_code:
        journal_keys.add((ent_code, icp_code))

# Read ICM pairs
wb_icm = openpyxl.load_workbook(r"g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx", data_only=True)
icm_pairs = set()
for sheet in wb_icm.worksheets:
    for r in range(1, min(50, sheet.max_row + 1)):
        v1 = str(sheet.cell(r, 1).value or "").strip().lower()
        v2 = str(sheet.cell(r, 2).value or "").strip().lower()
        if v1 == "entity" and v2 == "partner":
            for dr in range(r+1, sheet.max_row + 1):
                ent_raw = str(sheet.cell(dr, 1).value or "").strip()
                prt_raw = str(sheet.cell(dr, 2).value or "").strip()
                if not ent_raw and not prt_raw: continue
                m = re.match(r"(\d{6})", ent_raw)
                ent_code = m.group(1) if m else ""
                prt_code = extract_icp_code(prt_raw)
                if ent_code and prt_code:
                    icm_pairs.add((ent_code, prt_code))
            break

# For each ICM pair, check entity-side (direct) and partner-side (reverse) matching
print("MATCHING RESULTS:")
print("=" * 80)
print()

entity_matches = 0
partner_matches = 0
both_matches = 0
neither_matches = 0

for ent, prt in sorted(icm_pairs):
    # Entity-side: (ent, prt) in journal?
    ent_match = (ent, prt) in journal_keys
    
    # Partner-side: reverse the roles
    prt_digit = extract_6digit_from_icp(prt)
    prt_entity = normalize_entity_code(prt_digit)
    rev_plain = f"ICP_{ent}"
    rev_e = f"ICP_E{ent}"
    
    par_match = (prt_entity, rev_plain) in journal_keys or (prt_entity, rev_e) in journal_keys
    
    if ent_match and par_match:
        both_matches += 1
        tag = "BOTH"
    elif ent_match:
        entity_matches += 1
        tag = "ENTITY-ONLY"
    elif par_match:
        partner_matches += 1
        tag = "PARTNER-ONLY"
    else:
        neither_matches += 1
        tag = "NONE"
    
    print(f"  ({ent}, {prt}) => Entity({ent},{prt})={'FOUND' if ent_match else 'MISSING'}  Partner({prt_entity},{rev_plain}/{rev_e})={'FOUND' if par_match else 'MISSING'}  [{tag}]")

print()
print(f"Summary:")
print(f"  Both sides match:    {both_matches}")
print(f"  Entity-only match:   {entity_matches}")
print(f"  Partner-only match:  {partner_matches}")
print(f"  Neither matches:     {neither_matches}")
print(f"  Total ICM pairs:     {len(icm_pairs)}")

# Also check: which journal keys DON'T match any ICM pair?
print()
print("Journal keys NOT matching any ICM pair (entity-side):")
for jk in sorted(journal_keys):
    if jk not in icm_pairs:
        # Is it a partner-side key from an ICM pair?
        # It would be a partner key if there exists an ICM pair where
        # reverse maps to this journal key
        is_reverse = False
        for ent, prt in icm_pairs:
            prt_entity = normalize_entity_code(extract_6digit_from_icp(prt))
            rev_plain = f"ICP_{ent}"
            if jk == (prt_entity, rev_plain):
                is_reverse = True
                break
        print(f"  {jk} {'(is partner-side of ICM pair)' if is_reverse else '(ORPHAN - not in ICM at all)'}")
