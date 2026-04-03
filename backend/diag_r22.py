"""Diagnose: Check what entities exist in both journal files and what the output contains."""
import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import (
    extract_entity_code_journal, extract_icp_code, extract_account_code,
    JOURNAL_DATA_START, J_ENTITY, J_ACCT, J_ICP, J_DEBIT, J_CREDIT,
    is_detail_row, read_journal_report, match_journal_to_icm,
    read_icm_data, read_icm_headers, ICM_INPUT_HEADER_ROW, parse_report_inputs
)

f = open(r'G:\FCCS\backend\diag_r22.txt', 'w', encoding='utf-8')

# ── Check both journal files ──
journals_to_check = {
    "Update files (used as parent)": r'G:\FCCS\Update files\Journal Report.xlsx',
    "Report22 inputs J(1)": r'G:\FCCS\backend\uploads\reports\22\inputs\Journal Report (1).xlsx',
}

for label, jpath in journals_to_check.items():
    f.write(f"\n{'='*80}\n")
    f.write(f"JOURNAL: {label}\n  Path: {jpath}\n")
    f.write(f"{'='*80}\n")
    
    try:
        wb = openpyxl.load_workbook(jpath, data_only=True)
        ws = wb.active
        
        # Show first few rows to understand structure
        f.write(f"  Sheet: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}\n")
        f.write(f"\n  Header area (rows 28-32):\n")
        for r in range(28, 33):
            vals = []
            for c in range(1, min(ws.max_column+1, 20)):
                v = ws.cell(r, c).value
                if v is not None:
                    vals.append(f"C{c}='{str(v)[:50]}'")
            if vals:
                f.write(f"    Row {r}: {', '.join(vals)}\n")
        
        # Extract unique entities and ICPs
        entities = set()
        icps = set()
        count = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row), start=JOURNAL_DATA_START):
            vals = [cell.value for cell in row]
            if not is_detail_row(vals):
                continue
            
            entity_raw = str(vals[J_ENTITY - 1] or "").strip()
            icp_raw = str(vals[J_ICP - 1] or "").strip()
            
            ent_code = extract_entity_code_journal(entity_raw)
            icp_code = extract_icp_code(icp_raw)
            
            if ent_code:
                entities.add((ent_code, entity_raw[:60]))
            if icp_code:
                icps.add(icp_code)
            count += 1
        
        f.write(f"\n  Total detail rows: {count}\n")
        f.write(f"\n  Unique entities ({len(entities)}):\n")
        for code, raw in sorted(entities):
            f.write(f"    {code}: {raw}\n")
        
        f.write(f"\n  Has entity '001001': {'YES' if any(e[0]=='001001' for e in entities) else 'NO'}\n")
        f.write(f"  Has entity '001033': {'YES' if any(e[0]=='001033' for e in entities) else 'NO'}\n")
        f.write(f"  Has entity '022001': {'YES' if any(e[0]=='022001' for e in entities) else 'NO'}\n")
        f.write(f"  Has entity '013024': {'YES' if any(e[0]=='013024' for e in entities) else 'NO'}\n")
        
    except Exception as e:
        f.write(f"  ERROR: {e}\n")

# ── Check ICM data ──
f.write(f"\n\n{'='*80}\n")
f.write("ICM INPUT FILE:\n")
f.write(f"{'='*80}\n")
icm_path = r'G:\FCCS\backend\uploads\reports\22\inputs\Intercompany Balances IC Matching Report (1).xlsx'
wb_icm = openpyxl.load_workbook(icm_path, data_only=True)
candidates = []
for sheet in wb_icm.worksheets:
    c1 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 1).value or "").strip()
    c2 = str(sheet.cell(ICM_INPUT_HEADER_ROW, 2).value or "").strip()
    if c1.lower() == "entity" and c2.lower() == "partner":
        candidates.append(sheet)
ws_icm = min(candidates, key=lambda s: s.max_column) if candidates else wb_icm.active
data_rows = read_icm_data(ws_icm)
f.write(f"  Total ICM data rows: {len(data_rows)}\n")

# ── Run matching with the parent journal used ──
f.write(f"\n\n{'='*80}\n")
f.write("MATCHING RESULTS (parent journal = Update files\\Journal Report.xlsx):\n")
f.write(f"{'='*80}\n")
parent_path = r'G:\FCCS\Update files\Journal Report.xlsx'
primary, fallback = read_journal_report(parent_path)
primary_updates, fallback_updates = match_journal_to_icm(data_rows, primary, fallback)

f.write(f"\n  Primary updates ({len(primary_updates)} entries):\n")
for k, v in sorted(primary_updates.items()):
    f.write(f"    {k!r} -> {v}\n")

f.write(f"\n  Fallback updates ({len(fallback_updates)} entries):\n")
for k, v in sorted(fallback_updates.items()):
    f.write(f"    {k!r} -> {v}\n")

# ── Check output file ──
f.write(f"\n\n{'='*80}\n")
f.write("OUTPUT FILE CHECK (first rows with parent values):\n")
f.write(f"{'='*80}\n")
from app.ic_processor import BLK_PAR, FIXED_ENT_COLS, ICM_OUTPUT_DATA_START
out_path = r'G:\FCCS\backend\uploads\reports\22\outputs\ICM_Output_22_final.xlsx'
wb_out = openpyxl.load_workbook(out_path, data_only=True)
ws_out = wb_out.active
for r in range(ICM_OUTPUT_DATA_START, ws_out.max_row + 1):
    entity = str(ws_out.cell(r, 1).value or "").strip()
    partner = str(ws_out.cell(r, 2).value or "").strip()
    
    # Check parent block for any non-None values
    has_vals = False
    vals_str = []
    for i, (code, _, _, tag) in enumerate(FIXED_ENT_COLS):
        v = ws_out.cell(r, BLK_PAR[0] + i).value
        if v is not None:
            has_vals = True
            vals_str.append(f"{code}={v}")
    
    if has_vals:
        f.write(f"  Row {r}: {entity[:50]} / {partner[:50]}\n")
        f.write(f"    Parent entity-side: {', '.join(vals_str)}\n")

f.close()
print("Written to diag_r22.txt")
