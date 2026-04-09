import os
import sys
import tempfile
import unittest
import uuid

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "venv", "Lib", "site-packages"))
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl

from app.ic_processor import (
    ICM_OUTPUT_DATA_START,
    ICM_OUTPUT_HEADER_ROW,
    JOURNAL_DATA_START,
    _block_positions,
    build_covered_pairs,
    canonicalize_updates_to_rows,
    build_consolidated_rows,
    classify_pair_direction,
    extract_account_code,
    match_journal_to_icm,
    normalize_to_numeric,
    parse_journal_line,
    process_icm_report_v1 as process_icm_report,
    read_ic_elimination_report,
    read_journal_report,
    write_output,
)


def _journal_workbook(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    header_row = JOURNAL_DATA_START - 1
    ws.cell(header_row, 3).value = "Entity"
    ws.cell(header_row, 4).value = "Account"
    ws.cell(header_row, 5).value = "Intercompany"
    ws.cell(header_row, 15).value = "Debit"
    ws.cell(header_row, 16).value = "Credit"
    for idx, row in enumerate(rows, start=JOURNAL_DATA_START):
        ws.cell(idx, 1).value = row.get("label", f"Line {idx}")
        ws.cell(idx, 3).value = row.get("entity")
        ws.cell(idx, 4).value = row.get("account")
        ws.cell(idx, 5).value = row.get("icp")
        ws.cell(idx, 15).value = row.get("debit")
        ws.cell(idx, 16).value = row.get("credit")
    ws.cell(JOURNAL_DATA_START + len(rows), 1).value = "Grand Total"
    wb.save(path)


def _icm_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(4, 1).value = "Entity"
    ws.cell(4, 2).value = "Partner"
    ws.cell(4, 3).value = "534018 - 534018:Interest expense - related parties Entity"
    ws.cell(4, 4).value = "534018 - 534018:Interest expense - related parties Partner"
    ws.cell(5, 1).value = "007009:QD UK Holdings LP adjusted"
    ws.cell(5, 2).value = "ICP_E117100:QD Europe S.a.r.l. (H) ICP"
    ws.cell(5, 3).value = 0
    ws.cell(5, 4).value = 0
    wb.save(path)


def _report_inputs_workbook(path, plug_code="188600", elim_code="534018"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = f"Plug Account: {plug_code}"
    ws.cell(3, 1).value = f"{elim_code} - {elim_code}:Interest expense - related parties"
    wb.save(path)


def _ic_elim_workbook(path, rows, plug_code="188600"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(32, 1).value = "Entity"
    ws.cell(32, 2).value = "Partner"
    ws.cell(32, 3).value = "534018 - 534018:Interest expense - related parties Entity"
    ws.cell(32, 4).value = "534018 - 534018:Interest expense - related parties Partner"
    ws.cell(32, 5).value = "Total"
    for idx, row in enumerate(rows, start=33):
        ws.cell(idx, 1).value = row.get("entity")
        ws.cell(idx, 2).value = row.get("partner")
        ws.cell(idx, 3).value = row.get("entity_value")
        ws.cell(idx, 4).value = row.get("partner_value")
        ws.cell(idx, 5).value = row.get("total")
    ws.cell(296, 1).value = f"Plug Account: {plug_code} - {plug_code}:Intercompany Balances Plug A/c"
    ws.cell(297, 1).value = "Eliminating Accounts"
    ws.cell(298, 1).value = "534018 - 534018:Interest expense - related parties"
    wb.create_sheet("Sheet2")
    wb.save(path)


class IcProcessorBidirectionalTests(unittest.TestCase):
    def _mktemp_path(self, stem):
        path = os.path.join(os.path.dirname(__file__), f"{stem}_{uuid.uuid4().hex}.xlsx")
        self.addCleanup(lambda p=path: os.path.exists(p) and os.remove(p))
        return path

    def test_normalization_helpers(self):
        self.assertEqual(normalize_to_numeric("E117100"), "117100")
        self.assertEqual(normalize_to_numeric("ICP_E117100"), "117100")
        self.assertEqual(normalize_to_numeric("ICP_007009"), "007009")
        self.assertEqual(
            classify_pair_direction("E117100:QD UK Holdings LP adjusted"),
            "entity_to_partner",
        )
        self.assertEqual(
            classify_pair_direction("007009:QD UK Holdings LP adjusted"),
            "partner_to_entity",
        )
        self.assertEqual(
            extract_account_code("534018 - 534018:Interest expense - related parties"),
            "534018",
        )

    def test_build_covered_pairs_marks_reverse_for_entity_to_partner_row(self):
        covered = build_covered_pairs([
            {
                "entity_code": "117100",
                "partner_num": "007009",
                "row_direction": "entity_to_partner",
            }
        ])
        self.assertIn(("117100", "007009"), covered)
        self.assertIn(("007009", "117100"), covered)

        covered = build_covered_pairs([
            {
                "entity_code": "001032",
                "partner_num": "001033",
                "row_direction": "partner_to_entity",
            }
        ])
        self.assertIn(("001032", "001033"), covered)
        self.assertIn(("001033", "001032"), covered)

    def test_build_consolidated_rows_deduplicates_exact_pair_rows(self):
        rows = build_consolidated_rows(
            [
                {
                    "row_num": 5,
                    "entity": "E117100 - QD UK Holdings LP adjusted",
                    "partner": "ICP_007009 - QD Europe Chancery S.a.r.l. ICP",
                    "entity_code": "117100",
                    "partner_num": "007009",
                    "row_direction": "entity_to_partner",
                    "is_synthetic": False,
                },
                {
                    "row_num": 9,
                    "entity": "E117100 - QD UK Holdings LP adjusted",
                    "partner": "ICP_007009 - QD Europe Chancery S.a.r.l. ICP",
                    "entity_code": "117100",
                    "partner_num": "007009",
                    "row_direction": "entity_to_partner",
                    "is_synthetic": False,
                },
            ],
            {("117100", "007009")},
            {},
            {},
            {},
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["row_num"], 5)

    def test_canonicalize_updates_to_rows_folds_reverse_key_to_canonical_pair(self):
        canonical = canonicalize_updates_to_rows(
            {
                ("007009", "117100", "188600"): 10,
                ("117100", "007009", "188600"): 5,
            },
            [
                {
                    "entity_code": "117100",
                    "partner_num": "007009",
                }
            ],
        )
        self.assertEqual(canonical, {("117100", "007009", "188600"): 15})

    def test_parse_journal_line_skips_incomplete_rows(self):
        indices = {"entity": 2, "acct": 3, "icp": 4, "debit": 14, "credit": 15}
        vals = [None] * 16
        vals[2] = "E117100:QD UK Holdings LP adjusted"
        vals[3] = ""
        vals[4] = "ICP_007009:QD Europe Chancery S.a.r.l. ICP"
        vals[14] = 10
        vals[15] = 0
        self.assertIsNone(parse_journal_line(vals, indices, 31))

    def test_read_ic_elimination_report_parses_pair_report_shape(self):
        path = self._mktemp_path("ic_elim")
        _ic_elim_workbook(
            path,
            [
                {
                    "entity": "E117100 - QD UK Holdings LP adjusted",
                    "partner": "ICP_007009 - QD Europe Chancery S.a.r.l. ICP",
                    "entity_value": 100,
                    "partner_value": -25,
                    "total": 75,
                }
            ],
            plug_code="188600",
        )

        updates, pair_keys, entity_labels, icp_labels, pair_labels, raw_lines = read_ic_elimination_report(
            path, plug_code="188600"
        )
        self.assertEqual(updates[("117100", "007009", "188600")], 75)
        self.assertIn(("117100", "007009"), pair_keys)
        self.assertEqual(entity_labels["117100"], "E117100 - QD UK Holdings LP adjusted")
        self.assertEqual(icp_labels["007009"], "ICP_007009 - QD Europe Chancery S.a.r.l. ICP")
        self.assertEqual(pair_labels[("117100", "007009")]["entity"], "E117100 - QD UK Holdings LP adjusted")
        self.assertTrue(raw_lines)

    def test_identical_journal_rows_are_preserved_and_summed(self):
        path = self._mktemp_path("parent")
        rows = [
            {
                "entity": "E117100:QD UK Holdings LP adjusted",
                "account": "534018 - 534018:Interest expense - related parties",
                "icp": "ICP_007009:QD Europe Chancery S.a.r.l. ICP",
                "debit": 10,
                "credit": 0,
            },
            {
                "entity": "E117100:QD UK Holdings LP adjusted",
                "account": "534018 - 534018:Interest expense - related parties",
                "icp": "ICP_007009:QD Europe Chancery S.a.r.l. ICP",
                "debit": 10,
                "credit": 0,
            },
        ]
        _journal_workbook(path, rows)

        lookup = read_journal_report(path)
        key = ("117100", "007009", "534018")
        self.assertEqual(len(lookup[key]), 2)
        self.assertEqual(match_journal_to_icm([], lookup)[key], 20)

    def test_write_output_places_forward_and_reverse_once_and_computes_totals(self):
        wb = openpyxl.Workbook()
        ws_icm = wb.active
        icm_header_map = {
            ("534018", "Entity"): 3,
            ("534018", "Partner"): 4,
        }
        ent_cols = [("534018", "534018:Interest expense - related parties", "S1", "Entity")]
        par_cols = [("534018", "534018:Interest expense - related parties", "S1", "Partner")]
        data_rows = [{
            "row_num": 5,
            "entity": "007009:QD UK Holdings LP adjusted",
            "partner": "ICP_E117100:QD Europe S.a.r.l. (H) ICP",
            "entity_code": "007009",
            "partner_code": "ICP_E117100",
            "partner_num": "117100",
            "row_direction": "partner_to_entity",
            "is_synthetic": False,
        }]
        parent_updates = {
            ("007009", "117100", "534018"): 25,
            ("117100", "007009", "534018"): 40,
        }

        output_path = self._mktemp_path("out")
        write_output(
            ws_icm,
            data_rows,
            icm_header_map,
            [parent_updates, {}, {}],
            output_path,
            plug_code=None,
            ent_cols=ent_cols,
            par_cols=par_cols,
        )
        out_wb = openpyxl.load_workbook(output_path, data_only=True)
        out_ws = out_wb.active

        blk_base = _block_positions(3, len(ent_cols), len(par_cols), has_plug=False)
        blk_par = _block_positions(blk_base["spacer"] + 1, len(ent_cols), len(par_cols), has_plug=True)
        blk_cont = _block_positions(blk_par["spacer"] + 1, len(ent_cols), len(par_cols), has_plug=True)
        plug_section_start = blk_cont["spacer"] + 1
        col_final = plug_section_start + 3

        row = ICM_OUTPUT_DATA_START
        self.assertEqual(out_ws.cell(row, blk_par["ent_start"]).value, 25)
        self.assertEqual(out_ws.cell(row, blk_par["par_start"]).value, 40)
        self.assertEqual(out_ws.cell(row, blk_par["total"]).value, 65)
        self.assertEqual(out_ws.cell(row, col_final).value, 65)

    def test_exact_orientation_row_wins_over_reverse_partner_side(self):
        wb = openpyxl.Workbook()
        ws_icm = wb.active
        icm_header_map = {
            ("534018", "Entity"): 3,
            ("534018", "Partner"): 4,
        }
        ent_cols = [("534018", "534018:Interest expense - related parties", "S1", "Entity")]
        par_cols = [("534018", "534018:Interest expense - related parties", "S1", "Partner")]
        data_rows = [
            {
                "row_num": 5,
                "entity": "007009:QD UK Holdings LP adjusted",
                "partner": "ICP_E117100:QD Europe S.a.r.l. (H) ICP",
                "entity_code": "007009",
                "partner_code": "ICP_E117100",
                "partner_num": "117100",
                "row_direction": "partner_to_entity",
                "is_synthetic": False,
            },
            {
                "row_num": None,
                "entity": "E117100 - QD UK Holdings LP adjusted",
                "partner": "ICP_007009 - QD Europe Chancery S.a.r.l. ICP",
                "entity_code": "117100",
                "partner_code": "ICP_007009",
                "partner_num": "007009",
                "row_direction": "entity_to_partner",
                "is_synthetic": True,
            },
        ]
        parent_updates = {
            ("007009", "117100", "534018"): 25,
            ("117100", "007009", "534018"): 40,
        }

        output_path = self._mktemp_path("exact_wins")
        write_output(
            ws_icm,
            data_rows,
            icm_header_map,
            [parent_updates, {}, {}],
            output_path,
            plug_code=None,
            ent_cols=ent_cols,
            par_cols=par_cols,
        )
        out_wb = openpyxl.load_workbook(output_path, data_only=True)
        out_ws = out_wb.active

        blk_base = _block_positions(3, len(ent_cols), len(par_cols), has_plug=False)
        blk_par = _block_positions(blk_base["spacer"] + 1, len(ent_cols), len(par_cols), has_plug=True)
        blk_cont = _block_positions(blk_par["spacer"] + 1, len(ent_cols), len(par_cols), has_plug=True)
        col_final = (blk_cont["spacer"] + 1) + 3

        reverse_row = ICM_OUTPUT_DATA_START
        forward_row = ICM_OUTPUT_DATA_START + 1
        self.assertEqual(out_ws.cell(reverse_row, blk_par["ent_start"]).value, 25)
        self.assertIsNone(out_ws.cell(reverse_row, blk_par["par_start"]).value)
        self.assertEqual(out_ws.cell(reverse_row, col_final).value, 25)
        self.assertEqual(out_ws.cell(forward_row, blk_par["ent_start"]).value, 40)
        self.assertEqual(out_ws.cell(forward_row, blk_par["par_start"]).value, 25)
        self.assertEqual(out_ws.cell(forward_row, col_final).value, 40)

    def test_process_icm_report_adds_synthetic_reverse_row_and_filters_accounts(self):
        icm_path = self._mktemp_path("icm")
        parent_path = self._mktemp_path("parent")
        contrib_path = self._mktemp_path("contrib")
        plug_path = self._mktemp_path("plug")
        inputs_path = self._mktemp_path("report_inputs")
        output_path = self._mktemp_path("output")

        _icm_workbook(icm_path)
        _report_inputs_workbook(inputs_path)
        _journal_workbook(
            parent_path,
            [
                {
                    "entity": "E117100:QD UK Holdings LP adjusted",
                    "account": "534018 - 534018:Interest expense - related parties",
                    "icp": "ICP_007009:QD Europe Chancery S.a.r.l. ICP",
                    "debit": 100,
                    "credit": 0,
                },
                {
                    "entity": "E117100:QD UK Holdings LP adjusted",
                    "account": "999999 - 999999:Ignore me",
                    "icp": "ICP_007009:QD Europe Chancery S.a.r.l. ICP",
                    "debit": 999,
                    "credit": 0,
                },
            ],
        )
        _journal_workbook(contrib_path, [])
        _journal_workbook(plug_path, [])

        process_icm_report(
            icm_path,
            {
                "parent_journal": parent_path,
                "contribution_journal": contrib_path,
                "plugaccount_journal": plug_path,
            },
            output_path,
            report_inputs_path=inputs_path,
        )

        wb = openpyxl.load_workbook(output_path, data_only=True)
        ws = wb.active
        rows = {}
        for r in range(ICM_OUTPUT_DATA_START, ws.max_row + 1):
            entity = str(ws.cell(r, 1).value or "").strip()
            partner = str(ws.cell(r, 2).value or "").strip()
            if entity or partner:
                rows[(entity, partner)] = r

        synthetic_key = (
            "E117100 - QD UK Holdings LP adjusted",
            "ICP_007009 - QD Europe Chancery S.a.r.l. ICP",
        )
        reverse_key = (
            "007009:QD UK Holdings LP adjusted",
            "ICP_E117100:QD Europe S.a.r.l. (H) ICP",
        )

        self.assertIn(synthetic_key, rows)
        self.assertNotIn(reverse_key, rows)

        blk_base = _block_positions(3, 1, 1, has_plug=False)
        blk_par = _block_positions(blk_base["spacer"] + 1, 1, 1, has_plug=True)
        blk_cont = _block_positions(blk_par["spacer"] + 1, 1, 1, has_plug=True)
        col_final = (blk_cont["spacer"] + 1) + 3

        synthetic_row = rows[synthetic_key]
        self.assertEqual(ws.cell(synthetic_row, blk_par["ent_start"]).value, 100)
        self.assertEqual(ws.cell(synthetic_row, col_final).value, 100)

        self.assertIn(ws.cell(synthetic_row, blk_par["par_start"]).value, (0, None))

    def test_process_icm_report_merges_reciprocal_pairs_into_one_entity_row(self):
        icm_path = self._mktemp_path("icm_merge")
        parent_path = self._mktemp_path("parent_merge")
        contrib_path = self._mktemp_path("contrib_merge")
        plug_path = self._mktemp_path("plug_merge")
        inputs_path = self._mktemp_path("report_inputs_merge")
        output_path = self._mktemp_path("output_merge")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(4, 1).value = "Entity"
        ws.cell(4, 2).value = "Partner"
        ws.cell(4, 3).value = "534018 - 534018:Interest expense - related parties Entity"
        ws.cell(4, 4).value = "534018 - 534018:Interest expense - related parties Partner"
        ws.cell(5, 1).value = "007009:QD UK Holdings LP adjusted"
        ws.cell(5, 2).value = "ICP_E117100:QD Europe S.a.r.l. (H) ICP"
        ws.cell(6, 1).value = "E117100:QD UK Holdings LP adjusted"
        ws.cell(6, 2).value = "ICP_007009:QD Europe Chancery S.a.r.l. ICP"
        wb.save(icm_path)

        _report_inputs_workbook(inputs_path)
        _journal_workbook(
            parent_path,
            [
                {
                    "entity": "E117100:QD UK Holdings LP adjusted",
                    "account": "534018 - 534018:Interest expense - related parties",
                    "icp": "ICP_007009:QD Europe Chancery S.a.r.l. ICP",
                    "debit": 100,
                    "credit": 0,
                },
                {
                    "entity": "007009:QD UK Holdings LP adjusted",
                    "account": "534018 - 534018:Interest expense - related parties",
                    "icp": "ICP_E117100:QD Europe S.a.r.l. (H) ICP",
                    "debit": 0,
                    "credit": 25,
                },
            ],
        )
        _journal_workbook(contrib_path, [])
        _journal_workbook(plug_path, [])

        process_icm_report(
            icm_path,
            {
                "parent_journal": parent_path,
                "contribution_journal": contrib_path,
                "plugaccount_journal": plug_path,
            },
            output_path,
            report_inputs_path=inputs_path,
        )

        wb = openpyxl.load_workbook(output_path, data_only=True)
        ws = wb.active
        rows = {}
        for r in range(ICM_OUTPUT_DATA_START, ws.max_row + 1):
            entity = str(ws.cell(r, 1).value or "").strip()
            partner = str(ws.cell(r, 2).value or "").strip()
            if entity or partner:
                rows[(entity, partner)] = r

        forward_key = (
            "E117100 - QD UK Holdings LP adjusted",
            "ICP_007009 - QD Europe Chancery S.a.r.l. ICP",
        )
        reverse_key = (
            "007009 - QD UK Holdings LP adjusted",
            "ICP_E117100 - QD Europe S.a.r.l. (H) ICP",
        )
        self.assertIn(forward_key, rows)
        self.assertNotIn(reverse_key, rows)

        blk_base = _block_positions(3, 1, 1, has_plug=False)
        blk_par = _block_positions(blk_base["spacer"] + 1, 1, 1, has_plug=True)
        col_final = (_block_positions(blk_par["spacer"] + 1, 1, 1, has_plug=True)["spacer"] + 1) + 3

        row = rows[forward_key]
        self.assertEqual(ws.cell(row, blk_par["ent_start"]).value, 100)
        self.assertEqual(ws.cell(row, blk_par["par_start"]).value, -25)
        self.assertEqual(ws.cell(row, blk_par["total"]).value, 75)
        self.assertEqual(ws.cell(row, col_final).value, 75)

    def test_pair_specific_labels_override_ambiguous_numeric_code_labels(self):
        icm_path = self._mktemp_path("icm_labels")
        parent_path = self._mktemp_path("parent_labels")
        contrib_path = self._mktemp_path("contrib_labels")
        plug_path = self._mktemp_path("plug_labels")
        inputs_path = self._mktemp_path("report_inputs_labels")
        output_path = self._mktemp_path("output_labels")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(4, 1).value = "Entity"
        ws.cell(4, 2).value = "Partner"
        ws.cell(4, 3).value = "534018 - 534018:Interest expense - related parties Entity"
        ws.cell(4, 4).value = "534018 - 534018:Interest expense - related parties Partner"
        ws.cell(5, 1).value = "007009 - QD UK Holdings LP adjusted"
        ws.cell(5, 2).value = "ICP_E117100 - QD UK Holdings LP adjusted ICP"
        wb.save(icm_path)

        _report_inputs_workbook(inputs_path)
        _journal_workbook(
            parent_path,
            [
                {
                    "entity": "007009:QD UK Holdings LP adjusted",
                    "account": "534018 - 534018:Interest expense - related parties",
                    "icp": "ICP_E117100:QD Europe S.a.r.l. (H) ICP",
                    "debit": 0,
                    "credit": 17882306.93,
                },
            ],
        )
        _journal_workbook(contrib_path, [])
        _journal_workbook(plug_path, [])

        process_icm_report(
            icm_path,
            {
                "parent_journal": parent_path,
                "contribution_journal": contrib_path,
                "plugaccount_journal": plug_path,
            },
            output_path,
            report_inputs_path=inputs_path,
        )

        wb = openpyxl.load_workbook(output_path, data_only=True)
        ws = wb.active
        found = None
        for r in range(ICM_OUTPUT_DATA_START, ws.max_row + 1):
            entity = str(ws.cell(r, 1).value or "").strip()
            partner = str(ws.cell(r, 2).value or "").strip()
            if entity == "007009 - QD UK Holdings LP adjusted":
                found = (r, partner)
                break

        self.assertIsNotNone(found)
        self.assertEqual(found[1], "ICP_E117100 - QD Europe S.a.r.l. (H) ICP")

    def test_process_icm_report_adds_raw_ic_elimination_detail_sheet(self):
        icm_path = self._mktemp_path("icm_plug")
        parent_path = self._mktemp_path("parent_plug")
        contrib_path = self._mktemp_path("contrib_plug")
        plug_path = self._mktemp_path("plug_plug")
        inputs_path = self._mktemp_path("report_inputs_plug")
        output_path = self._mktemp_path("output_plug")

        _icm_workbook(icm_path)
        _report_inputs_workbook(inputs_path, plug_code="188600", elim_code="534018")
        _journal_workbook(parent_path, [])
        _journal_workbook(contrib_path, [])
        _ic_elim_workbook(
            plug_path,
            [
                {
                    "entity": "E117100 - QD UK Holdings LP adjusted",
                    "partner": "ICP_007009 - QD Europe Chancery S.a.r.l. ICP",
                    "entity_value": 11,
                    "partner_value": None,
                    "total": 11,
                },
            ],
        )

        process_icm_report(
            icm_path,
            {
                "parent_journal": parent_path,
                "contribution_journal": contrib_path,
                "plugaccount_journal": plug_path,
            },
            output_path,
            report_inputs_path=inputs_path,
        )

        wb = openpyxl.load_workbook(output_path, data_only=True)
        self.assertIn("IC Elim Detail", wb.sheetnames)
        ws = wb["IC Elim Detail"]
        self.assertEqual(ws.cell(2, 1).value, "IC Elim Row 33")
        self.assertEqual(ws.cell(2, 5).value, "117100")
        self.assertEqual(ws.cell(2, 7).value, "534018")
        self.assertEqual(ws.cell(2, 9).value, "007009")

    def test_process_icm_report_includes_pairs_from_ic_elimination_report(self):
        icm_path = self._mktemp_path("icm_icelim")
        parent_path = self._mktemp_path("parent_icelim")
        contrib_path = self._mktemp_path("contrib_icelim")
        plug_path = self._mktemp_path("plug_icelim")
        inputs_path = self._mktemp_path("report_inputs_icelim")
        output_path = self._mktemp_path("output_icelim")

        _icm_workbook(icm_path)
        _report_inputs_workbook(inputs_path, plug_code="188600", elim_code="534018")
        _journal_workbook(parent_path, [])
        _journal_workbook(contrib_path, [])
        _ic_elim_workbook(
            plug_path,
            [
                {
                    "entity": "001001 - QDRE Investment Company Q.S.C. (H)",
                    "partner": "ICP_E100000 - QDRE Investment Company Q.C.S.C. ICP",
                    "entity_value": None,
                    "partner_value": -125,
                    "total": -125,
                },
            ],
            plug_code="188600",
        )

        process_icm_report(
            icm_path,
            {
                "parent_journal": parent_path,
                "contribution_journal": contrib_path,
                "plugaccount_journal": plug_path,
            },
            output_path,
            report_inputs_path=inputs_path,
        )

        wb = openpyxl.load_workbook(output_path, data_only=True)
        ws = wb["ICM Matched"]
        found = None
        for r in range(ICM_OUTPUT_DATA_START, ws.max_row + 1):
            entity = str(ws.cell(r, 1).value or "").strip()
            partner = str(ws.cell(r, 2).value or "").strip()
            if entity == "001001 - QDRE Investment Company Q.S.C. (H)":
                found = (r, partner)
                break

        self.assertIsNotNone(found)
        self.assertEqual(found[1], "ICP_E100000 - QDRE Investment Company Q.C.S.C. ICP")


if __name__ == "__main__":
    unittest.main()
