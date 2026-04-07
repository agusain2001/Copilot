"""
Count how many times each Parent journal transaction appears in the ICM Output.
Each of the 35 transactions should appear exactly ONCE.
CHECKING V2 OUTPUT.
"""
import sys, re
sys.path.insert(0, '.')
import openpyxl
from app.ic_processor import (
    extract_entity_code_journal, normalize_entity_code, extract_icp_code,
    extract_account_code, apply_sign, to_float, get_journal_indices,
    JOURNAL_DATA_START, is_detail_row
)
from collections import defaultdict

f = open("tmp_count_result.txt", "w", encoding="utf-8")

# Read Parent journal — get every individual transaction
wb_j = openpyxl.load_workbook(r"uploads\reports\31\inputs\Parent report.xlsx", data_only=True)
ws_j = wb_j.active
indices = get_journal_indices(ws_j)

transactions = []
for row_num, row in enumerate(ws_j.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws_j.max_row), start=JOURNAL_DATA_START):
    vals = [cell.value for cell in row]
    if not vals: continue
    label = str(vals[0] or "").strip()
    if label == "Grand Total": break
    if not is_detail_row(vals, indices): continue
    try:
        entity_raw = str(vals[indices["entity"]] or "").strip()
        account_raw = str(vals[indices["acct"]] or "").strip()
        icp_raw = str(vals[indices["icp"]] or "").strip()
        debit = to_float(vals[indices["debit"]])
        credit = to_float(vals[indices["credit"]])
    except IndexError:
        continue
    ent = normalize_entity_code(extract_entity_code_journal(entity_raw))
    icp = extract_icp_code(icp_raw)
    acct = extract_account_code(account_raw)
    signed = apply_sign(debit, credit, acct)
    if ent and icp and acct and signed != 0:
        transactions.append({
            "row": row_num, "entity": ent, "icp": icp, "account": acct,
            "debit": debit, "credit": credit, "signed": signed,
            "entity_raw": entity_raw,
        })

f.write(f"Parent journal: {len(transactions)} transactions\n\n")

# Read ICM output
wb_o = openpyxl.load_workbook(r"uploads\reports\31\outputs\ICM_Output_31_v3.xlsx", data_only=True)
ws_o = wb_o.active

# Build column map for Parent block only (cols 17-29)
parent_entity_cols = {17: "534018", 18: "433002", 19: "111006", 20: "120030", 21: "121015"}
parent_partner_cols = {23: "433002", 24: "534018", 25: "111006", 26: "120030", 27: "121015"}

# For each output row, collect all Parent block values
output_parent_values = []  # (output_row, entity, partner, side, acct, value)

for r in range(33, ws_o.max_row + 1):
    ent_disp = str(ws_o.cell(r, 1).value or "").strip()
    prt_disp = str(ws_o.cell(r, 2).value or "").strip()
    if not ent_disp or not prt_disp: continue
    
    # Extract codes
    m = re.match(r"(\d{6})", ent_disp)
    if m:
        ent_code = m.group(1)
    else:
        m = re.match(r"E(\d+)", ent_disp)
        ent_code = m.group(1) if m else ""
    
    m = re.match(r"(ICP_\w+)", prt_disp)
    prt_code = m.group(1) if m else ""
    
    if not ent_code or not prt_code: continue
    
    # Entity-side values
    for c, acct in parent_entity_cols.items():
        v = ws_o.cell(r, c).value
        if v is not None and v != "" and v != 0:
            output_parent_values.append((r, ent_code, prt_code, "entity", acct, to_float(v)))
    
    # Partner-side values
    prt_digit = prt_code[4:] if prt_code.startswith("ICP_") else prt_code
    if prt_digit.startswith("E") and len(prt_digit) > 1 and prt_digit[1:].isdigit():
        prt_entity = prt_digit[1:]
    else:
        prt_entity = prt_digit
    
    for c, acct in parent_partner_cols.items():
        v = ws_o.cell(r, c).value
        if v is not None and v != "" and v != 0:
            output_parent_values.append((r, prt_entity, f"ICP_{ent_code}", "partner", acct, to_float(v)))

f.write(f"Total Parent block non-zero cells in output: {len(output_parent_values)}\n")
f.write(f"Expected: {len(transactions)} (one per transaction)\n\n")

# Map each output value back to a journal transaction
for ov in sorted(output_parent_values, key=lambda x: (x[0], x[3], x[4])):
    r, ent, icp, side, acct, val = ov
    # Find matching journal transaction
    matches = [t for t in transactions if t["entity"] == ent and t["icp"] == icp and t["account"] == acct]
    
    # Also try with ICP_E prefix
    icp_e = icp.replace("ICP_", "ICP_E") if not icp.startswith("ICP_E") else icp
    matches_e = [t for t in transactions if t["entity"] == ent and t["icp"] == icp_e and t["account"] == acct]
    
    all_matches = matches + matches_e
    
    status = "OK" if all_matches else "NO MATCH"
    match_info = ""
    if all_matches:
        m = all_matches[0]
        match_info = f"JnlRow={m['row']}, expect={m['signed']:.2f}"
    
    f.write(f"  OutRow={r:3d} side={side:7s} key=({ent}, {icp}, {acct}) val={val:>15.2f} [{status}] {match_info}\n")

# Count how many times each transaction is used
f.write(f"\n\n--- Transaction usage count ---\n")
txn_usage = defaultdict(int)
for ov in output_parent_values:
    r, ent, icp, side, acct, val = ov
    for t in transactions:
        if t["entity"] == ent and t["account"] == acct:
            # Check ICP match (with or without E prefix)
            if t["icp"] == icp or t["icp"] == icp.replace("ICP_", "ICP_E"):
                txn_usage[t["row"]] += 1

total_uses = 0
for t in transactions:
    count = txn_usage.get(t["row"], 0)
    total_uses += count
    status = "OK" if count == 1 else ("MISSING" if count == 0 else f"DUPLICATE x{count}")
    f.write(f"  JnlRow {t['row']:3d}: ({t['entity']}, {t['icp']}, {t['account']}) = {t['signed']:>15.2f} -> used {count} times [{status}]\n")

f.write(f"\nTotal transaction uses: {total_uses} (should be {len(transactions)})\n")
f.write(f"Duplicated: {sum(1 for v in txn_usage.values() if v > 1)}\n")
f.write(f"Missing: {sum(1 for t in transactions if txn_usage.get(t['row'], 0) == 0)}\n")

f.close()
print("Done - see tmp_count_result.txt")
