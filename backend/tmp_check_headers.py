import openpyxl

path = r"g:\FCCS\backend\uploads\reports\9\inputs\Intercompany Balances IC Matching Report (1).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

for sheet in wb.worksheets:
    for r in [4, 32]:
        c1 = str(sheet.cell(r, 1).value or "").strip()
        c2 = str(sheet.cell(r, 2).value or "").strip()
        print(f"Sheet '{sheet.title}' Row {r}: C1='{c1}' C2='{c2}'")
