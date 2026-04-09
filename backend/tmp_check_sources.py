"""Check report_inputs.xlsx for elimination accounts and also check 
if 155020, Plug_InvSh exist anywhere in the ICM or report_inputs."""
import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl

# Check report_inputs from report 22
ri_path = r'G:\FCCS\backend\uploads\reports\22\inputs\report_inputs.xlsx'
wb = openpyxl.load_workbook(ri_path, data_only=True)
ws = wb.active

print("report_inputs.xlsx (Report 22):")
print(f"  Row 1 headers: {[str(c.value or '')[:50] for c in ws[1]]}")
print(f"\n  All rows:")
for r in range(1, ws.max_row + 1):
    row_vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    non_empty = [v for v in row_vals if v]
    if non_empty:
        print(f"  Row {r}: {row_vals[:5]}")

# Check ICM source for 155020 and Plug_InvSh in ALL cells (not just headers)
print("\n\nSearching ICM source for 155020 and Plug_InvSh in all cells...")
icm_path = r'G:\FCCS\Update files\Intercompany Balances IC Matching Report (1).xlsx'
wb2 = openpyxl.load_workbook(icm_path, data_only=True)
for sheet in wb2.worksheets:
    for r in range(1, min(50, sheet.max_row + 1)):
        for c in range(1, sheet.max_column + 1):
            v = str(sheet.cell(r, c).value or '')
            if '155020' in v or 'Plug_InvSh' in v or 'plug_inv' in v.lower():
                print(f"  Sheet '{sheet.title}', Row {r}, Col {c}: {v[:80]}")

# Check if those accounts appear in the ICM_Output_NEW_RUN which the user provided
print("\n\nSearching ICM_Output_NEW_RUN.xlsx for 155020 and Plug_InvSh...")
wb3 = openpyxl.load_workbook(r'G:\FCCS\Update files\ICM_Output_NEW_RUN.xlsx', data_only=True)
ws3 = wb3.active
for r in [32]:  # header row
    for c in range(1, ws3.max_column + 1):
        v = str(ws3.cell(r, c).value or '')
        if '155020' in v or 'Plug_InvSh' in v:
            print(f"  Row {r}, Col {c}: {v[:80]}")

# Check the generate_proper_output.py to see if those were defined there
print("\n\nChecking ICM_Output_WITH_MISSING.xlsx headers...")
wb4 = openpyxl.load_workbook(r'G:\FCCS\Update files\ICM_Output_WITH_MISSING.xlsx', data_only=True)
ws4 = wb4.active
for c in range(1, ws4.max_column + 1):
    v = str(ws4.cell(32, c).value or '')
    if '155020' in v or 'Plug_InvSh' in v or '189501' in v:
        print(f"  Col {c}: {v[:80]}")
