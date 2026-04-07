"""Check how 001032/001033 appear in ICM source and FIXED reference — write to file."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
from app.ic_processor import (
    detect_icm_header_row, extract_entity_code_icm,
    extract_icp_code, normalize_to_numeric
)

out = open("tmp_pair_detail.txt", "w", encoding="utf-8")

icm_path = r"g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx"
wb = openpyxl.load_workbook(icm_path, data_only=True)

for sheet in wb.worksheets:
    hdr_row, data_start = detect_icm_header_row(sheet)
    out.write(f"Sheet: {sheet.title}, header={hdr_row}, data_start={data_start}, max_col={sheet.max_column}\n")
    
    for r in range(data_start, min(sheet.max_row + 1, 400)):
        ent = str(sheet.cell(r, 1).value or "").strip()
        prt = str(sheet.cell(r, 2).value or "").strip()
        ent_code = extract_entity_code_icm(ent)
        prt_icp = extract_icp_code(prt)
        prt_num = normalize_to_numeric(prt_icp)
        
        if ent_code in ("001032", "001033") or prt_num in ("001032", "001033"):
            out.write(f"\n  ICM Row {r}: entity='{ent}' partner='{prt}'\n")
            out.write(f"    entity_code={ent_code} partner_icp={prt_icp} partner_num={prt_num}\n")
            # Show all non-None values
            for c in range(3, sheet.max_column + 1):
                v = sheet.cell(r, c).value
                h = str(sheet.cell(hdr_row, c).value or "")[:60]
                if v is not None and str(v).strip() not in ("", "0"):
                    out.write(f"    Col {c} ({h}): {v}\n")

# Reference output
out.write("\n\n=== FIXED REFERENCE ===\n")
ref_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_FIXED.xlsx"
wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
ws_ref = wb_ref.active

for r in range(33, ws_ref.max_row + 1):
    ent = str(ws_ref.cell(r, 1).value or "").strip()
    prt = str(ws_ref.cell(r, 2).value or "").strip()
    if "001033" in ent or "001032" in ent:
        out.write(f"\nREF Row {r}: Entity='{ent}' Partner='{prt}'\n")
        for c in range(3, ws_ref.max_column + 1):
            v = ws_ref.cell(r, c).value
            if v is not None and v != 0 and str(v).strip() not in ("", "0"):
                hdr = str(ws_ref.cell(32, c).value or f"Col{c}")[:60]
                out.write(f"  Col {c} ({hdr}): {v}\n")
        if all(ws_ref.cell(r, c).value in (None, 0, "", " ") for c in range(3, ws_ref.max_column + 1)):
            out.write("  [ALL VALUES ZERO/EMPTY]\n")

out.close()
print("Done")
