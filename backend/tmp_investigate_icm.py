"""Check how 001032/001033 appear in the ICM source file."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
from app.ic_processor import (
    detect_icm_header_row, read_icm_data, extract_entity_code_icm,
    extract_icp_code, normalize_to_numeric
)

icm_path = r"g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx"
wb = openpyxl.load_workbook(icm_path, data_only=True)

for sheet in wb.worksheets:
    hdr_row, data_start = detect_icm_header_row(sheet)
    print(f"Sheet: {sheet.title}, header={hdr_row}, data_start={data_start}")
    
    for r in range(data_start, sheet.max_row + 1):
        ent = str(sheet.cell(r, 1).value or "").strip()
        prt = str(sheet.cell(r, 2).value or "").strip()
        ent_code = extract_entity_code_icm(ent)
        prt_icp = extract_icp_code(prt)
        prt_num = normalize_to_numeric(prt_icp)
        
        if ent_code in ("001032", "001033") or prt_num in ("001032", "001033"):
            print(f"  Row {r}: entity='{ent[:60]}' partner='{prt[:60]}'")
            print(f"    entity_code={ent_code} partner_icp={prt_icp} partner_num={prt_num}")

# Also check: what does the FIXED reference show for this pair in detail
ref_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_FIXED.xlsx"
wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
ws_ref = wb_ref.active

for r in range(33, ws_ref.max_row + 1):
    ent = str(ws_ref.cell(r, 1).value or "").strip()
    prt = str(ws_ref.cell(r, 2).value or "").strip()
    if "001033" in ent or "001032" in ent:
        print(f"\nREF Row {r}: Entity='{ent}' Partner='{prt}'")
        for c in range(3, ws_ref.max_column + 1):
            v = ws_ref.cell(r, c).value
            if v is not None and v != 0:
                hdr = str(ws_ref.cell(32, c).value or f"Col{c}")[:60]
                print(f"  Col {c} ({hdr}): {v}")
