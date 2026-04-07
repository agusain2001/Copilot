"""Verify partnerless rows have no values, and partnered rows still match."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl, logging
logging.basicConfig(level=logging.WARNING)

output_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_VERIFY3.xlsx"
icm_path = r"g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx"
journal_paths = {
    "parent_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx",
    "contribution_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Contribution report.xlsx",
    "plugaccount_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Journal Report (4).xlsx",
}
report_inputs_path = r"g:\FCCS\backend\uploads\reports\31\inputs\report Inputs.xlsx"

from app.ic_processor import process_icm_report
process_icm_report(icm_path, journal_paths, output_path, report_inputs_path)

wb = openpyxl.load_workbook(output_path, data_only=True)
ws = wb.active

out = open("tmp_partnerless_check.txt", "w", encoding="utf-8")

out.write("=== PARTNERLESS ROWS CHECK ===\n")
partnerless_with_values = 0
partnerless_clean = 0
partnered_count = 0

for r in range(33, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or "").strip()
    prt = str(ws.cell(r, 2).value or "").strip()
    if not ent and not prt:
        continue
    
    has_data = False
    for c in range(3, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v is not None and v != 0 and str(v).strip() not in ("", "0"):
            has_data = True
            break
    
    if ent and not prt:
        if has_data:
            partnerless_with_values += 1
            out.write(f"\n  [WARNING] Row {r}: Entity='{ent[:50]}' Partner=EMPTY — HAS VALUES!\n")
            for c in range(3, min(ws.max_column + 1, 20)):
                v = ws.cell(r, c).value
                if v is not None and v != 0:
                    hdr = str(ws.cell(32, c).value or f"Col{c}")[:50]
                    out.write(f"    Col {c} ({hdr}): {v}\n")
        else:
            partnerless_clean += 1
            out.write(f"  [OK] Row {r}: Entity='{ent[:50]}' Partner=EMPTY — NO values\n")
    elif ent and prt:
        partnered_count += 1

out.write(f"\n\n=== SUMMARY ===\n")
out.write(f"Partnered rows (have values): {partnered_count}\n")
out.write(f"Partnerless rows (CLEAN, no values): {partnerless_clean}\n")
out.write(f"Partnerless rows with values (VIOLATIONS): {partnerless_with_values}\n")

# Also compare partnered rows against reference
ref_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_FIXED.xlsx"
wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
ws_ref = wb_ref.active

def to_num(v):
    if v is None or str(v).strip() in ("", " "): return 0.0
    try: return round(float(v), 2)
    except: return v

def pair_map(ws):
    m = {}
    for r in range(33, ws.max_row + 1):
        e = str(ws.cell(r, 1).value or "").strip()
        p = str(ws.cell(r, 2).value or "").strip()
        if (e and p): m[(e, p)] = r  # Only count paired rows
    return m

ref_pm = pair_map(ws_ref)
new_pm = pair_map(ws)
common = set(ref_pm) & set(new_pm)

mismatches = 0
for pair in sorted(common):
    rr = ref_pm[pair]
    nr = new_pm[pair]
    for c in range(3, max(ws_ref.max_column, ws.max_column) + 1):
        rv = to_num(ws_ref.cell(rr, c).value)
        nv = to_num(ws.cell(nr, c).value)
        if rv != nv:
            mismatches += 1

out.write(f"\n=== PARTNERED ROW VALUE COMPARISON ===\n")
out.write(f"Common paired rows: {len(common)}\n")
out.write(f"Mismatches (partnered only): {mismatches}\n")
out.close()
print("Done")
