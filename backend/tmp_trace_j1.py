"""Trace exact journal entries for entity 001001/001033 to understand
what gets matched as partner-side values."""
import sys
sys.path.insert(0, r'G:\FCCS\backend')
import openpyxl
from app.ic_processor import (extract_entity_code_journal, extract_icp_code,
                               extract_account_code, is_detail_row,
                               J_ENTITY, J_ACCT, J_ICP, J_DEBIT, J_CREDIT,
                               JOURNAL_DATA_START)

j1_path = r'G:\FCCS\Update files\Journal Report.xlsx'
wb = openpyxl.load_workbook(j1_path, data_only=True)
ws = wb.active

print("=== Journal 1 (Parent) entries involving 001001 or 001033 ===\n")

look_for_entities = {'001001', '001033', '001032'}
look_for_icps = {'ICP_001001', 'ICP_001033', 'ICP_001032'}

found = []
for row_num in range(JOURNAL_DATA_START, ws.max_row + 1):
    vals = [ws.cell(row_num, c).value for c in range(1, 20)]
    label = str(vals[0] or "").strip()
    if label == "Grand Total":
        break
    if not is_detail_row(vals):
        continue
    
    entity_raw  = str(vals[J_ENTITY - 1] or "").strip()
    account_raw = str(vals[J_ACCT - 1] or "").strip()
    icp_raw     = str(vals[J_ICP - 1] or "").strip()
    debit       = vals[J_DEBIT - 1]
    credit      = vals[J_CREDIT - 1]
    
    ent_code = extract_entity_code_journal(entity_raw)
    icp_code = extract_icp_code(icp_raw)
    acct_code = extract_account_code(account_raw)
    
    if ent_code in look_for_entities or icp_code in look_for_icps:
        found.append({
            'row': row_num,
            'entity_raw': entity_raw[:45],
            'ent_code': ent_code,
            'acct_raw': account_raw[:35],
            'acct_code': acct_code,
            'icp_raw': icp_raw[:50],
            'icp_code': icp_code,
            'debit': debit,
            'credit': credit,
        })

for f in found:
    print(f"Row {f['row']}:")
    print(f"  Entity:  {f['entity_raw']}  code={f['ent_code']}")
    print(f"  Account: {f['acct_raw']}  code={f['acct_code']}")
    print(f"  ICP:     {f['icp_raw']}  code={f['icp_code']}")
    print(f"  Debit={f['debit']}  Credit={f['credit']}")
    print()

# Check: does ICP_001001 appear in J1 at all?
print("\n=== All ICP codes containing '001001' in Journal 1 ===")
for row_num in range(JOURNAL_DATA_START, ws.max_row + 1):
    vals = [ws.cell(row_num, c).value for c in range(1, 20)]
    if str(vals[0] or "").strip() == "Grand Total":
        break
    icp_raw = str(vals[J_ICP - 1] or "").strip()
    if '001001' in icp_raw:
        ent_raw = str(vals[J_ENTITY - 1] or "").strip()
        acct_raw = str(vals[J_ACCT - 1] or "").strip()
        print(f"  Row {row_num}: Entity={ent_raw[:30]}  ICP={icp_raw[:50]}  Acct={acct_raw[:30]}")

# Now check what would match for ICM row Entity=001001, Partner=ICP_001033
print("\n=== Vice-versa match for (Entity=001001, Partner=ICP_001033) ===")
print("  Looking for journal keys: (001033, ICP_001001, any_account)")
for f in found:
    if f['ent_code'] == '001033' and f['icp_code'] == 'ICP_001001':
        print(f"  MATCH: Row {f['row']} acct={f['acct_code']} D={f['debit']} C={f['credit']}")
