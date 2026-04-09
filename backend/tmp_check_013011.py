"""Quick verify: Check entity 013011 430015 values in the output for Parent block."""
import openpyxl

output_path = r"g:\FCCS\backend\uploads\reports\26\outputs\ICM_Output_26.xlsx"
wb = openpyxl.load_workbook(output_path, data_only=True)
ws = wb.active

# Find 430015 Entity column in Parent block (should be around col 72 based on earlier check)
for c in range(60, 100):
    h = str(ws.cell(32, c).value or '')
    if '430015' in h:
        print(f"Col {c}: '{h}'")
        for r in range(33, ws.max_row + 1):
            e = str(ws.cell(r, 1).value or '')
            p = str(ws.cell(r, 2).value or '')
            v = ws.cell(r, c).value
            if v is not None and '013011' in e:
                print(f"  Row {r}: E='{e[:20]}', P='{p[:25]}', Val={v}")
        print()
