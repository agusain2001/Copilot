"""
Run the full processor and check the output for entity-side vs partner-side values.
"""
import sys
sys.path.insert(0, '.')
import logging
logging.basicConfig(level=logging.INFO, format='%(name)s - %(levelname)s - %(message)s')

from app.ic_processor import process_icm_report

icm_path = r'uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx'
journal_paths = {
    'parent_journal': r'uploads\reports\31\inputs\Parent report.xlsx',
    'contribution_journal': r'uploads\reports\31\inputs\Contribution report.xlsx',
}
report_inputs = r'uploads\reports\31\inputs\report Inputs.xlsx'
output_path = r'uploads\reports\31\outputs\ICM_Output_31_test.xlsx'

result = process_icm_report(icm_path, journal_paths, output_path, report_inputs)
print(f"\nOutput saved to: {result}")

# Now analyze the output
import openpyxl
wb = openpyxl.load_workbook(output_path, data_only=True)
ws = wb.active

# Find rows with data
print("\n=== ANALYZING OUTPUT ===")
for r in range(33, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or '').strip()
    prt = str(ws.cell(r, 2).value or '').strip()
    if not ent and not prt:
        continue
    
    # Check if any cell beyond column 2 has a non-zero value
    has_data = False
    for c in range(3, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v is not None and v != 0 and v != '':
            has_data = True
            break
    
    if has_data:
        # Get a sample of values
        vals = []
        for c in range(3, min(ws.max_column + 1, 40)):
            v = ws.cell(r, c).value
            if v is not None and v != 0:
                vals.append((c, v))
        print(f"Row {r}: E={ent[:30]:30s} P={prt[:30]:30s} data_cols: {vals[:10]}")
