"""Analyze ICM_Output_31_FIXED.xlsx to understand expected structure & values,
then run the processor and compare."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl

# ── 1. Analyze reference output ─────────────────────────────────────────────
ref_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_FIXED.xlsx"
wb = openpyxl.load_workbook(ref_path, data_only=True)
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Header row (row 32)
print("\n=== HEADER ROW (row 32) ===")
headers = []
for c in range(1, ws.max_column + 1):
    v = ws.cell(32, c).value
    if v:
        headers.append((c, str(v).strip()[:80]))
for col, h in headers:
    print(f"  Col {col:3d}: {h}")

# Data rows
print("\n=== DATA ROWS (starting row 33) ===")
data_count = 0
entities = set()
for r in range(33, ws.max_row + 1):
    ent = ws.cell(r, 1).value
    prt = ws.cell(r, 2).value
    if ent or prt:
        data_count += 1
        entities.add(str(ent or ""))
        if data_count <= 5:
            # Print first 5 rows with all non-None values
            vals = []
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(r, c).value
                if v is not None:
                    vals.append(f"C{c}={v}")
            print(f"  Row {r}: {', '.join(vals)}")

print(f"\nTotal data rows: {data_count}")
print(f"Unique entities: {len(entities)}")

# Section labels (row 29)
print("\n=== SECTION LABELS (row 29) ===")
for c in range(1, ws.max_column + 1):
    v = ws.cell(29, c).value
    if v:
        print(f"  Col {c}: {v}")

# ── 2. Run processor and compare ────────────────────────────────────────────
print("\n\n" + "=" * 65)
print("  RUNNING PROCESSOR")
print("=" * 65)

from app.ic_processor import process_icm_report

icm_path = r"g:\FCCS\backend\uploads\reports\31\inputs\IC Elimination Report_188800_Intercompany Balances Plug A_c_1156_Intercompany Report 1.xlsx"
journal_paths = {
    "parent_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Parent report.xlsx",
    "contribution_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Contribution report.xlsx",
    "plugaccount_journal": r"g:\FCCS\backend\uploads\reports\31\inputs\Journal Report (4).xlsx",
}
report_inputs_path = r"g:\FCCS\backend\uploads\reports\31\inputs\report Inputs.xlsx"
output_path = r"g:\FCCS\backend\uploads\reports\31\outputs\ICM_Output_31_NEW.xlsx"

import logging
logging.basicConfig(level=logging.INFO, format="%(message)s")

process_icm_report(icm_path, journal_paths, output_path, report_inputs_path)

# ── 3. Compare outputs ──────────────────────────────────────────────────────
print("\n\n" + "=" * 65)
print("  COMPARING OUTPUTS")
print("=" * 65)

wb_new = openpyxl.load_workbook(output_path, data_only=True)
ws_new = wb_new.active

wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
ws_ref = wb_ref.active

print(f"Reference: {ws_ref.max_row} rows x {ws_ref.max_column} cols")
print(f"New:       {ws_new.max_row} rows x {ws_new.max_column} cols")

# Compare row counts
ref_data = 0
new_data = 0
for r in range(33, ws_ref.max_row + 1):
    if ws_ref.cell(r, 1).value or ws_ref.cell(r, 2).value:
        ref_data += 1
for r in range(33, ws_new.max_row + 1):
    if ws_new.cell(r, 1).value or ws_new.cell(r, 2).value:
        new_data += 1
print(f"Reference data rows: {ref_data}")
print(f"New data rows:       {new_data}")

# Build entity-partner pair maps for both
def get_pair_map(ws):
    pairs = {}
    for r in range(33, ws.max_row + 1):
        ent = str(ws.cell(r, 1).value or "").strip()
        prt = str(ws.cell(r, 2).value or "").strip()
        if ent or prt:
            pairs[(ent, prt)] = r
    return pairs

ref_pairs = get_pair_map(ws_ref)
new_pairs = get_pair_map(ws_new)

only_in_ref = set(ref_pairs.keys()) - set(new_pairs.keys())
only_in_new = set(new_pairs.keys()) - set(ref_pairs.keys())
common = set(ref_pairs.keys()) & set(new_pairs.keys())

if only_in_ref:
    print(f"\nPairs ONLY in reference ({len(only_in_ref)}):")
    for p in sorted(only_in_ref):
        print(f"  {p}")
if only_in_new:
    print(f"\nPairs ONLY in new ({len(only_in_new)}):")
    for p in sorted(only_in_new):
        print(f"  {p}")

# Compare values for common pairs
print(f"\nCommon pairs: {len(common)}")
mismatches = 0
for pair in sorted(common):
    ref_r = ref_pairs[pair]
    new_r = new_pairs[pair]
    max_c = max(ws_ref.max_column, ws_new.max_column)
    for c in range(3, max_c + 1):
        ref_v = ws_ref.cell(ref_r, c).value
        new_v = ws_new.cell(new_r, c).value
        # Compare numerically
        def to_num(v):
            if v is None or str(v).strip() in ("", " "): return 0.0
            try: return float(v)
            except: return v
        rv = to_num(ref_v)
        nv = to_num(new_v)
        if rv != nv:
            mismatches += 1
            if mismatches <= 30:
                ref_hdr = ws_ref.cell(32, c).value or f"Col{c}"
                print(f"  MISMATCH {pair} col {c} ({str(ref_hdr)[:40]}): ref={rv} new={nv}")

print(f"\nTotal value mismatches: {mismatches}")
