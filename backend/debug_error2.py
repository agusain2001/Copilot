"""Debug script to find the 'list index out of range' error."""
import sys, os, traceback
sys.path.insert(0, os.path.dirname(__file__))

import logging
logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s')

import openpyxl

icm_path = r"g:\FCCS\New folder (3)\Intercompany Report_1150_Intercompany Report.xlsx"
parent_journal = r"g:\FCCS\New folder (3)\Journal Report (5).xlsx"
contribution_journal = r"g:\FCCS\New folder (3)\Journal Report (6).xlsx"
plug_journal = r"g:\FCCS\Update files\Journal Report (4).xlsx"
report_inputs = r"g:\FCCS\AI\report_inputs.xlsx"

print("=== Testing ICM file load ===")
try:
    wb = openpyxl.load_workbook(icm_path, data_only=True)
    print(f"ICM loaded OK. Sheets: {wb.sheetnames}")
    for ws in wb.worksheets:
        print(f"  Sheet '{ws.title}': rows={ws.max_row}, cols={ws.max_column}")
        # Check first few rows
        for r in range(1, min(6, ws.max_row + 1)):
            vals = [ws.cell(r, c).value for c in range(1, min(6, ws.max_column + 1))]
            print(f"    Row {r}: {vals}")
except Exception as e:
    print(f"ICM load ERROR: {e}")
    traceback.print_exc()

print("\n=== Testing Journal 5 (Parent) ===")
try:
    wb = openpyxl.load_workbook(parent_journal, data_only=True)
    ws = wb.active
    print(f"Loaded OK. Sheet '{ws.title}': rows={ws.max_row}, cols={ws.max_column}")
    # Check row 30 and 31 (JOURNAL_DATA_START=31)
    for r in range(28, min(35, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
        print(f"  Row {r}: {vals}")
except Exception as e:
    print(f"Journal 5 load ERROR: {e}")
    traceback.print_exc()

print("\n=== Testing Journal 6 (Contribution) ===")
try:
    wb = openpyxl.load_workbook(contribution_journal, data_only=True)
    ws = wb.active
    print(f"Loaded OK. Sheet '{ws.title}': rows={ws.max_row}, cols={ws.max_column}")
    for r in range(28, min(35, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
        print(f"  Row {r}: {vals}")
except Exception as e:
    print(f"Journal 6 load ERROR: {e}")
    traceback.print_exc()

print("\n=== Testing Journal 4 (Plug) ===")
try:
    wb = openpyxl.load_workbook(plug_journal, data_only=True)
    ws = wb.active
    print(f"Loaded OK. Sheet '{ws.title}': rows={ws.max_row}, cols={ws.max_column}")
    for r in range(28, min(35, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
        print(f"  Row {r}: {vals}")
except Exception as e:
    print(f"Journal 4 load ERROR: {e}")
    traceback.print_exc()

print("\n=== Testing report_inputs ===")
try:
    wb = openpyxl.load_workbook(report_inputs, data_only=True)
    ws = wb.active
    print(f"Loaded OK. Sheet '{ws.title}': rows={ws.max_row}, cols={ws.max_column}")
    for r in range(1, min(6, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(10, ws.max_column + 1))]
        print(f"  Row {r}: {vals}")
except Exception as e:
    print(f"report_inputs load ERROR: {e}")
    traceback.print_exc()

print("\n=== Now testing is_detail_row with journal data ===")
from app.ic_processor import is_detail_row, J_ENTITY, J_ACCT, J_ICP, J_DEBIT, J_CREDIT, JOURNAL_DATA_START

for jname, jpath in [("Parent(5)", parent_journal), ("Contribution(6)", contribution_journal), ("Plug(4)", plug_journal)]:
    print(f"\n--- {jname}: {jpath} ---")
    wb = openpyxl.load_workbook(jpath, data_only=True)
    ws = wb.active
    print(f"  max_row={ws.max_row}, max_column={ws.max_column}")
    
    # Check if the journal has enough columns for J_CREDIT (col 16)
    print(f"  Required columns: J_ENTITY={J_ENTITY}, J_ACCT={J_ACCT}, J_ICP={J_ICP}, J_DEBIT={J_DEBIT}, J_CREDIT={J_CREDIT}")
    
    for r in range(JOURNAL_DATA_START, min(JOURNAL_DATA_START + 5, ws.max_row + 1)):
        row_cells = list(ws.iter_rows(min_row=r, max_row=r))[0]
        vals = [cell.value for cell in row_cells]
        print(f"  Row {r} ({len(vals)} cols): {vals[:20]}")
        
        # Test is_detail_row
        try:
            result = is_detail_row(vals)
            print(f"    is_detail_row = {result}")
        except IndexError as e:
            print(f"    is_detail_row IndexError: {e}")
            print(f"    vals length = {len(vals)}, need at least {max(J_ENTITY, J_ACCT, J_ICP)} elements")
