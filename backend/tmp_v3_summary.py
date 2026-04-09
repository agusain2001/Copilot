"""Quick summary of all Parent block rows in v3 output for screenshot comparison."""
import sys, re
sys.path.insert(0, '.')
import openpyxl
from app.ic_processor import to_float

wb = openpyxl.load_workbook(r"uploads\reports\31\outputs\ICM_Output_31_v3.xlsx", data_only=True)
ws = wb.active

f = open("tmp_v3_summary.txt", "w", encoding="utf-8")
f.write(f"{'Row':>4s} | {'Entity':40s} | {'Partner':40s} | {'534 Ent':>14s} | {'433 Ent':>14s} | {'111 Ent':>14s} | {'120 Ent':>14s} | {'121 Ent':>14s} | {'Var1':>14s} | {'433 Par':>14s} | {'534 Par':>14s} | {'111 Par':>14s} | {'120 Par':>14s} | {'121 Par':>14s} | {'Var2':>14s} | {'Total':>14s}\n")
f.write("-" * 300 + "\n")

for r in range(33, ws.max_row + 1):
    ent = str(ws.cell(r, 1).value or "").strip()
    prt = str(ws.cell(r, 2).value or "").strip()
    if not ent: continue
    
    # Check if this row has Parent block data (cols 17-29)
    has_parent = False
    for c in range(17, 30):
        v = ws.cell(r, c).value
        if v is not None and v != 0 and v != "":
            has_parent = True
            break
    
    if not has_parent:
        continue
    
    def gv(col):
        v = ws.cell(r, col).value
        if v is None or v == "" or v == 0:
            return ""
        return f"{to_float(v):>14.2f}"
    
    f.write(f"{r:4d} | {ent[:40]:40s} | {prt[:40]:40s} | {gv(17):>14s} | {gv(18):>14s} | {gv(19):>14s} | {gv(20):>14s} | {gv(21):>14s} | {gv(22):>14s} | {gv(23):>14s} | {gv(24):>14s} | {gv(25):>14s} | {gv(26):>14s} | {gv(27):>14s} | {gv(28):>14s} | {gv(29):>14s}\n")

f.close()
print("Done - see tmp_v3_summary.txt")
