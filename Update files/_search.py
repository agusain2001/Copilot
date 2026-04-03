import openpyxl
wb = openpyxl.load_workbook(r'g:\FCCS\Update files\ICM_Output_CORRECT.xlsx', data_only=True)
ws = wb['ICM Matched']
found = False
for r in range(33, ws.max_row + 1):
    for c in range(3, ws.max_column + 1):
        v = ws.cell(r, c).value
        try:
            fv = float(v)
            if abs(fv) == 1576708.0 or abs(fv) == 1576708:
                print('Found', fv, 'at row', r, 'col', c)
                found = True
        except:
            pass
if not found:
    print('Value 1576708 not found anywhere in the output')
