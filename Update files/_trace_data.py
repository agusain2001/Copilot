"""
Trace specific journal entries from the screenshots and compare 
with what ended up in the output file.
"""
import openpyxl
import re
from collections import defaultdict

# Read the Parent Journal to find the specific entries user is showing
wb_j = openpyxl.load_workbook(r'g:\FCCS\Update files\Journal Report.xlsx', data_only=True)
ws_j = wb_j.active

print("=" * 80)
print("JOURNAL ENTRIES FOR ENTITY 001033 + ACCOUNT 165003")
print("=" * 80)

for row_num in range(31, ws_j.max_row + 1):
    entity_raw = str(ws_j.cell(row_num, 3).value or "").strip()
    acct_raw = str(ws_j.cell(row_num, 4).value or "").strip()
    icp_raw = str(ws_j.cell(row_num, 5).value or "").strip()
    
    # Check if entity contains 001033 or E118000 (group), and account is 165003
    if ('001033' in entity_raw or 'E118000' in entity_raw) and '165003' in acct_raw:
        label = str(ws_j.cell(row_num, 1).value or "").strip()
        debit = ws_j.cell(row_num, 15).value
        credit = ws_j.cell(row_num, 16).value
        print(f"\nRow {row_num}: Label={label}")
        print(f"  Entity: {entity_raw}")
        print(f"  Account: {acct_raw}")
        print(f"  ICP: {icp_raw}")
        print(f"  Debit: {debit}")
        print(f"  Credit: {credit}")

# Also check for ALL entries related to 001033
print("\n" + "=" * 80)
print("ALL JOURNAL ENTRIES FOR ENTITY 001033 (any account)")
print("=" * 80)

for row_num in range(31, ws_j.max_row + 1):
    entity_raw = str(ws_j.cell(row_num, 3).value or "").strip()
    acct_raw = str(ws_j.cell(row_num, 4).value or "").strip()
    icp_raw = str(ws_j.cell(row_num, 5).value or "").strip()
    
    if '001033' in entity_raw and acct_raw:
        # Extract account code
        m = re.match(r"(\d{6})", acct_raw.strip())
        acct_code = m.group(1) if m else ""
        # Only care about IC elimination accounts
        if acct_code in {"165001","165002","165003","165004","165005","187052",
                         "189001","189014","189015","189501","224000","224001",
                         "224003","224009","224024","188800"}:
            label = str(ws_j.cell(row_num, 1).value or "").strip()
            debit = ws_j.cell(row_num, 15).value
            credit = ws_j.cell(row_num, 16).value
            print(f"\nRow {row_num}: Label={label}")
            print(f"  Entity: {entity_raw}")
            print(f"  Account: {acct_raw} (code: {acct_code})")
            print(f"  ICP: {icp_raw}")
            print(f"  Debit: {debit}, Credit: {credit}")

# Also check E118000 (group entity for SBG)
print("\n" + "=" * 80)
print("ALL JOURNAL ENTRIES FOR GROUP ENTITY E118000")
print("=" * 80)

for row_num in range(31, ws_j.max_row + 1):
    entity_raw = str(ws_j.cell(row_num, 3).value or "").strip()
    acct_raw = str(ws_j.cell(row_num, 4).value or "").strip()
    icp_raw = str(ws_j.cell(row_num, 5).value or "").strip()
    
    if 'E118000' in entity_raw and acct_raw:
        label = str(ws_j.cell(row_num, 1).value or "").strip()
        debit = ws_j.cell(row_num, 15).value
        credit = ws_j.cell(row_num, 16).value
        print(f"\nRow {row_num}: Label={label}")
        print(f"  Entity: {entity_raw}")
        print(f"  Account: {acct_raw}")
        print(f"  ICP: {icp_raw}")
        print(f"  Debit: {debit}, Credit: {credit}")

# Now check what appears in the output for entity 001033
print("\n" + "=" * 80)
print("OUTPUT FILE: ROWS FOR ENTITY 001033")
print("=" * 80)

wb_out = openpyxl.load_workbook(r'g:\FCCS\Update files\ICM_Output_CORRECT.xlsx', data_only=True)
ws_out = wb_out['ICM Matched']

for row_num in range(33, ws_out.max_row + 1):
    entity = str(ws_out.cell(row_num, 1).value or "").strip()
    partner = str(ws_out.cell(row_num, 2).value or "").strip()
    
    if '001033' in entity or '001033' in partner:
        print(f"\nRow {row_num}: Entity={entity}")
        print(f"  Partner={partner}")
        # Print all non-None, non-zero values
        for col in range(3, ws_out.max_column + 1):
            v = ws_out.cell(row_num, col).value
            if v is not None and v != 0:
                header = str(ws_out.cell(32, col).value or "")[:50]
                from openpyxl.utils import get_column_letter
                cl = get_column_letter(col)
                print(f"  Col {cl}({col}) [{header}]: {v}")

# Also check ICM source for 001033
print("\n" + "=" * 80)
print("ICM SOURCE: ROWS FOR ENTITY 001033")
print("=" * 80)

wb_icm = openpyxl.load_workbook(r'g:\FCCS\Update files\Intercompany Balances IC Matching Report (1).xlsx', data_only=True)
ws_icm = wb_icm['Sheet1']

for row_num in range(33, ws_icm.max_row + 1):
    entity = str(ws_icm.cell(row_num, 1).value or "").strip()
    partner = str(ws_icm.cell(row_num, 2).value or "").strip()
    
    if '001033' in entity:
        has_data = False
        for col in range(3, ws_icm.max_column + 1):
            v = ws_icm.cell(row_num, col).value
            if v is not None and v != 0 and str(v).strip() not in ("", " "):
                has_data = True
                break
        print(f"\nRow {row_num}: Entity={entity[:50]}")
        print(f"  Partner={partner[:50]}")
        if has_data:
            for col in range(3, ws_icm.max_column + 1):
                v = ws_icm.cell(row_num, col).value
                if v is not None and v != 0 and str(v).strip() not in ("", " "):
                    header = str(ws_icm.cell(32, col).value or "")[:40]
                    print(f"  Col {col} [{header}]: {v}")
        else:
            print(f"  [All base values are zero/blank]")
