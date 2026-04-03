"""
Deep investigation: The user's screenshots show journal entries like QDCONSOLADJ191 for entity 001033.
Let's check exactly how the journal entries were matched and whether they should have been included.

Key finding from the screenshots:
- QDCONSOLADJ191: entity=001033, acct=165003, ICP=ICP_001001, Debit=26,340,466.19  (user's screenshot shows different number!)
- QDCONSOLADJ191: entity=001033, acct=165003, ICP=ICP_001032, Debit=88,717,872.52  (user's screenshot)
- QDCONSOLADJ255: entity=E118000, acct=165003, ICP=ICP_001001, Credit=15,76,708     (user's screenshot)

But ACTUAL data from journal:
- Row 267: entity=001033, acct=165003, ICP=ICP_001001, Debit=20,340,466.09 (NOT 26,340,466!)
- Row 268: entity=001033, acct=165003, ICP=ICP_001032, Credit=88,717,872.52 (CREDIT not Debit!)
- Row 471: entity=E118000, acct=165003, ICP=ICP_001001, Credit=1,576,708

The user's screenshots may show the CONTRIBUTION journal, not the PARENT journal!
Let me check Journal Report (2).xlsx
"""
import openpyxl

# Check contribution journal for these entries
print("=" * 80)
print("CHECKING CONTRIBUTION JOURNAL for entities 001033 / E118000")
print("=" * 80)

wb = openpyxl.load_workbook(r'g:\FCCS\Update files\Journal Report (2).xlsx', data_only=True)
ws = wb.active
for row in range(31, ws.max_row + 1):
    entity = str(ws.cell(row, 3).value or "").strip()
    acct = str(ws.cell(row, 4).value or "").strip()
    icp = str(ws.cell(row, 5).value or "").strip()
    if '001033' in entity or 'E118000' in entity or '001032' in entity:
        label = str(ws.cell(row, 1).value or "").strip()
        debit = ws.cell(row, 15).value
        credit = ws.cell(row, 16).value
        print(f"Row {row}: {label}")
        print(f"  Entity: {entity}")
        print(f"  Account: {acct}")
        print(f"  ICP: {icp}")
        print(f"  Debit: {debit}, Credit: {credit}")
        print()

# Check plug journal too
print("=" * 80)
print("CHECKING PLUG JOURNAL for entities 001033 / E118000")
print("=" * 80)

wb = openpyxl.load_workbook(r'g:\FCCS\Update files\Journal Report (4).xlsx', data_only=True)
ws = wb.active
for row in range(31, ws.max_row + 1):
    entity = str(ws.cell(row, 3).value or "").strip()
    if '001033' in entity or 'E118000' in entity or '001032' in entity:
        label = str(ws.cell(row, 1).value or "").strip()
        acct = str(ws.cell(row, 4).value or "").strip()
        icp = str(ws.cell(row, 5).value or "").strip()
        debit = ws.cell(row, 15).value
        credit = ws.cell(row, 16).value
        print(f"Row {row}: {label}")
        print(f"  Entity: {entity}")
        print(f"  Account: {acct}")
        print(f"  ICP: {icp}")
        print(f"  Debit: {debit}, Credit: {credit}")
        print()

# Now check: are "FCCS_No Intercompany" entries being WRONGLY included?
print("=" * 80)
print("CHECKING: FCCS_No Intercompany entries in Parent Journal")
print("=" * 80)

wb = openpyxl.load_workbook(r'g:\FCCS\Update files\Journal Report.xlsx', data_only=True)
ws = wb.active
no_icp_count = 0
with_icp_count = 0
no_icp_included = 0

elim_accounts = {"165001","165002","165003","165004","165005","187052",
                 "189001","189014","189015","189501","224000","224001",
                 "224003","224009","224024","188800"}
import re
for row in range(31, ws.max_row + 1):
    entity_raw = str(ws.cell(row, 3).value or "").strip()
    acct_raw = str(ws.cell(row, 4).value or "").strip()
    icp_raw = str(ws.cell(row, 5).value or "").strip()
    
    if not entity_raw or not acct_raw:
        continue
    
    m = re.match(r"(\d{6})", acct_raw)
    acct_code = m.group(1) if m else ""
    
    if acct_code not in elim_accounts:
        continue
    
    if 'No Intercompany' in icp_raw or 'FCCS_No Intercompany' in icp_raw:
        no_icp_count += 1
    elif icp_raw.startswith('ICP_'):
        with_icp_count += 1

print(f"Journal lines with IC elimination accounts:")
print(f"  With ICP partner: {with_icp_count}")
print(f"  With 'No Intercompany': {no_icp_count}")
print(f"\nNOTE: 'No Intercompany' entries should NOT be included in ICM matching!")
print(f"These are non-IC adjustments that don't belong in the IC matching output.")

# Check what my script does with "No Intercompany" entries
print("\n" + "=" * 80)
print("CHECKING: Does my script filter out 'No Intercompany'?")
print("=" * 80)

# Re-read generate script logic
# extract_icp_code looks for ICP_ prefix
# If ICP code is "FCCS_No" or empty, the match key will be ("001033", "", "165003")
# This would match ICM row where entity=001033 and partner="" (blank)
# Let's check:
icp_test = "FCCS_No Intercompany:No Intercompany"
m = re.match(r"(ICP_\w+)", icp_test)
icp_code = m.group(1) if m else ""
print(f"extract_icp_code('{icp_test}') = '{icp_code}'")
print(f"This means 'No Intercompany' entries get icp_code='' (empty)")
print(f"They would only match ICM rows where partner is also empty...")
print()

# Check ICM source for entity=001033 with EMPTY partner
print("ICM rows where entity=001033 and partner is EMPTY:")
wb_icm = openpyxl.load_workbook(r'g:\FCCS\Update files\Intercompany Balances IC Matching Report (1).xlsx', data_only=True)
ws_icm = wb_icm['Sheet1']
for row in range(33, ws_icm.max_row + 1):
    entity = str(ws_icm.cell(row, 1).value or "").strip()
    partner = str(ws_icm.cell(row, 2).value or "").strip()
    if '001033' in entity and not partner:
        print(f"  Row {row}: Entity={entity[:50]}, Partner='{partner}'")
        for col in range(3, ws_icm.max_column + 1):
            v = ws_icm.cell(row, col).value
            if v is not None and v != 0 and str(v).strip() not in ("", " "):
                header = str(ws_icm.cell(32, col).value or "")[:40]
                print(f"    Col {col} [{header}]: {v}")
