"""
Proper ICM Output Generator
============================
Reads the actual ICM report (headers at row 32, data at row 33+),
matches journal entries, and produces a correct output file.
"""

import re
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ── Configuration ──────────────────────────────────────────────────────────────
ICM_PATH = r'g:\FCCS\Update files\Intercompany Balances IC Matching Report (1).xlsx'
PARENT_JOURNAL = r'g:\FCCS\Update files\Journal Report.xlsx'
CONTRIB_JOURNAL = r'g:\FCCS\Update files\Journal Report (2).xlsx'
PLUG_JOURNAL = r'g:\FCCS\Update files\Journal Report (4).xlsx'
REPORT_INPUTS = r'g:\FCCS\AI\report_inputs.xlsx'
OUTPUT_PATH = r'g:\FCCS\Update files\ICM_Output_CORRECT.xlsx'

# The 15 elimination accounts from report_inputs.xlsx
PLUG_CODE = "188800"
ELIM_ACCOUNTS = {
    "165001", "165002", "165003", "165004", "165005",
    "187052", "189001", "189014", "189015", "189501",
    "224000", "224001", "224003", "224009", "224024",
}

# S1 = Asset accounts (sign: Debit - Credit), S2 = Liability accounts (sign: Credit - Debit)
S1_CODES = {"165001", "165002", "165003", "165004", "165005", "187052", "189001", "189014", "189015", "189501"}
S2_CODES = {"224000", "224001", "224003", "224009", "224024"}

# Entity-side columns: S1 accounts tagged Entity + S2 accounts tagged Partner
ENT_COLS = [
    ("165001", "165001:Loans provided to related parties",                             "S1", "Entity"),
    ("165002", "165002:Expenses recharged on behalf of a related party",               "S1", "Entity"),
    ("165003", "165003:Advances provided for related parties",                         "S1", "Entity"),
    ("165004", "165004:Payments to subcontractors & suppliers on behalf of other parties", "S1", "Entity"),
    ("165005", "165005:Retention receivable-Intercompany (Current)",                   "S1", "Entity"),
    ("187052", "187052:Retention receivable-Intercomp",                                "S1", "Entity"),
    ("189001", "189001:Inter company account - receivables",                           "S1", "Entity"),
    ("189014", "189014:Discount on intercompany loan",                                 "S1", "Entity"),
    ("189015", "189015:Intercompany receivable A/C for outside QD group",              "S1", "Entity"),
    ("189501", "189501:Due from Related Party - Non-current",                          "S1", "Entity"),
    ("224000", "224000:Due to related parties",                                        "S2", "Partner"),
    ("224001", "224001:Loans provided from related parties",                           "S2", "Partner"),
    ("224003", "224003:Advances provided from related parties",                        "S2", "Partner"),
    ("224009", "224009:Inter company accounts - Payables",                             "S2", "Partner"),
    ("224024", "224024:Intercompany payable A/C for outside QD group",                 "S2", "Partner"),
]

# Partner-side columns: S1 accounts tagged Partner + S2 accounts tagged Entity (reverse perspective)
PAR_COLS = [
    ("165001", "165001:Loans provided to related parties",                             "S1", "Partner"),
    ("165002", "165002:Expenses recharged on behalf of a related party",               "S1", "Partner"),
    ("165003", "165003:Advances provided for related parties",                         "S1", "Partner"),
    ("165004", "165004:Payments to subcontractors & suppliers on behalf of other parties", "S1", "Partner"),
    ("165005", "165005:Retention receivable-Intercompany (Current)",                   "S1", "Partner"),
    ("187052", "187052:Retention receivable-Intercomp",                                "S1", "Partner"),
    ("189001", "189001:Inter company account - receivables",                           "S1", "Partner"),
    ("189014", "189014:Discount on intercompany loan",                                 "S1", "Partner"),
    ("189015", "189015:Intercompany receivable A/C for outside QD group",              "S1", "Partner"),
    ("189501", "189501:Due from Related Party - Non-current",                          "S1", "Partner"),
    ("224000", "224000:Due to related parties",                                        "S2", "Entity"),
    ("224001", "224001:Loans provided from related parties",                           "S2", "Entity"),
    ("224003", "224003:Advances provided from related parties",                        "S2", "Entity"),
    ("224009", "224009:Inter company accounts - Payables",                             "S2", "Entity"),
    ("224024", "224024:Intercompany payable A/C for outside QD group",                 "S2", "Entity"),
]

NUM_ACCTS = len(ENT_COLS)  # 15

# ── Column layout ──────────────────────────────────────────────────────────────
def _block_positions(start):
    """Return (ent_start, var1, par_start, var2, total, next_start) for a block."""
    ent_start = start
    var1      = start + NUM_ACCTS
    par_start = var1 + 1
    var2      = par_start + NUM_ACCTS
    total     = var2 + 1
    spacer    = total + 1
    return ent_start, var1, par_start, var2, total, spacer

BLK_BASE  = _block_positions(3)
BLK_PAR   = _block_positions(BLK_BASE[5] + 1)
BLK_CONT  = _block_positions(BLK_PAR[5] + 1)
COL_PLUG  = BLK_CONT[5] + 1
COL_PLUG_SPACER = COL_PLUG + 1
COL_FINAL = COL_PLUG_SPACER + 1
TOTAL_COLS = COL_FINAL

# ── Output metadata constants ─────────────────────────────────────────────────
OUTPUT_HEADER_ROW = 32
OUTPUT_DATA_START = 33
SECTION_LABEL_ROW = 29

# We'll also write metadata rows 33-56 but shifted for FCCS-style layout
META_START = 33  # row where metadata dimensions go (Cube, Scenario, etc.)

# ── Styling ────────────────────────────────────────────────────────────────────
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=9)
HEADER_FILL    = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
ID_FILL        = PatternFill(start_color="C8DCF0", end_color="C8DCF0", fill_type="solid")
VARIANCE_FILL  = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
TOTAL_FILL     = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
MATCH_FILL     = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
GROUP_FILL     = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
SECTION_FILL   = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
NO_FILL        = PatternFill(fill_type=None)
ID_FONT        = Font(bold=True, size=11)
DATA_FONT      = Font(size=11)
SECTION_FONT   = Font(bold=True, color="FFFFFF", size=11)
NUM_FORMAT     = r"###,##0;\-###,##0"

DATA_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ═══════════════════════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def extract_entity_code(raw):
    """Extract 6-digit entity code from ICM text like '001001 - QDRE...'"""
    raw = str(raw or "").strip()
    m = re.match(r"(\d{6})", raw)
    return m.group(1) if m else ""

def extract_icp_code(raw):
    """Extract ICP code like 'ICP_001051' from text"""
    raw = str(raw or "").strip()
    m = re.match(r"(ICP_\w+)", raw)
    return m.group(1) if m else ""

def extract_account_code(raw):
    """Extract 6-digit account code from journal account field"""
    raw = str(raw or "").strip()
    # Try patterns like "].[(code)]:" or "]:(code):"
    m = re.search(r"\]\.\[?(\w+)\]?:", raw)
    if m:
        val = m.group(1)
        if not val.isdigit():
            m2 = re.search(r"]:(\d{6}):", raw)
            if m2: return m2.group(1)
        return val
    # Try direct "123456:..."
    m = re.match(r"(\d{6})", raw)
    return m.group(1) if m else ""

def extract_entity_code_journal(raw):
    """Extract entity code from journal entity field like '013014:30 GS LLP' or 'E108000:...'"""
    raw = str(raw or "").strip()
    m = re.match(r"(E?\d+\w*)", raw)
    return m.group(1) if m else ""

def to_float(val):
    if val is None or str(val).strip() in ("", " "):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def apply_sign(debit, credit, account_code):
    """Apply correct sign based on account classification."""
    code = str(account_code or "").strip()
    s = code[0] if code and code[0].isdigit() else "0"
    if s in ("1", "5"):
        return debit - credit  # S1: Debit - Credit
    elif s in ("2", "3", "4"):
        return credit - debit  # S2: Credit - Debit
    return debit - credit

def is_detail_row(vals):
    """Check if a journal row has meaningful entity/account/ICP data."""
    entity = str(vals[2] or "").strip()  # Col C (index 2)
    acct   = str(vals[3] or "").strip()  # Col D (index 3)
    icp    = str(vals[4] or "").strip()  # Col E (index 4)
    return bool(entity or acct or icp)

# ═══════════════════════════════════════════════════════════════════════════════
# READ ICM DATA (from the actual FCCS ICM report)
# ═══════════════════════════════════════════════════════════════════════════════

def find_header_row(ws):
    """Find the row where 'Entity' and 'Partner' headers are."""
    for row_num in range(1, min(50, ws.max_row + 1)):
        a_val = str(ws.cell(row_num, 1).value or "").strip().lower()
        b_val = str(ws.cell(row_num, 2).value or "").strip().lower()
        if a_val == "entity" and b_val == "partner":
            return row_num
    return None

def read_icm_file(filepath):
    """Read the ICM report file, returning headers map and data rows."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Find the sheet with headers - try all sheets
    best_ws = None
    best_header_row = None
    for ws in wb.worksheets:
        hr = find_header_row(ws)
        if hr is not None:
            if best_ws is None or ws.max_column < best_ws.max_column:
                best_ws = ws
                best_header_row = hr
    
    if best_ws is None:
        raise ValueError("Cannot find header row with 'Entity'/'Partner' in ICM file")
    
    ws = best_ws
    header_row = best_header_row
    data_start = header_row + 1
    
    print(f"  ICM sheet: '{ws.title}', headers at row {header_row}, data from row {data_start}")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    
    # Build column map: (account_code, tag) -> column_index
    col_map = {}
    for cell in ws[header_row]:
        h = str(cell.value or "").strip()
        if not h or h.lower() in ("entity", "partner"):
            continue
        # Extract account code
        m = re.match(r"(\d{6})\s*-\s*", h)
        if m:
            code = m.group(1)
            tag = "Partner" if re.search(r"\bPartner\s*$", h) else "Entity"
            col_map[(code, tag)] = cell.column
    
    print(f"  Mapped {len(col_map)} account columns")
    
    # Read data rows
    data_rows = []
    for row_num in range(data_start, ws.max_row + 1):
        entity_raw = str(ws.cell(row_num, 1).value or "").strip()
        partner_raw = str(ws.cell(row_num, 2).value or "").strip()
        if not entity_raw and not partner_raw:
            continue
        
        data_rows.append({
            "row_num": row_num,
            "entity": entity_raw,
            "partner": partner_raw,
            "entity_code": extract_entity_code(entity_raw),
            "partner_code": extract_icp_code(partner_raw),
        })
    
    print(f"  Read {len(data_rows)} data rows")
    return ws, col_map, data_rows

# ═══════════════════════════════════════════════════════════════════════════════
# READ JOURNAL REPORT
# ═══════════════════════════════════════════════════════════════════════════════

def read_journal(filepath, is_plug_journal=False):
    """Read a journal report file and build lookup dictionaries."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Get consolidation type from row 9
    consol = str(ws.cell(9, 3).value or "").strip()
    print(f"  Journal: {os.path.basename(filepath)}")
    print(f"    Consolidation: {consol}")
    print(f"    Rows: {ws.max_row}")
    
    # Headers at row 30, data from row 31
    # Col C(3)=Entity, D(4)=Account, E(5)=Intercompany, O(15)=Debit, P(16)=Credit
    J_ENTITY = 3
    J_ACCT = 4
    J_ICP = 5
    J_DEBIT = 15
    J_CREDIT = 16
    
    lines = []
    for row_num in range(31, ws.max_row + 1):
        vals = [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]
        
        label = str(vals[0] or "").strip()
        if label == "Grand Total":
            break
        
        # Check if it's a detail row (has entity/account/icp)
        entity_raw = str(vals[J_ENTITY - 1] or "").strip()
        acct_raw   = str(vals[J_ACCT - 1] or "").strip()
        icp_raw    = str(vals[J_ICP - 1] or "").strip()
        
        if not (entity_raw or acct_raw or icp_raw):
            continue
        
        entity_code = extract_entity_code_journal(entity_raw)
        account_code = extract_account_code(acct_raw)
        icp_code = extract_icp_code(icp_raw)
        debit = to_float(vals[J_DEBIT - 1])
        credit = to_float(vals[J_CREDIT - 1])
        
        # For plug journal: map elimination accounts to plug code
        if is_plug_journal:
            if account_code in ELIM_ACCOUNTS or not account_code[:1].isdigit():
                account_code = PLUG_CODE
        
        is_group = bool(re.match(r"^E\d+", entity_code))
        
        lines.append({
            "entity_code": entity_code,
            "is_group": is_group,
            "account_code": account_code,
            "icp_code": icp_code,
            "debit": debit,
            "credit": credit,
        })
    
    print(f"    Parsed {len(lines)} journal lines")
    
    # Build lookup dictionaries
    primary_lookup = defaultdict(list)
    fallback_lookup = defaultdict(list)
    
    for line in lines:
        if not line["is_group"]:
            primary_lookup[(line["entity_code"], line["icp_code"], line["account_code"])].append(line)
        else:
            fallback_lookup[(line["entity_code"], line["icp_code"], line["account_code"])].append(line)
    
    # Compute net values
    primary_updates = {}
    fallback_updates = {}
    
    for (e, p, a), jlines in primary_lookup.items():
        net = sum(apply_sign(j["debit"], j["credit"], a) for j in jlines)
        if net != 0:
            primary_updates[(e, p, a)] = net
    
    for (ge, p, a), jlines in fallback_lookup.items():
        net = sum(apply_sign(j["debit"], j["credit"], a) for j in jlines)
        if net != 0:
            fallback_updates[(ge, p, a)] = net
    
    print(f"    Primary updates: {len(primary_updates)}, Fallback updates: {len(fallback_updates)}")
    return primary_updates, fallback_updates

# ═══════════════════════════════════════════════════════════════════════════════
# WRITE OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

def _style_header_cell(cell, text):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _style_id_cell(cell, text):
    cell.value = text
    cell.font = ID_FONT
    cell.fill = ID_FILL
    cell.border = DATA_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

def _style_data_cell(cell, value, fill=None):
    cell.value = value
    cell.font = DATA_FONT
    cell.fill = fill if fill is not None else NO_FILL
    cell.border = DATA_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if isinstance(value, (int, float)):
        cell.number_format = NUM_FORMAT

def write_output(ws_icm, col_map, data_rows, parent_updates, contrib_updates, plug_updates):
    """Write the complete output workbook."""
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "ICM Matched"
    
    # Column widths
    spacer_cols = {BLK_BASE[5], BLK_PAR[5], BLK_CONT[5], COL_PLUG_SPACER}
    for i in range(TOTAL_COLS):
        col_num = i + 1
        ws.column_dimensions[get_column_letter(col_num)].width = 4 if col_num in spacer_cols else 17.5714
    
    ws.row_dimensions[OUTPUT_HEADER_ROW].height = 78.75
    
    # Section labels
    def _section_label(start, end, text):
        ws.merge_cells(start_row=SECTION_LABEL_ROW, start_column=start,
                       end_row=SECTION_LABEL_ROW, end_column=end)
        c = ws.cell(SECTION_LABEL_ROW, start)
        c.value = text
        c.fill = SECTION_FILL
        c.font = SECTION_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    
    _section_label(BLK_PAR[0], BLK_PAR[4], "Parent Input")
    _section_label(BLK_CONT[0], BLK_CONT[4], "Contribution Input")
    _section_label(COL_PLUG, COL_FINAL, "Plug Account")
    
    # Column headers
    _style_header_cell(ws.cell(OUTPUT_HEADER_ROW, 1), "Entity")
    _style_header_cell(ws.cell(OUTPUT_HEADER_ROW, 2), "Partner")
    
    def _write_block_headers(ent_start, par_start, var1_col, var2_col, total_col):
        for i, (code, desc, _, tag) in enumerate(ENT_COLS):
            _style_header_cell(ws.cell(OUTPUT_HEADER_ROW, ent_start + i), f"{code} - {desc} {tag}")
        _style_header_cell(ws.cell(OUTPUT_HEADER_ROW, var1_col), "Variance")
        for i, (code, desc, _, tag) in enumerate(PAR_COLS):
            _style_header_cell(ws.cell(OUTPUT_HEADER_ROW, par_start + i), f"{code} - {desc} {tag}")
        _style_header_cell(ws.cell(OUTPUT_HEADER_ROW, var2_col), "Variance")
        _style_header_cell(ws.cell(OUTPUT_HEADER_ROW, total_col), "Total")
    
    _write_block_headers(BLK_BASE[0], BLK_BASE[2], BLK_BASE[1], BLK_BASE[3], BLK_BASE[4])
    _write_block_headers(BLK_PAR[0], BLK_PAR[2], BLK_PAR[1], BLK_PAR[3], BLK_PAR[4])
    _write_block_headers(BLK_CONT[0], BLK_CONT[2], BLK_CONT[1], BLK_CONT[3], BLK_CONT[4])
    
    _style_header_cell(ws.cell(OUTPUT_HEADER_ROW, COL_PLUG), f"{PLUG_CODE}:Intercompany Balances Plug A/c")
    _style_header_cell(ws.cell(OUTPUT_HEADER_ROW, COL_FINAL), "Total")
    
    # Helper: read value from ICM source
    def _icm_num(src_row, code, tag):
        col_idx = col_map.get((code, tag))
        if not col_idx:
            return None
        v = ws_icm.cell(src_row, col_idx).value
        try:
            if v in (None, "", " "):
                return None
            fv = float(v)
            # Round to integer if close (ICM stores integers)
            rv = round(fv)
            return rv if abs(fv - rv) < 0.01 else fv
        except:
            return None
    
    # S1/S2 code lists for variance computation
    s1_ent_codes = [c[0] for c in ENT_COLS if c[2] == "S1"]  # 165xxx, 187xxx, 189xxx
    s2_par_codes = [c[0] for c in ENT_COLS if c[2] == "S2"]  # 224xxx
    s1_par_codes = [c[0] for c in PAR_COLS if c[2] == "S1"]  # 165xxx, 187xxx, 189xxx
    s2_ent_codes = [c[0] for c in PAR_COLS if c[2] == "S2"]  # 224xxx
    
    # Unpack journal updates
    parent_primary, parent_fallback = parent_updates
    contrib_primary, contrib_fallback = contrib_updates
    plug_primary, plug_fallback = plug_updates
    
    # Track consumed keys
    consumed_parent_fb  = set()
    consumed_contrib_fb = set()
    consumed_plug_fb    = set()
    consumed_parent_pri  = set()
    consumed_contrib_pri = set()
    consumed_plug_pri    = set()
    
    def _extract_val(raw):
        if raw is None:
            return None, NO_FILL
        if isinstance(raw, tuple):
            return raw[1], GROUP_FILL
        return raw, MATCH_FILL
    
    base_data_count = 0
    parent_data_count = 0
    contrib_data_count = 0
    plug_data_count = 0
    
    for out_idx, icm_row in enumerate(data_rows):
        src_r = icm_row["row_num"]
        out_r = OUTPUT_DATA_START + out_idx
        ent   = icm_row["entity_code"]
        prt   = icm_row["partner_code"]
        
        _style_id_cell(ws.cell(out_r, 1), icm_row["entity"])
        _style_id_cell(ws.cell(out_r, 2), icm_row["partner"])
        
        # Derived keys for vice-versa matching
        prt_as_entity = prt[4:] if prt.startswith("ICP_") else prt
        ent_as_icp    = f"ICP_{ent}"
        
        def _write_block(b_ent_s, b_par_s, b_var1, b_var2, b_total,
                         is_base, primary=None, fallback=None,
                         consumed_fb=None, consumed_pri=None):
            """Write one block of columns."""
            ent_vals = {}
            par_vals = {}
            local_has_data = False
            
            # ── Entity-side columns ──
            for i, (code, _, series, tag) in enumerate(ENT_COLS):
                if is_base:
                    v = _icm_num(src_r, code, tag)
                    _style_data_cell(ws.cell(out_r, b_ent_s + i), v)
                    ent_vals[code] = v
                    if v is not None: local_has_data = True
                else:
                    # Direct match: journal(entity, partner, account)
                    pri_key = (ent, prt, code)
                    raw = None
                    if primary and pri_key in primary and (consumed_pri is None or pri_key not in consumed_pri):
                        raw = primary[pri_key]
                        if consumed_pri is not None:
                            consumed_pri.add(pri_key)
                    # Fallback: group entity match
                    if raw is None and fallback and consumed_fb is not None:
                        for fb_key, fb_val in fallback.items():
                            if fb_key[1] == prt and fb_key[2] == code and fb_key not in consumed_fb:
                                raw = ('GROUP', fb_val)
                                consumed_fb.add(fb_key)
                                break
                    v, fill = _extract_val(raw)
                    _style_data_cell(ws.cell(out_r, b_ent_s + i), v,
                                     fill if v is not None else None)
                    ent_vals[code] = v
                    if v is not None: local_has_data = True
            
            # ── Variance 1: sum(S1 entity) - sum(S2 partner) ──
            sum_s1_ent = sum(to_float(ent_vals.get(c)) for c in s1_ent_codes)
            sum_s2_par = sum(to_float(ent_vals.get(c)) for c in s2_par_codes)
            var1 = round(sum_s1_ent - sum_s2_par)
            _style_data_cell(ws.cell(out_r, b_var1), var1, VARIANCE_FILL)
            ws.cell(out_r, b_var1).number_format = NUM_FORMAT
            
            # ── Partner-side columns (Vice-Versa) ──
            for i, (code, _, series, tag) in enumerate(PAR_COLS):
                if is_base:
                    v = _icm_num(src_r, code, tag)
                    _style_data_cell(ws.cell(out_r, b_par_s + i), v)
                    par_vals[code] = v
                    if v is not None: local_has_data = True
                else:
                    # Vice-versa: journal(partner_as_entity, entity_as_icp, account)
                    pri_key_par = (prt_as_entity, ent_as_icp, code)
                    raw_par = None
                    if primary and pri_key_par in primary and (consumed_pri is None or pri_key_par not in consumed_pri):
                        raw_par = primary[pri_key_par]
                        if consumed_pri is not None:
                            consumed_pri.add(pri_key_par)
                    v, fill = _extract_val(raw_par)
                    _style_data_cell(ws.cell(out_r, b_par_s + i), v,
                                     fill if v is not None else None)
                    par_vals[code] = v
                    if v is not None: local_has_data = True
            
            # ── Variance 2: sum(S1 partner) - sum(S2 entity) ──
            sum_s1_par = sum(to_float(par_vals.get(c)) for c in s1_par_codes)
            sum_s2_ent = sum(to_float(par_vals.get(c)) for c in s2_ent_codes)
            var2 = round(sum_s1_par - sum_s2_ent)
            _style_data_cell(ws.cell(out_r, b_var2), var2, VARIANCE_FILL)
            ws.cell(out_r, b_var2).number_format = NUM_FORMAT
            
            # ── Total ──
            total = var1 + var2
            _style_data_cell(ws.cell(out_r, b_total), total, TOTAL_FILL)
            ws.cell(out_r, b_total).number_format = NUM_FORMAT
            
            return total, local_has_data
        
        # Write the three main blocks
        base_total, base_has = _write_block(
            BLK_BASE[0], BLK_BASE[2], BLK_BASE[1], BLK_BASE[3], BLK_BASE[4],
            is_base=True)
        
        parent_total, parent_has = _write_block(
            BLK_PAR[0], BLK_PAR[2], BLK_PAR[1], BLK_PAR[3], BLK_PAR[4],
            is_base=False, primary=parent_primary, fallback=parent_fallback,
            consumed_fb=consumed_parent_fb, consumed_pri=consumed_parent_pri)
        
        contrib_total, contrib_has = _write_block(
            BLK_CONT[0], BLK_CONT[2], BLK_CONT[1], BLK_CONT[3], BLK_CONT[4],
            is_base=False, primary=contrib_primary, fallback=contrib_fallback,
            consumed_fb=consumed_contrib_fb, consumed_pri=consumed_contrib_pri)
        
        if base_has: base_data_count += 1
        if parent_has: parent_data_count += 1
        if contrib_has: contrib_data_count += 1
        
        # ── Plug Account ──
        plug_val = 0.0
        plug_pri_key = (ent, prt, PLUG_CODE)
        raw = None
        if plug_pri_key in plug_primary and plug_pri_key not in consumed_plug_pri:
            raw = plug_primary[plug_pri_key]
            consumed_plug_pri.add(plug_pri_key)
        if raw is None and plug_fallback:
            for fb_key, fb_val in plug_fallback.items():
                if fb_key[1] == prt and fb_key[2] == PLUG_CODE and fb_key not in consumed_plug_fb:
                    raw = ('GROUP', fb_val)
                    consumed_plug_fb.add(fb_key)
                    break
        v, fill = _extract_val(raw)
        if v is not None:
            plug_val = to_float(v)
            _style_data_cell(ws.cell(out_r, COL_PLUG), v, fill)
            plug_data_count += 1
        else:
            _style_data_cell(ws.cell(out_r, COL_PLUG), None)
        
        # ── Final Total ──
        final_total = base_total + parent_total + contrib_total + plug_val
        _style_data_cell(ws.cell(out_r, COL_FINAL), final_total, TOTAL_FILL)
        ws.cell(out_r, COL_FINAL).number_format = NUM_FORMAT
    
    # Write metadata rows (above the data)
    meta_fields = [
        ("Cube", "Consol"),
        ("Scenario", "Actual - Actual"),
        ("Year", "FY24 - 2024"),
        ("Period", "Dec - Dec"),
        ("View", "FCCS_YTD - YTD"),
        ("Consolidation", "FCCS_Entity Total - Entity Total"),
        ("Currency", "QAR_Reporting - QAR"),
        ("Entity", "ILvl0Descendants(E100000)"),
        ("Partner", "ILvl0Descendants(E100000)"),
        ("Movement", "FCCS_Movements - Total Movements"),
        ("Data Source", "FCCS_Total Data Source - Total Data Source"),
        ("Project", "Total Project - Total Project"),
        ("Sub Code", "Total Sub Code - Total Sub Code"),
        ("Future1", "Total Future1 - Total Future1"),
        ("Future2", "Total Future2 - Total Future2"),
        ("Suppress Matches", "false"),
        ("Suppress Reversals", "false"),
        ("Suppress Blank Columns", "false"),
        ("Tolerance Value", "0"),
        ("Tolerance Percent", "0"),
        ("Scale Factor", "0"),
        ("Decimal Override", "0"),
        ("Intercompany Partner Display", "With ICP prefix"),
        ("Accounts in Rows", "false"),
    ]
    
    # Write metadata to columns A, starting before the header
    # These go in rows just before data, or in a standard position
    # For now, let's put them in rows starting from row 5 (similar to original ICM)
    for i, (field, value) in enumerate(meta_fields):
        r = 5 + i
        ws.cell(r, 1).value = field
    
    out_wb.save(OUTPUT_PATH)
    
    print(f"\n{'='*60}")
    print(f"OUTPUT SAVED: {OUTPUT_PATH}")
    print(f"{'='*60}")
    print(f"Total data rows written: {len(data_rows)}")
    print(f"Rows with Base ICM data: {base_data_count}")
    print(f"Rows with Parent Input data: {parent_data_count}")
    print(f"Rows with Contribution Input data: {contrib_data_count}")
    print(f"Rows with Plug Account data: {plug_data_count}")
    print(f"Total output columns: {TOTAL_COLS}")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  ICM Output Generator — Proper Version")
    print("=" * 60)
    
    # 1. Read ICM source data
    print("\n[1] Reading ICM source file...")
    ws_icm, col_map, data_rows = read_icm_file(ICM_PATH)
    
    # 2. Read journal reports
    print("\n[2] Reading Parent Journal...")
    parent_updates = read_journal(PARENT_JOURNAL)
    
    print("\n[3] Reading Contribution Journal...")
    contrib_updates = read_journal(CONTRIB_JOURNAL)
    
    print("\n[4] Reading Plug Account Journal...")
    plug_updates = read_journal(PLUG_JOURNAL, is_plug_journal=True)
    
    # 3. Generate output
    print("\n[5] Generating output...")
    write_output(ws_icm, col_map, data_rows, parent_updates, contrib_updates, plug_updates)

if __name__ == "__main__":
    main()
