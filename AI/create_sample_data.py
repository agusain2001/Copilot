"""
Generates:
  1. Sample_ICM_Report.xlsx   — Entity + Partner + Series1 + Series2 accounts (like real ICM)
  2. Sample_Journal_1.xlsx    — Journal with simple A/B/C companies
  3. Sample_ICM_Output.xlsx   — Result of IC matching (Entity+Partner+Account lookup, sign, variance)
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# ── Output paths ──────────────────────────────────────────────────────────
SAMPLE_ICM    = r"g:\AI\Sample_ICM_v2.xlsx"
SAMPLE_J1     = r"g:\AI\Sample_Journal_v2.xlsx"
SAMPLE_OUTPUT = r"g:\AI\Sample_ICM_Output.xlsx"

# ── ONLY Series 1 (Asset) and Series 2 (Liability) — matching the real ICM ──
ACCOUNTS = [
    # (code,   description,                 series)
    ("165003", "Advances to related parties",  1),
    ("189001", "IC Receivables",               1),
    ("224000", "Due to related parties",       2),
    ("224009", "IC Payables",                  2),
]
S1 = [c for c, _, s in ACCOUNTS if s == 1]   # ["165003","189001"]
S2 = [c for c, _, s in ACCOUNTS if s == 2]   # ["224000","224009"]

# ── Simple company names ──────────────────────────────────────────────────
ENT = {
    "001001": "Company A",
    "001002": "Company B",
    "001003": "Company C",
}
ICP = {
    "ICP_001002": "Company B ICP",
    "ICP_001003": "Company C ICP",
    "ICP_001001": "Company A ICP",
}

# ── Styles ────────────────────────────────────────────────────────────────
B  = Font(bold=True)
WB = Font(bold=True, color="FFFFFF", size=10)
S1_HDR = PatternFill(start_color="003366", fill_type="solid")   # dark blue
S2_HDR = PatternFill(start_color="336600", fill_type="solid")   # dark green
V_HDR  = PatternFill(start_color="663300", fill_type="solid")   # brown
YEL    = PatternFill(start_color="FFFF99", fill_type="solid")   # yellow match
TBL    = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))

def hcell(ws, r, c, val, fill):
    cell = ws.cell(r, c, val)
    cell.font = WB; cell.fill = fill; cell.border = TBL
    cell.alignment = Alignment(wrap_text=True, horizontal="center")
    return cell


# ═════════════════════════════════════════════════════════════════════════
# 1.  SAMPLE ICM REPORT
# ═════════════════════════════════════════════════════════════════════════
#
# Row 4 — header  (Entity | Partner | 165003 | 189001 | 224000 | 224009 | Variance)
# Rows 5-14 — 10 entity-partner pairs with NO pre-filled account values
#             (account values come from the journal matching later)
#
ICM_PAIRS = [
    ("001001","ICP_001002"),   # A → B
    ("001001","ICP_001003"),   # A → C
    ("001002","ICP_001001"),   # B → A
    ("001002","ICP_001003"),   # B → C
    ("001003","ICP_001001"),   # C → A
    ("001003","ICP_001002"),   # C → B
    ("001001","ICP_001002"),   # A → B  (appears twice — different journal lines)
    ("001002","ICP_001001"),   # B → A  (appears twice)
    ("001001","ICP_001003"),   # A → C  (appears twice)
    ("001003","ICP_001002"),   # C → B  (appears twice)
]

def create_icm():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"

    # Row 4 header
    ws.cell(4,1,"Entity").font = B
    ws.cell(4,2,"Partner").font = B
    col = 3
    for code, desc, series in ACCOUNTS:
        tag  = "Entity" if series == 1 else "Partner"
        fill = S1_HDR if series == 1 else S2_HDR
        hcell(ws, 4, col, f"{code} - {code}:{desc} {tag}", fill)
        col += 1
    hcell(ws, 4, col, "Variance", V_HDR)

    # Rows 5-14: entity-partner pairs, account cells are empty
    for i, (ent_code, prt_icp) in enumerate(ICM_PAIRS):
        r = 5 + i
        ws.cell(r, 1, f"{ent_code} - {ENT[ent_code]}")
        ws.cell(r, 2, f"{prt_icp} - {ICP[prt_icp]}")
        # account columns left genuinely blank

    # widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    for c in range(3, col+2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 28

    wb.save(SAMPLE_ICM)
    print(f"✅ ICM → {SAMPLE_ICM}")


# ═════════════════════════════════════════════════════════════════════════
# 2.  SAMPLE JOURNAL REPORT
# ═════════════════════════════════════════════════════════════════════════
#
# Row 30 — header (Label|Status|Entity col3|Account col4|Intercompany col5|...|Debit col15|Credit col16)
# Rows 31-40 — 10 IC journal rows (original intercompany entries)
# Rows 41-50 — 5 Elimination journal rows (consolidation eliminations — reverse/reduce the IC balance)
#
# Sign convention:
#   Series 1 (Asset)      Debit = + (positive)   Credit = − (negative)
#   Series 2 (Liability)  Debit = − (negative)   Credit = + (positive)
#
# Elimination logic:
#   Original:    A debits  165003 for 50,000  (A has a receivable from B)
#   Elimination: A credits 165003 for 20,000  (partially eliminate the receivable)
#   Net signed:  165003 = +50,000 − 20,000 = +30,000  (remaining IC balance)
#
JOURNAL_ROWS = [
    # ── Original IC entries ──────────────────────────────────────────────
    # label     entity   account   icp           debit     credit        note
    ("JRN001", "001001", "165003", "ICP_001002", 50000,    None,  "IC"),  # A→B  Asset debit
    ("JRN001", "001002", "224000", "ICP_001001", None,     50000, "IC"),  # B→A  Liability credit
    ("JRN002", "001001", "189001", "ICP_001003", 120000,   None,  "IC"),  # A→C  IC Receivable
    ("JRN002", "001003", "224009", "ICP_001001", None,     120000,"IC"),  # C→A  IC Payable
    ("JRN003", "001002", "165003", "ICP_001003", 75000,    None,  "IC"),  # B→C  Asset debit
    ("JRN003", "001003", "224000", "ICP_001002", None,     75000, "IC"),  # C→B  Liability credit
    ("JRN004", "001001", "165003", "ICP_001002", 30000,    None,  "IC"),  # A→B  2nd IC entry
    ("JRN004", "001002", "224000", "ICP_001001", None,     30000, "IC"),  # B→A  2nd IC entry
    ("JRN005", "001003", "189001", "ICP_001001", 20000,    None,  "IC"),  # C→A  IC Receivable
    ("JRN005", "001001", "224009", "ICP_001003", None,     20000, "IC"),  # A→C  IC Payable

    # ── Elimination entries (partial — flip Debit/Credit to reduce IC balance) ──
    # Same (Entity + Partner + Account), opposite side — net reduces the balance
    ("ELIM001","001001","165003", "ICP_001002", None,     20000, "ELIM"),  # Eliminate part of A→B receivable
    ("ELIM001","001002","224000", "ICP_001001", 20000,    None,  "ELIM"),  # Eliminate part of B→A payable
    ("ELIM002","001001","189001", "ICP_001003", None,     50000, "ELIM"),  # Eliminate part of A→C receivable
    ("ELIM002","001003","224009", "ICP_001001", 50000,    None,  "ELIM"),  # Eliminate part of C→A payable
    ("ELIM003","001002","165003", "ICP_001003", None,     25000, "ELIM"),  # Eliminate part of B→C receivable
    ("ELIM003","001003","224000", "ICP_001002", 25000,    None,  "ELIM"),  # Eliminate part of C→B payable
    ("ELIM004","001003","189001", "ICP_001001", None,     20000, "ELIM"),  # Fully eliminate C→A receivable
    ("ELIM004","001001","224009", "ICP_001003", 20000,    None,  "ELIM"),  # Fully eliminate A→C payable
    ("ELIM005","001001","165003", "ICP_001002", None,     10000, "ELIM"),  # Eliminate more of A→B
    ("ELIM005","001002","224000", "ICP_001001", 10000,    None,  "ELIM"),  # Eliminate more of B→A
]

def create_journal():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"

    # Meta rows 1-29
    ws.cell(1,1,"Journal Report (Sample)")
    ws.cell(2,1,"Application: FCCS")
    ws.cell(3,1,"Year: 2025  |  Period: Dec")
    ws.cell(5,1,"Sign rule: Series 1 (Asset) — Debit = positive | Series 2 (Liability) — Credit = positive")

    # Row 30: column headers
    HDRS = {1:"Label",2:"Status",3:"Entity",4:"Account",5:"Intercompany",
            6:"Movement",7:"Data Source",8:"Journal Type",9:"Journal Group",
            10:"Journal ID",11:"Journal",12:"Description",13:"Reversing",
            14:"Journal Description",15:"Debit",16:"Credit",17:"Net"}
    for c, h in HDRS.items():
        hcell(ws, 30, c, h, S1_HDR)

    # ── Row colors for IC vs Elimination ──
    ELIM_FILL = PatternFill(start_color="FFE0CC", fill_type="solid")  # light orange for elim rows

    # Rows 31 onwards: IC entries then Elimination entries
    for i, row_data in enumerate(JOURNAL_ROWS):
        lbl, ent, acc, icp_code, dbt, crd, jtype = row_data
        r = 31 + i
        acc_desc = next(d for c,d,s in ACCOUNTS if c == acc)
        series   = next(s for c,d,s in ACCOUNTS if c == acc)
        sign_note = "Debit = +" if series == 1 else "Credit = +"
        is_elim  = (jtype == "ELIM")

        ws.cell(r, 1,  lbl)
        ws.cell(r, 2,  "Posted")
        ws.cell(r, 3,  f"{ent}:{ENT[ent]}")
        ws.cell(r, 4,  f"{acc}:{acc}:{acc_desc}")
        ws.cell(r, 5,  f"{icp_code}:{ICP[icp_code]}")
        ws.cell(r, 6,  "Periodic")
        ws.cell(r, 7,  "FCCS_Journal Input")
        ws.cell(r, 8,  "Elimination" if is_elim else "Manual")
        ws.cell(r, 9,  "ELIM_Sample" if is_elim else "IC_Sample")
        ws.cell(r, 10, lbl)
        ws.cell(r, 11, f"{'ELIM' if is_elim else 'IC'} {ENT[ent]} ↔ {ICP[icp_code]}")
        ws.cell(r, 12, acc_desc)
        ws.cell(r, 13, "No")
        ws.cell(r, 14, f"Series {series} | {sign_note} | {'ELIMINATION' if is_elim else 'IC entry'}")
        ws.cell(r, 15, dbt)
        ws.cell(r, 16, crd)
        ws.cell(r, 17, (dbt or 0) - (crd or 0))

        # Highlight elimination rows in orange
        if is_elim:
            for c in range(1, 18):
                if ws.cell(r, c).value is not None:
                    ws.cell(r, c).fill = ELIM_FILL

    # Grand Total
    gr = 31 + len(JOURNAL_ROWS)
    ws.cell(gr, 1, "Grand Total").font = B
    ws.cell(gr, 15, sum(row[4] for row in JOURNAL_ROWS if row[4])).font = B
    ws.cell(gr, 16, sum(row[5] for row in JOURNAL_ROWS if row[5])).font = B

    # widths
    for c, w in [(1,12),(2,10),(3,25),(4,35),(5,30)]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    for c in range(6, 18):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 18

    wb.save(SAMPLE_J1)
    print(f"✅ Journal → {SAMPLE_J1}")


# ═════════════════════════════════════════════════════════════════════════
# 3.  IC MATCHING  (same logic as ic_matching_v2.py)
# ═════════════════════════════════════════════════════════════════════════
#
# Steps per journal:
#   1. Check (Entity + Partner + Account) combination
#   2. Find amount type (Debit or Credit)
#   3. Map sign (+/−) based on account type
#        Series 1 → net = debit − credit  (+debit, −credit)
#        Series 2 → net = credit − debit  (+credit, −debit)
#   4. Write signed value into Account column in output
#   5. Calculate Variance = Σ Series1 − Σ Series2
#

def get_sign(net, series):
    """Apply sign based on account type:
       Series 1 (Asset/Expense): Debit=+, Credit=−  → net = debit−credit (already correct)
       Series 2 (Liability/Equity/Revenue): Credit=+, Debit=−  → flip net sign
    """
    return net if series == 1 else -net

def run_matching():
    print("\n── IC Matching ──")

    # A. Read ICM pairs
    wb_icm = openpyxl.load_workbook(SAMPLE_ICM, data_only=True)
    ws_icm = wb_icm.active
    icm_rows = []
    for row in range(5, ws_icm.max_row + 1):
        ev = str(ws_icm.cell(row,1).value or "").strip()
        pv = str(ws_icm.cell(row,2).value or "").strip()
        if not ev and not pv:
            continue
        em = re.match(r"(\d{6})", ev)
        pm = re.match(r"(ICP_\w+)", pv)
        icm_rows.append({
            "row":     row,
            "ent_raw": ev,
            "prt_raw": pv,
            "ent":     em.group(1) if em else "",
            "prt":     pm.group(1) if pm else "",
        })
    print(f"  ICM rows: {len(icm_rows)}")

    # B. Read Journal — build (entity, icp, account) → signed_net
    wb_j = openpyxl.load_workbook(SAMPLE_J1, data_only=True)
    ws_j = wb_j.active
    lookup = defaultdict(float)

    for row in range(31, ws_j.max_row + 1):
        lbl = str(ws_j.cell(row,1).value or "").strip()
        if lbl == "Grand Total":
            break
        ent_raw = str(ws_j.cell(row,3).value or "")
        acc_raw = str(ws_j.cell(row,4).value or "")
        icp_raw = str(ws_j.cell(row,5).value or "")

        em = re.match(r"(\d{6})", ent_raw)
        am = re.match(r"(\d{6})", acc_raw)
        im = re.match(r"(ICP_\w+)", icp_raw)
        if not (em and am and im):
            continue

        ent_code = em.group(1)
        acc_code = am.group(1)
        icp_code = im.group(1)
        dbt = float(ws_j.cell(row,15).value or 0)
        crd = float(ws_j.cell(row,16).value or 0)
        net = dbt - crd

        # Map sign by series
        series = next((s for c,_,s in ACCOUNTS if c == acc_code), 0)
        signed = get_sign(net, series)

        lookup[(ent_code, icp_code, acc_code)] += signed

    print(f"  Journal (entity,icp,account) keys: {len(lookup)}")
    for k,v in lookup.items():
        print(f"    {k} → {v:,.0f}")

    # C. Build output
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "IC Matched (Sample)"

    # Header row 1
    out_ws.cell(1,1,"Entity").font = B
    out_ws.cell(1,2,"Partner").font = B
    acct_col = {}
    col = 3
    for code, desc, series in ACCOUNTS:
        fill = S1_HDR if series == 1 else S2_HDR
        hcell(out_ws, 1, col, f"{code}\n{desc}\n({'Series '+str(series)})", fill)
        acct_col[code] = col
        col += 1
    hcell(out_ws, 1, col, "Variance\n(Σ S1 − Σ S2)", V_HDR)
    var_col = col

    # Data rows
    for icm in icm_rows:
        r   = icm["row"]
        ent = icm["ent"]
        prt = icm["prt"]

        out_ws.cell(r, 1, icm["ent_raw"])
        out_ws.cell(r, 2, icm["prt_raw"])

        s1_sum = s2_sum = 0
        for code, _, series in ACCOUNTS:
            key = (ent, prt, code)
            val = lookup.get(key)
            if val is not None and val != 0:
                cell = out_ws.cell(r, acct_col[code], val)
                cell.fill = YEL
                if series == 1: s1_sum += val
                else:           s2_sum += val

        variance = s1_sum - s2_sum
        if variance != 0:
            out_ws.cell(r, var_col, variance)

    # Widths
    out_ws.column_dimensions["A"].width = 30
    out_ws.column_dimensions["B"].width = 30
    for c in range(3, var_col+1):
        out_ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 20
    out_ws.row_dimensions[1].height = 50

    out_wb.save(SAMPLE_OUTPUT)
    print(f"\n✅ Output → {SAMPLE_OUTPUT}")


# ═════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("="*55)
    create_icm()
    create_journal()
    run_matching()
    print("="*55)
    print("Done. Blue header = Series 1 | Green = Series 2 | Brown = Variance | Yellow = matched")
