"""
IC Matching Report Enrichment Script
=====================================
Reads:
  1. ICM Report  (Intercompany Balances IC Matching Report (1).xlsx)
  2. Journal Report (Journal Report (1).xlsx)

Match logic:
  - PRIMARY KEY: (entity_code, icp_code, account_code)
    Used when entity codes align directly between Journal and ICM (e.g., 001021, 001033)
  - FALLBACK KEY: (icp_code, account_code)
    Used when journal uses E-prefix group entity codes that don't map 1:1 to ICM entities.
    These are flagged in the log for user review.

For each Entity+Partner row in the ICM:
  - Finds matching journal lines
  - Net = Debit - Credit
  - Writes net into the correct elimination account column (Entity-side = first col occurrence)
  - Highlights updated cells in yellow

Outputs:
  ICM_Output.xlsx  (new file, same layout as ICM, with journal amounts filled in)
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

# ── File paths (static) ────────────────────────────────────────────────────
ICM_FILE     = r"g:\AI\Intercompany Balances IC Matching Report (1).xlsx"
JOURNAL_FILE = r"g:\AI\Journal Report (1).xlsx"
OUTPUT_FILE  = r"g:\AI\ICM_Output.xlsx"

# ── Row constants ──────────────────────────────────────────────────────────
ICM_HEADER_ROW     = 4
ICM_DATA_START     = 5
JOURNAL_HEADER_ROW = 30
JOURNAL_DATA_START = 31

# Journal column indices (1-based)
J_ENTITY = 3
J_ACCT   = 4
J_ICP    = 5
J_DEBIT  = 15
J_CREDIT = 16

HIGHLIGHT_FILL   = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
UNMATCHED_FILL   = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def extract_icp_code(raw: str) -> str:
    """'ICP_001051 - Name ICP'  or  'ICP_001051:Name ICP'  →  'ICP_001051'"""
    m = re.match(r"(ICP_\w+)", str(raw or "").strip())
    return m.group(1) if m else ""


def extract_account_code(raw: str) -> str:
    """
    '[165000].[189501]:189501:...'  →  '189501'
    '224000:224000:...'             →  '224000'
    '224000 - 224000:...'           →  '224000'
    """
    raw = str(raw or "").strip()
    m = re.search(r"\]\.\[?(\d{6})\]?:", raw)
    if m:
        return m.group(1)
    m = re.match(r"(\d{6})", raw)
    return m.group(1) if m else ""


def extract_entity_code_icm(raw: str) -> str:
    """'001001 - Company Name'  →  '001001'"""
    m = re.match(r"(\d{6})", str(raw or "").strip())
    return m.group(1) if m else ""


def extract_entity_code_journal(raw: str) -> str:
    """
    'E117100:Name'   →  'E117100'  (E-prefix consolidated entity)
    '001033:Name'    →  '001033'   (direct code)
    """
    raw = str(raw or "").strip()
    m = re.match(r"(E\d+|\d{6})", raw)
    return m.group(1) if m else ""


def to_float(val) -> float:
    if val is None or str(val).strip() in ("", " "):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def is_detail_row(vals: list) -> bool:
    """Return True if this is a real journal detail row (not subtotal/empty)."""
    entity = str(vals[J_ENTITY - 1] or "").strip()
    acct   = str(vals[J_ACCT - 1]   or "").strip()
    icp    = str(vals[J_ICP - 1]    or "").strip()
    return bool(entity or acct or icp)


# ═══════════════════════════════════════════════════════════════════════════
# STEP 1 – Read ICM Report
# ═══════════════════════════════════════════════════════════════════════════

def read_icm_report(filepath: str):
    print(f"[Step 1] Reading ICM report: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Parse header row → { account_code: [col_idx, ...] }
    headers = [cell.value for cell in ws[ICM_HEADER_ROW]]
    account_cols: dict[str, list[int]] = defaultdict(list)
    for col_idx, h in enumerate(headers, start=1):
        code = extract_account_code(str(h) if h else "")
        if re.match(r"^\d{6}$", code):
            account_cols[code].append(col_idx)

    print(f"  Elimination accounts detected: {sorted(account_cols.keys())}")

    # Parse data rows
    data_rows = []
    for row_num, row in enumerate(
            ws.iter_rows(min_row=ICM_DATA_START, max_row=ws.max_row), start=ICM_DATA_START):
        vals = [cell.value for cell in row]
        entity_raw  = str(vals[0] or "").strip()
        partner_raw = str(vals[1] or "").strip() if len(vals) > 1 else ""
        if not entity_raw and not partner_raw:
            continue
        data_rows.append({
            "row_num":      row_num,
            "entity":       entity_raw,
            "partner":      partner_raw,
            "entity_code":  extract_entity_code_icm(entity_raw),
            "partner_code": extract_icp_code(partner_raw),
            "values":       vals,
        })

    print(f"  Data rows loaded: {len(data_rows)}")
    return headers, account_cols, data_rows, ws, wb


# ═══════════════════════════════════════════════════════════════════════════
# STEP 2 – Read Journal Report
# ═══════════════════════════════════════════════════════════════════════════

def read_journal_report(filepath: str):
    print(f"\n[Step 2] Reading Journal report: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    lines = []
    for row_num, row in enumerate(
            ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row), start=JOURNAL_DATA_START):
        vals = [cell.value for cell in row]
        label = str(vals[0] or "").strip()
        if label == "Grand Total":
            break
        if not is_detail_row(vals):
            continue

        entity_raw   = str(vals[J_ENTITY - 1] or "").strip()
        account_raw  = str(vals[J_ACCT - 1]   or "").strip()
        icp_raw      = str(vals[J_ICP - 1]    or "").strip()
        entity_code  = extract_entity_code_journal(entity_raw)
        account_code = extract_account_code(account_raw)
        icp_code     = extract_icp_code(icp_raw)
        debit        = to_float(vals[J_DEBIT - 1])
        credit       = to_float(vals[J_CREDIT - 1])

        lines.append({
            "row_num":      row_num,
            "label":        label,
            "entity_raw":   entity_raw,
            "entity_code":  entity_code,
            "is_group":     entity_code.startswith("E"),   # E-prefix = consolidated group
            "account_code": account_code,
            "icp_code":     icp_code,
            "debit":        debit,
            "credit":       credit,
        })

    print(f"  Detail lines loaded: {len(lines)}")

    # Build two lookups:
    #   primary:  (entity_code, icp_code, account_code)  → direct match
    #   fallback: (icp_code, account_code) with is_group=True → group-level match
    primary_lookup  = defaultdict(list)
    fallback_lookup = defaultdict(list)

    for line in lines:
        if not line["is_group"]:
            primary_lookup[(line["entity_code"], line["icp_code"], line["account_code"])].append(line)
        else:
            fallback_lookup[(line["icp_code"], line["account_code"])].append(line)

    print(f"  Primary (entity+ICP+account) keys: {len(primary_lookup)}")
    print(f"  Fallback (ICP+account, group entities) keys: {len(fallback_lookup)}")
    return lines, primary_lookup, fallback_lookup


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3 – Match & Update
# ═══════════════════════════════════════════════════════════════════════════

def match_and_update(data_rows, account_cols, primary_lookup, fallback_lookup):
    print("\n[Step 3] Matching ICM rows against Journal entries...")
    updates     = {}   # (row_num, col_idx) → net_amount
    match_log   = []
    unmatched   = []   # ICP+account combos not found anywhere

    for icm_row in data_rows:
        ent_code     = icm_row["entity_code"]
        partner_code = icm_row["partner_code"]
        if not partner_code:
            continue

        for acct_code, col_list in account_cols.items():
            target_col = col_list[0]          # Entity-side column (first occurrence)

            # ── Attempt primary match (entity + ICP + account) ──────────
            key_primary = (ent_code, partner_code, acct_code)
            if key_primary in primary_lookup:
                journal_matches = primary_lookup[key_primary]
                total_debit  = sum(j["debit"]  for j in journal_matches)
                total_credit = sum(j["credit"] for j in journal_matches)
                net          = total_debit - total_credit
                if net != 0:
                    updates[(icm_row["row_num"], target_col)] = net
                    match_log.append(
                        f"PRIMARY  | ICM Row {icm_row['row_num']:4d} | "
                        f"Entity: {ent_code} | Partner: {partner_code} | "
                        f"Acct: {acct_code} | Net: {net:,.2f} → Col {target_col}"
                    )
                continue   # primary match consumed, skip fallback

            # ── Attempt fallback match (ICP + account, group entity) ────
            key_fallback = (partner_code, acct_code)
            if key_fallback in fallback_lookup:
                journal_matches = fallback_lookup[key_fallback]
                total_debit  = sum(j["debit"]  for j in journal_matches)
                total_credit = sum(j["credit"] for j in journal_matches)
                net          = total_debit - total_credit
                labels       = list({j["label"] for j in journal_matches})
                if net != 0:
                    # Mark as group-level — still write it, flag it
                    updates[(icm_row["row_num"], target_col)] = ("GROUP", net)
                    match_log.append(
                        f"FALLBACK | ICM Row {icm_row['row_num']:4d} | "
                        f"Entity: {ent_code} | Partner: {partner_code} | "
                        f"Acct: {acct_code} | Net: {net:,.2f} (group adj {labels}) → Col {target_col} [REVIEW]"
                    )

    primary_count  = sum(1 for v in updates.values() if not isinstance(v, tuple))
    fallback_count = sum(1 for v in updates.values() if isinstance(v, tuple))
    print(f"  Primary matches:  {primary_count} cells")
    print(f"  Fallback matches: {fallback_count} cells (group entities — flagged for review)")
    return updates, match_log


# ═══════════════════════════════════════════════════════════════════════════
# STEP 4 – Write Output
# ═══════════════════════════════════════════════════════════════════════════

def write_output(icm_wb, icm_ws, updates: dict, output_path: str):
    print(f"\n[Step 4] Writing output to: {output_path}")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = icm_ws.title

    # Copy all cells from original ICM
    for row in icm_ws.iter_rows():
        for cell in row:
            new_cell = out_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.font:
                try:
                    new_cell.font = Font(
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        size=cell.font.size,
                        name=cell.font.name,
                    )
                except Exception:
                    pass

    # Copy column widths
    for col_letter, col_dim in icm_ws.column_dimensions.items():
        out_ws.column_dimensions[col_letter].width = col_dim.width

    # Apply updates
    for (row_num, col_idx), value in updates.items():
        cell = out_ws.cell(row=row_num, column=col_idx)
        if isinstance(value, tuple) and value[0] == "GROUP":
            cell.value = value[1]
            cell.fill  = UNMATCHED_FILL   # light red = needs review
        else:
            cell.value = value
            cell.fill  = HIGHLIGHT_FILL   # yellow = matched

    out_wb.save(output_path)
    yellow = sum(1 for v in updates.values() if not isinstance(v, tuple))
    red    = sum(1 for v in updates.values() if isinstance(v, tuple))
    print(f"  Saved. Yellow cells (direct match): {yellow} | Red cells (group, needs review): {red}")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("  IC Matching Report Enrichment Script")
    print("=" * 65)

    headers, account_cols, data_rows, icm_ws, icm_wb = read_icm_report(ICM_FILE)
    lines, primary_lookup, fallback_lookup = read_journal_report(JOURNAL_FILE)
    updates, match_log = match_and_update(
        data_rows, account_cols, primary_lookup, fallback_lookup
    )

    print("\n── Match Log ──────────────────────────────────────────────────")
    for msg in match_log:
        print(f"  {msg}")

    write_output(icm_wb, icm_ws, updates, OUTPUT_FILE)

    print("\n" + "=" * 65)
    print(f"  Done!  Output: {OUTPUT_FILE}")
    print("  Yellow = direct entity match  |  Red = group entity (review)")
    print("=" * 65)


if __name__ == "__main__":
    main()
