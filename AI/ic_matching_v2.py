"""
IC Matching Report — N-Journal with 2-Variance per Journal
===========================================================
Reads:
  1. ICM Report  (Intercompany Balances IC Matching Report (1).xlsx)
  2. N Journal files (auto-discovered from JOURNAL_FOLDER)

Output layout per journal (matching real ICM structure):
  [Entity] [Partner] |
  Journal1: [S1-Entity cols] [S2-Partner cols] [Variance1] [S1-Partner cols] [S2-Entity cols] [Variance2] |
  Journal2: ...same...  |  JournalN: ...

Variance1 = Sum(S1-Entity)  - Sum(S2-Partner)
Variance2 = Sum(S1-Partner) - Sum(S2-Entity)

Sign convention:
  Series 1 (Asset)    & 5 (Expense)          : Debit = +, Credit = -
  Series 2 (Liability), 3 (Equity), 4 (Rev.) : Credit = +, Debit = -
"""

import re
import os
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict, OrderedDict

# ── File paths ──────────────────────────────────────────────────────────────
ICM_FILE        = r"g:\AI\Intercompany Balances IC Matching Report (1).xlsx"
JOURNAL_FOLDER  = r"g:\AI"
JOURNAL_PATTERN = "Journal Report*.xlsx"
OUTPUT_FILE     = r"g:\AI\ICM_Output_v2.xlsx"

# ── Row/Column constants ─────────────────────────────────────────────────────
ICM_HEADER_ROW     = 4
ICM_DATA_START     = 5
JOURNAL_DATA_START = 31
J_ENTITY = 3; J_ACCT = 4; J_ICP = 5; J_DEBIT = 15; J_CREDIT = 16

# ── Styling ──────────────────────────────────────────────────────────────────
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL_S1 = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FILL_S2 = PatternFill(start_color="336600", end_color="336600", fill_type="solid")
HEADER_FILL_VAR= PatternFill(start_color="663300", end_color="663300", fill_type="solid")
HEADER_FILL_EX = PatternFill(start_color="555555", end_color="555555", fill_type="solid")
MATCH_FILL     = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
GROUP_FILL     = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
THIN_BORDER    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 1 — UTILITY / SIGN FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def extract_icp_code(raw: str) -> str:
    m = re.match(r"(ICP_\w+)", str(raw or "").strip())
    return m.group(1) if m else ""


def extract_account_code(raw: str) -> str:
    """
    Handles:
      '[165000].[189501]:189501:...'  → '189501'
      '[FCCS_Group].[Plug_InvSh]:...' → 'Plug_InvSh'
      '224000:224000:...'             → '224000'
      '224000 - 224000:...'           → '224000'
    """
    raw = str(raw or "").strip()
    m = re.search(r"\]\.\[?(\w+)\]?:", raw)
    if m:
        val = m.group(1)
        if not val.isdigit():
            m2 = re.search(r"]:(\d{6}):", raw)
            if m2:
                return m2.group(1)
        return val
    m = re.match(r"(\d{6})", raw)
    return m.group(1) if m else ""


def extract_entity_code_icm(raw: str) -> str:
    m = re.match(r"(\d{6})", str(raw or "").strip())
    return m.group(1) if m else ""


def extract_entity_code_journal(raw: str) -> str:
    raw = str(raw or "").strip()
    m = re.match(r"(E?\d+\w*)", raw)
    return m.group(1) if m else ""


def to_float(val) -> float:
    if val is None or str(val).strip() in ("", " "):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def is_detail_row(vals: list) -> bool:
    entity = str(vals[J_ENTITY - 1] or "").strip()
    acct   = str(vals[J_ACCT - 1]   or "").strip()
    icp    = str(vals[J_ICP - 1]    or "").strip()
    return bool(entity or acct or icp)


def classify_account(code: str) -> str:
    """
    S1 = debit-positive  (series 1 Asset, series 5 Expense)
    S2 = credit-positive (series 2 Liability, 3 Equity, 4 Revenue)
    """
    code = str(code or "").strip()
    if not code or not code[0].isdigit():
        return "EXTRA"
    s = code[0]
    if s in ("1", "5"):
        return "S1"
    elif s in ("2", "3", "4"):
        return "S2"
    return "EXTRA"


def apply_sign(debit: float, credit: float, account_code: str) -> float:
    """
    Apply sign based on account series:
      Series 1 & 5 → debit - credit  (Debit = positive)
      Series 2,3,4 → credit - debit  (Credit = positive)
    """
    code = str(account_code or "").strip()
    s = code[0] if code and code[0].isdigit() else "0"
    if s in ("1", "5"):
        return debit - credit
    elif s in ("2", "3", "4"):
        return credit - debit
    return debit - credit


def get_account_description(raw_header: str) -> str:
    raw = str(raw_header or "").strip()
    raw = re.sub(r"\s+(Entity|Partner)\s*$", "", raw)
    return raw


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 2 — COLUMN MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════

def read_icm_headers(ws):
    """
    Read ICM header row and split accounts into 4 tagged sub-lists:

      Set 1 (Entity-side):
        s1_entity  — S1 accounts tagged 'Entity'  (Asset/Expense receivable side)
        s2_partner — S2 accounts tagged 'Partner' (Liability/Equity payable side)
        → Variance 1 = Σ(s1_entity) − Σ(s2_partner)

      Set 2 (Partner-side / mirror):
        s1_partner — S1 accounts tagged 'Partner'
        s2_entity  — S2 accounts tagged 'Entity'
        → Variance 2 = Σ(s1_partner) − Σ(s2_entity)

    Returns: known_accounts, s1_entity, s2_partner, s1_partner, s2_entity
    """
    headers = [cell.value for cell in ws[ICM_HEADER_ROW]]
    known_accounts = OrderedDict()
    s1_entity  = []
    s2_partner = []
    s1_partner = []
    s2_entity  = []

    SKIP = {"Entity", "Partner", "Variance", "Total",
            "PARENT INPUT", "QAR_Reporting", "Final amount", "final variance"}
    seen = set()

    for h in headers:
        hs = str(h or "").strip()
        if not hs or hs in SKIP:
            continue
        code = extract_account_code(hs)
        if not code:
            continue

        # Detect Entity / Partner suffix
        if re.search(r"\bEntity\b", hs):
            tag = "Entity"
        elif re.search(r"\bPartner\b", hs):
            tag = "Partner"
        else:
            tag = "Entity"   # default

        pair = (code, tag)
        if pair in seen:
            continue
        seen.add(pair)

        if code not in known_accounts:
            known_accounts[code] = get_account_description(hs)

        cls = classify_account(code)
        if cls == "S1":
            if tag == "Entity":  s1_entity.append(code)
            else:                s1_partner.append(code)
        elif cls == "S2":
            if tag == "Partner": s2_partner.append(code)
            else:                s2_entity.append(code)

    return known_accounts, s1_entity, s2_partner, s1_partner, s2_entity


def build_output_columns(icm_accounts, all_journal_accounts,
                         s1_entity, s2_partner, s1_partner, s2_entity):
    """
    Build two column lists for the output, using ONLY accounts defined in the ICM.
    Journal-only accounts are logged but excluded from the output.

      entity_cols  = [(code, desc, series), ...]  → S1-Entity + S2-Partner
      partner_cols = [(code, desc, series), ...]  → S1-Partner + S2-Entity
    """
    def col(code):
        return (code, icm_accounts.get(code, code), classify_account(code))

    entity_cols  = [col(c) for c in s1_entity]  + [col(c) for c in s2_partner]
    partner_cols = [col(c) for c in s1_partner] + [col(c) for c in s2_entity]

    # Log journal-only accounts (excluded from output — not in ICM)
    icm_known = set(icm_accounts.keys())
    extra = sorted(all_journal_accounts - icm_known)
    if extra:
        print(f"  Journal-only accounts (not in ICM, excluded): {extra}")

    return entity_cols, partner_cols



def discover_journals(folder=JOURNAL_FOLDER, pattern=JOURNAL_PATTERN):
    paths = sorted(glob.glob(os.path.join(folder, pattern)))
    print(f"  Discovered {len(paths)} journal file(s):")
    for p in paths:
        print(f"    {os.path.basename(p)}")
    return paths


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 3 — JOURNAL READING
# ═══════════════════════════════════════════════════════════════════════════

def read_journal_report(filepath: str):
    """
    Parse a single journal Excel file.
    Returns: lines, primary_lookup, fallback_lookup, all_accounts
      primary_lookup  : { (entity_code, icp_code, account_code) → [line dicts] }
      fallback_lookup : { (icp_code, account_code) → [line dicts] }  for E-prefix group entities
    """
    print(f"  Reading: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    lines = []
    all_accounts = set()

    for row_num, row in enumerate(
            ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row),
            start=JOURNAL_DATA_START):
        vals = [cell.value for cell in row]
        label = str(vals[0] or "").strip()
        if label == "Grand Total":
            break
        if not is_detail_row(vals):
            continue

        entity_raw   = str(vals[J_ENTITY - 1] or "").strip()
        account_raw  = str(vals[J_ACCT - 1]   or "").strip()
        icp_raw      = str(vals[J_ICP - 1]    or "").strip()
        entity_code  = extract_entity_code_journal(entity_raw)
        account_code = extract_account_code(account_raw)
        icp_code     = extract_icp_code(icp_raw)
        debit        = to_float(vals[J_DEBIT - 1])
        credit       = to_float(vals[J_CREDIT - 1])

        if account_code:
            all_accounts.add(account_code)

        lines.append({
            "row_num":      row_num,
            "label":        label,
            "entity_raw":   entity_raw,
            "entity_code":  entity_code,
            "is_group":     bool(re.match(r"^E\d+", entity_code)),
            "account_code": account_code,
            "icp_code":     icp_code,
            "debit":        debit,
            "credit":       credit,
        })

    primary_lookup  = defaultdict(list)
    fallback_lookup = defaultdict(list)
    for line in lines:
        if not line["is_group"]:
            primary_lookup[(line["entity_code"], line["icp_code"], line["account_code"])].append(line)
        else:
            fallback_lookup[(line["icp_code"], line["account_code"])].append(line)

    print(f"    Lines: {len(lines)} | Accounts: {sorted(all_accounts)}")
    print(f"    Primary keys: {len(primary_lookup)} | Fallback keys: {len(fallback_lookup)}")
    return lines, primary_lookup, fallback_lookup, all_accounts


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 4 — MATCHING
# ═══════════════════════════════════════════════════════════════════════════

def match_journal_to_icm(data_rows, account_code_list, primary_lookup, fallback_lookup):
    """
    For each (Entity, Partner) in ICM and each account code:
      Step 1: Check (Entity + Partner + Account) combination — primary match
      Step 2: Find amount type (Debit or Credit)
      Step 3: Map sign based on account series using apply_sign()
      Fallback: Group entities (E-prefix) matched on (Partner + Account) only

    Returns: updates dict, match_log list
    """
    updates   = {}
    match_log = []

    for icm_row in data_rows:
        ent_code     = icm_row["entity_code"]
        partner_code = icm_row["partner_code"]
        if not partner_code:
            continue

        for acct_code in account_code_list:
            # ── Primary match: Entity + Partner + Account ──
            key_primary = (ent_code, partner_code, acct_code)
            if key_primary in primary_lookup:
                jlines = primary_lookup[key_primary]
                # Step 2+3: find debit/credit, then apply sign
                net = sum(apply_sign(j["debit"], j["credit"], acct_code) for j in jlines)
                if net != 0:
                    updates[(ent_code, partner_code, acct_code)] = net
                    match_log.append(
                        f"PRIMARY  | E:{ent_code} P:{partner_code} A:{acct_code} "
                        f"({classify_account(acct_code)}) | Net:{net:,.2f}"
                    )
                continue

            # ── Fallback: group entity (E-prefix) ──
            key_fallback = (partner_code, acct_code)
            if key_fallback in fallback_lookup:
                jlines = fallback_lookup[key_fallback]
                net = sum(apply_sign(j["debit"], j["credit"], acct_code) for j in jlines)
                if net != 0:
                    updates[(ent_code, partner_code, acct_code)] = ("GROUP", net)
                    labels = list({j["label"] for j in jlines})
                    match_log.append(
                        f"FALLBACK | E:{ent_code} P:{partner_code} A:{acct_code} "
                        f"| Net:{net:,.2f} ({labels})"
                    )

    primary_count  = sum(1 for v in updates.values() if not isinstance(v, tuple))
    fallback_count = sum(1 for v in updates.values() if isinstance(v, tuple))
    print(f"    Matches: {primary_count} primary + {fallback_count} fallback")
    return updates, match_log


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 5 — ICM DATA READER
# ═══════════════════════════════════════════════════════════════════════════

def read_icm_data(ws):
    """Read all (Entity, Partner) rows from the ICM sheet."""
    data_rows = []
    for row_num, row in enumerate(
            ws.iter_rows(min_row=ICM_DATA_START, max_row=ws.max_row), start=ICM_DATA_START):
        vals = [cell.value for cell in row]
        entity_raw  = str(vals[0] or "").strip()
        partner_raw = str(vals[1] or "").strip() if len(vals) > 1 else ""
        if not entity_raw and not partner_raw:
            continue
        data_rows.append({
            "row_num":      row_num,
            "entity":       entity_raw,
            "partner":      partner_raw,
            "entity_code":  extract_entity_code_icm(entity_raw),
            "partner_code": extract_icp_code(partner_raw),
        })
    return data_rows


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 6 — OUTPUT WRITER
# ═══════════════════════════════════════════════════════════════════════════

def write_output(data_rows, entity_cols, partner_cols,
                 all_updates_list, journal_names, output_path):
    """
    Write output Excel with 2 variances per journal set:

    Per journal set layout:
      [S1-Entity cols] [S2-Partner cols] [Variance 1] [S1-Partner cols] [S2-Entity cols] [Variance 2]

    Variance 1 = Σ(S1-Entity values)  − Σ(S2-Partner values)
    Variance 2 = Σ(S1-Partner values) − Σ(S2-Entity values)

    Step 4 of matching: add signed value into account column in output.
    Step 5: calculate variance.
    """
    print(f"\n[Step 4+5] Writing output: {output_path}")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "ICM Matched"

    # Columns per journal set: entity_cols + Var1 + partner_cols + Var2
    en = len(entity_cols)
    pn = len(partner_cols)
    cols_per_set = en + 1 + pn + 1

    # Pre-compute which codes go into each variance
    s1_entity_codes  = [c for c, _, s in entity_cols  if s == "S1"]
    s2_partner_codes = [c for c, _, s in entity_cols  if s == "S2"]
    s1_partner_codes = [c for c, _, s in partner_cols if s == "S1"]
    s2_entity_codes  = [c for c, _, s in partner_cols if s == "S2"]

    # ── Fixed columns: Entity, Partner ────────────────────────────────────
    out_ws.cell(ICM_HEADER_ROW, 1, "Entity").font  = Font(bold=True)
    out_ws.cell(ICM_HEADER_ROW, 2, "Partner").font = Font(bold=True)

    # ── Header rows for each journal set ──────────────────────────────────
    for set_idx, jname in enumerate(journal_names):
        base = 3 + set_idx * cols_per_set

        # Journal name (row above header)
        cell = out_ws.cell(ICM_HEADER_ROW - 1, base, jname)
        cell.font = Font(bold=True, size=11)

        col = base
        # --- Entity-side account headers (S1-Entity + S2-Partner) ---
        for code, desc, series in entity_cols:
            fill = HEADER_FILL_S1 if series == "S1" else HEADER_FILL_S2
            tag  = "(Entity)"  if series == "S1" else "(Partner)"
            c = out_ws.cell(ICM_HEADER_ROW, col, f"{code} - {desc} {tag}")
            c.font = HEADER_FONT; c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(wrap_text=True, horizontal="center")
            col += 1

        # Variance 1
        c = out_ws.cell(ICM_HEADER_ROW, col, "Variance 1\n(S1 Entity − S2 Partner)")
        c.font = HEADER_FONT; c.fill = HEADER_FILL_VAR
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True, horizontal="center")
        col += 1

        # --- Partner-side account headers (S1-Partner + S2-Entity) ---
        for code, desc, series in partner_cols:
            fill = HEADER_FILL_S1 if series == "S1" else HEADER_FILL_S2
            tag  = "(Partner)" if series == "S1" else "(Entity)"
            c = out_ws.cell(ICM_HEADER_ROW, col, f"{code} - {desc} {tag}")
            c.font = HEADER_FONT; c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(wrap_text=True, horizontal="center")
            col += 1

        # Variance 2
        c = out_ws.cell(ICM_HEADER_ROW, col, "Variance 2\n(S1 Partner − S2 Entity)")
        c.font = HEADER_FONT; c.fill = HEADER_FILL_VAR
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    # ── Data rows ──────────────────────────────────────────────────────────
    for icm_row in data_rows:
        r            = icm_row["row_num"]
        ent_code     = icm_row["entity_code"]
        partner_code = icm_row["partner_code"]

        out_ws.cell(r, 1, icm_row["entity"])
        out_ws.cell(r, 2, icm_row["partner"])

        for set_idx, updates in enumerate(all_updates_list):
            base = 3 + set_idx * cols_per_set
            col  = base

            ev = {}   # entity-side values  {code: signed_value}
            pv = {}   # partner-side values {code: signed_value}

            # --- Write Entity-side (S1-Entity + S2-Partner) ---
            for code, desc, series in entity_cols:
                val = updates.get((ent_code, partner_code, code))
                if val is not None:
                    v = val[1] if isinstance(val, tuple) else val
                    cell = out_ws.cell(r, col, v)
                    cell.fill = GROUP_FILL if isinstance(val, tuple) else MATCH_FILL
                    ev[code] = v
                col += 1

            # Variance 1 = Σ S1-Entity minus Σ S2-Partner
            var1 = (sum(to_float(ev.get(c, 0)) for c in s1_entity_codes) -
                    sum(to_float(ev.get(c, 0)) for c in s2_partner_codes))
            if var1 != 0:
                out_ws.cell(r, col, var1)
            col += 1

            # --- Write Partner-side (S1-Partner + S2-Entity) ---
            for code, desc, series in partner_cols:
                val = updates.get((ent_code, partner_code, code))
                if val is not None:
                    v = val[1] if isinstance(val, tuple) else val
                    cell = out_ws.cell(r, col, v)
                    cell.fill = GROUP_FILL if isinstance(val, tuple) else MATCH_FILL
                    pv[code] = v
                col += 1

            # Variance 2 = Σ S1-Partner minus Σ S2-Entity
            var2 = (sum(to_float(pv.get(c, 0)) for c in s1_partner_codes) -
                    sum(to_float(pv.get(c, 0)) for c in s2_entity_codes))
            if var2 != 0:
                out_ws.cell(r, col, var2)

    # ── Column widths ──────────────────────────────────────────────────────
    out_ws.column_dimensions["A"].width = 45
    out_ws.column_dimensions["B"].width = 45
    for col_idx in range(3, 3 + len(journal_names) * cols_per_set + 1):
        ltr = openpyxl.utils.get_column_letter(col_idx)
        out_ws.column_dimensions[ltr].width = 20
    out_ws.row_dimensions[ICM_HEADER_ROW].height = 50

    out_wb.save(output_path)
    print(f"  Saved: {len(journal_names)} journal(s) × {cols_per_set} cols each")
    print(f"         ({en} Entity-side + Var1 + {pn} Partner-side + Var2)")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("  IC Matching — N-Journal, 2-Variance per Journal")
    print("=" * 65)

    # ── Step 1: ICM structure ────────────────────────────────────────────
    print("\n[Step 1] Reading ICM report...")
    wb_icm = openpyxl.load_workbook(ICM_FILE, data_only=True)
    ws_icm = wb_icm.active

    icm_accounts, s1_entity, s2_partner, s1_partner, s2_entity = read_icm_headers(ws_icm)
    data_rows = read_icm_data(ws_icm)

    print(f"  Accounts (unique): {list(icm_accounts.keys())}")
    print(f"  S1-Entity  : {s1_entity}")
    print(f"  S2-Partner : {s2_partner}")
    print(f"  S1-Partner : {s1_partner}")
    print(f"  S2-Entity  : {s2_entity}")
    print(f"  ICM data rows: {len(data_rows)}")

    # ── Step 2: Read journals ────────────────────────────────────────────
    print("\n[Step 2] Reading journal files...")
    all_updates    = []
    all_match_logs = []
    all_j_accounts = set()
    journal_names  = []

    journal_files = discover_journals()
    if not journal_files:
        print("  ERROR: No journal files found. Check JOURNAL_FOLDER/JOURNAL_PATTERN.")
        return

    for jfile in journal_files:
        jname = os.path.basename(jfile).replace(".xlsx", "")
        journal_names.append(jname)
        lines, primary, fallback, j_accts = read_journal_report(jfile)
        all_j_accounts.update(j_accts)

        all_codes = list(icm_accounts.keys()) + list(j_accts - set(icm_accounts.keys()))
        updates, mlog = match_journal_to_icm(data_rows, all_codes, primary, fallback)
        all_updates.append(updates)
        all_match_logs.append(mlog)

    # ── Step 3: Build output column layout ──────────────────────────────
    print("\n[Step 3] Building output column layout...")
    entity_cols, partner_cols = build_output_columns(
        icm_accounts, all_j_accounts,
        s1_entity, s2_partner, s1_partner, s2_entity
    )
    print(f"  Entity-side  cols ({len(entity_cols)}): {[c for c,_,_ in entity_cols]}")
    print(f"  Partner-side cols ({len(partner_cols)}): {[c for c,_,_ in partner_cols]}")

    # ── Steps 4+5: Write output ──────────────────────────────────────────
    write_output(data_rows, entity_cols, partner_cols,
                 all_updates, journal_names, OUTPUT_FILE)

    # ── Match summary ────────────────────────────────────────────────────
    print("\n-- Match Summary " + "-" * 47)
    for jname, mlog in zip(journal_names, all_match_logs):
        print(f"\n  [{jname}]  {len(mlog)} matches")
        for msg in mlog[:3]:
            print(f"    {msg}")
        if len(mlog) > 3:
            print(f"    ... and {len(mlog) - 3} more")

    print("\n" + "=" * 65)
    print(f"  Done!  →  {OUTPUT_FILE}")
    print("  Yellow = primary match | Pink = group entity fallback")
    print("  Blue header = S1 (debit+) | Green = S2 (credit+) | Brown = Variance")
    print("  Per journal: [S1-Entity][S2-Partner][Var1][S1-Partner][S2-Entity][Var2]")
    print("=" * 65)


if __name__ == "__main__":
    main()
