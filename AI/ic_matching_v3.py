"""
IC Matching Report — N-Journal with 2-Variance per Journal (ICM-Style Output)
==============================================================================
Reads:
  1. ICM Report  (Intercompany Balances IC Matching Report (1).xlsx)
  2. N Journal files (auto-discovered from JOURNAL_FOLDER)

Output layout — ONE SHEET PER JOURNAL, styled to match the ICM report:
  Row 4 : Headers (uniform navy, white text, 9pt, wrap, height=78.75)
  Rows 5+: Data rows matching ICM row numbering

Per sheet column layout:
  [Entity] [Partner] |
  [S1-Entity cols] [S2-Partner cols] [Variance 1] |
  [S1-Partner cols] [S2-Entity cols] [Variance 2] |
  [Total]

Variance 1 = Sum(S1-Entity)  - Sum(S2-Partner)
Variance 2 = Sum(S1-Partner) - Sum(S2-Entity)
Total      = Variance 1 + Variance 2

Sign convention:
  Series 1 (Asset)    & 5 (Expense)          : Debit = +, Credit = -
  Series 2 (Liability), 3 (Equity), 4 (Rev.) : Credit = +, Debit = -

Cell colours match the ICM report exactly:
  Entity/Partner columns : FFC8DCF0  (light blue)
  Variance columns       : FFDCDCDC  (light grey)
  Total column           : FFC8C8C8  (medium grey)
  Headers (all cols)     : FF003366  (uniform navy)
  Primary match          : FFFF99    (yellow highlight)
  Group fallback match   : FFD9D9    (pink highlight)
"""

import re
import os
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict

# ── File paths ──────────────────────────────────────────────────────────────
ICM_FILE        = r"g:\AI\Intercompany Balances IC Matching Report (1).xlsx"
JOURNAL_FOLDER  = r"g:\AI"
JOURNAL_PATTERN = "Journal Report*.xlsx"
OUTPUT_FILE     = r"g:\AI\ICM_Output_v3.xlsx"

# ── Journal display labels (shown in row 3 above each journal block) ──────────
JOURNAL_LABELS = {
    "Journal Report (1)": "Journal Report 1 - Parent Input Data",
    "Journal Report (2)": "Journal Report 2 - Contribution Input Data",
    "Journal Report (4)": "Journal Report 4 - Plug Account Data",
}

# ── Row/Column constants ─────────────────────────────────────────────────────
ICM_HEADER_ROW     = 4
ICM_DATA_START     = 5
JOURNAL_DATA_START = 31
J_ENTITY = 3; J_ACCT = 4; J_ICP = 5; J_DEBIT = 15; J_CREDIT = 16

# ── Styling — matches ICM report exactly ─────────────────────────────────────
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=9)
HEADER_FILL    = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
ID_FILL        = PatternFill(start_color="C8DCF0", end_color="C8DCF0", fill_type="solid")
VARIANCE_FILL  = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
TOTAL_FILL     = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
MATCH_FILL     = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
GROUP_FILL     = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
JOURNAL_FILL   = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
ID_FONT        = Font(bold=True, size=11)
DATA_FONT      = Font(size=11)
NUM_FORMAT     = r"###,##0;\-###,##0"

# Data border: thin L/R/T only (no bottom), matching ICM style
DATA_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
)
# Header border: thin all sides
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 1 — UTILITY / SIGN FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def extract_icp_code(raw: str) -> str:
    m = re.match(r"(ICP_\w+)", str(raw or "").strip())
    return m.group(1) if m else ""


def extract_account_code(raw: str) -> str:
    """
    Handles:
      '[165000].[189501]:189501:...'  → '189501'
      '[FCCS_Group].[Plug_InvSh]:...' → 'Plug_InvSh'
      '224000:224000:...'             → '224000'
      '224000 - 224000:...'           → '224000'
    """
    raw = str(raw or "").strip()
    m = re.search(r"\]\.\[?(\w+)\]?:", raw)
    if m:
        val = m.group(1)
        if not val.isdigit():
            m2 = re.search(r"]:(\d{6}):", raw)
            if m2:
                return m2.group(1)
        return val
    m = re.match(r"(\d{6})", raw)
    return m.group(1) if m else ""


def extract_entity_code_icm(raw: str) -> str:
    m = re.match(r"(\d{6})", str(raw or "").strip())
    return m.group(1) if m else ""


def extract_entity_code_journal(raw: str) -> str:
    raw = str(raw or "").strip()
    m = re.match(r"(E?\d+\w*)", raw)
    return m.group(1) if m else ""


def to_float(val) -> float:
    if val is None or str(val).strip() in ("", " "):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def is_detail_row(vals: list) -> bool:
    entity = str(vals[J_ENTITY - 1] or "").strip()
    acct   = str(vals[J_ACCT - 1]   or "").strip()
    icp    = str(vals[J_ICP - 1]    or "").strip()
    return bool(entity or acct or icp)


def classify_account(code: str) -> str:
    """
    S1 = debit-positive  (series 1 Asset, series 5 Expense)
    S2 = credit-positive (series 2 Liability, 3 Equity, 4 Revenue)
    """
    code = str(code or "").strip()
    if not code or not code[0].isdigit():
        return "EXTRA"
    s = code[0]
    if s in ("1", "5"):
        return "S1"
    elif s in ("2", "3", "4"):
        return "S2"
    return "EXTRA"


def apply_sign(debit: float, credit: float, account_code: str) -> float:
    """
    Apply sign based on account series:
      Series 1 & 5 → debit - credit  (Debit = positive)
      Series 2,3,4 → credit - debit  (Credit = positive)
    """
    code = str(account_code or "").strip()
    s = code[0] if code and code[0].isdigit() else "0"
    if s in ("1", "5"):
        return debit - credit
    elif s in ("2", "3", "4"):
        return credit - debit
    return debit - credit


def get_account_description(raw_header: str) -> str:
    raw = str(raw_header or "").strip()
    raw = re.sub(r"\s+(Entity|Partner)\s*$", "", raw)
    return raw


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 2 — COLUMN MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════

def read_icm_headers(ws):
    """
    Read ICM header row and split accounts into 4 tagged sub-lists.
    Also captures the full original header text for each (code, tag) pair
    so it can be reproduced exactly in the output.

    Returns: known_accounts, s1_entity, s2_partner, s1_partner, s2_entity, full_headers
      full_headers : { (code, tag) → original cell string from ICM row 4 }
    """
    headers = [cell.value for cell in ws[ICM_HEADER_ROW]]
    known_accounts = OrderedDict()
    s1_entity  = []
    s2_partner = []
    s1_partner = []
    s2_entity  = []
    full_headers = {}   # (code, tag) → full original header string

    SKIP = {"Entity", "Partner", "Variance", "Total",
            "PARENT INPUT", "QAR_Reporting", "Final amount", "final variance"}
    seen = set()

    for h in headers:
        hs = str(h or "").strip()
        if not hs or hs in SKIP:
            continue
        code = extract_account_code(hs)
        if not code:
            continue

        # Detect Entity / Partner suffix
        if re.search(r"\bEntity\b", hs):
            tag = "Entity"
        elif re.search(r"\bPartner\b", hs):
            tag = "Partner"
        else:
            tag = "Entity"   # default

        pair = (code, tag)
        if pair in seen:
            continue
        seen.add(pair)

        # Store full original header text for exact reproduction
        full_headers[(code, tag)] = hs

        if code not in known_accounts:
            known_accounts[code] = get_account_description(hs)

        cls = classify_account(code)
        if cls == "S1":
            if tag == "Entity":  s1_entity.append(code)
            else:                s1_partner.append(code)
        elif cls == "S2":
            if tag == "Partner": s2_partner.append(code)
            else:                s2_entity.append(code)

    return known_accounts, s1_entity, s2_partner, s1_partner, s2_entity, full_headers


def build_output_columns(icm_accounts, all_journal_accounts,
                         s1_entity, s2_partner, s1_partner, s2_entity):
    """
    Build two column lists for the output.
    ICM-defined accounts follow the standard Entity/Partner side structure.
    Journal-only accounts (not in ICM) are appended as extra columns on the
    entity side with series 'EXTRA' — they appear in output but do NOT
    contribute to Variance 1 or Variance 2 calculations.

      entity_cols  = [(code, desc, series), ...]  → S1-Entity + S2-Partner + EXTRA
      partner_cols = [(code, desc, series), ...]  → S1-Partner + S2-Entity
    """
    def col(code):
        return (code, icm_accounts.get(code, code), classify_account(code))

    entity_cols  = [col(c) for c in s1_entity]  + [col(c) for c in s2_partner]
    partner_cols = [col(c) for c in s1_partner] + [col(c) for c in s2_entity]

    # Journal-only accounts: include as extra entity-side columns so their
    # values are visible in the output (tagged "EXTRA" → excluded from variances)
    icm_known = set(icm_accounts.keys())
    extra = sorted(all_journal_accounts - icm_known)
    if extra:
        print(f"  Journal-only accounts (added as extra cols): {extra}")
        for code in extra:
            entity_cols.append((code, code, "EXTRA"))

    return entity_cols, partner_cols


def discover_journals(folder=JOURNAL_FOLDER, pattern=JOURNAL_PATTERN):
    paths = sorted(glob.glob(os.path.join(folder, pattern)))
    print(f"  Discovered {len(paths)} journal file(s):")
    for p in paths:
        print(f"    {os.path.basename(p)}")
    return paths


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 3 — JOURNAL READING
# ═══════════════════════════════════════════════════════════════════════════

def read_journal_report(filepath: str):
    """
    Parse a single journal Excel file.
    Returns: lines, primary_lookup, fallback_lookup, all_accounts
      primary_lookup  : { (entity_code, icp_code, account_code) → [line dicts] }
      fallback_lookup : { (icp_code, account_code) → [line dicts] }
    """
    print(f"  Reading: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    lines = []
    all_accounts = set()

    for row_num, row in enumerate(
            ws.iter_rows(min_row=JOURNAL_DATA_START, max_row=ws.max_row),
            start=JOURNAL_DATA_START):
        vals = [cell.value for cell in row]
        label = str(vals[0] or "").strip()
        if label == "Grand Total":
            break
        if not is_detail_row(vals):
            continue

        entity_raw   = str(vals[J_ENTITY - 1] or "").strip()
        account_raw  = str(vals[J_ACCT - 1]   or "").strip()
        icp_raw      = str(vals[J_ICP - 1]    or "").strip()
        entity_code  = extract_entity_code_journal(entity_raw)
        account_code = extract_account_code(account_raw)
        icp_code     = extract_icp_code(icp_raw)
        debit        = to_float(vals[J_DEBIT - 1])
        credit       = to_float(vals[J_CREDIT - 1])

        if account_code:
            all_accounts.add(account_code)

        lines.append({
            "row_num":      row_num,
            "label":        label,
            "entity_raw":   entity_raw,
            "entity_code":  entity_code,
            "is_group":     bool(re.match(r"^E\d+", entity_code)),
            "account_code": account_code,
            "icp_code":     icp_code,
            "debit":        debit,
            "credit":       credit,
        })

    primary_lookup  = defaultdict(list)
    fallback_lookup = defaultdict(list)
    for line in lines:
        if not line["is_group"]:
            primary_lookup[(line["entity_code"], line["icp_code"], line["account_code"])].append(line)
        else:
            fallback_lookup[(line["icp_code"], line["account_code"])].append(line)

    print(f"    Lines: {len(lines)} | Accounts: {sorted(all_accounts)}")
    print(f"    Primary keys: {len(primary_lookup)} | Fallback keys: {len(fallback_lookup)}")
    return lines, primary_lookup, fallback_lookup, all_accounts


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 4 — MATCHING
# ═══════════════════════════════════════════════════════════════════════════

def match_journal_to_icm(data_rows, account_code_list, primary_lookup, fallback_lookup):
    """
    For each (Entity, Partner) in ICM and each account code:
      Step 1: Check (Entity + Partner + Account) — primary match
      Step 2: Apply sign based on account series
      Fallback: Group entities (E-prefix) matched on (Partner + Account) only

    Returns: updates dict, match_log list
    """
    updates   = {}
    match_log = []

    for icm_row in data_rows:
        ent_code     = icm_row["entity_code"]
        partner_code = icm_row["partner_code"]
        if not partner_code:
            continue

        for acct_code in account_code_list:
            key_primary = (ent_code, partner_code, acct_code)
            if key_primary in primary_lookup:
                jlines = primary_lookup[key_primary]
                net = sum(apply_sign(j["debit"], j["credit"], acct_code) for j in jlines)
                if net != 0:
                    updates[(ent_code, partner_code, acct_code)] = net
                    match_log.append(
                        f"PRIMARY  | E:{ent_code} P:{partner_code} A:{acct_code} "
                        f"({classify_account(acct_code)}) | Net:{net:,.2f}"
                    )
                continue

            key_fallback = (partner_code, acct_code)
            if key_fallback in fallback_lookup:
                jlines = fallback_lookup[key_fallback]
                net = sum(apply_sign(j["debit"], j["credit"], acct_code) for j in jlines)
                if net != 0:
                    updates[(ent_code, partner_code, acct_code)] = ("GROUP", net)
                    labels = list({j["label"] for j in jlines})
                    match_log.append(
                        f"FALLBACK | E:{ent_code} P:{partner_code} A:{acct_code} "
                        f"| Net:{net:,.2f} ({labels})"
                    )

    primary_count  = sum(1 for v in updates.values() if not isinstance(v, tuple))
    fallback_count = sum(1 for v in updates.values() if isinstance(v, tuple))
    print(f"    Matches: {primary_count} primary + {fallback_count} fallback")
    return updates, match_log


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 5 — ICM DATA READER
# ═══════════════════════════════════════════════════════════════════════════

def read_icm_data(ws):
    """Read all (Entity, Partner) rows from the ICM sheet."""
    data_rows = []
    for row_num, row in enumerate(
            ws.iter_rows(min_row=ICM_DATA_START, max_row=ws.max_row), start=ICM_DATA_START):
        vals = [cell.value for cell in row]
        entity_raw  = str(vals[0] or "").strip()
        partner_raw = str(vals[1] or "").strip() if len(vals) > 1 else ""
        if not entity_raw and not partner_raw:
            continue
        data_rows.append({
            "row_num":      row_num,
            "entity":       entity_raw,
            "partner":      partner_raw,
            "entity_code":  extract_entity_code_icm(entity_raw),
            "partner_code": extract_icp_code(partner_raw),
        })
    return data_rows


# ═══════════════════════════════════════════════════════════════════════════
# MODULE 6 — OUTPUT WRITER (ICM-style)
# ═══════════════════════════════════════════════════════════════════════════

def _style_header_cell(cell, text):
    """Apply ICM-style header formatting."""
    cell.value     = text
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.border    = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_id_cell(cell, text):
    """Apply ICM-style Entity/Partner ID column formatting."""
    cell.value     = text
    cell.font      = ID_FONT
    cell.fill      = ID_FILL
    cell.border    = DATA_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _style_data_cell(cell, value, fill=None):
    """Apply ICM-style account value cell formatting."""
    cell.value         = value
    cell.font          = DATA_FONT
    cell.fill          = fill if fill is not None else PatternFill()
    cell.border        = DATA_BORDER
    cell.alignment     = Alignment(vertical="top", wrap_text=True)
    if value is not None:
        cell.number_format = NUM_FORMAT


def _style_variance_cell(cell, value):
    """Apply ICM-style variance cell formatting (light grey)."""
    cell.value         = value
    cell.font          = DATA_FONT
    cell.fill          = VARIANCE_FILL
    cell.border        = DATA_BORDER
    cell.alignment     = Alignment(vertical="top", wrap_text=True)
    cell.number_format = NUM_FORMAT


def _style_total_cell(cell, value):
    """Apply ICM-style total cell formatting (medium grey)."""
    cell.value         = value
    cell.font          = DATA_FONT
    cell.fill          = TOTAL_FILL
    cell.border        = DATA_BORDER
    cell.alignment     = Alignment(vertical="top", wrap_text=True)
    cell.number_format = NUM_FORMAT


def write_output(data_rows, entity_cols, partner_cols,
                 all_updates_list, journal_names, full_headers, output_path):
    """
    Write output Excel — ONE sheet, all journals side by side, ICM-style formatting.

    Layout:
      Row 3 : Journal name labels (one per journal set)
      Row 4 : Headers (Entity/Partner once, then per-journal account headers)
      Rows 5+: Data rows

    Per journal column block (cols_per_set columns):
      [S1-Entity cols] [S2-Partner cols] [Variance 1]
      [S1-Partner cols] [S2-Entity cols] [Variance 2] [Total]

    Variance 1 = Σ(S1-Entity) − Σ(S2-Partner)
    Variance 2 = Σ(S1-Partner) − Σ(S2-Entity)
    Total      = Variance 1 + Variance 2
    """
    print(f"\n[Step 4+5] Writing ICM-style output: {output_path}")

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "ICM Matched"

    en = len(entity_cols)
    pn = len(partner_cols)
    # per journal: en entity-side + 1 Var1 + pn partner-side + 1 Var2 + 1 Total
    cols_per_set = en + 1 + pn + 1 + 1
    # +1 gap column between each journal set
    stride     = cols_per_set + 1
    total_cols = 2 + len(journal_names) * stride - 1   # last set has no trailing gap

    s1_entity_codes  = [c for c, _, s in entity_cols  if s == "S1"]
    s2_partner_codes = [c for c, _, s in entity_cols  if s == "S2"]
    s1_partner_codes = [c for c, _, s in partner_cols if s == "S1"]
    s2_entity_codes  = [c for c, _, s in partner_cols if s == "S2"]

    # ── Column widths (alternating across all columns) ──────────────────
    for i in range(total_cols):
        ltr = get_column_letter(i + 1)
        ws.column_dimensions[ltr].width = 17.5714 if i % 2 == 0 else 16.7143

    # ── Row heights ──────────────────────────────────────────────────────
    ws.row_dimensions[ICM_HEADER_ROW - 1].height = 20    # journal name row
    ws.row_dimensions[ICM_HEADER_ROW].height     = 78.75

    # ── Fixed headers: Entity, Partner (cols 1-2) ────────────────────────
    _style_header_cell(ws.cell(ICM_HEADER_ROW, 1), " Entity ")
    _style_header_cell(ws.cell(ICM_HEADER_ROW, 2), " Partner ")

    # ── Per-journal headers and data ─────────────────────────────────────
    for set_idx, (jname, updates) in enumerate(zip(journal_names, all_updates_list)):
        base  = 3 + set_idx * stride   # first data column for this journal
        label = JOURNAL_LABELS.get(jname, jname)

        # Row 3: journal name label — green background, white bold text
        name_cell = ws.cell(ICM_HEADER_ROW - 1, base, label)
        name_cell.font      = Font(bold=True, color="FFFFFF", size=11)
        name_cell.fill      = JOURNAL_FILL
        name_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row 4: account headers for this journal
        col = base
        for code, desc, series in entity_cols:
            if series == "S1":
                tag = "Entity"
            elif series == "S2":
                tag = "Partner"
            else:
                tag = ""   # EXTRA accounts — no Entity/Partner suffix
            text = full_headers.get((code, tag), f"{code} - {desc} {tag}".strip())
            _style_header_cell(ws.cell(ICM_HEADER_ROW, col), text)
            col += 1

        _style_header_cell(ws.cell(ICM_HEADER_ROW, col), "Variance "); col += 1

        for code, desc, series in partner_cols:
            tag  = "Partner" if series == "S1" else "Entity"
            text = full_headers.get((code, tag), f"{code} - {desc} {tag}")
            _style_header_cell(ws.cell(ICM_HEADER_ROW, col), text)
            col += 1

        _style_header_cell(ws.cell(ICM_HEADER_ROW, col), "Variance "); col += 1
        _style_header_cell(ws.cell(ICM_HEADER_ROW, col), "Total ");    col += 1

        # ── Data rows ────────────────────────────────────────────────────
        for icm_row in data_rows:
            r            = icm_row["row_num"]
            ent_code     = icm_row["entity_code"]
            partner_code = icm_row["partner_code"]

            # Entity / Partner ID cells (written once — first journal sets them)
            if set_idx == 0:
                _style_id_cell(ws.cell(r, 1), icm_row["entity"])
                _style_id_cell(ws.cell(r, 2), icm_row["partner"])

            col = base
            ev = {}
            pv = {}

            for code, desc, series in entity_cols:
                val = updates.get((ent_code, partner_code, code))
                if val is not None:
                    v    = val[1] if isinstance(val, tuple) else val
                    fill = GROUP_FILL if isinstance(val, tuple) else MATCH_FILL
                    _style_data_cell(ws.cell(r, col), v, fill)
                    ev[code] = v
                else:
                    _style_data_cell(ws.cell(r, col), None)
                col += 1

            var1 = (sum(to_float(ev.get(c, 0)) for c in s1_entity_codes) -
                    sum(to_float(ev.get(c, 0)) for c in s2_partner_codes))
            _style_variance_cell(ws.cell(r, col), var1); col += 1

            for code, desc, series in partner_cols:
                val = updates.get((ent_code, partner_code, code))
                if val is not None:
                    v    = val[1] if isinstance(val, tuple) else val
                    fill = GROUP_FILL if isinstance(val, tuple) else MATCH_FILL
                    _style_data_cell(ws.cell(r, col), v, fill)
                    pv[code] = v
                else:
                    _style_data_cell(ws.cell(r, col), None)
                col += 1

            var2 = (sum(to_float(pv.get(c, 0)) for c in s1_partner_codes) -
                    sum(to_float(pv.get(c, 0)) for c in s2_entity_codes))
            _style_variance_cell(ws.cell(r, col), var2); col += 1

            _style_total_cell(ws.cell(r, col), var1 + var2); col += 1

        print(f"  Journal '{jname}': {cols_per_set} cols "
              f"({en} entity + Var1 + {pn} partner + Var2 + Total)")

    out_wb.save(output_path)
    print(f"  Saved: 1 sheet, {len(journal_names)} journal(s), "
          f"{total_cols} total cols -> {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("  IC Matching — ICM-Style Output (v3)")
    print("=" * 65)

    # ── Step 1: ICM structure ────────────────────────────────────────────
    print("\n[Step 1] Reading ICM report...")
    wb_icm = openpyxl.load_workbook(ICM_FILE, data_only=True)
    ws_icm = wb_icm.active

    icm_accounts, s1_entity, s2_partner, s1_partner, s2_entity, full_headers = \
        read_icm_headers(ws_icm)
    data_rows = read_icm_data(ws_icm)

    print(f"  Accounts (unique): {list(icm_accounts.keys())}")
    print(f"  S1-Entity  : {s1_entity}")
    print(f"  S2-Partner : {s2_partner}")
    print(f"  S1-Partner : {s1_partner}")
    print(f"  S2-Entity  : {s2_entity}")
    print(f"  ICM data rows: {len(data_rows)}")
    print(f"  Full headers captured: {len(full_headers)}")

    # ── Step 2: Read journals ────────────────────────────────────────────
    print("\n[Step 2] Reading journal files...")
    all_updates    = []
    all_match_logs = []
    all_j_accounts = set()
    journal_names  = []

    journal_files = discover_journals()
    if not journal_files:
        print("  ERROR: No journal files found. Check JOURNAL_FOLDER/JOURNAL_PATTERN.")
        return

    for jfile in journal_files:
        jname = os.path.basename(jfile).replace(".xlsx", "")
        journal_names.append(jname)
        lines, primary, fallback, j_accts = read_journal_report(jfile)
        all_j_accounts.update(j_accts)

        all_codes = list(icm_accounts.keys()) + list(j_accts - set(icm_accounts.keys()))
        updates, mlog = match_journal_to_icm(data_rows, all_codes, primary, fallback)
        all_updates.append(updates)
        all_match_logs.append(mlog)

    # ── Step 3: Build output column layout ──────────────────────────────
    print("\n[Step 3] Building output column layout...")
    entity_cols, partner_cols = build_output_columns(
        icm_accounts, all_j_accounts,
        s1_entity, s2_partner, s1_partner, s2_entity
    )
    print(f"  Entity-side  cols ({len(entity_cols)}): {[c for c,_,_ in entity_cols]}")
    print(f"  Partner-side cols ({len(partner_cols)}): {[c for c,_,_ in partner_cols]}")

    # ── Steps 4+5: Write ICM-style output ───────────────────────────────
    write_output(data_rows, entity_cols, partner_cols,
                 all_updates, journal_names, full_headers, OUTPUT_FILE)

    # ── Match summary ────────────────────────────────────────────────────
    print("\n-- Match Summary " + "-" * 47)
    for jname, mlog in zip(journal_names, all_match_logs):
        print(f"\n  [{jname}]  {len(mlog)} matches")
        for msg in mlog[:3]:
            print(f"    {msg}")
        if len(mlog) > 3:
            print(f"    ... and {len(mlog) - 3} more")

    print("\n" + "=" * 65)
    print(f"  Done!  ->  {OUTPUT_FILE}")
    print("  Layout : 1 sheet, all journals side by side, ICM-style format")
    print("  Colours: Navy headers | Light-blue ID cols | Grey variance | Grey total")
    print("  Matches: Yellow = primary | Pink = group entity fallback")
    print("=" * 65)


if __name__ == "__main__":
    main()
